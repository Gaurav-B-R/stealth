import os
import re
import hmac
import json
import uuid
import secrets
import logging
import hashlib
import threading
from typing import List, Optional
from urllib.parse import quote, urlparse

import requests
from jose import jwt as jose_jwt, JWTError
from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, Request, Response, UploadFile, status
from fastapi.concurrency import run_in_threadpool
from sqlalchemy import and_, func, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from datetime import timedelta, datetime, date, timezone as dt_timezone
from pydantic import BaseModel, EmailStr, Field

from app.database import get_db, SessionLocal
from app import models
from app import money
from app import brand_marks
from app import enterprise_catalog as catalog
from app import enterprise_access as access
from app import enterprise_team as team_svc
from app import enterprise_time as ent_time
from app import enterprise_dates as ent_dates
from app import enterprise_client_fields as client_fields
from app import enterprise_duplicates as dupes
from app import enterprise_billing as billing
from app import enterprise_credits as credits
from app import enterprise_coupons
from app import enterprise_payments
from app import enterprise_finance as finance
from app import enterprise_ai
from app import enterprise_writing
from app import enterprise_copilot
from app import enterprise_interview
from app import ai_guardrails
from app import ai_usage
from app import enterprise_storage
from app import enterprise_calendar_files as cal_files
from app import enterprise_notifications as notif
from app.utils import gemini_service
from app.routers.ai_chat import ChatSessionAttachment
from app.auth import (
    authenticate_user,
    create_access_token,
    get_current_active_user,
    get_password_hash,
    validate_password_strength,
    ACCESS_TOKEN_EXPIRE_MINUTES,
    SECRET_KEY as AUTH_SECRET_KEY,
    ALGORITHM as AUTH_ALGORITHM,
    set_auth_cookie,
)
from app.utils.rate_limiter import (
    check_ip_rate_limit,
    extract_client_ip,
    is_request_ip_whitelisted,
)
from app.utils.turnstile import is_turnstile_enabled, verify_turnstile_token
from app.utils.html_sanitizer import sanitize_email_html, html_to_text
from app.email_service import send_enterprise_team_invite_email, send_enterprise_client_email
from app.email_service import send_enterprise_inbound_reply_alert_email
from app import enterprise_inbound_email as inbound_email
from app.email_service import send_enterprise_interview_invite_email, send_enterprise_interview_code_email
from app.email_service import send_enterprise_interview_report_email
from app.email_service import send_enterprise_document_request_email, send_enterprise_document_request_code_email
from app.email_service import send_enterprise_portal_share_email, send_enterprise_portal_code_email
from app.email_service import send_enterprise_copilot_invite_email, send_enterprise_copilot_code_email
from app.email_service import send_enterprise_lead_form_link_email, send_enterprise_new_lead_email
from app.email_service import send_enterprise_payment_request_email
from app.email_service import send_enterprise_payment_dispute_alert_email
from app.email_service import generate_verification_token, DEFAULT_PUBLIC_BASE_URL
from app.email_service import send_enterprise_support_request_email, send_enterprise_demo_request_email
from app.email_service import send_feature_request_confirmation
from app.email_service import send_enterprise_welcome_email
from app.email_service import send_email_otp
from app.email_service import (
    send_enterprise_owner_transfer_code_email,
    send_enterprise_owner_transfer_notice_email,
)
from app.utils.token_security import hash_token, token_matches
from app import enterprise_step_up as step_up

RAZORPAY_API_BASE = os.getenv("RAZORPAY_API_BASE", "https://api.razorpay.com/v1").rstrip("/")

router = APIRouter(prefix="/api/enterprise", tags=["enterprise"])
logger = logging.getLogger(__name__)

ENTERPRISE_ROLE_ADMIN = "admin"
ENTERPRISE_ROLE_EDITOR = "editor"
ENTERPRISE_ROLE_VIEWER = "viewer"
ENTERPRISE_ALLOWED_ROLES = {ENTERPRISE_ROLE_ADMIN, ENTERPRISE_ROLE_EDITOR, ENTERPRISE_ROLE_VIEWER}

ENTERPRISE_SUBDOMAIN_MIN_LENGTH = 3
ENTERPRISE_SUBDOMAIN_MAX_LENGTH = 32
ENTERPRISE_SUBDOMAIN_PATTERN = re.compile(r"^[a-z0-9](?:[a-z0-9-]{1,30}[a-z0-9])$")
ENTERPRISE_RESERVED_SUBDOMAINS = {
    "www",
    "app",
    "api",
    "admin",
    "auth",
    "portal",
    "enterprise",
    "mail",
    "status",
    "support",
    "docs",
    "blog",
    "cdn",
    "m",
    "ftp",
    "smtp",
    "imap",
    "pop",
    "rilono",
}
ENTERPRISE_ROOT_DOMAIN = (os.getenv("ENTERPRISE_ROOT_DOMAIN", "rilono.com").strip().lower() or "rilono.com").lstrip(".")
ENTERPRISE_PORTAL_SCHEME = os.getenv("ENTERPRISE_PORTAL_SCHEME", "https").strip().lower() or "https"
ENTERPRISE_PORTAL_PORT = os.getenv("ENTERPRISE_PORTAL_PORT", "").strip().lstrip(":")
ENTERPRISE_GENERIC_SUBDOMAINS = {
    item.strip().lower()
    for item in os.getenv("ENTERPRISE_GENERIC_SUBDOMAINS", "enterprise,portal").split(",")
    if item.strip()
}
ENTERPRISE_PASSWORD_SETUP_BASE_URL = (
    os.getenv("BASE_URL", DEFAULT_PUBLIC_BASE_URL).strip() or DEFAULT_PUBLIC_BASE_URL
).rstrip("/")
ENTERPRISE_INVITE_PASSWORD_SETUP_EXPIRES_HOURS = max(
    1,
    int(os.getenv("ENTERPRISE_INVITE_PASSWORD_SETUP_EXPIRES_HOURS", "72")),
)
ENTERPRISE_BRAND_MARK_PATH = "/api/enterprise/public/brand-mark/"
# Random stock photos we used to hand out as default logos. Kept only so the orgs still
# carrying one in the database get a real mark instead — never re-issued.
_LEGACY_PLACEHOLDER_LOGO_HOSTS = {"picsum.photos", "www.picsum.photos"}
ENTERPRISE_LOGIN_RATE_LIMIT = int(os.getenv("ENTERPRISE_LOGIN_RATE_LIMIT", os.getenv("LOGIN_RATE_LIMIT", "12")))
ENTERPRISE_LOGIN_RATE_WINDOW_SECONDS = int(
    os.getenv("ENTERPRISE_LOGIN_RATE_WINDOW_SECONDS", os.getenv("LOGIN_RATE_WINDOW_SECONDS", "300"))
)
ENTERPRISE_TEAM_INVITE_RATE_LIMIT = int(os.getenv("ENTERPRISE_TEAM_INVITE_RATE_LIMIT", "20"))
ENTERPRISE_TEAM_INVITE_RATE_WINDOW_SECONDS = int(os.getenv("ENTERPRISE_TEAM_INVITE_RATE_WINDOW_SECONDS", "3600"))
ENTERPRISE_CLIENT_EMAIL_RATE_LIMIT = int(os.getenv("ENTERPRISE_CLIENT_EMAIL_RATE_LIMIT", "60"))
ENTERPRISE_CLIENT_EMAIL_RATE_WINDOW_SECONDS = int(os.getenv("ENTERPRISE_CLIENT_EMAIL_RATE_WINDOW_SECONDS", "3600"))
ENTERPRISE_BULK_EMAIL_RATE_LIMIT = int(os.getenv("ENTERPRISE_BULK_EMAIL_RATE_LIMIT", "10"))
ENTERPRISE_BULK_EMAIL_RATE_WINDOW_SECONDS = int(os.getenv("ENTERPRISE_BULK_EMAIL_RATE_WINDOW_SECONDS", "3600"))
# Composer attachment uploads: generous for real use (10 files per email), bounded
# enough that an abused account can't grow storage without limit.
ENTERPRISE_EMAIL_ATTACH_RATE_LIMIT = int(os.getenv("ENTERPRISE_EMAIL_ATTACH_RATE_LIMIT", "120"))
ENTERPRISE_EMAIL_ATTACH_RATE_WINDOW_SECONDS = int(
    os.getenv("ENTERPRISE_EMAIL_ATTACH_RATE_WINDOW_SECONDS", "3600")
)
# Shown when an authenticated account belongs to no workspace — a personal B2C student account,
# or someone who was never invited. Every raise site sits BEHIND a verified password (see
# enterprise_login), so naming the B2C/B2B split here discloses nothing a caller doesn't already
# own. Enterprise is self-serve (POST /signup), so the advice is invite-or-create, never sales.
ENTERPRISE_NO_WORKSPACE_DETAIL = (
    "This email isn't part of any Rilono workspace. A personal Rilono student account is separate "
    "and can't sign in here. Ask your workspace admin to invite this email, or create a free "
    "workspace with a different email."
)
# The offboarded teammate's version of that 403: they had access and it was switched off, so the
# way back is their own admin, not signup. Kept verbatim in sync with the string
# _require_enterprise_membership already raises for this state.
ENTERPRISE_ACCESS_REVOKED_DETAIL = (
    "Your access to this workspace has been turned off. Ask a workspace admin to restore it."
)
# Shown when workspace signup is attempted with an email that already has a Rilono account of
# EITHER kind. The old copy said "please sign in instead", which is wrong (and a dead end) for the
# common case: a B2C student account can't sign in to the portal, so the real fix is a different
# email. Unlike the two above, this one IS pre-authentication and unavoidably confirms the email
# is taken — so it must stay vague about WHICH product owns it.
ENTERPRISE_SIGNUP_EMAIL_TAKEN_DETAIL = (
    "This email already has a Rilono account, so it can't create a new workspace. If it's your "
    "personal Rilono student account, sign up with a different work email — student accounts and "
    "workspaces are always separate. If your workspace already exists, sign in instead."
)
ENTERPRISE_LOGO_URL_MAX_LENGTH = 2048
ENTERPRISE_STUDENT_NAME_MAX_LENGTH = 160
ENTERPRISE_STUDY_COUNTRY_CODE_MAX_LENGTH = 12
ENTERPRISE_STUDY_COUNTRY_NAME_MAX_LENGTH = 120
ENTERPRISE_VISA_TYPE_MAX_LENGTH = 120
ENTERPRISE_INTAKE_MAX_LENGTH = 120
ENTERPRISE_STUDY_DESTINATION_OPTIONS = [
    {
        "code": "US",
        "name": "United States",
        "flag_emoji": "🇺🇸",
        "iconic_place": "Statue of Liberty",
        "visa_types": [
            "F-1 Student Visa",
            "J-1 Exchange Visitor Visa",
            "M-1 Vocational Student Visa",
        ],
        "intakes_by_visa": {
            "F-1 Student Visa": ["Spring", "Summer", "Fall"],
            "J-1 Exchange Visitor Visa": ["Spring", "Fall"],
            "M-1 Vocational Student Visa": ["Spring", "Summer", "Fall"],
        },
    },
    {
        "code": "CA",
        "name": "Canada",
        "flag_emoji": "🇨🇦",
        "iconic_place": "Niagara Falls",
        "visa_types": [
            "Study Permit",
        ],
        "intakes_by_visa": {
            "Study Permit": ["January", "May", "September"],
        },
    },
    {
        "code": "UK",
        "name": "United Kingdom",
        "flag_emoji": "🇬🇧",
        "iconic_place": "Big Ben",
        "visa_types": [
            "UK Student Visa",
            "Child Student Visa",
        ],
        "intakes_by_visa": {
            "UK Student Visa": ["January", "September"],
            "Child Student Visa": ["January", "September"],
        },
    },
    {
        "code": "AU",
        "name": "Australia",
        "flag_emoji": "🇦🇺",
        "iconic_place": "Sydney Opera House",
        "visa_types": [
            "Subclass 500 Student Visa",
            "Subclass 590 Student Guardian Visa",
        ],
        "intakes_by_visa": {
            "Subclass 500 Student Visa": ["February", "July", "November"],
            "Subclass 590 Student Guardian Visa": ["February", "July", "November"],
        },
    },
    {
        "code": "DE",
        "name": "Germany",
        "flag_emoji": "🇩🇪",
        "iconic_place": "Brandenburg Gate",
        "visa_types": [
            "National Visa (Type D) - Study",
            "Student Applicant Visa",
        ],
        "intakes_by_visa": {
            "National Visa (Type D) - Study": ["Summer Semester", "Winter Semester"],
            "Student Applicant Visa": ["Summer Semester", "Winter Semester"],
        },
    },
    {
        "code": "FR",
        "name": "France",
        "flag_emoji": "🇫🇷",
        "iconic_place": "Eiffel Tower",
        "visa_types": [
            "Long-Stay Student Visa (VLS-TS « Étudiant »)",
            "Temporary Long-Stay Student Visa (VLS-T)",
            "Entrance-Exam Visa (Court Séjour « Étudiant-Concours »)",
            "Short-Stay Study Visa (Schengen Type C)",
        ],
        "intakes_by_visa": {
            "Long-Stay Student Visa (VLS-TS « Étudiant »)": ["January", "September"],
            "Temporary Long-Stay Student Visa (VLS-T)": ["January", "September"],
            "Entrance-Exam Visa (Court Séjour « Étudiant-Concours »)": ["January", "September"],
            "Short-Stay Study Visa (Schengen Type C)": ["January", "September"],
        },
    },
    {
        "code": "IE",
        "name": "Ireland",
        "flag_emoji": "🇮🇪",
        "iconic_place": "Cliffs of Moher",
        "visa_types": [
            "D Study Visa",
        ],
        "intakes_by_visa": {
            "D Study Visa": ["January", "September"],
        },
    },
    {
        "code": "NZ",
        "name": "New Zealand",
        "flag_emoji": "🇳🇿",
        "iconic_place": "Milford Sound",
        "visa_types": [
            "Fee Paying Student Visa",
        ],
        "intakes_by_visa": {
            "Fee Paying Student Visa": ["February", "July"],
        },
    },
    {
        "code": "ES",
        "name": "Spain",
        "flag_emoji": "🇪🇸",
        "iconic_place": "Sagrada Família",
        "visa_types": [
            "Long-Stay Study Visa (Type D) – Higher Education",
            "Long-Stay Study Visa (Type D) – Language / Training Activity",
            "Long-Stay Study Visa (Type D) – Secondary / Student Mobility",
            "Short-Stay Study Visa (Schengen Type C)",
        ],
        "intakes_by_visa": {
            "Long-Stay Study Visa (Type D) – Higher Education": ["February", "September", "October"],
            "Long-Stay Study Visa (Type D) – Language / Training Activity": ["January", "February", "September", "October"],
            "Long-Stay Study Visa (Type D) – Secondary / Student Mobility": ["January", "September"],
            "Short-Stay Study Visa (Schengen Type C)": ["January", "September"],
        },
    },
    {
        "code": "NL",
        "name": "Netherlands",
        "flag_emoji": "🇳🇱",
        "iconic_place": "Kinderdijk Windmills",
        "visa_types": [
            "Study Residence Permit (MVV/TEV) – Higher Education",
            "Study Residence Permit – Secondary / MBO",
            "Exchange Student Residence Permit",
            "Orientation Year (Zoekjaar) Permit",
        ],
        "intakes_by_visa": {
            "Study Residence Permit (MVV/TEV) – Higher Education": ["September", "February"],
            "Study Residence Permit – Secondary / MBO": ["September", "February"],
            "Exchange Student Residence Permit": ["September", "February"],
            "Orientation Year (Zoekjaar) Permit": [],
        },
    },
    {
        "code": "AE",
        "name": "United Arab Emirates",
        "flag_emoji": "🇦🇪",
        "iconic_place": "Burj Khalifa",
        "visa_types": [
            "Student Residence Visa",
            "Study / Training Visit Visa",
            "Golden Residence – Outstanding Student",
            "Parent-Sponsored Student Residence",
        ],
        "intakes_by_visa": {
            "Student Residence Visa": ["January", "May", "September"],
            "Study / Training Visit Visa": ["January", "May", "September"],
            "Golden Residence – Outstanding Student": ["January", "May", "September"],
            "Parent-Sponsored Student Residence": ["January", "May", "September"],
        },
    },
    {
        "code": "PL",
        "name": "Poland",
        "flag_emoji": "🇵🇱",
        "iconic_place": "Wawel Royal Castle",
        "visa_types": [
            "National Visa (Type D) – Studies",
            "Temporary Residence Permit for Studies (Karta Pobytu)",
            "Schengen Visa (Type C) – Short Study",
            "Temporary Residence Permit – Graduate Job Search",
        ],
        "intakes_by_visa": {
            "National Visa (Type D) – Studies": ["October", "February"],
            "Temporary Residence Permit for Studies (Karta Pobytu)": ["October", "February"],
            "Schengen Visa (Type C) – Short Study": [],
            "Temporary Residence Permit – Graduate Job Search": [],
        },
    },
    {
        "code": "SG",
        "name": "Singapore",
        "flag_emoji": "🇸🇬",
        "iconic_place": "Marina Bay Sands",
        "visa_types": [
            "Student's Pass (IHL Track)",
            "Student's Pass (PEI / EduTrust Track)",
            "Short-Term Visit Pass – Short Course",
        ],
        "intakes_by_visa": {
            "Student's Pass (IHL Track)": ["January", "August"],
            "Student's Pass (PEI / EduTrust Track)": ["January", "April", "July", "October"],
            "Short-Term Visit Pass – Short Course": [],
        },
    },
    {
        "code": "IT",
        "name": "Italy",
        "flag_emoji": "🇮🇹",
        "iconic_place": "Colosseum",
        "visa_types": [
            "National Study Visa (Type D) – Immatricolazione Università",
            "National Study Visa (Type D) – PhD / AFAM / Non-Degree",
            "Short-Stay Study Visa (Schengen Type C)",
        ],
        "intakes_by_visa": {
            "National Study Visa (Type D) – Immatricolazione Università": ["February", "September", "October"],
            "National Study Visa (Type D) – PhD / AFAM / Non-Degree": ["February", "September", "October", "November"],
            "Short-Stay Study Visa (Schengen Type C)": ["January", "June", "September"],
        },
    },
    {
        "code": "SE",
        "name": "Sweden",
        "flag_emoji": "🇸🇪",
        "iconic_place": "Stockholm City Hall",
        "visa_types": [
            "Residence Permit for Studies at Higher Education",
            "Residence Permit for Doctoral Studies",
            "Post-Study Job-Seeking Residence Permit",
            "Short-Stay Study Visa (Schengen Type C)",
        ],
        "intakes_by_visa": {
            "Residence Permit for Studies at Higher Education": ["January", "August"],
            "Residence Permit for Doctoral Studies": [],
            "Post-Study Job-Seeking Residence Permit": [],
            "Short-Stay Study Visa (Schengen Type C)": ["January", "August"],
        },
    },
]
ENTERPRISE_STUDY_DESTINATION_MAP = {
    str(item["code"]).strip().upper(): item for item in ENTERPRISE_STUDY_DESTINATION_OPTIONS
}
ENTERPRISE_INTAKE_START_MONTH_HINTS = {
    "spring": 1,
    "summer": 5,
    "fall": 8,
    "winter": 11,
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
    "summer semester": 4,
    "winter semester": 10,
}


class EnterpriseLoginRequest(BaseModel):
    email: EmailStr
    password: str
    cf_turnstile_token: Optional[str] = None


class EnterpriseOnboardingRequest(BaseModel):
    company_name: str = Field(..., min_length=2, max_length=120)
    subdomain_slug: str = Field(
        ...,
        min_length=ENTERPRISE_SUBDOMAIN_MIN_LENGTH,
        max_length=ENTERPRISE_SUBDOMAIN_MAX_LENGTH,
    )


class EnterpriseTeamAddUserRequest(BaseModel):
    email: EmailStr
    # `role` stays for older clients (admin|editor|viewer). `role_key` is the richer field:
    # a system preset key, or "custom:<id>" for one of the workspace's own roles.
    role: str = Field(default=ENTERPRISE_ROLE_VIEWER, min_length=2, max_length=40)
    role_key: Optional[str] = Field(default=None, max_length=40)
    custom_role_id: Optional[int] = None
    data_scope: Optional[str] = Field(default=None, max_length=12)
    branch_ids: Optional[List[int]] = None
    primary_branch_id: Optional[int] = None
    job_title: Optional[str] = Field(default=None, max_length=120)
    phone: Optional[str] = Field(default=None, max_length=40)
    full_name: Optional[str] = Field(default=None, max_length=120)


class EnterpriseTeamRoleUpdateRequest(BaseModel):
    role: str = Field(..., min_length=2, max_length=40)


class EnterpriseBrandingUpdateRequest(BaseModel):
    company_name: Optional[str] = Field(default=None, min_length=2, max_length=120)
    logo_url: Optional[str] = Field(default=None, max_length=ENTERPRISE_LOGO_URL_MAX_LENGTH)
    generate_random_logo: bool = False
    # Company location (org records). country_code also drives the portal display currency.
    country_code: Optional[str] = Field(default=None, max_length=8)
    state_region: Optional[str] = Field(default=None, max_length=80)


class EnterpriseStudentCreateRequest(BaseModel):
    student_name: str = Field(..., min_length=2, max_length=ENTERPRISE_STUDENT_NAME_MAX_LENGTH)
    study_country_code: str = Field(..., min_length=2, max_length=ENTERPRISE_STUDY_COUNTRY_CODE_MAX_LENGTH)
    visa_type: str = Field(..., min_length=2, max_length=ENTERPRISE_VISA_TYPE_MAX_LENGTH)
    intake: str = Field(..., min_length=2, max_length=ENTERPRISE_INTAKE_MAX_LENGTH)


def _is_development_env() -> bool:
    return os.getenv("ENVIRONMENT", "production").strip().lower() == "development"


def _is_turnstile_required() -> bool:
    return is_turnstile_enabled() and not _is_development_env()


def _enforce_rate_limit_or_429(
    request: Request,
    scope: str,
    limit: int,
    window_seconds: int,
    extra_key: str | None = None,
) -> None:
    allowed, retry_after = check_ip_rate_limit(
        request=request,
        scope=scope,
        limit=limit,
        window_seconds=window_seconds,
        extra_key=extra_key,
    )
    if not allowed:
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
            headers={"Retry-After": str(retry_after)},
        )


class EnterpriseDemoRequestBody(BaseModel):
    full_name: str = Field(..., min_length=2, max_length=120)
    work_email: EmailStr
    company: Optional[str] = Field(default=None, max_length=160)
    phone: Optional[str] = Field(default=None, max_length=40)
    team_size: Optional[str] = Field(default=None, max_length=40)
    students_count: Optional[str] = Field(default=None, max_length=40)
    message: Optional[str] = Field(default=None, max_length=2000)
    source: Optional[str] = Field(default=None, max_length=200)


@router.post("/demo")
def enterprise_public_demo_request(
    payload: EnterpriseDemoRequestBody,
    request: Request,
    db: Session = Depends(get_db),
):
    """PUBLIC (no auth) 'book a demo' lead from the enterprise landing page. Stores the
    lead (so none is lost) and best-effort emails the sales inbox."""
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.demo_request", limit=8, window_seconds=3600,
    )
    name = (payload.full_name or "").strip()
    if len(name) < 2:
        raise HTTPException(status_code=400, detail="Please enter your name.")

    lead = models.EnterpriseDemoRequest(
        full_name=name[:120],
        work_email=str(payload.work_email).strip()[:200],
        company=(payload.company or "").strip()[:160] or None,
        phone=(payload.phone or "").strip()[:40] or None,
        team_size=(payload.team_size or "").strip()[:40] or None,
        students_count=(payload.students_count or "").strip()[:40] or None,
        message=(payload.message or "").strip()[:2000] or None,
        source=(payload.source or "").strip()[:200] or None,
        ip_address=(extract_client_ip(request) if request else None),
        status="new",
    )
    db.add(lead)
    db.commit()

    try:
        send_enterprise_demo_request_email(
            full_name=lead.full_name,
            work_email=lead.work_email,
            company=lead.company or "",
            phone=lead.phone or "",
            team_size=lead.team_size or "",
            students_count=lead.students_count or "",
            message=lead.message or "",
        )
    except Exception:
        logger.exception("Demo-request notification email failed (lead=%s)", lead.id)

    return {"message": "Thanks! Our team will email you shortly to schedule your demo."}


def _brand_mark_url(spec: str) -> str:
    # Same-origin, like uploaded logos: one value that is correct on the apex, on every
    # portal subdomain and on a developer's localhost, with no environment baked in.
    return f"{ENTERPRISE_BRAND_MARK_PATH}{spec}.png"


def _extract_brand_mark_spec(raw_logo_url: str | None) -> str | None:
    """The design slug out of one of our own generated-mark URLs, absolute or relative.
    None for anything else — an uploaded logo, an external URL, or a spec we no longer
    render (a retired emblem must not 404 an org's logo)."""
    path = urlparse(str(raw_logo_url or "").strip()).path
    if not path.startswith(ENTERPRISE_BRAND_MARK_PATH) or not path.endswith(".png"):
        return None
    spec = path[len(ENTERPRISE_BRAND_MARK_PATH):-len(".png")]
    return spec if brand_marks.parse_spec(spec) else None


def _build_default_enterprise_logo_url(
    *,
    organization_id: int | None,
    company_name: str | None,
    subdomain_slug: str | None,
    randomize: bool = False,
    current_logo_url: str | None = None,
) -> str:
    """A drawn study-abroad emblem — mortarboard, globe, campus — over a brand gradient.
    The whole design is the filename, so nothing is stored and the same organization
    keeps the same logo across requests and deploys."""
    if randomize:
        spec = brand_marks.random_spec(exclude=_extract_brand_mark_spec(current_logo_url))
    else:
        seed_source = "|".join([
            f"org-{int(organization_id)}" if organization_id is not None else "org-pending",
            (company_name or "").strip().lower(),
            (subdomain_slug or "").strip().lower(),
        ])
        spec = brand_marks.spec_for_seed(seed_source)
    return _brand_mark_url(spec)


def _normalize_enterprise_logo_url_or_400(raw_logo_url: str | None) -> str | None:
    value = str(raw_logo_url or "").strip()
    if not value:
        return None
    if len(value) > ENTERPRISE_LOGO_URL_MAX_LENGTH:
        raise HTTPException(
            status_code=400,
            detail=f"Logo URL must be at most {ENTERPRISE_LOGO_URL_MAX_LENGTH} characters.",
        )

    parsed = urlparse(value)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise HTTPException(
            status_code=400,
            detail="Logo URL must start with http:// or https://",
        )
    return value


def _resolve_enterprise_logo_url(organization: models.EnterpriseOrganization) -> str:
    raw_logo = str(getattr(organization, "logo_url", "") or "").strip()
    if raw_logo:
        # Uploaded logos are stored privately and served via our own public streaming
        # route — a same-origin relative URL that works on every portal subdomain.
        if raw_logo.startswith("/api/enterprise/public/org-logo/"):
            return raw_logo
        parsed = urlparse(raw_logo)
        if parsed.path.startswith(ENTERPRISE_BRAND_MARK_PATH):
            # A generated mark is re-issued against this deployment's base URL: the design
            # is the only part of the stored value worth keeping, and the origin baked into
            # it may predate a BASE_URL change or come from another environment. A spec we
            # no longer draw falls through to a fresh one rather than serving a dead 404.
            mark_spec = _extract_brand_mark_spec(raw_logo)
            if mark_spec:
                return _brand_mark_url(mark_spec)
        elif parsed.scheme in {"http", "https"} and parsed.netloc:
            # Orgs created before we drew our own marks still hold a random stock photo
            # here. Fall through and give them an emblem instead of a cat.
            if parsed.netloc.split(":")[0].lower() not in _LEGACY_PLACEHOLDER_LOGO_HOSTS:
                return raw_logo

    return _build_default_enterprise_logo_url(
        organization_id=getattr(organization, "id", None),
        company_name=getattr(organization, "company_name", None),
        subdomain_slug=getattr(organization, "subdomain_slug", None),
        randomize=False,
    )


def _absolute_enterprise_logo_url(organization: models.EnterpriseOrganization) -> str:
    """The logo for consumers that render outside our origin — email clients and the
    Chrome extension, where a same-origin path has nothing to resolve against. Every
    email template drops the <img> entirely unless the src starts with http(s)."""
    logo_url = _resolve_enterprise_logo_url(organization)
    if logo_url.startswith(("http://", "https://")):
        return logo_url
    return f"{ENTERPRISE_PASSWORD_SETUP_BASE_URL}{logo_url}"


def _normalize_enterprise_role(raw_role: str | None) -> str:
    normalized = str(raw_role or "").strip().lower()
    if normalized in {"owner", "org_admin", "organization_admin"}:
        return ENTERPRISE_ROLE_ADMIN
    if normalized in {"edit", "write"}:
        return ENTERPRISE_ROLE_EDITOR
    if normalized in {"read", "view"}:
        return ENTERPRISE_ROLE_VIEWER
    if normalized in ENTERPRISE_ALLOWED_ROLES:
        return normalized
    return ENTERPRISE_ROLE_VIEWER


def _parse_enterprise_role_or_400(raw_role: str | None) -> str:
    normalized = str(raw_role or "").strip().lower()
    if normalized in {"owner", "org_admin", "organization_admin"}:
        return ENTERPRISE_ROLE_ADMIN
    if normalized in {"edit", "write"}:
        return ENTERPRISE_ROLE_EDITOR
    if normalized in {"read", "view"}:
        return ENTERPRISE_ROLE_VIEWER
    if normalized in ENTERPRISE_ALLOWED_ROLES:
        return normalized
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Invalid role. Use admin, editor, or viewer.",
    )


def _enterprise_permissions_for_role(role: str) -> dict[str, bool]:
    """The three legacy booleans the deployed SPA reads, resolved from the caller's real access.

    Kept at this name and signature on purpose: ~80 endpoints emit `"permissions":
    _enterprise_permissions_for_role(role)`, and `role` at every one of them is the
    EnterpriseRoleContext the gate returns. Answering from the attached AccessContext therefore
    upgrades all of them at once — and, more importantly, stops two different producers of the
    same key from disagreeing (a granular capability set collapsed one way here and another way
    in /me would flip admin-only panels on for someone who shouldn't see them).

    A plain string still works, for the handful of callers that pass a stored role directly.
    """
    ctx = getattr(role, "ctx", None)
    if ctx is not None:
        return access.legacy_permissions(ctx)
    normalized = _normalize_enterprise_role(role)
    return {
        "can_view_data": True,
        "can_edit_data": normalized in {ENTERPRISE_ROLE_ADMIN, ENTERPRISE_ROLE_EDITOR},
        "can_manage_users": normalized == ENTERPRISE_ROLE_ADMIN,
    }


def _blocked_enterprise_permissions() -> dict[str, bool]:
    return {
        "can_view_data": False,
        "can_edit_data": False,
        "can_manage_users": False,
    }


def _is_development_env() -> bool:
    return os.getenv("ENVIRONMENT", "production").strip().lower() == "development"


def _request_port_for_local_enterprise_url(request: Request | None) -> str | None:
    if ENTERPRISE_PORTAL_PORT:
        return ENTERPRISE_PORTAL_PORT
    if not request or not _is_development_env():
        return None
    port = request.url.port
    if not port or int(port) in {80, 443}:
        return None
    return str(port)


def _build_enterprise_portal_url(
    subdomain_slug: str | None,
    request: Request | None = None,
) -> str | None:
    subdomain = str(subdomain_slug or "").strip().lower()
    if not subdomain:
        return None
    host = f"{subdomain}.{ENTERPRISE_ROOT_DOMAIN}"
    port = _request_port_for_local_enterprise_url(request)
    if port:
        host = f"{host}:{port}"
    return f"{ENTERPRISE_PORTAL_SCHEME}://{host}/enterprise"


def _extract_enterprise_subdomain_from_request(request: Request | None) -> str | None:
    # SECURITY NOTE: x-forwarded-host is client-controllable unless the edge proxy overwrites
    # it. The value here only drives the *cosmetic* subdomain guard (_enforce_request_subdomain_
    # matches_org) — the acting org is ALWAYS resolved from the caller's membership, never from
    # the host — so a spoofed header cannot cross tenants, only bypass a "use your org URL" 403.
    # Harden at the infra layer: configure the proxy (Render/Cloudflare) to STRIP any inbound
    # X-Forwarded-Host and set it itself, so clients can't inject it. (Kept XFH-first here to
    # match app-wide host resolution in main.py:_request_host and avoid breaking prod routing.)
    if request is None:
        return None

    host = (
        request.headers.get("x-forwarded-host")
        or request.headers.get("host")
        or request.url.netloc
        or ""
    ).split(",")[0].strip().lower()
    if not host:
        return None
    host = host.split(":")[0].strip()
    if not host:
        return None

    root_domain = ENTERPRISE_ROOT_DOMAIN.lower()
    if host == root_domain or host == f"www.{root_domain}":
        return None

    suffix = f".{root_domain}"
    if not host.endswith(suffix):
        return None

    subdomain_part = host[: -len(suffix)].strip(".")
    if not subdomain_part:
        return None
    if subdomain_part in ENTERPRISE_GENERIC_SUBDOMAINS:
        return None
    return subdomain_part


def _enforce_request_subdomain_matches_org(
    request: Request | None,
    organization: models.EnterpriseOrganization | None,
) -> None:
    if not request or not organization:
        return
    request_subdomain = _extract_enterprise_subdomain_from_request(request)
    if not request_subdomain:
        return
    org_subdomain = str(organization.subdomain_slug or "").strip().lower()
    if not org_subdomain:
        return
    if request_subdomain != org_subdomain:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"Use your organization URL: {_build_enterprise_portal_url(org_subdomain, request)}",
        )


def _enforce_payload_subdomain_matches_request(
    request: Request | None,
    payload_subdomain: str,
) -> None:
    request_subdomain = _extract_enterprise_subdomain_from_request(request)
    if not request_subdomain:
        return
    if request_subdomain != payload_subdomain:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Organization URL must match the current subdomain. "
                f"You are currently on {request_subdomain}.{ENTERPRISE_ROOT_DOMAIN}."
            ),
        )


def _parse_enterprise_subdomain_or_400(raw_subdomain: str | None) -> str:
    normalized = str(raw_subdomain or "").strip().lower()
    if not normalized:
        raise HTTPException(status_code=400, detail="Organization URL is required.")
    if len(normalized) < ENTERPRISE_SUBDOMAIN_MIN_LENGTH or len(normalized) > ENTERPRISE_SUBDOMAIN_MAX_LENGTH:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Organization URL must be {ENTERPRISE_SUBDOMAIN_MIN_LENGTH}-"
                f"{ENTERPRISE_SUBDOMAIN_MAX_LENGTH} characters."
            ),
        )
    if normalized in ENTERPRISE_RESERVED_SUBDOMAINS:
        raise HTTPException(
            status_code=400,
            detail="This organization URL is reserved. Please choose a different one.",
        )
    if not ENTERPRISE_SUBDOMAIN_PATTERN.fullmatch(normalized):
        raise HTTPException(
            status_code=400,
            detail=(
                "Organization URL can only contain lowercase letters, numbers, and hyphens, "
                "and cannot start or end with a hyphen."
            ),
        )
    return normalized


def _assert_subdomain_available(
    db: Session,
    subdomain_slug: str,
    *,
    exclude_organization_id: int | None = None,
) -> None:
    query = (
        db.query(models.EnterpriseOrganization)
        .filter(func.lower(models.EnterpriseOrganization.subdomain_slug) == subdomain_slug.lower())
    )
    if exclude_organization_id is not None:
        query = query.filter(models.EnterpriseOrganization.id != int(exclude_organization_id))
    existing = query.first()
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"The URL {subdomain_slug}.{ENTERPRISE_ROOT_DOMAIN} is already taken.",
        )


def _get_active_enterprise_membership(
    db: Session,
    user_id: int,
) -> tuple[models.EnterpriseOrganizationMember | None, models.EnterpriseOrganization | None]:
    membership = (
        db.query(models.EnterpriseOrganizationMember)
        .filter(
            models.EnterpriseOrganizationMember.user_id == int(user_id),
            models.EnterpriseOrganizationMember.is_active.is_(True),
        )
        .order_by(models.EnterpriseOrganizationMember.id.asc())
        .first()
    )
    if not membership:
        return None, None
    organization = (
        db.query(models.EnterpriseOrganization)
        .filter(models.EnterpriseOrganization.id == membership.organization_id)
        .first()
    )
    if not organization:
        return None, None
    return membership, organization


def _has_enterprise_access(db: Session, user: models.User) -> bool:
    email = (user.email or "").strip().lower()
    if not email:
        return False
    credential = (
        db.query(models.EnterpriseCredential.id)
        .filter(
            models.EnterpriseCredential.email == email,
            models.EnterpriseCredential.is_active.is_(True),
        )
        .first()
    )
    if credential:
        return True

    if user.id is None:
        return False

    membership = (
        db.query(models.EnterpriseOrganizationMember.id)
        .filter(
            models.EnterpriseOrganizationMember.user_id == int(user.id),
            models.EnterpriseOrganizationMember.is_active.is_(True),
        )
        .first()
    )
    return bool(membership)


def _enforce_enterprise_access_or_403(db: Session, user: models.User) -> None:
    """Gate the endpoints that only need "is this account in a workspace at all?".

    Callers must have verified the password (or hold a session) first — the messages below
    distinguish account states, which is only safe once the caller has proved ownership.
    """
    if _has_enterprise_access(db, user):
        return
    # Same distinction _require_enterprise_membership makes, so an offboarded teammate is sent
    # to their admin instead of being told to create a workspace they already had.
    if _has_deactivated_enterprise_membership(db, getattr(user, "id", None)):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=ENTERPRISE_ACCESS_REVOKED_DETAIL,
        )
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail=ENTERPRISE_NO_WORKSPACE_DETAIL,
    )


_CURRENCY_SYMBOLS = {
    "INR": "₹", "USD": "$", "GBP": "£", "EUR": "€", "CAD": "CA$",
    "AUD": "A$", "AED": "AED ", "SGD": "S$", "JPY": "¥",
}


def _org_display_currency(organization) -> dict:
    """The portal's DISPLAY currency, derived from the company's country in Settings.
    Billing stays in INR — this only converts what the UI shows, at the live exchange
    rate (Frankfurter via /api/pricing/exchange-rates, 24h-cached, safe fallbacks).
    India (or unset) → INR; unlisted countries → USD like the B2C pricing page."""
    from app.routers import pricing as pricing_fx

    country_code = (getattr(organization, "country_code", None) or "").strip().upper()
    code = pricing_fx._COUNTRY_TO_CURRENCY.get(country_code, "USD") if country_code else "INR"
    if code == "INR":
        return {"code": "INR", "symbol": "₹", "rate_from_inr": 1.0, "source": "native", "provider_date": None}

    rate = None
    source = "fallback"
    provider_date = None
    try:
        payload = pricing_fx.get_exchange_rates(refresh=False)
        rates = payload.get("rates") or {}
        inr = float(rates.get("INR") or 0.0)
        target = float(rates.get(code) or 0.0)
        if inr > 0 and target > 0:
            rate = target / inr
            source = payload.get("source") or "live"
            provider_date = payload.get("provider_date")
    except Exception:
        logger.exception("Display-currency rate lookup failed (org_id=%s)", getattr(organization, "id", None))
    if not rate:
        fallback = pricing_fx.FALLBACK_RATES
        rate = float(fallback.get(code, 1.0)) / float(fallback.get("INR", 83.2))
    return {
        "code": code,
        "symbol": _CURRENCY_SYMBOLS.get(code, code + " "),
        "rate_from_inr": round(rate, 6),
        "source": source,
        "provider_date": provider_date,
    }


def _build_dpa_consent_state(organization) -> dict:
    """Whether the org is still on a superseded Data Processing Agreement.

    ``dpa_accepted_version`` is stamped at signup (and on every re-acceptance) so we can
    detect organizations that agreed to an older DPA and re-prompt an admin. Legacy rows
    created before the column existed have a NULL version and also need re-consent.
    """
    accepted_version = (getattr(organization, "dpa_accepted_version", None) or "").strip() or None
    return {
        "current_version": LEGAL_DPA_VERSION,
        "accepted_version": accepted_version,
        "accepted_at": getattr(organization, "dpa_accepted_at", None),
        "reconsent_required": accepted_version != LEGAL_DPA_VERSION,
    }


def _build_enterprise_context(
    db: Session,
    user: models.User,
    request: Request | None = None,
) -> dict:
    membership, organization = _get_active_enterprise_membership(db, user.id)
    if not membership or not organization:
        return {
            "onboarding_required": True,
            "organization": None,
            "membership": None,
            "permissions": _blocked_enterprise_permissions(),
            "access": access.blocked_access_payload(),
            "branches": [],
            "dpa": {
                "current_version": LEGAL_DPA_VERSION,
                "accepted_version": None,
                "accepted_at": None,
                # Nothing to re-accept until an organization exists.
                "reconsent_required": False,
            },
        }

    normalized_role = _normalize_enterprise_role(membership.role)
    ctx = access.resolve_access_context(db, membership, organization)
    company_name = (organization.company_name or "").strip()
    subdomain_slug = (organization.subdomain_slug or "").strip().lower()
    logo_url = _resolve_enterprise_logo_url(organization)
    onboarding_required = not company_name or not subdomain_slug

    subscription_summary = None
    credits_summary = None
    if not onboarding_required:
        try:
            subscription_summary = _serialize_subscription_state(
                billing.build_subscription_state(db, organization.id)
            )
        except Exception:
            logger.exception("Failed to build subscription state for org_id=%s", organization.id)
        try:
            credits_summary = credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization))
        except Exception:
            logger.exception("Failed to build credit wallet state for org_id=%s", organization.id)

    return {
        "onboarding_required": onboarding_required,
        "organization": {
            "id": organization.id,
            "company_name": company_name or organization.company_name,
            "subdomain_slug": subdomain_slug or None,
            "logo_url": logo_url,
            "portal_url": _build_enterprise_portal_url(subdomain_slug, request),
            "created_at": organization.created_at,
            "country_code": organization.country_code,
            "state_region": organization.state_region,
            "display_currency": _org_display_currency(organization),
        },
        "membership": {
            # Still the legacy admin|editor|viewer string: the deployed SPA prints it, and a
            # mid-deploy tab must not suddenly render "branch_manager". The UI reads the real
            # label from access.role_label.
            "role": normalized_role,
            "is_active": bool(membership.is_active),
            "joined_at": membership.created_at,
            "job_title": getattr(membership, "job_title", None),
        },
        "subscription": subscription_summary,
        "credits": credits_summary,
        "permissions": (
            access.legacy_permissions(ctx)
            if not onboarding_required
            else _blocked_enterprise_permissions()
        ),
        "access": (
            access.access_payload(ctx)
            if not onboarding_required
            else access.blocked_access_payload()
        ),
        # Offices ride along on /me so the Add-client modal — which opens from any view — can
        # populate its office picker without every screen having to fetch the team endpoint.
        "branches": _org_branch_options(db, organization, ctx) if not onboarding_required else [],
        "dpa": _build_dpa_consent_state(organization),
    }


def _org_branch_options(db: Session, organization, ctx) -> list[dict]:
    """Active offices this member may file a client under: the whole org, or just their own."""
    try:
        from app import enterprise_team as team

        rows = team.list_branches(db, organization.id)
        if ctx is not None and ctx.scope_kind == "branch":
            rows = [b for b in rows if int(b.id) in ctx.branch_ids]
        return [
            {
                "id": b.id,
                "name": b.name,
                "code": b.code,
                "city": b.city,
                "is_default": bool(b.is_default),
            }
            for b in rows
        ]
    except Exception:
        logger.exception("Failed to list offices for org_id=%s", getattr(organization, "id", None))
        return []


class EnterpriseRoleContext(str):
    """The legacy role string, with the caller's resolved AccessContext riding along.

    Subclasses `str` deliberately. Every one of the ~100 existing call sites does
    `_, organization, role = _require_enterprise_membership(...)` and then uses `role` as a
    string — `role == ENTERPRISE_ROLE_ADMIN`, `role in {...}`, `"role": role` in a JSON payload.
    All of that keeps working untouched, while new code reads `role.ctx` for granular
    capabilities and record scope. That is what makes granular access controls landable on a
    live app without rewriting a hundred handlers in one commit.

    `ctx` is optional and `__reduce__` is defined so `copy`/`deepcopy`/`pickle` of a response
    payload can't blow up on the custom `__new__`.
    """

    ctx: "access.AccessContext | None"

    def __new__(cls, legacy_role: str, ctx=None):
        obj = super().__new__(cls, legacy_role)
        obj.ctx = ctx
        return obj

    def __reduce__(self):
        return (EnterpriseRoleContext, (str(self), self.ctx))


def _require_enterprise_membership(
    *,
    db: Session,
    user: models.User,
    request: Request | None = None,
    require_manage_users: bool = False,
    require_edit_data: bool = False,
    require_capability: "str | tuple[str, ...] | None" = None,
    require_any: "tuple[str, ...] | None" = None,
) -> tuple[models.EnterpriseOrganizationMember, models.EnterpriseOrganization, EnterpriseRoleContext]:
    """Resolve the caller's membership, organization and access context, or 403.

    `require_manage_users` and `require_edit_data` are the two legacy flags. They keep their old
    meaning on purpose:

      * `require_manage_users` maps to "is genuinely an owner or admin of this workspace" — NOT
        to holding the `team.manage` capability. Those two are different things now: an owner can
        delegate `team.manage` to an office coordinator so they can invite staff, and that person
        must not thereby be able to issue a refund, repoint the payout bank account or rename the
        workspace URL. Endpoints that SHOULD be delegable carry an explicit `require_capability`
        instead.
      * `require_edit_data` maps to `clients.edit`, which is what it always meant.

    `require_capability` requires every listed capability; `require_any` requires at least one.
    """
    membership, organization = _get_active_enterprise_membership(db, user.id)
    if not membership or not organization:
        # Distinguish "never onboarded" from "your access was switched off", so an offboarded
        # teammate isn't told to complete a setup step they can't reach.
        if _has_deactivated_enterprise_membership(db, user.id):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ENTERPRISE_ACCESS_REVOKED_DETAIL,
            )
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Enterprise onboarding is required before accessing this feature.",
        )

    subdomain_slug = (organization.subdomain_slug or "").strip().lower()
    if not subdomain_slug:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Complete enterprise onboarding by setting your organization URL first.",
        )
    _enforce_request_subdomain_matches_org(request, organization)

    ctx = access.resolve_access_context(db, membership, organization)
    role = EnterpriseRoleContext(access.legacy_role_for(ctx.role_key, ctx.capabilities), ctx)

    if require_manage_users and not ctx.is_admin_like:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only organization admins can manage users.",
        )
    if require_edit_data and not ctx.has("clients.edit"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admins or editors can modify student records.",
        )

    needed = (require_capability,) if isinstance(require_capability, str) else tuple(require_capability or ())
    for capability in needed:
        if not ctx.has(capability):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=access.denied_detail(capability),
            )
    if require_any and not any(ctx.has(capability) for capability in require_any):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=access.denied_detail(require_any[0]),
        )

    return membership, organization, role


def _has_deactivated_enterprise_membership(db: Session, user_id: int) -> bool:
    if not user_id:
        return False
    return bool(
        db.query(models.EnterpriseOrganizationMember.id)
        .filter(
            models.EnterpriseOrganizationMember.user_id == int(user_id),
            models.EnterpriseOrganizationMember.is_active.is_(False),
        )
        .first()
    )


def _touch_member_last_active(membership_id: int) -> None:
    """Stamp `last_active_at` at most once every few minutes, on its own connection.

    Managers need "is this seat actually being used" — `users.last_login_at` only moves on
    sign-in. Two deliberate choices: this runs on a SEPARATE short-lived connection rather than
    the request session (a `commit()` mid-gate would expire `membership`/`organization` and force
    a re-SELECT in every handler, and would open a transaction boundary before handlers that take
    row locks), and it is called only from the couple of endpoints the SPA hits on boot rather
    than from the gate, so it never contends on a hot path. Telemetry must never fail a request.
    """
    if not membership_id:
        return
    try:
        cutoff = datetime.utcnow() - timedelta(minutes=5)
        from app.database import engine as _engine
        from sqlalchemy import text as _sql_text

        with _engine.begin() as conn:
            conn.execute(
                _sql_text(
                    "UPDATE enterprise_organization_members SET last_active_at = :now "
                    "WHERE id = :id AND (last_active_at IS NULL OR last_active_at < :cutoff)"
                ),
                {"now": datetime.utcnow(), "id": int(membership_id), "cutoff": cutoff},
            )
    except Exception:
        logger.debug("last_active_at stamp skipped (member_id=%s)", membership_id, exc_info=True)


def _serialize_team_member(
    membership: models.EnterpriseOrganizationMember,
    user: models.User,
) -> dict:
    return {
        "user_id": user.id,
        "email": user.email,
        "full_name": user.full_name,
        "role": _normalize_enterprise_role(membership.role),
        "is_active": bool(membership.is_active),
        "joined_at": membership.created_at,
        "last_login_at": user.last_login_at,
    }


def _list_organization_members(
    db: Session,
    organization_id: int,
    *,
    include_inactive: bool = False,
    ctx=None,
) -> list[dict]:
    """The team roster.

    Delegates to enterprise_team, which resolves roles, offices and caseload counts for every
    member in a handful of bulk queries. The rows are a strict SUPERSET of what
    `_serialize_team_member` used to return — the legacy `role` string is still there, so the
    currently-deployed frontend keeps working while the new Team screen reads the richer fields.
    """
    from app import enterprise_team as team

    try:
        return team.list_members(
            db, int(organization_id), include_inactive=include_inactive, ctx=ctx
        )
    except Exception:
        # The roster backs the client-assignment dropdown as well as the Team screen, so fall
        # back to the original minimal shape rather than failing those pages outright.
        logger.exception("Falling back to basic member list for org_id=%s", organization_id)
        rows = (
            db.query(models.EnterpriseOrganizationMember, models.User)
            .join(
                models.User,
                models.User.id == models.EnterpriseOrganizationMember.user_id,
            )
            .filter(
                models.EnterpriseOrganizationMember.organization_id == int(organization_id),
                models.EnterpriseOrganizationMember.is_active.is_(True),
            )
            .order_by(models.EnterpriseOrganizationMember.created_at.asc())
            .all()
        )
        return [_serialize_team_member(member, user) for member, user in rows]


# _active_admin_count lived here. Its last caller was the legacy role endpoint, which now
# delegates to team_svc.update_member_access and so inherits `assert_admin_remains`
# (app/enterprise_team.py) — a stricter version of the same invariant that also counts rows
# written before `role_key` existed.


def _enterprise_student_options_payload() -> dict:
    return {
        "countries": [
            {
                "code": item["code"],
                "name": item["name"],
                "flag_emoji": str(item.get("flag_emoji") or "").strip(),
                "iconic_place": str(item.get("iconic_place") or "").strip(),
                "visa_types": list(item["visa_types"]),
                "intakes_by_visa": {
                    str(visa_name): _materialize_future_intakes(intakes or [])
                    for visa_name, intakes in (item.get("intakes_by_visa") or {}).items()
                },
            }
            for item in ENTERPRISE_STUDY_DESTINATION_OPTIONS
        ]
    }


def _resolve_intake_start_month(intake_label: str) -> int:
    normalized = str(intake_label or "").strip().lower()
    if not normalized:
        return 9
    if normalized in ENTERPRISE_INTAKE_START_MONTH_HINTS:
        return int(ENTERPRISE_INTAKE_START_MONTH_HINTS[normalized])

    for key, month in ENTERPRISE_INTAKE_START_MONTH_HINTS.items():
        if key in normalized:
            return int(month)
    return 9


def _materialize_future_intakes(intake_labels: list[str]) -> list[str]:
    now_utc = datetime.utcnow()
    current_month = int(now_utc.month)
    current_year = int(now_utc.year)
    sortable_results: list[tuple[int, int, str, str]] = []
    seen: set[str] = set()

    for raw_label in intake_labels:
        base_label = str(raw_label or "").strip()
        if not base_label:
            continue
        start_month = _resolve_intake_start_month(base_label)
        year = current_year if start_month > current_month else (current_year + 1)
        future_label = f"{base_label} {year}"
        dedupe_key = future_label.lower()
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        sortable_results.append((year, start_month, base_label.lower(), future_label))

    sortable_results.sort(key=lambda item: (item[0], item[1], item[2]))
    return [item[3] for item in sortable_results]


def _parse_enterprise_student_payload_or_400(
    payload: EnterpriseStudentCreateRequest,
) -> tuple[str, str, str, str, str]:
    student_name = str(payload.student_name or "").strip()
    if len(student_name) < 2:
        raise HTTPException(status_code=400, detail="Student name must be at least 2 characters.")
    if len(student_name) > ENTERPRISE_STUDENT_NAME_MAX_LENGTH:
        raise HTTPException(
            status_code=400,
            detail=f"Student name must be at most {ENTERPRISE_STUDENT_NAME_MAX_LENGTH} characters.",
        )

    study_country_code = str(payload.study_country_code or "").strip().upper()
    country_meta = ENTERPRISE_STUDY_DESTINATION_MAP.get(study_country_code)
    if not country_meta:
        raise HTTPException(status_code=400, detail="Please select a valid study destination country.")

    visa_type_input = str(payload.visa_type or "").strip()
    if not visa_type_input:
        raise HTTPException(status_code=400, detail="Please select a visa type.")

    canonical_visa_type = ""
    for visa_option in country_meta["visa_types"]:
        if str(visa_option).strip().lower() == visa_type_input.lower():
            canonical_visa_type = str(visa_option).strip()
            break
    if not canonical_visa_type:
        raise HTTPException(
            status_code=400,
            detail="Selected visa type does not match the selected study destination.",
        )

    intake_input = str(payload.intake or "").strip()
    if not intake_input:
        raise HTTPException(status_code=400, detail="Please select an intake.")

    intakes_by_visa = country_meta.get("intakes_by_visa") or {}
    allowed_base_intakes = list(intakes_by_visa.get(canonical_visa_type) or [])
    allowed_intakes = _materialize_future_intakes(allowed_base_intakes)
    canonical_intake = ""
    for intake_option in allowed_intakes:
        if str(intake_option).strip().lower() == intake_input.lower():
            canonical_intake = str(intake_option).strip()
            break
    if not canonical_intake:
        for raw_base_intake in allowed_base_intakes:
            base_intake = str(raw_base_intake).strip()
            if base_intake and base_intake.lower() == intake_input.lower():
                canonical_list = _materialize_future_intakes([base_intake])
                canonical_intake = canonical_list[0] if canonical_list else base_intake
                break
    if not canonical_intake:
        raise HTTPException(
            status_code=400,
            detail="Selected intake does not match the selected country and visa type.",
        )

    return (
        student_name,
        study_country_code,
        str(country_meta["name"]).strip(),
        canonical_visa_type,
        canonical_intake,
    )


def _serialize_enterprise_student(student: models.EnterpriseStudent) -> dict:
    return {
        "id": student.id,
        "student_name": student.student_name,
        "study_country_code": student.study_country_code,
        "study_country_name": student.study_country_name,
        "visa_type": student.visa_type,
        "intake": student.intake,
        "created_at": student.created_at,
    }


def _list_enterprise_students(db: Session, organization_id: int) -> list[dict]:
    rows = (
        db.query(models.EnterpriseStudent)
        .filter(models.EnterpriseStudent.organization_id == int(organization_id))
        .order_by(models.EnterpriseStudent.created_at.desc(), models.EnterpriseStudent.id.desc())
        .all()
    )
    return [_serialize_enterprise_student(item) for item in rows]


def seed_enterprise_user() -> None:
    """
    Sync enterprise credential rows from PostgreSQL into users table for auth/session reuse.
    """
    db = SessionLocal()
    try:
        credentials = (
            db.query(models.EnterpriseCredential)
            .filter(models.EnterpriseCredential.is_active.is_(True))
            .all()
        )
        if not credentials:
            return

        changes_made = False
        for entry in credentials:
            email = (entry.email or "").strip().lower()
            password_hash = (entry.password_hash or "").strip()
            if not email or not password_hash:
                continue

            existing = db.query(models.User).filter(models.User.email == email).first()
            if existing:
                existing.hashed_password = password_hash
                existing.is_active = True
                existing.email_verified = True
                existing.is_admin = True
                if not existing.full_name:
                    existing.full_name = (entry.full_name or "Enterprise Admin").strip()
                if not existing.university:
                    existing.university = "Enterprise Account"
                if existing.accepted_terms_privacy_at is None:
                    existing.accepted_terms_privacy_at = datetime.utcnow()
                changes_made = True
                continue

            user = models.User(
                email=email,
                username=None,
                hashed_password=password_hash,
                full_name=(entry.full_name or "Enterprise Admin").strip(),
                university="Enterprise Account",
                is_active=True,
                email_verified=True,
                is_admin=True,
                accepted_terms_privacy_at=datetime.utcnow(),
            )
            db.add(user)
            changes_made = True

        if changes_made:
            db.commit()
    finally:
        db.close()


def backfill_enterprise_account_flag(db: Session) -> None:
    """Idempotent backfill of users.is_enterprise_account for accounts that predate the B2B/B2C
    product separation. Flags enterprise-CREATED accounts (so they're blocked from the B2C app)
    using RELIABLE positive signals only — never a "B2C footprint" heuristic, which is unsafe on
    this shared users table (a one-time journey migration blanket-set destination_country_code /
    onboarding_completed_at on every row, and a dormant password B2C signup has no footprint at
    all — so footprint-guessing both misses old enterprise owners and wrongly locks out real
    consumers who were later invited to a team).

    Policy ("block only enterprise-created"; keep B2C access for anyone who was a B2C user first):
      * Workspace OWNERS (EnterpriseOrganization.created_by_user_id) are flagged UNCONDITIONALLY:
        enterprise signup rejects an existing email, so an owner's account was created BY the
        enterprise product and was never a prior B2C user.
      * MEMBERS are flagged only when the account's `university` equals their org's company_name —
        the exact value the invite flow writes when it CREATES a brand-new teammate. The invite
        REUSE branch never rewrites an existing user's university, so a person who was a B2C user
        first keeps their own university (or NULL) and is left untouched → keeps consumer access.
        Membership rows are considered regardless of is_active: origin doesn't change when a
        teammate is removed (their going-forward flag would have survived deactivation too), and
        the university==company_name gate keeps this false-positive-safe.

    Only ever SETS the flag (additive/idempotent); never clears it and never touches
    admins/developers. Safe to run on every startup.

    KNOWN RESIDUAL (accepted): a member is matched against the org's CURRENT company_name. If an
    org renamed its company AFTER a historical teammate was invite-created (their frozen university
    holds the OLD name, and there is no name history to recover it), the backfill won't flag that
    teammate — they keep B2C access. This is bounded to pre-separation renamed orgs; every account
    created from now on is flagged at creation, which is the real source of truth.
    """
    organizations = db.query(
        models.EnterpriseOrganization.id,
        models.EnterpriseOrganization.company_name,
        models.EnterpriseOrganization.created_by_user_id,
    ).all()
    if not organizations:
        return

    owner_ids: set[int] = set()
    company_by_org: dict[int, str] = {}
    for org_id, company_name, created_by in organizations:
        company_by_org[org_id] = (company_name or "").strip()
        if created_by is not None:
            owner_ids.add(int(created_by))

    # For each member (active OR not — membership indicates enterprise origin, which removal from
    # the team doesn't undo), the set of company names of the org(s) they belong to.
    member_companies: dict[int, set[str]] = {}
    for user_id, org_id in (
        db.query(
            models.EnterpriseOrganizationMember.user_id,
            models.EnterpriseOrganizationMember.organization_id,
        )
        .all()
    ):
        if user_id is None:
            continue
        company = company_by_org.get(org_id)
        if company:
            member_companies.setdefault(int(user_id), set()).add(company)

    candidate_ids = owner_ids | set(member_companies.keys())
    if not candidate_ids:
        return

    changed = 0
    for user in db.query(models.User).filter(models.User.id.in_(candidate_ids)).all():
        if getattr(user, "is_enterprise_account", False):
            continue
        if user.is_admin or user.is_developer:
            continue
        is_owner = user.id in owner_ids
        # Positive enterprise-origin marker for members: the account's university was set to the
        # org's company name — which only the invite CREATE path does (reuse leaves it as-is).
        university = (user.university or "").strip()
        member_created_by_org = bool(university) and university in member_companies.get(user.id, set())
        if is_owner or member_created_by_org:
            user.is_enterprise_account = True
            changed += 1

    if changed:
        db.commit()
        logger.info(
            "Backfilled is_enterprise_account=True for %d enterprise-origin account(s).", changed
        )


@router.post("/login")
def enterprise_login(
    payload: EnterpriseLoginRequest,
    request: Request,
    response: Response,
    db: Session = Depends(get_db),
):
    # Deliberately a plain `def`, not `async def`. This handler never awaits — it does a
    # DB-backed rate-limit check, a blocking Turnstile HTTP call (up to 10s), a bcrypt verify
    # (~170ms of pure CPU) and several sync DB queries. On the event loop those would freeze
    # every other request the worker is serving; as a sync handler FastAPI runs the whole
    # thing in its threadpool, so one person signing in can no longer stall everyone else.
    _enforce_rate_limit_or_429(
        request=request,
        scope="enterprise.login",
        limit=ENTERPRISE_LOGIN_RATE_LIMIT,
        window_seconds=ENTERPRISE_LOGIN_RATE_WINDOW_SECONDS,
    )

    turnstile_token = (payload.cf_turnstile_token or "").strip()
    ip_whitelisted = is_request_ip_whitelisted(request)
    if is_turnstile_enabled() and not ip_whitelisted:
        if turnstile_token:
            client_ip = extract_client_ip(request) if request else None
            if not verify_turnstile_token(turnstile_token, client_ip):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Security verification failed. Please try again.",
                )
        elif _is_turnstile_required():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Security verification is required",
            )

    login_email = payload.email.lower().strip()

    # Password first. Every message below names a specific account state, so it may only be
    # disclosed to a caller who has already proved they own the account — otherwise the form
    # becomes an oracle for "which emails have a workspace". Mirrors the B2C rule in
    # routers/auth.py, where the product-separation check likewise runs post-authentication.
    user = authenticate_user(db, login_email, payload.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid email or password.",
        )

    if not user.is_active:
        raise HTTPException(status_code=403, detail="Account is deactivated.")

    _enforce_enterprise_access_or_403(db, user)

    user.last_login_at = datetime.utcnow()
    db.commit()

    access_token = create_access_token(
        data={"sub": user.email},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
    )
    set_auth_cookie(request, response, access_token)

    _, organization = _get_active_enterprise_membership(db, user.id)
    _enforce_request_subdomain_matches_org(request, organization)

    context = _build_enterprise_context(db, user, request)
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": user.id,
            "email": user.email,
            "full_name": user.full_name,
            "is_admin": user.is_admin,
        },
        **context,
    }


@router.get("/me")
def enterprise_me(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _enforce_enterprise_access_or_403(db, current_user)
    _, organization = _get_active_enterprise_membership(db, current_user.id)
    _enforce_request_subdomain_matches_org(request, organization)
    return {
        "user": {
            "id": current_user.id,
            "email": current_user.email,
            "full_name": current_user.full_name,
            "is_admin": bool(current_user.is_admin),
            "is_developer": bool(current_user.is_developer),
            "heard_about_answered": getattr(current_user, "heard_about_us_at", None) is not None,
        },
        **_build_enterprise_context(db, current_user, request),
    }


@router.post("/dpa/accept")
def enterprise_accept_dpa(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Record an admin's acceptance of the CURRENT Data Processing Agreement.

    Used when ``LEGAL_DPA_VERSION`` is bumped and existing organizations are still on a
    superseded version: the portal shows a blocking banner until an admin accepts here.
    Idempotent — re-accepting the version already on file is a no-op that returns the
    same state, so a double-click cannot rewrite the original proof-of-consent timestamp.
    """
    _, organization, _ = _require_enterprise_membership(
        db=db,
        user=current_user,
        request=request,
        require_capability="org.legal_accept",
    )

    already_current = (
        (organization.dpa_accepted_version or "").strip() == LEGAL_DPA_VERSION
    )
    if not already_current:
        organization.dpa_accepted_at = datetime.utcnow()
        organization.dpa_accepted_version = LEGAL_DPA_VERSION
        organization.dpa_accepted_by_user_id = current_user.id
        db.commit()
        db.refresh(organization)
        logger.info(
            "Enterprise DPA re-accepted (org_id=%s, user_id=%s, version=%s)",
            organization.id,
            current_user.id,
            LEGAL_DPA_VERSION,
        )

    return {
        "accepted": True,
        "already_current": already_current,
        "dpa": _build_dpa_consent_state(organization),
    }


class EnterpriseHeardAboutRequest(BaseModel):
    source: str = Field(..., min_length=1, max_length=40)
    detail: Optional[str] = Field(default=None, max_length=200)


@router.get("/heard-about")
def enterprise_heard_about_status(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Options + answered state for the enterprise post-signup 'How did you hear' prompt."""
    from app import acquisition
    _enforce_enterprise_access_or_403(db, current_user)
    return {
        "answered": getattr(current_user, "heard_about_us_at", None) is not None,
        "selected": getattr(current_user, "heard_about_us", None),
        "options": acquisition.HEARD_ABOUT_OPTIONS,
    }


@router.post("/heard-about")
def enterprise_submit_heard_about(
    payload: EnterpriseHeardAboutRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Save the self-reported 'How did you hear about us?' answer for an enterprise user."""
    from app import acquisition
    _enforce_enterprise_access_or_403(db, current_user)
    if (payload.source or "").strip().lower() == "skip":
        current_user.heard_about_us_at = datetime.utcnow()  # mark asked; don't nag again
        db.commit()
        return {"ok": True, "selected": None}
    vid = acquisition.apply_self_reported_source(current_user, payload.source, payload.detail)
    if not vid:
        raise HTTPException(status_code=400, detail="Please choose a valid option.")
    db.commit()
    return {"ok": True, "selected": vid}


@router.get("/students/options")
def enterprise_student_options(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="catalog.view")
    return {
        "organization_id": organization.id,
        "permissions": _enterprise_permissions_for_role(role),
        "options": _enterprise_student_options_payload(),
    }


@router.get("/students")
def enterprise_students(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="clients.view")
    students = _list_enterprise_students(db, organization.id)
    return {
        "organization_id": organization.id,
        "permissions": _enterprise_permissions_for_role(role),
        "students_count": len(students),
        "students": students,
    }


@router.post("/students")
def enterprise_add_student(
    payload: EnterpriseStudentCreateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db,
        user=current_user,
        request=request,
        require_capability="clients.create",
    )
    student_name, study_country_code, study_country_name, visa_type, intake = _parse_enterprise_student_payload_or_400(payload)

    student = models.EnterpriseStudent(
        organization_id=organization.id,
        student_name=student_name,
        study_country_code=study_country_code,
        study_country_name=study_country_name,
        visa_type=visa_type,
        intake=intake,
        created_by_user_id=current_user.id,
    )
    db.add(student)
    db.commit()
    db.refresh(student)

    students_count = int(
        db.query(models.EnterpriseStudent)
        .filter(models.EnterpriseStudent.organization_id == int(organization.id))
        .count()
    )

    return {
        "message": "Student added successfully.",
        "permissions": _enterprise_permissions_for_role(role),
        "students_count": students_count,
        "student": _serialize_enterprise_student(student),
    }


@router.post("/onboarding")
def enterprise_onboarding(
    payload: EnterpriseOnboardingRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _enforce_enterprise_access_or_403(db, current_user)
    existing_membership, existing_org = _get_active_enterprise_membership(db, current_user.id)
    company_name = (payload.company_name or "").strip()
    if len(company_name) < 2:
        raise HTTPException(status_code=400, detail="Company name must be at least 2 characters.")
    subdomain_slug = _parse_enterprise_subdomain_or_400(payload.subdomain_slug)
    _enforce_payload_subdomain_matches_request(request, subdomain_slug)

    if existing_membership and existing_org:
        if (existing_org.subdomain_slug or "").strip():
            _enforce_request_subdomain_matches_org(request, existing_org)
            return {
                "message": "Enterprise organization already configured.",
                **_build_enterprise_context(db, current_user, request),
            }

        _assert_subdomain_available(
            db=db,
            subdomain_slug=subdomain_slug,
            exclude_organization_id=existing_org.id,
        )
        existing_org.company_name = company_name
        existing_org.subdomain_slug = subdomain_slug
        if not str(existing_org.logo_url or "").strip():
            existing_org.logo_url = _build_default_enterprise_logo_url(
                organization_id=existing_org.id,
                company_name=company_name,
                subdomain_slug=subdomain_slug,
                randomize=True,
            )
        db.commit()
        return {
            "message": "Enterprise organization setup complete.",
            **_build_enterprise_context(db, current_user, request),
        }

    _assert_subdomain_available(db=db, subdomain_slug=subdomain_slug)

    organization = models.EnterpriseOrganization(
        company_name=company_name,
        subdomain_slug=subdomain_slug,
        logo_url=_build_default_enterprise_logo_url(
            organization_id=None,
            company_name=company_name,
            subdomain_slug=subdomain_slug,
            randomize=True,
        ),
        created_by_user_id=current_user.id,
    )
    db.add(organization)
    db.flush()

    membership = models.EnterpriseOrganizationMember(
        organization_id=organization.id,
        user_id=current_user.id,
        # See the matching note in enterprise_signup: role_key must be set explicitly or
        # the column default ("viewer") makes the workspace creator a viewer of it.
        role_key=access.ROLE_OWNER,
        role=ENTERPRISE_ROLE_ADMIN,
        is_active=True,
        invited_by_user_id=current_user.id,
    )
    db.add(membership)
    db.commit()

    return {
        "message": "Enterprise organization setup complete.",
        **_build_enterprise_context(db, current_user, request),
    }


@router.get("/team")
def enterprise_team_members(
    request: Request,
    include_inactive: bool = False,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    membership, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="team.view"
    )
    from app import enterprise_team as team

    # Deactivated members are only listed for someone who can actually do something about them.
    show_inactive = bool(include_inactive) and role.ctx.has("team.manage")
    members = _list_organization_members(db, organization.id, include_inactive=show_inactive, ctx=role.ctx)
    _touch_member_last_active(membership.id)
    branches = [
        team.serialize_branch(b) for b in team.list_branches(db, organization.id)
    ]
    return {
        "organization": {
            "id": organization.id,
            "company_name": organization.company_name,
            "subdomain_slug": (organization.subdomain_slug or "").strip().lower() or None,
            "logo_url": _resolve_enterprise_logo_url(organization),
            "portal_url": _build_enterprise_portal_url(organization.subdomain_slug, request),
        },
        # `current_role` and `permissions` keep their legacy shape — the deployed SPA reads
        # `d.permissions.can_manage_users` directly and a missing key would blank the page.
        "current_role": role,
        "permissions": _enterprise_permissions_for_role(role),
        "access": access.access_payload(role.ctx),
        "members": members,
        "branches": branches,
        "role_presets": access.capability_registry_payload().get("role_presets", []),
    }


@router.patch("/organization/branding")
def enterprise_update_branding(
    payload: EnterpriseBrandingUpdateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, _ = _require_enterprise_membership(
        db=db,
        user=current_user,
        request=request,
        require_capability="settings.manage",
    )

    has_update = False
    new_company_name: str | None = None
    if payload.company_name is not None:
        new_company_name = (payload.company_name or "").strip()
        if len(new_company_name) < 2:
            raise HTTPException(status_code=400, detail="Company name must be at least 2 characters.")
        has_update = True

    if payload.generate_random_logo:
        organization.logo_url = _build_default_enterprise_logo_url(
            organization_id=organization.id,
            company_name=new_company_name or organization.company_name,
            subdomain_slug=organization.subdomain_slug,
            randomize=True,
            current_logo_url=organization.logo_url,
        )
        has_update = True
    elif payload.logo_url is not None:
        normalized_logo = _normalize_enterprise_logo_url_or_400(payload.logo_url)
        if normalized_logo:
            organization.logo_url = normalized_logo
        else:
            organization.logo_url = _build_default_enterprise_logo_url(
                organization_id=organization.id,
                company_name=new_company_name or organization.company_name,
                subdomain_slug=organization.subdomain_slug,
                randomize=True,
                current_logo_url=organization.logo_url,
            )
        has_update = True

    if new_company_name is not None:
        organization.company_name = new_company_name

    if payload.country_code is not None:
        country_code = (payload.country_code or "").strip().upper()
        if country_code and (len(country_code) != 2 or not country_code.isalpha()):
            raise HTTPException(status_code=400, detail="Enter a valid 2-letter country code.")
        organization.country_code = country_code or None
        has_update = True
    if payload.state_region is not None:
        organization.state_region = (payload.state_region or "").strip()[:80] or None
        has_update = True

    if not has_update:
        raise HTTPException(status_code=400, detail="Nothing to update.")

    db.commit()

    context = _build_enterprise_context(db, current_user, request)
    return {
        "message": "Organization branding updated successfully.",
        "organization": context.get("organization"),
    }


ENTERPRISE_LOGO_MAX_BYTES = 2 * 1024 * 1024  # 2 MB upload cap
_ENTERPRISE_LOGO_FILENAME_RE = re.compile(r"^logo-[0-9a-f]{32}\.png$")


def _normalize_logo_image_or_400(data: bytes) -> bytes:
    """Validate an uploaded logo and re-encode it. Only a real PNG/JPEG/WebP raster is
    accepted; it is downscaled to ≤512px and re-encoded to a fresh PNG, so nothing but a
    clean, metadata-free raster we produced ourselves is ever stored or served."""
    import io
    try:
        from PIL import Image
        probe = Image.open(io.BytesIO(data))
        probe.verify()  # detects truncated/corrupt files (invalidates the handle)
        img = Image.open(io.BytesIO(data))
        if (img.format or "").upper() not in {"PNG", "JPEG", "WEBP"}:
            raise ValueError("unsupported format")
        img = img.convert("RGBA")
        img.thumbnail((512, 512))
        out = io.BytesIO()
        img.save(out, format="PNG", optimize=True)
        return out.getvalue()
    except Exception:
        raise HTTPException(status_code=400, detail="Please upload a valid PNG, JPG, or WebP image.")


@router.post("/organization/logo")
async def enterprise_upload_org_logo(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Upload a custom organization logo. Stored encrypted like other enterprise assets and
    served back through the unguessable public logo route below."""
    _, organization, _ = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="settings.manage"
    )
    if not enterprise_storage.is_configured():
        raise HTTPException(status_code=503, detail="Logo storage is not configured.")

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="The file is empty.")
    if len(data) > ENTERPRISE_LOGO_MAX_BYTES:
        raise HTTPException(status_code=400, detail="Logo is too large. Maximum size is 2 MB.")
    # Pillow re-encode is CPU-bound; offload it and the R2 PUT below off the event loop so a
    # logo upload can't stall other requests on this worker (this handler must stay async for
    # `await file.read()` above).
    png_bytes = await run_in_threadpool(_normalize_logo_image_or_400, data)

    filename = f"logo-{uuid.uuid4().hex}.png"
    try:
        await run_in_threadpool(
            enterprise_storage.store_document,
            f"enterprise/{organization.id}/branding/{filename}", png_bytes, content_type="image/png",
        )
    except Exception:
        logger.exception("Failed to store org logo (org_id=%s)", organization.id)
        raise HTTPException(status_code=502, detail="Could not store the logo right now. Please try again.")

    # Best-effort cleanup of the previously uploaded logo (generated/external URLs untouched).
    old_match = re.match(
        r"^/api/enterprise/public/org-logo/(\d+)/(logo-[0-9a-f]{32}\.png)$",
        str(organization.logo_url or ""),
    )
    if old_match and int(old_match.group(1)) == organization.id:
        await run_in_threadpool(
            enterprise_storage.delete_document,
            f"enterprise/{organization.id}/branding/{old_match.group(2)}",
        )

    organization.logo_url = f"/api/enterprise/public/org-logo/{organization.id}/{filename}"
    db.commit()

    context = _build_enterprise_context(db, current_user, request)
    return {"message": "Logo updated.", "organization": context.get("organization")}


@router.get("/public/org-logo/{org_id}/{filename}")
def enterprise_public_org_logo(org_id: int, filename: str):
    """Serve an uploaded organization logo. Unauthenticated by design: the filename embeds an
    unguessable 128-bit token, and a logo is public branding (never client data). Only files
    matching our own generated pattern under the org's branding prefix can be addressed."""
    if not _ENTERPRISE_LOGO_FILENAME_RE.fullmatch(str(filename or "")):
        raise HTTPException(status_code=404, detail="Not found")
    try:
        data = enterprise_storage.fetch_document(f"enterprise/{int(org_id)}/branding/{filename}")
    except Exception:
        raise HTTPException(status_code=404, detail="Not found")
    return Response(
        content=data,
        media_type="image/png",
        headers={"Cache-Control": "public, max-age=31536000, immutable"},
    )


@router.get("/public/brand-mark/{spec}.png")
def enterprise_public_brand_mark(spec: str):
    """Draw the generated logo for an organization that has not uploaded one.

    Unauthenticated and stateless by design: the spec in the path *is* the design, so
    there is nothing to look up and nothing about the organization to leak. Same spec,
    same bytes, forever — hence the immutable cache."""
    if brand_marks.parse_spec(spec) is None:
        raise HTTPException(status_code=404, detail="Not found")
    try:
        png_bytes = brand_marks.render_png(spec)
    except Exception:
        logger.exception("Failed to render brand mark (spec=%s)", spec)
        raise HTTPException(status_code=500, detail="Could not render the logo.")
    return Response(
        content=png_bytes,
        media_type="image/png",
        headers={"Cache-Control": "public, max-age=31536000, immutable"},
    )


# ---------------------------------------------------------------------------
# In-portal notifications (topbar bell)
# ---------------------------------------------------------------------------

class EnterpriseNotificationsReadRequest(BaseModel):
    ids: Optional[list[int]] = None
    all: bool = False


@router.get("/notifications")
def enterprise_notifications_list(
    request: Request,
    limit: int = 30,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """The signed-in member's notifications (newest first) + unread count."""
    _, organization, _role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="notifications.view")
    take = max(1, min(int(limit or 30), 50))
    rows = (
        db.query(models.EnterpriseNotification)
        .filter(
            models.EnterpriseNotification.organization_id == organization.id,
            models.EnterpriseNotification.recipient_user_id == current_user.id,
        )
        .order_by(models.EnterpriseNotification.created_at.desc(), models.EnterpriseNotification.id.desc())
        .limit(take)
        .all()
    )
    unread = (
        db.query(func.count(models.EnterpriseNotification.id))
        .filter(
            models.EnterpriseNotification.organization_id == organization.id,
            models.EnterpriseNotification.recipient_user_id == current_user.id,
            models.EnterpriseNotification.is_read.is_(False),
        )
        .scalar() or 0
    )
    member_names = _org_member_name_map(db, organization.id)
    return {
        "unread_count": int(unread),
        "notifications": [
            {
                "id": n.id,
                "type": n.type,
                "title": n.title,
                "body": n.body,
                "actor_name": member_names.get(n.actor_user_id) if n.actor_user_id else None,
                "reference_type": n.reference_type,
                "reference_id": n.reference_id,
                "is_read": bool(n.is_read),
                "created_at": _iso(n.created_at),
            }
            for n in rows
        ],
    }


@router.post("/notifications/read")
def enterprise_notifications_mark_read(
    payload: EnterpriseNotificationsReadRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Mark the given notifications (or all of them) as read for this member."""
    _, organization, _role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="notifications.view")
    q = db.query(models.EnterpriseNotification).filter(
        models.EnterpriseNotification.organization_id == organization.id,
        models.EnterpriseNotification.recipient_user_id == current_user.id,
        models.EnterpriseNotification.is_read.is_(False),
    )
    if not payload.all:
        ids = [int(i) for i in (payload.ids or []) if i]
        if not ids:
            return {"updated": 0}
        q = q.filter(models.EnterpriseNotification.id.in_(ids))
    updated = q.update({models.EnterpriseNotification.is_read: True}, synchronize_session=False)
    db.commit()
    return {"updated": int(updated)}


@router.post("/team/users")
def enterprise_team_add_user(
    payload: EnterpriseTeamAddUserRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db,
        user=current_user,
        request=request,
        require_capability="team.manage",
    )
    _enforce_rate_limit_or_429(
        request=request,
        scope="enterprise.team_invite",
        limit=ENTERPRISE_TEAM_INVITE_RATE_LIMIT,
        window_seconds=ENTERPRISE_TEAM_INVITE_RATE_WINDOW_SECONDS,
        extra_key=f"org:{organization.id}:user:{current_user.id}",
    )

    target_email = (payload.email or "").strip().lower()
    if not target_email:
        raise HTTPException(status_code=400, detail="Email is required.")

    target_role = _parse_enterprise_role_or_400(payload.role)

    user = db.query(models.User).filter(models.User.email == target_email).first()

    # Pre-check the seat limit when this invite would consume a NEW active seat.
    existing_membership = None
    if user:
        existing_membership = (
            db.query(models.EnterpriseOrganizationMember)
            .filter(
                models.EnterpriseOrganizationMember.organization_id == organization.id,
                models.EnterpriseOrganizationMember.user_id == user.id,
            )
            .first()
        )
    consumes_new_seat = not (existing_membership and existing_membership.is_active)
    if consumes_new_seat:
        billing.enforce_seat_limit_or_402(db, organization.id)

    created_user = False
    if not user:
        # The invitee sets their own password via the emailed one-time link, so we
        # seed a random unusable hash. Access is granted by their org membership.
        fallback_name = target_email.split("@")[0].replace(".", " ").replace("_", " ").title()
        user = models.User(
            email=target_email,
            username=None,
            hashed_password=get_password_hash(secrets.token_urlsafe(24)),
            full_name=(payload.full_name or fallback_name or "Enterprise User").strip(),
            university=organization.company_name,
            is_active=True,
            email_verified=True,
            # Account created BY the org for a brand-new teammate → enterprise-origin, so it's
            # blocked from the B2C consumer app. (When the invite REUSES an existing user row
            # below, we deliberately leave the flag untouched so a prior B2C user keeps access.)
            is_enterprise_account=True,
            accepted_terms_privacy_at=datetime.utcnow(),
        )
        db.add(user)
        db.flush()
        created_user = True
    else:
        if payload.full_name and not user.full_name:
            user.full_name = payload.full_name.strip()
        if not user.is_active:
            user.is_active = True
    if existing_membership:
        existing_membership.is_active = True
        existing_membership.role = target_role
        existing_membership.invited_by_user_id = current_user.id
        existing_membership.status = team_svc.MEMBER_STATUS_ACTIVE
        existing_membership.deactivated_at = None
        existing_membership.invited_at = existing_membership.invited_at or datetime.utcnow()
        membership_row = existing_membership
    else:
        membership_row = models.EnterpriseOrganizationMember(
            organization_id=organization.id,
            user_id=user.id,
            role=target_role,
            is_active=True,
            invited_by_user_id=current_user.id,
            status=team_svc.MEMBER_STATUS_ACTIVE,
            invited_at=datetime.utcnow(),
        )
        db.add(membership_row)
    db.flush()

    # Role, record scope and offices, all validated against what the INVITER may hand out (they
    # can't grant a capability they lack, or widen someone's scope past their own — otherwise an
    # invite is a way to mint a second, more powerful identity for yourself).
    invite_access = {
        k: v for k, v in {
            "role_key": payload.role_key,
            "custom_role_id": payload.custom_role_id,
            "data_scope": payload.data_scope,
            "branch_ids": payload.branch_ids,
            "primary_branch_id": payload.primary_branch_id,
        }.items() if v is not None
    }
    if not invite_access.get("role_key") and not invite_access.get("custom_role_id"):
        invite_access["role_key"] = access.LEGACY_ROLE_TO_ROLE_KEY.get(target_role, access.ROLE_VIEWER)
    try:
        team_svc.update_member_access(
            db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
            membership=membership_row, data=invite_access,
        )
    except HTTPException:
        db.rollback()
        raise
    if payload.job_title or payload.phone:
        team_svc.update_member_profile(
            db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
            membership=membership_row, user=user,
            data={"job_title": payload.job_title, "phone": payload.phone},
        )
    if not membership_row.primary_branch_id:
        membership_row.primary_branch_id = team_svc.ensure_default_branch(
            db, organization.id, actor_user_id=current_user.id,
            company_name=organization.company_name,
        ).id

    # Every invite issues a fresh one-time password setup link for this recipient.
    password_setup_token = generate_verification_token()
    user.password_reset_token = hash_token(password_setup_token)
    user.password_reset_token_expires = datetime.utcnow() + timedelta(
        hours=ENTERPRISE_INVITE_PASSWORD_SETUP_EXPIRES_HOURS
    )

    db.commit()

    notif.notify_org(
        db, organization.id, type="member_added",
        title=f"{current_user.full_name or current_user.email} added {user.full_name or target_email} to the team",
        body=f"Role: {target_role}",
        actor_user_id=current_user.id, reference_type="team", commit=True,
    )

    portal_url = _build_enterprise_portal_url(organization.subdomain_slug, request)
    password_setup_url = (
        f"{ENTERPRISE_PASSWORD_SETUP_BASE_URL}/reset-password"
        f"?token={quote(password_setup_token, safe='')}"
    )
    invite_email_sent = False
    try:
        invite_email_sent = send_enterprise_team_invite_email(
            invitee_email=target_email,
            invitee_name=user.full_name,
            organization_name=organization.company_name,
            role=target_role,
            portal_url=portal_url,
            set_password_url=password_setup_url,
            password_setup_expires_hours=ENTERPRISE_INVITE_PASSWORD_SETUP_EXPIRES_HOURS,
            invited_by_name=current_user.full_name,
            invited_by_email=current_user.email,
        )
    except Exception:
        logger.exception(
            "Failed to send enterprise invite email (org_id=%s, invitee=%s)",
            organization.id,
            target_email,
        )

    base_message = (
        "User added to organization."
        if not created_user
        else "User account created and added to organization."
    )
    if invite_email_sent:
        response_message = f"{base_message} Invitation email sent."
    else:
        response_message = (
            f"{base_message} Invite email could not be sent right now. "
            "Ask the user to request a fresh password setup link from the login screen."
        )

    return {
        "message": response_message,
        "created_user": created_user,
        "invite_email_sent": invite_email_sent,
        "invite_email_to": target_email,
        "organization_portal_url": portal_url,
        "members": _list_organization_members(db, organization.id, ctx=role.ctx),
    }


@router.patch("/team/users/{member_user_id}/role")
def enterprise_team_update_role(
    member_user_id: int,
    payload: EnterpriseTeamRoleUpdateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """LEGACY three-option role control. Retired from the SPA; kept for older clients.

    SECURITY: this used to write `membership.role_key` directly on nothing but a
    `team.manage` check, bypassing the whole escalation core in enterprise_team — so an
    office coordinator holding only the delegable `team.manage` grant could promote
    themselves to Admin in one request, with no audit row. It now requires `roles.manage`
    (as the supported /access endpoint does) and delegates the write to
    `team_svc.update_member_access`, inheriting assert_not_self, assert_owner_protected,
    assert_can_grant, assert_scope_allowed and assert_admin_remains, plus the audit rows
    and notifications. Nothing here may write a membership row on its own again.
    """
    _, organization, role = _require_enterprise_membership(
        db=db,
        user=current_user,
        request=request,
        require_capability="roles.manage",
    )

    target_role = _parse_enterprise_role_or_400(payload.role)
    membership = (
        db.query(models.EnterpriseOrganizationMember)
        .filter(
            models.EnterpriseOrganizationMember.organization_id == organization.id,
            models.EnterpriseOrganizationMember.user_id == int(member_user_id),
            models.EnterpriseOrganizationMember.is_active.is_(True),
        )
        .first()
    )
    if not membership:
        raise HTTPException(status_code=404, detail="Organization member not found.")

    # This is the LEGACY three-option control. It can only express admin / editor / viewer, so it
    # refuses to touch anyone on a richer role — otherwise an admin looking at the old dropdown
    # would see a Finance member rendered as "Viewer", pick "Editor" to "fix" it, and silently
    # convert them to a Counsellor, wiping their finance access and any per-person overrides. A
    # stray scroll-wheel over a focused <select> is enough to fire it.
    current_role_key = access.normalize_role_key(
        getattr(membership, "role_key", None), getattr(membership, "role", None)
    )
    if current_role_key not in set(access.LEGACY_ROLE_TO_ROLE_KEY.values()):
        raise HTTPException(
            status_code=409,
            detail=(
                f"{(target_user_name(db, member_user_id) or 'This member')} is on an advanced role. "
                "Use Edit access to change what they can do."
            ),
        )

    # Delegate the write. update_member_access runs the full guard chain (self-service,
    # owner-protection, grant-escalation, scope clamp, last-admin-remains), recomputes the
    # `role` mirror and role_key together, resets the scope override to the new role's own
    # default, and writes the audit rows — none of which this endpoint used to do.
    new_role_key = access.LEGACY_ROLE_TO_ROLE_KEY.get(target_role, access.ROLE_VIEWER)
    team_svc.update_member_access(
        db,
        organization=organization,
        actor_user=current_user,
        actor_ctx=role.ctx,
        membership=membership,
        data={"role_key": new_role_key},
    )
    db.commit()
    return {
        "message": "Role updated successfully.",
        "members": _list_organization_members(db, organization.id, ctx=role.ctx),
    }


def target_user_name(db: Session, user_id: int) -> str | None:
    row = db.query(models.User.full_name, models.User.email).filter(models.User.id == int(user_id)).first()
    if not row:
        return None
    return (row[0] or row[1] or None)


@router.delete("/team/users/{member_user_id}")
def enterprise_team_remove_user(
    member_user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db,
        user=current_user,
        request=request,
        require_capability="team.manage",
    )

    if int(member_user_id) == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot remove your own account.")

    membership = (
        db.query(models.EnterpriseOrganizationMember)
        .filter(
            models.EnterpriseOrganizationMember.organization_id == organization.id,
            models.EnterpriseOrganizationMember.user_id == int(member_user_id),
            models.EnterpriseOrganizationMember.is_active.is_(True),
        )
        .first()
    )
    if not membership:
        raise HTTPException(status_code=404, detail="Organization member not found.")
    user = db.query(models.User).filter(models.User.id == int(member_user_id)).first()
    if not user:
        raise HTTPException(status_code=404, detail="Organization member not found.")

    # Delegates to the same offboarding path as the newer /deactivate endpoint rather than just
    # flipping is_active. That flip on its own left the owner removable, the removed person's
    # session live, their caseload silently orphaned and nothing in the audit trail. This route
    # cannot pass a successor, so it will 409 with the counts when the member still owns records —
    # the old client surfaces that message in a toast, which is a far better outcome than quietly
    # detaching thirty clients from the only person who knew about them.
    result = team_svc.deactivate_member(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        membership=membership, user=user,
    )
    db.commit()
    return {
        "message": "User removed from organization.",
        "members": _list_organization_members(db, organization.id, ctx=role.ctx),
        **result,
    }


# ===========================================================================
# CRM: catalog, clients, notes, client emails, dashboard
# ===========================================================================

ENTERPRISE_CLIENT_NAME_MAX = 160
ENTERPRISE_NOTE_MAX = 5000
ENTERPRISE_EMAIL_SUBJECT_MAX = 200
ENTERPRISE_EMAIL_BODY_MAX = 20000
# The rich-text version of the same message: markup makes it several times longer
# than its plain-text rendition, so it gets its own (larger) ceiling.
ENTERPRISE_EMAIL_HTML_MAX = 120000
# Attachment ceilings. Resend caps a whole message at 40 MB and base64 inflates
# payloads by ~33%, so the total here stays well under that even with the HTML body.
ENTERPRISE_EMAIL_ATTACH_MAX_FILES = int(os.getenv("ENTERPRISE_EMAIL_ATTACH_MAX_FILES", "10"))
ENTERPRISE_EMAIL_ATTACH_MAX_TOTAL_BYTES = int(
    os.getenv("ENTERPRISE_EMAIL_ATTACH_MAX_TOTAL_BYTES", str(15 * 1024 * 1024))
)
# Draft attachments uploaded but never sent are swept on the next upload.
ENTERPRISE_EMAIL_ATTACH_DRAFT_TTL_HOURS = int(
    os.getenv("ENTERPRISE_EMAIL_ATTACH_DRAFT_TTL_HOURS", "48")
)
ENTERPRISE_SIGNUP_RATE_LIMIT = int(os.getenv("ENTERPRISE_SIGNUP_RATE_LIMIT", "6"))
ENTERPRISE_SIGNUP_RATE_WINDOW_SECONDS = int(os.getenv("ENTERPRISE_SIGNUP_RATE_WINDOW_SECONDS", "900"))
ENTERPRISE_BULK_EMAIL_MAX_RECIPIENTS = int(os.getenv("ENTERPRISE_BULK_EMAIL_MAX_RECIPIENTS", "200"))

# Version (Last Updated date) of the Terms & Conditions / Privacy Policy a user
# accepts at signup. Defined centrally in app.legal so B2C, OAuth, and enterprise
# signups all record the same value. Keep app.legal in sync with
# LEGAL_LAST_UPDATED.terms / .privacy in static/app.js.
from app.legal import LEGAL_TERMS_PRIVACY_VERSION, LEGAL_DPA_VERSION, FINANCE_ATTESTATION_VERSION


class EnterpriseSignupRequest(BaseModel):
    company_name: str = Field(..., min_length=2, max_length=120)
    subdomain_slug: str = Field(
        ...,
        min_length=ENTERPRISE_SUBDOMAIN_MIN_LENGTH,
        max_length=ENTERPRISE_SUBDOMAIN_MAX_LENGTH,
    )
    full_name: str = Field(..., min_length=2, max_length=120)
    email: EmailStr
    password: str = Field(..., min_length=8, max_length=200)
    accepted_terms_privacy: bool = False
    accepted_dpa: bool = False
    marketing_emails_consent: bool = False
    cf_turnstile_token: Optional[str] = None
    # 6-digit email-verification code from /signup/send-code (required to create the workspace).
    email_otp: Optional[str] = Field(default=None, max_length=10)


class EnterpriseClientIntakeFields(BaseModel):
    """The lead-intake record a consultancy keeps on every client — shared verbatim by the
    create and update payloads so the two can never drift. Every field is optional: only
    name, destination and visa type are ever required."""
    whatsapp_number: Optional[str] = Field(default=None, max_length=40)
    current_city: Optional[str] = Field(default=None, max_length=80)
    gender: Optional[str] = Field(default=None, max_length=30)
    guardian_name: Optional[str] = Field(default=None, max_length=120)
    guardian_relation: Optional[str] = Field(default=None, max_length=30)
    guardian_phone: Optional[str] = Field(default=None, max_length=40)
    study_level: Optional[str] = Field(default=None, max_length=40)
    field_of_study: Optional[str] = Field(default=None, max_length=120)
    admission_stage: Optional[str] = Field(default=None, max_length=40)
    prior_refusal_history: Optional[str] = Field(default=None, max_length=40)
    prior_refusal_notes: Optional[str] = Field(default=None, max_length=2000)
    highest_qualification: Optional[str] = Field(default=None, max_length=40)
    qualification_score: Optional[str] = Field(default=None, max_length=20)
    qualification_scale: Optional[str] = Field(default=None, max_length=30)
    year_of_passing: Optional[str] = Field(default=None, max_length=8)
    backlogs_count: Optional[str] = Field(default=None, max_length=4)
    work_experience_band: Optional[str] = Field(default=None, max_length=20)
    english_test_status: Optional[str] = Field(default=None, max_length=30)
    english_test_type: Optional[str] = Field(default=None, max_length=30)
    english_test_score: Optional[str] = Field(default=None, max_length=20)
    english_test_date: Optional[str] = Field(default=None, max_length=20)
    aptitude_test_type: Optional[str] = Field(default=None, max_length=20)
    aptitude_test_score: Optional[str] = Field(default=None, max_length=20)
    budget_band: Optional[str] = Field(default=None, max_length=30)
    funding_source: Optional[str] = Field(default=None, max_length=30)
    lead_source: Optional[str] = Field(default=None, max_length=40)
    lead_source_detail: Optional[str] = Field(default=None, max_length=120)
    branch_name: Optional[str] = Field(default=None, max_length=80)
    next_followup_date: Optional[str] = Field(default=None, max_length=20)
    # Promotional-contact opt-in, per channel. Sent as a list of channel keys.
    marketing_consent_channels: Optional[List[str]] = None
    # Consent to share the profile with universities / partner institutions abroad.
    institution_share_consent: Optional[bool] = None


class EnterpriseClientCreateRequest(EnterpriseClientIntakeFields):
    full_name: str = Field(..., min_length=2, max_length=ENTERPRISE_CLIENT_NAME_MAX)
    visa_category: Optional[str] = Field(default="student", max_length=40)
    destination_country_code: str = Field(..., min_length=2, max_length=12)
    visa_type: str = Field(..., min_length=2, max_length=160)
    intake: Optional[str] = Field(default=None, max_length=120)
    email: Optional[EmailStr] = None
    phone: Optional[str] = Field(default=None, max_length=40)
    nationality: Optional[str] = Field(default=None, max_length=80)
    date_of_birth: Optional[str] = Field(default=None, max_length=20)
    passport_number: Optional[str] = Field(default=None, max_length=60)
    passport_expiry: Optional[str] = Field(default=None, max_length=20)
    priority: Optional[str] = Field(default=None, max_length=20)
    status: Optional[str] = Field(default=None, max_length=30)
    target_date: Optional[str] = Field(default=None, max_length=20)
    application_reference: Optional[str] = Field(default=None, max_length=120)
    assigned_to_user_id: Optional[int] = None
    # Which office owns this case. Offices are real records now; `branch_name` above is the
    # server-written display snapshot and is ignored on the way in.
    branch_id: Optional[int] = None
    initial_note: Optional[str] = Field(default=None, max_length=ENTERPRISE_NOTE_MAX)
    # Staff attestation that the client consented to having their data processed
    # through Rilono. Enforced in the UI; recorded here as proof-of-consent.
    client_consent_confirmed: bool = False
    # Set by the second submit, after the duplicate warning has been shown and read. The
    # check still runs — the answer is used to leave a note on the new file instead of
    # refusing it. See _assert_not_duplicate_client.
    allow_duplicate: bool = False


class EnterpriseClientUpdateRequest(EnterpriseClientIntakeFields):
    full_name: Optional[str] = Field(default=None, min_length=2, max_length=ENTERPRISE_CLIENT_NAME_MAX)
    visa_category: Optional[str] = Field(default=None, max_length=40)
    destination_country_code: Optional[str] = Field(default=None, max_length=12)
    visa_type: Optional[str] = Field(default=None, max_length=160)
    intake: Optional[str] = Field(default=None, max_length=120)
    email: Optional[str] = Field(default=None, max_length=200)
    phone: Optional[str] = Field(default=None, max_length=40)
    nationality: Optional[str] = Field(default=None, max_length=80)
    date_of_birth: Optional[str] = Field(default=None, max_length=20)
    passport_number: Optional[str] = Field(default=None, max_length=60)
    passport_expiry: Optional[str] = Field(default=None, max_length=20)
    priority: Optional[str] = Field(default=None, max_length=20)
    status: Optional[str] = Field(default=None, max_length=30)
    target_date: Optional[str] = Field(default=None, max_length=20)
    application_reference: Optional[str] = Field(default=None, max_length=120)
    assigned_to_user_id: Optional[int] = None
    branch_id: Optional[int] = None
    # As on the create payload: proceed even though the edited identity now collides
    # with another file.
    allow_duplicate: bool = False
    # Concurrency baseline. `expected_values` is {field: value-as-rendered} for the fields
    # being written — the precise check, which lets two people edit different fields of the
    # same client without colliding. `expected_version` is the coarse whole-row fallback for
    # a caller that cannot produce a baseline. Both optional; see _assert_client_write_is_current.
    expected_values: Optional[dict] = None
    expected_version: Optional[int] = None
    # Set once the user has seen the conflict banner and chosen to save regardless. Kept
    # separate from `expected_version` so "I never sent a version" and "I saw the conflict
    # and decided" stay distinguishable in the logs and in intent.
    force_overwrite: bool = False


class EnterpriseClientStatusRequest(BaseModel):
    status: str = Field(..., min_length=2, max_length=30)
    # The stage the caller's screen was showing. A move is refused only when somebody else
    # has ALREADY moved the case since — an unrelated edit to the same client is not a
    # conflict with a stage move.
    expected_status: Optional[str] = None
    expected_version: Optional[int] = None


class EnterpriseClientNoteRequest(BaseModel):
    body: str = Field(..., min_length=1, max_length=ENTERPRISE_NOTE_MAX)


class EnterpriseClientEmailRequest(BaseModel):
    subject: str = Field(..., min_length=1, max_length=ENTERPRISE_EMAIL_SUBJECT_MAX)
    body: str = Field(..., min_length=1, max_length=ENTERPRISE_EMAIL_BODY_MAX)
    # Rich-text body from the composer. Sanitized server-side before it is stored or
    # sent — `body` remains the authoritative plain-text rendition either way.
    body_html: Optional[str] = Field(default=None, max_length=ENTERPRISE_EMAIL_HTML_MAX)
    # Draft attachments already uploaded via /clients/{id}/email/attachments.
    attachment_ids: list[int] = Field(default_factory=list, max_length=100)
    # Documents already on file for this client, attached by reference (copied on send).
    document_ids: list[int] = Field(default_factory=list, max_length=100)


class EnterpriseBulkEmailRequest(BaseModel):
    client_ids: list[int] = Field(..., min_length=1)
    subject: str = Field(..., min_length=1, max_length=ENTERPRISE_EMAIL_SUBJECT_MAX)
    body: str = Field(..., min_length=1, max_length=ENTERPRISE_EMAIL_BODY_MAX)


class EnterpriseBillingCheckoutRequest(BaseModel):
    plan: str = Field(..., min_length=2, max_length=30)
    billing_cycle: str = Field(default="monthly", max_length=12)
    coupon_code: Optional[str] = Field(default=None, max_length=40)


class EnterpriseBillingVerifyRequest(BaseModel):
    razorpay_order_id: str = Field(..., min_length=6, max_length=64)
    razorpay_payment_id: str = Field(..., min_length=6, max_length=64)
    razorpay_signature: str = Field(..., min_length=6, max_length=256)


class EnterpriseCreditTopupRequest(BaseModel):
    package: str = Field(..., min_length=2, max_length=30)
    coupon_code: Optional[str] = Field(default=None, max_length=40)
    # Presentment currency HINT. The server maps it to a price from app/money.py; the
    # client never sends an amount. See _resolve_charge_currency.
    currency: Optional[str] = Field(default=None, max_length=3)


class EnterpriseInfraCheckoutRequest(BaseModel):
    currency: Optional[str] = Field(default=None, max_length=3)


class EnterpriseCouponValidateRequest(BaseModel):
    code: str = Field(..., min_length=1, max_length=40)
    context: str = Field(default="credits", max_length=12)  # credits | billing
    package: Optional[str] = Field(default=None, max_length=30)
    plan: Optional[str] = Field(default=None, max_length=30)
    billing_cycle: Optional[str] = Field(default="monthly", max_length=12)


class EnterpriseCreditVerifyRequest(BaseModel):
    razorpay_order_id: str = Field(..., min_length=6, max_length=64)
    razorpay_payment_id: str = Field(..., min_length=6, max_length=64)
    razorpay_signature: str = Field(..., min_length=6, max_length=256)


def _parse_iso_date_or_400(
    value: Optional[str], field_label: str, *, direction: Optional[str] = None
) -> Optional[date]:
    """Parse and BOUND a staff-entered date. Every client-record and calendar date lands here.

    The sanity window is judged against the server's date rather than the org's: it spans 25
    years, so a zone offset can't move a date across it, and this is called from paths that
    have no organization in hand. `direction` is for the rare field whose direction is
    unambiguous (a date of birth) — see `enterprise_dates` for why most fields get none.
    """
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    try:
        parsed = datetime.strptime(raw[:10], "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail=f"{field_label} must be a valid date (YYYY-MM-DD).")
    return ent_dates.validate(parsed, field_label, today=date.today(), direction=direction)


def _serialize_subscription_state(state: dict) -> dict:
    return {
        "plan": state["plan"],
        "plan_label": state["plan_label"],
        "status": state["status"],
        "is_trial": state["is_trial"],
        "is_sandbox": state.get("is_sandbox", state["is_trial"]),
        "trial_expired": state["trial_expired"],
        "trial_days_left": state["trial_days_left"],
        "trial_ends_at": state["trial_ends_at"],
        "current_period_end": state["current_period_end"],
        "max_clients": state["max_clients"],
        "max_seats": state["max_seats"],
        "clients_used": state["clients_used"],
        "seats_used": state["seats_used"],
        "can_add_client": state["can_add_client"],
        "can_add_seat": state["can_add_seat"],
        "grandfathered": state.get("grandfathered", False),
        "grace_ends_at": state.get("grace_ends_at"),
        "over_cap": state.get("over_cap", False),
        # A paid tier whose period ran out. Without these the UI shows a lapsed customer a
        # fresh sandbox and never tells them why they lost their seats and credits.
        "plan_lapsed": state.get("plan_lapsed", False),
        "auto_renews": state.get("auto_renews", False),
        "has_mandate": state.get("has_mandate", False),
        "cancel_at_period_end": state.get("cancel_at_period_end", False),
        "mandate_status": state.get("mandate_status"),
        "lapsed_plan": state.get("lapsed_plan"),
        "lapsed_plan_label": state.get("lapsed_plan_label"),
        "lapsed_at": state.get("lapsed_at"),
        # Tier economics, so the plan chip and the credits panel need no second request.
        "included_credits": state.get("included_credits", 0),
        "credits_recur": state.get("credits_recur", False),
        "currency": state.get("currency", billing.CURRENCY),
        "monthly_minor": state.get("monthly_minor", 0),
        "monthly_display": state.get("monthly_display"),
        "tax_label": state.get("tax_label"),
        "tax_percent": state.get("tax_percent"),
        "total_minor": state.get("total_minor", 0),
        "total_display": state.get("total_display"),
        "is_free_platform": state.get("is_free_platform", False),
    }


def _country_brief(country_code: str | None) -> dict:
    country = catalog.get_country(country_code)
    if not country:
        return {
            "code": str(country_code or "").upper(),
            "name": str(country_code or "").upper(),
            "flag_emoji": "🌐",
            "landmark": "",
            "gradient_from": "#6366f1",
            "gradient_to": "#8b5cf6",
            "accent": "#eef2ff",
        }
    return {
        "code": country["code"],
        "name": country["name"],
        "flag_emoji": country["flag_emoji"],
        "landmark": country["landmark"],
        "gradient_from": country["gradient_from"],
        "gradient_to": country["gradient_to"],
        "accent": country["accent"],
    }


def _stage_brief(status_key: str | None, country_code: str | None = None) -> dict:
    # Worded for the client's own destination — a UAE case's stage 3 reads "Entry Permit
    # Filed", not "Application Submitted". Falls back to the generic wording.
    stage = (catalog.stage_brief(country_code, status_key)
             or catalog.CLIENT_STAGE_MAP.get(catalog.normalize_stage(status_key)))
    return {
        "key": stage["key"],
        "label": stage["label"],
        "color": stage["color"],
        "is_open": stage["is_open"],
        "is_terminal": stage["is_terminal"],
    }


def _category_label(category_key: str | None) -> str:
    item = catalog.VISA_CATEGORY_MAP.get(str(category_key or "").strip().lower())
    return item["label"] if item else str(category_key or "").title()


def _org_member_name_map(db: Session, organization_id: int) -> dict[int, str]:
    rows = (
        db.query(models.EnterpriseOrganizationMember.user_id, models.User.full_name, models.User.email)
        .join(models.User, models.User.id == models.EnterpriseOrganizationMember.user_id)
        .filter(
            models.EnterpriseOrganizationMember.organization_id == int(organization_id),
            models.EnterpriseOrganizationMember.is_active.is_(True),
        )
        .all()
    )
    return {int(uid): (full_name or email or "Team member") for uid, full_name, email in rows}


def _is_active_org_member(db: Session, organization_id: int, user_id: int) -> bool:
    return bool(
        db.query(models.EnterpriseOrganizationMember.id)
        .filter(
            models.EnterpriseOrganizationMember.organization_id == int(organization_id),
            models.EnterpriseOrganizationMember.user_id == int(user_id),
            models.EnterpriseOrganizationMember.is_active.is_(True),
        )
        .first()
    )


def _iso(value) -> Optional[str]:
    if value is None:
        return None
    try:
        return value.isoformat()
    except Exception:
        return str(value)


def _stage_key_or_400(raw_stage: str | None) -> str:
    """Validate a caller-supplied pipeline stage, or 400 naming the value.

    Never normalize_stage() on a WRITE path: it coerces anything it doesn't recognise to
    new_lead, so a stale browser tab or an importer sending a retired key would silently
    rewind the case to the top of the pipeline, and there is no stage history to detect or
    reverse it with. Reads still normalize — a row must render whatever it holds."""
    key = str(raw_stage or "").strip().lower()
    if key not in catalog.CLIENT_STAGE_KEYS:
        raise HTTPException(
            status_code=400,
            detail=f"'{str(raw_stage or '').strip()}' is not a pipeline stage. Refresh the page and try again.",
        )
    return key


# Where the ADMISSIONS milestone on the intake record lands when a case is MOVED to each
# pipeline stage. Moving the case to a university stage is the counselor stating that the
# phase has been reached, so the milestone follows the move (see _apply_status_change) —
# but only on the move. The column stays staff-editable in its own right, because it also
# records what the pipeline cannot say: a walk-in already holding an admit whose visa case
# has legitimately not started yet.
#
# Keyed by stage KEY, never by position: inserting a stage must not re-map the others.
# Everything from the visa phase onwards holds at doc_in_hand — the visa pipeline says
# nothing further about admissions, and inventing a value would mean a key no other surface
# knows. A stage absent from this map (on_hold) leaves the recorded value untouched.
#
# applications_sent is the one stage that does not settle on its mapped value: it spans both
# "applied" and "admitted" (an admit is in hand, the offer is not accepted yet), and the
# shortlist rows already record which it is. The value here is the no-shortlist floor —
# _admission_stage_for() upgrades it, so "admitted" stays reachable.
_ADMISSION_STAGE_BY_PIPELINE_STAGE = {
    catalog.STAGE_NEW_LEAD: "exploring",
    catalog.STAGE_SHORTLISTING: "shortlisting",
    catalog.STAGE_APPLICATIONS_SENT: "applied",
    catalog.STAGE_OFFER_ACCEPTED: "offer_accepted",
    catalog.STAGE_DOCUMENTS: "doc_in_hand",
    catalog.STAGE_SUBMITTED: "doc_in_hand",
    catalog.STAGE_APPOINTMENT: "doc_in_hand",
    catalog.STAGE_DECISION: "doc_in_hand",
    catalog.STAGE_APPROVED: "doc_in_hand",
    catalog.STAGE_REJECTED: "doc_in_hand",
}


def _has_admitted_university(db: Session, client: models.EnterpriseClient) -> bool:
    """Whether any university on this client's shortlist has come back with an admit."""
    client_id = getattr(client, "id", None)
    if not client_id:
        return False
    return (
        db.query(models.EnterpriseClientUniversity.id)
        .filter(
            models.EnterpriseClientUniversity.client_id == int(client_id),
            models.EnterpriseClientUniversity.status == "admitted",
        )
        .first()
    ) is not None


def _admission_stage_for(
    new_status: str,
    client: models.EnterpriseClient,
    db: Session,
) -> Optional[str]:
    """The admissions milestone implied by a pipeline stage, or None if the stage implies
    nothing and the recorded value should stand.

    `db` is required, not optional: the applications stage has to read the shortlist, and a
    caller who omitted the session would silently downgrade an admit to "applied"."""
    derived = _ADMISSION_STAGE_BY_PIPELINE_STAGE.get(new_status)
    # Applications are out, but an admit may already be in hand and simply not accepted —
    # which is exactly what "admitted" records, and the only place that fact lives.
    if new_status == catalog.STAGE_APPLICATIONS_SENT and _has_admitted_university(db, client):
        return "admitted"
    return derived


def _apply_status_change(
    client: models.EnterpriseClient,
    new_status: str,
    db: Session,
    *,
    creating: bool = False,
) -> None:
    """Set client.status while maintaining held_from_status: putting a case On Hold
    remembers the stage it was held FROM (so the UI can show its real position and
    offer one-click Resume); moving to any other stage clears the marker.

    An ACTUAL move — old stage != new stage — also advances the admissions milestone to
    match, because moving the case is the counselor stating the phase has been reached. A
    request that re-sends the stage the client already holds is not a move and leaves the
    milestone exactly as recorded: `admission_stage` is staff-owned, it carries facts no
    pipeline stage implies, and the Edit-client form sends the stage on every save, so
    re-deriving here would rewrite the record on an unrelated edit.

    Every arrival is also stamped into stage_visits, which is what lets the journey tracker
    distinguish a stage the case worked through from one it jumped over. That stamp is taken
    on EVERY call, including the re-send that isn't a move — the point is to record where the
    case has stood, and an unrelated edit is still evidence it stands there now.

    `creating` marks the opening stage of a brand-new client, where the stages before it were
    never worked rather than merely unrecorded — see _record_stage_visit.

    `db` is required — see _admission_stage_for()."""
    old = client.status
    old_held = getattr(client, "held_from_status", None)
    if new_status == catalog.STAGE_ON_HOLD:
        if old != catalog.STAGE_ON_HOLD:
            client.held_from_status = old if (old in catalog.CLIENT_STAGE_KEYS and old != catalog.STAGE_ON_HOLD) else None
    else:
        client.held_from_status = None
    client.status = new_status
    _record_stage_visit(
        client, new_status, previous=old, previous_held=old_held, creating=creating,
    )
    if old == new_status:
        return

    derived = _admission_stage_for(new_status, client, db)
    # The manual-only values (deferred) describe a state no stage implies, so derivation
    # must never overwrite one — a deferred case still sits at the stage it reached.
    if derived and not client_fields.is_manual_choice("admission_stage", client.admission_stage):
        client.admission_stage = derived


def _parse_stage_data(client: models.EnterpriseClient) -> tuple[dict, bool]:
    """(stored, unreadable) — the stage-record object exactly as stored, and whether the
    column could not be read as an object at all.

    Write paths must use this rather than _load_stage_data: an unreadable column parses to
    an empty dict, and building the next save on that would replace EVERY stage's records
    with the one bucket being written. `stored` is returned unfiltered so buckets this code
    doesn't recognise are carried through a save instead of dropped."""
    raw = getattr(client, "stage_data", None)
    if not raw:
        return {}, False
    try:
        parsed = json.loads(raw)
    except Exception:
        logger.error(
            "Client stage_data is not valid JSON (client_id=%s) — stored value left untouched",
            getattr(client, "id", None), exc_info=True,
        )
        return {}, True
    if not isinstance(parsed, dict):
        logger.error(
            "Client stage_data is not an object (client_id=%s, type=%s) — stored value left untouched",
            getattr(client, "id", None), type(parsed).__name__,
        )
        return {}, True
    return parsed, False


def _load_stage_data(client: models.EnterpriseClient) -> dict:
    """Parse the client's per-stage record JSON. Always returns a dict of dicts, so a
    damaged column still renders as "nothing recorded" rather than breaking the page."""
    stored, _unreadable = _parse_stage_data(client)
    return {k: v for k, v in stored.items() if isinstance(v, dict)}


# --- Stage progress: which stages a case actually worked through -------------------------
# The pipeline stage says where a case IS. On its own it cannot say how it got there, and
# "everything before the current stage is complete" is a guess that reads as a fact — a case
# opened straight at Collecting Documents, or jumped from New Lead to Awaiting Decision, has
# green ticks on stages nobody ever touched. stage_visits records the arrivals so the journey
# can mark those stages skipped instead.
#
# The one rule everything here serves: never report a skip that cannot be proved. A stage is
# skipped only when the record can speak for it AND the case went past it without working it.
# Anything older than the record is unknown, and unknown reads as it always did — complete.


def _load_stage_progress(client: models.EnterpriseClient) -> dict:
    """The case's recorded stage history: {"from": <stage_key|None>, "visits": {key: iso|None}}.

    `from` is the earliest stage the record can speak for. Whatever happened before it happened
    before the record existed, so it is UNKNOWN — never skipped. `visits` maps a stage to when
    the case first arrived there; a null timestamp means it was already standing there when the
    record began.

    An absent or damaged column returns an empty record, which callers must read as "history
    unknown" rather than "nothing was reached"."""
    blank = {"from": None, "visits": {}}
    raw = getattr(client, "stage_visits", None)
    if not raw:
        return blank
    try:
        parsed = json.loads(raw)
    except Exception:
        logger.error(
            "Client stage_visits is not valid JSON (client_id=%s) — treated as untracked",
            getattr(client, "id", None), exc_info=True,
        )
        return blank
    if not isinstance(parsed, dict):
        logger.error(
            "Client stage_visits is not an object (client_id=%s, type=%s) — treated as untracked",
            getattr(client, "id", None), type(parsed).__name__,
        )
        return blank
    visits = parsed.get("visits")
    if isinstance(visits, dict):
        started = parsed.get("from")
        return {
            "from": (str(started) if started else None),
            # Unrecognised keys are carried rather than dropped, for the same reason stage_data
            # keeps its unknown buckets: a stage retired from the catalog still happened.
            "visits": {str(k): (str(v) if v else None) for k, v in visits.items()},
        }
    # A record written before this shape existed is the bare visit map. The earliest stage it
    # names is as far back as it can honestly speak for.
    flat = {str(k): (str(v) if v else None) for k, v in parsed.items()}
    linear = _linear_stage_keys(client)
    placed = [k for k in linear if k in flat]
    return {"from": (placed[0] if placed else None), "visits": flat}


def _load_stage_visits(client: models.EnterpriseClient) -> dict:
    """{stage_key: iso timestamp | None} — when this case first arrived at each stage it stood
    at. Null where it was already there when the record began."""
    return _load_stage_progress(client)["visits"]


def _stages_reached(client: models.EnterpriseClient) -> list[str]:
    """Every stage this case demonstrably worked through: the ones it stood at, plus any stage
    a counselor filled the case record in for. That second half matters — recording a stage's
    details without ever parking the case there is real work, and work that was done must never
    be reported back to the org as skipped."""
    reached = dict.fromkeys(_load_stage_visits(client))
    for stage_key, values in _load_stage_data(client).items():
        if any(str(value or "").strip() for value in values.values()):
            reached.setdefault(stage_key, None)
    return list(reached)


def _linear_stage_keys(client: models.EnterpriseClient) -> list[str]:
    """This destination's pipeline in order, minus the two stages that sit OFF it. A refusal
    and a hold are states a case is in, not steps it walks through."""
    return [
        stage["key"]
        for stage in catalog.stages_for(getattr(client, "destination_country_code", None))
        if stage["key"] not in (catalog.STAGE_REJECTED, catalog.STAGE_ON_HOLD)
    ]


def _stage_anchor(status: str | None, held_from: str | None) -> str:
    """The linear stage a case with this status stands at. Empty when it stands off the path —
    refused, or held from a position that was never recorded."""
    key = str(status or "").strip().lower()
    if key == catalog.STAGE_ON_HOLD:
        return str(held_from or "").strip().lower()
    return key


def _stages_behind(
    client: models.EnterpriseClient, status: str | None, held_from: str | None,
) -> list[str]:
    """The stages this case has already gone PAST. Only these can have been skipped: a stage
    ahead of the case has not been passed over, it simply has not happened yet."""
    linear = _linear_stage_keys(client)
    key = str(status or "").strip().lower()
    if key == catalog.STAGE_REJECTED:
        # A refused case is closed where it stopped, and the tracker has always drawn the whole
        # pipeline behind it. Only the part up to the furthest stage it actually reached can be
        # judged: nothing beyond that was passed over — it never happened at all, and calling it
        # skipped would report a refusal as somebody's choice.
        reached = set(_stages_reached(client))
        last = max((i for i, k in enumerate(linear) if k in reached), default=-1)
        return linear[: last + 1]
    if key == catalog.STAGE_ON_HOLD:
        key = str(held_from or "").strip().lower()   # "" for a hold with no recorded position
    if key not in linear:
        return []
    return linear[: linear.index(key)]


def _skipped_stage_keys(client: models.EnterpriseClient) -> list[str]:
    """The stages this case went straight past: the ones behind it, within the span its record
    can speak for, that it never worked.

    Empty whenever the record is missing or cannot place itself — an unknown past is not a
    skip. This is the ONE derivation; the staff journey, the client portal and Deep Scan all
    render from it rather than each re-deriving what "skipped" means."""
    progress = _load_stage_progress(client)
    linear = _linear_stage_keys(client)
    started = progress["from"]
    if not progress["visits"] or started not in linear:
        return []
    floor = linear.index(started)
    reached = set(_stages_reached(client))
    behind = _stages_behind(client, client.status, getattr(client, "held_from_status", None))
    return [
        key for key in behind
        if key not in reached and key in linear and linear.index(key) >= floor
    ]


def _record_stage_visit(
    client: models.EnterpriseClient,
    stage_key: str,
    *,
    previous: str | None = None,
    previous_held: str | None = None,
    creating: bool = False,
) -> None:
    """Stamp `stage_key` as a stage this case has stood at.

    Append-only, first arrival wins: reopening a refusal or resuming a hold must not rewrite the
    date the case originally reached a stage, and moving a case backwards must never erase one
    it already worked through.

    The first write also fixes how far back the record can speak — the difference between a case
    that skipped a stage and one that simply predates the record:
      * a brand-new client (`creating`) has no past at all. It begins at its opening stage, so
        the whole pipeline is answerable and a walk-in opened halfway down really did skip
        everything above it;
      * an existing case starts from where it already stands, which also counts as reached;
      * an existing case whose position cannot be placed — a hold taken before held_from_status
        existed, or a stage key the catalog no longer carries — starts HERE. Under-reporting a
        skip is a smaller lie than inventing one.
    """
    progress = _load_stage_progress(client)
    visits = progress["visits"]
    started = progress["from"]
    if not visits:
        linear = _linear_stage_keys(client)
        if creating:
            started = linear[0] if linear else stage_key
        else:
            anchor = _stage_anchor(previous, previous_held)
            if anchor in linear:
                started = anchor
                visits.setdefault(anchor, None)
            else:
                started = stage_key if stage_key in linear else None
    # Offset-qualified on purpose: this string is rendered by the browser with new Date(), which
    # reads a bare timestamp as LOCAL time and would show the arrival a day early east of UTC.
    visits.setdefault(stage_key, _iso(datetime.now(dt_timezone.utc)))
    client.stage_visits = json.dumps({"from": started, "visits": visits})


# The intake record, grouped the way it is captured and displayed. Text/date/int fields
# are echoed verbatim; CHOICE fields store an option key from enterprise_client_fields and
# are echoed with a *_label alongside so the UI never has to own the wording.
_CLIENT_INTAKE_TEXT_FIELDS = (
    "whatsapp_number", "current_city", "guardian_name", "guardian_phone",
    "field_of_study", "prior_refusal_notes", "qualification_score",
    "english_test_score", "aptitude_test_score", "lead_source_detail", "branch_name",
)
_CLIENT_INTAKE_CHOICE_FIELDS = (
    "gender", "guardian_relation", "study_level", "admission_stage",
    "prior_refusal_history", "highest_qualification", "qualification_scale",
    "work_experience_band", "english_test_status", "english_test_type",
    "aptitude_test_type", "budget_band", "funding_source", "lead_source",
)
_CLIENT_INTAKE_DATE_FIELDS = (
    ("english_test_date", "English test date"),
    ("next_followup_date", "Next follow-up date"),
)
# (field, label, min, max) — a plain year / count, validated so a typo can't be stored.
_CLIENT_INTAKE_INT_FIELDS = (
    ("year_of_passing", "Year of passing", 1950, 2100),
    ("backlogs_count", "Backlogs / arrears", 0, 99),
)


def _serialize_client_intake(client: models.EnterpriseClient) -> dict:
    data: dict = {}
    for field in _CLIENT_INTAKE_TEXT_FIELDS:
        data[field] = getattr(client, field, None)
    for field in _CLIENT_INTAKE_CHOICE_FIELDS:
        value = getattr(client, field, None)
        data[field] = value
        data[f"{field}_label"] = client_fields.choice_label(field, value)
    for field, _label in _CLIENT_INTAKE_DATE_FIELDS:
        data[field] = _iso(getattr(client, field, None))
    for field, _label, _lo, _hi in _CLIENT_INTAKE_INT_FIELDS:
        data[field] = getattr(client, field, None)
    channels = getattr(client, "marketing_consent_channels", None)
    data["marketing_consent_channels"] = [c for c in str(channels or "").split(",") if c]
    data["marketing_consent_channel_labels"] = client_fields.marketing_channel_labels(channels)
    data["marketing_consent_at"] = _iso(getattr(client, "marketing_consent_at", None))
    data["institution_share_consent"] = bool(getattr(client, "institution_share_consent_at", None))
    data["institution_share_consent_at"] = _iso(getattr(client, "institution_share_consent_at", None))
    return data


# Friendly names for the fields a conflict banner is most likely to have to name. Anything
# not listed falls back to a prettified key, so a new intake field needs no maintenance here.
_CLIENT_FIELD_LABELS = {
    "full_name": "Full name", "email": "Email", "phone": "Phone",
    "whatsapp_number": "WhatsApp number", "current_city": "Current city",
    "nationality": "Nationality", "date_of_birth": "Date of birth",
    "passport_number": "Passport number", "passport_expiry": "Passport expiry",
    "destination_country_code": "Destination country", "visa_type": "Visa type",
    "intake": "Target intake", "status": "Pipeline stage", "priority": "Priority",
    "target_date": "Key date", "assigned_to_user_id": "Assigned counselor",
    "branch_id": "Branch / office", "admission_stage": "Admission stage",
    "next_followup_date": "Next follow-up", "application_reference": "Application reference",
}


def _norm_for_compare(value):
    """Collapse the several ways this API spells "empty" so they compare equal.

    The form posts "" for a cleared field while the serializer emits None, and a cleared
    multi-select is [] on one side and None on the other. Without this every such field
    would read as a conflict on every save.
    """
    if value is None or value == "" or value == []:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (list, tuple)):
        return sorted(str(v) for v in value)
    return str(value)


_OCC_CONTROL_KEYS = ("expected_version", "expected_values", "force_overwrite", "allow_duplicate")


def _conflicting_fields(stored: dict, expected: dict, writing) -> list[str]:
    """Fields this caller is about to write whose stored value has moved since they read it.

    This — not the row version — is the real conflict test. `expected` holds the values the
    caller's form was RENDERED from, so comparing it against what is stored now isolates
    exactly the changes somebody else made, to the fields this save would overwrite.

    Two things fall out of comparing against the caller's baseline rather than against their
    new values. Concurrent edits to DIFFERENT fields stop colliding: they merge cleanly,
    which is what makes a shared client file usable by two counsellors at once. And the
    fields reported back are genuinely the other person's changes — comparing `stored` to
    the caller's own new value would flag every key they typed and name their own work as
    the colleague's, which is worse than saying nothing.
    """
    labels: list[str] = []
    for key in writing:
        if key in _OCC_CONTROL_KEYS or key not in expected or key not in stored:
            continue
        if _norm_for_compare(stored.get(key)) == _norm_for_compare(expected.get(key)):
            continue
        labels.append(_CLIENT_FIELD_LABELS.get(key, key.replace("_", " ").capitalize()))
    return labels


def _assert_client_write_is_current(client, data: dict, *, force: bool = False) -> None:
    """Concurrency precondition for a client write.

    Prefers the field-level check when the caller sent the baseline it rendered from, and
    falls back to the whole-row version for a caller that only sent that (an older bundle,
    a script). A caller that sent neither keeps the previous last-write-wins behaviour, so
    this tightened the contract without a flag day.
    """
    if force:
        return
    # A save that writes nothing cannot overwrite anything. Opening the edit form, changing
    # your mind and pressing Save still reaches here with only the control keys; answering
    # that with "someone else changed this — save anyway?" would offer a destructive-sounding
    # choice over an empty payload.
    if not [k for k in data if k not in _OCC_CONTROL_KEYS]:
        return
    expected = data.get("expected_values")
    if isinstance(expected, dict) and expected:
        stored = _serialize_client(client, None, include_sensitive=True)
        writing = [k for k in data if k not in _OCC_CONTROL_KEYS]
        conflicts = _conflicting_fields(stored, expected, writing)
        if conflicts:
            _raise_stale_write(client, what="client", conflicts=conflicts)
        # Every field being written must be COVERED by the baseline, or the field-level
        # check silently passed on the ones it never looked at. The browser always sends a
        # baseline entry per written key; anything else falls back to the whole-row version
        # so a partial baseline cannot be used to slip an unchecked write past this.
        if all(k in expected for k in writing):
            return
    _assert_fresh_write(client, data.get("expected_version"), what="client")


def _raise_stale_write(row, *, what: str, conflicts=None):
    """The one 409 shape for a concurrent-edit conflict.

    `conflicts` names the other person's changes so the banner can be specific instead of
    "something changed"; `current_version` lets the UI resync without a second round trip.
    """
    raise HTTPException(
        status_code=409,
        detail={
            "code": "stale_write",
            "message": (
                f"Someone else saved changes to this {what} while you had it open. "
                "Your edits were not applied."
            ),
            "current_version": int(getattr(row, "version", 0) or 0),
            "conflicts": list(conflicts or []),
        },
    )


def _assert_fresh_write(row, expected_version, *, what: str, conflicts=None) -> None:
    """Whole-row optimistic-concurrency precondition.

    HTTP If-Match semantics, carried in the payload rather than a header because the SPA
    posts JSON through one `api()` helper. The client echoes back the `version` it rendered
    from; if the stored row has moved on since, somebody else saved in between.

    This is the COARSE check: it fires on any concurrent write, whether or not it touched
    the same fields. Use it where the payload is whole-object and therefore cannot be
    merged (a finance entry), or as the fallback for a caller that sent no baseline. Where
    the caller does send one, `_assert_client_write_is_current` is strictly better.

    `expected_version` is OPTIONAL on purpose: a caller that sends nothing keeps the
    previous last-write-wins behaviour instead of breaking.
    """
    if expected_version in (None, 0):
        return
    if int(getattr(row, "version", 0) or 0) == int(expected_version):
        return
    _raise_stale_write(row, what=what, conflicts=conflicts)


# The fields `clients.view_sensitive` governs. Kept as one tuple so the read path (what
# _serialize_client omits) and the write path (what _strip_unwritable_sensitive refuses to
# let a blind caller overwrite) can never drift apart.
_CLIENT_SENSITIVE_FIELDS = ("passport_number",)
# The same fields as the AI autofill record labels them. Needed because rows written before
# the record carried `attr` can only be matched on their human label.
_CLIENT_SENSITIVE_LABELS = frozenset({"Passport number"})


def _strip_unwritable_sensitive(client, data: dict, ctx) -> None:
    """Stop a caller who cannot READ a sensitive field from overwriting it. Mutates `data`.

    Omitting the passport number from the payload (above) means a form rendered for someone
    without `clients.view_sensitive` draws that input EMPTY. Left unguarded their next save
    would post the empty string and clear an encrypted column they were never shown, turning
    a read restriction into silent data loss.

    An EMPTY submitted value is dropped rather than refused. It carries no intent — it is
    just the blank input echoing back, which is exactly what an older cached bundle sends on
    every save — so rejecting it would 403 an unrelated phone-number correction with a
    permissions error naming a field the user cannot even see. Dropping the key leaves the
    stored value untouched, which is what they would have wanted either way.

    A NON-EMPTY value is a real attempt to set something they are not cleared to see, so it
    is refused loudly. Create is deliberately not guarded: there is no stored value to
    destroy, and blocking it would stop a restricted counsellor completing an ordinary intake.
    """
    if ctx is not None and ctx.has("clients.view_sensitive"):
        return
    for field in _CLIENT_SENSITIVE_FIELDS:
        if field not in data:
            continue
        submitted = (data.get(field) or "").strip() or None
        if submitted is None:
            data.pop(field, None)
            continue
        if submitted != (getattr(client, field, None) or None):
            raise HTTPException(
                status_code=403,
                detail=access.denied_detail("clients.view_sensitive"),
            )
        data.pop(field, None)   # unchanged echo — nothing to write


def _serialize_client(
    client: models.EnterpriseClient,
    member_names: dict[int, str] | None = None,
    *,
    include_sensitive: bool = False,
) -> dict:
    """Serialize a client for the API.

    `include_sensitive=False` OMITS the passport number instead of masking it. That distinction
    matters: the client edit form round-trips every field it is given, so a masked placeholder
    would be written straight back over an encrypted column the next time anyone saved an
    unrelated change. Omitting the key is safe because a PATCH only applies the fields it receives.

    THE DEFAULT IS CLOSED. `clients.view_sensitive` is a `dangerous`-flagged capability
    (passport number and other identity fields) that Viewer and Finance deliberately do not
    hold — and Finance resolves to workspace scope, so a caller who can walk /clients/{id}
    could otherwise collect every student's passport number. Hiding the field in the browser
    is not enforcement: the value is on the wire and visible in devtools. Every caller must
    therefore opt IN with `include_sensitive=role.ctx.has("clients.view_sensitive")`, and a
    call site that forgets fails safe (the field is omitted) instead of leaking. Mirrors the
    rule the AI surface already applies in enterprise_ai.py.
    """
    assigned_name = None
    if client.assigned_to_user_id and member_names is not None:
        assigned_name = member_names.get(int(client.assigned_to_user_id))
    stage_visits = _load_stage_visits(client)
    payload = {
        **_serialize_client_intake(client),
        "id": client.id,
        "full_name": client.full_name,
        "email": client.email,
        "phone": client.phone,
        "nationality": client.nationality,
        "date_of_birth": _iso(client.date_of_birth),
        "passport_number": client.passport_number,
        "passport_expiry": _iso(client.passport_expiry),
        "visa_category": client.visa_category,
        "visa_category_label": _category_label(client.visa_category),
        "destination_country_code": client.destination_country_code,
        "destination_country_name": client.destination_country_name,
        "country": _country_brief(client.destination_country_code),
        "visa_type": client.visa_type,
        "intake": client.intake,
        "application_reference": client.application_reference,
        "status": client.status,
        "stage": _stage_brief(client.status, client.destination_country_code),
        "held_from_status": getattr(client, "held_from_status", None),
        "held_from_stage": _stage_brief(client.held_from_status, client.destination_country_code) if getattr(client, "held_from_status", None) else None,
        "priority": client.priority,
        "target_date": _iso(client.target_date),
        # Per-stage case record: {"<stage_key>": {"<field_key>": value}}. Field definitions
        # come from the destination-aware catalog served by /catalog.
        "stage_data": _load_stage_data(client),
        # How the case got to where it is. `stages_skipped` is the whole contract for the
        # journey: a stage in it was gone past without being worked and is drawn as skipped
        # instead of complete; everything else behind the case reads exactly as it always did.
        # Deriving it here rather than in each renderer is what keeps the staff journey, the
        # client portal and Deep Scan from disagreeing about what a skip is.
        # `stage_visits` is only for showing WHEN a stage was reached.
        "stage_visits": stage_visits,
        "stages_skipped": _skipped_stage_keys(client),
        "assigned_to_user_id": client.assigned_to_user_id,
        "assigned_to_name": assigned_name,
        "branch_id": getattr(client, "branch_id", None),
        "created_at": _iso(client.created_at),
        "updated_at": _iso(client.updated_at),
        # Optimistic-concurrency token. The edit form holds this and sends it back on save;
        # see _assert_fresh_write.
        "version": int(getattr(client, "version", 0) or 0),
    }
    if not include_sensitive:
        for field in _CLIENT_SENSITIVE_FIELDS:
            payload.pop(field, None)
        payload["passport_hidden"] = True
    return payload


def _serialize_duplicate(match, *, in_scope: bool, member_names: dict[int, str]) -> dict:
    """Render one possible-duplicate match for the warning dialog.

    find_duplicates() searches the whole organization on purpose — the duplicate worth
    catching is usually the one another office opened — so this is where the caller's
    record scope is paid back. For a match they are not allowed to open we hand over no
    id (the dialog must not offer a link into a file that would 404 on them) and mask the
    contact details. The NAME stays: they just typed an email or number that matches it,
    so it tells them nothing they don't already know, and without it the warning is
    unactionable.
    """
    row = match.row
    email = row.email or None
    phone = row.phone or row.whatsapp_number or None
    assigned_name = None
    if row.assigned_to_user_id:
        assigned_name = member_names.get(int(row.assigned_to_user_id))
    return {
        "id": int(row.id) if in_scope else None,
        "in_scope": in_scope,
        "full_name": row.full_name,
        "email": (email if in_scope else (_mask_email(email) if email else None)),
        "phone": (phone if in_scope else dupes.mask_phone(phone)),
        "stage": _stage_brief(row.status, row.destination_country_code),
        "destination_country_name": row.destination_country_name,
        "branch_name": row.branch_name,
        "assigned_to_name": assigned_name,
        "created_at": _iso(row.created_at),
        "reasons": match.reasons,
        "reason_labels": [dupes.REASON_LABELS.get(r, r) for r in match.reasons],
    }


def _assert_not_duplicate_client(
    db: Session,
    *,
    organization: models.EnterpriseOrganization,
    ctx,
    subject,
    exclude_client_id: Optional[int] = None,
    allow: bool = False,
    hint: str = "Check the match before adding a second file.",
) -> list:
    """Soft-block a client that looks like one the organization already has.

    409 rather than 400: the request is well-formed and the caller may legitimately want
    it, they just have to say so twice. `allow=True` skips the block but NOT the search —
    the create path uses the matches it gets back to leave a note on the file, which is
    what makes a knowing duplicate distinguishable later from an accidental one.
    """
    matches = dupes.find_duplicates(
        db,
        organization_id=organization.id,
        subject=subject,
        exclude_client_id=exclude_client_id,
    )
    if not matches or allow:
        return matches

    member_names = _org_member_name_map(db, organization.id)
    payload = [
        _serialize_duplicate(m, in_scope=access.client_in_scope(m.row, ctx), member_names=member_names)
        for m in matches
    ]
    lead = (
        f"{payload[0]['full_name']} is already in your workspace."
        if len(payload) == 1
        else f"{len(payload)} clients in your workspace look like this person."
    )
    raise HTTPException(
        status_code=409,
        detail={"code": "duplicate_client", "message": f"{lead} {hint}", "duplicates": payload},
    )


def _duplicate_override_note(matches: list) -> str:
    """Timeline entry for a file opened in full knowledge that it matched another.

    Cheaper and more durable than a column: it survives on the record itself, so whoever
    picks the case up later can see the second file was a decision rather than an accident.
    """
    parts = [
        f"{m.row.full_name} (#{m.row.id}) — {', '.join(dupes.REASON_LABELS.get(r, r).lower() for r in m.reasons)}"
        for m in matches[:3]
    ]
    more = len(matches) - len(parts)
    return (
        "Opened as a separate file despite a possible duplicate: "
        + "; ".join(parts)
        + (f", and {more} more" if more > 0 else "")
        + "."
    )


def _serialize_note(note: models.EnterpriseClientNote) -> dict:
    return {
        "id": note.id,
        "client_id": note.client_id,
        "author_user_id": note.author_user_id,
        "author_name": note.author_name,
        "body": note.body,
        "created_at": _iso(note.created_at),
    }


def _serialize_email_attachment(row: models.EnterpriseClientEmailAttachment) -> dict:
    return {
        "id": row.id,
        "filename": row.original_filename,
        "file_size": row.file_size,
        "mime_type": row.mime_type,
        "source_document_id": row.source_document_id,
        "download_url": (
            f"/api/enterprise/clients/{row.client_id}/email/attachments/{row.id}/download"
            if row.client_id else None
        ),
        "created_at": _iso(row.created_at),
    }


def _serialize_client_email(row: models.EnterpriseClientEmail) -> dict:
    direction = getattr(row, "direction", None) or "outbound"
    # Inbound bodies are stored as plain text and rendered escaped — never hand the
    # dashboard HTML that originated outside the org.
    body_html = getattr(row, "body_html", None) if direction == "outbound" else None
    try:
        attachments = [_serialize_email_attachment(a) for a in (row.attachments or [])]
    except Exception:  # pragma: no cover - pre-migration DBs without the table
        attachments = []
    return {
        "id": row.id,
        "client_id": row.client_id,
        "to_email": row.to_email,
        "subject": row.subject,
        "body": row.body,
        "body_html": body_html,
        "status": row.status,
        "sent_by_name": row.sent_by_name,
        "error_message": row.error_message,
        "direction": direction,
        "from_email": getattr(row, "from_email", None),
        "attachments": attachments,
        "created_at": _iso(row.created_at),
    }


# ---------------------------------------------------------------------------
# RECORD-LEVEL SCOPE
#
# Capabilities answer "what may this person do"; scope answers "to which records". A member is
# either workspace-wide ("all"), limited to the offices they staff ("branch"), or limited to their
# own caseload ("assigned"). Two rules make this safe rather than decorative:
#
#   1. Scope is applied to the BASE query, before any caller-supplied filter, so a query
#      parameter can only ever narrow the result set — never widen it.
#   2. Out-of-scope records raise 404, not 403. A 403 would confirm that a record exists, which
#      is exactly what the office partition is meant to hide.
# ---------------------------------------------------------------------------

_SCOPE_UNSET = object()


def _scope_ctx(ctx):
    """Normalize the `ctx` argument, and refuse to silently run unscoped.

    Passing nothing is a programming error, not a licence to return the whole workspace: an
    endpoint that forgets `ctx=` would otherwise leak every office's clients with no symptom.
    This raises loudly instead — and, unlike a keyword-required parameter, it names the caller.
    """
    if ctx is _SCOPE_UNSET:
        import inspect

        caller = "unknown"
        try:
            frame = inspect.currentframe()
            caller = frame.f_back.f_back.f_code.co_name if frame and frame.f_back else "unknown"
        except Exception:
            pass
        logger.error("Enterprise client lookup ran without an access context (caller=%s)", caller)
        raise RuntimeError(
            f"_get_org_client_or_404 called without ctx= from {caller}; pass ctx=role.ctx"
        )
    return ctx


# The scope predicates themselves live in enterprise_access, next to the scope resolution they
# belong to: the router, the AI tool surface and the notification fan-out all have to apply the
# identical rule, and three copies of "which clients can this person see" is exactly the kind of
# drift that becomes a data leak the first time one of them is edited.
scope_client_query = access.scope_client_query
client_in_scope = access.client_in_scope


def scoped_client_ids_subq(db: Session, organization_id: int, ctx):
    """Scoped client ids, for the two tables that reference a client by loose `reference_id`
    (notifications, credit transactions) and so have nothing to join on."""
    query = db.query(models.EnterpriseClient.id).filter(
        models.EnterpriseClient.organization_id == int(organization_id)
    )
    # `.scalar_subquery()`, not `.subquery()`: the result is only ever used as the right-hand
    # side of an IN(...), and passing a plain Subquery there makes SQLAlchemy coerce it with a
    # deprecation warning.
    return scope_client_query(query, ctx).scalar_subquery()


def scope_child_query(query, child_model, ctx):
    """One join covers every client-keyed child table (notes, emails, documents, scans, …)."""
    if ctx is None or ctx.scope_kind == "all":
        return query
    return scope_client_query(
        query.join(models.EnterpriseClient, models.EnterpriseClient.id == child_model.client_id),
        ctx,
    )


def assert_client_in_scope(client, ctx) -> None:
    if not client_in_scope(client, ctx):
        raise HTTPException(status_code=404, detail="Client not found.")


def _get_org_client_or_404(
    db: Session,
    organization_id: int,
    client_id: int,
    *,
    ctx=_SCOPE_UNSET,
) -> models.EnterpriseClient:
    ctx = _scope_ctx(ctx)
    client = (
        db.query(models.EnterpriseClient)
        .filter(
            models.EnterpriseClient.id == int(client_id),
            models.EnterpriseClient.organization_id == int(organization_id),
        )
        .first()
    )
    if not client:
        raise HTTPException(status_code=404, detail="Client not found.")
    assert_client_in_scope(client, ctx)
    return client


@router.get("/catalog")
def enterprise_catalog(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="catalog.view")
    return catalog.build_catalog_payload(db=db)


@router.get("/clients")
def enterprise_list_clients(
    request: Request,
    status_filter: Optional[str] = None,
    category: Optional[str] = None,
    country: Optional[str] = None,
    assigned_to: Optional[int] = None,
    branch_id: Optional[int] = None,
    q: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="clients.view"
    )

    # Record scope is applied to the BASE query, before any of the caller's filters below, so a
    # query parameter can only ever narrow what comes back — never widen it.
    base = scope_client_query(
        db.query(models.EnterpriseClient).filter(
            models.EnterpriseClient.organization_id == organization.id
        ),
        role.ctx,
    )
    if branch_id:
        # An office filter on top of scope: intersects, so a branch-scoped member asking for
        # someone else's office simply gets nothing.
        base = base.filter(models.EnterpriseClient.branch_id == int(branch_id))

    query = base
    if status_filter and status_filter.strip():
        # Dropping an unrecognised stage here would return the whole workspace under a filter
        # chip the UI still draws as active — a stale bookmark reading as "every client is at
        # this stage". Same reason writes go through _stage_key_or_400: say what was wrong.
        query = query.filter(models.EnterpriseClient.status == _stage_key_or_400(status_filter))
    normalized_category = catalog.normalize_category(category) if category else None
    if normalized_category:
        query = query.filter(models.EnterpriseClient.visa_category == normalized_category)
    if country:
        query = query.filter(models.EnterpriseClient.destination_country_code == country.strip().upper())
    if assigned_to:
        query = query.filter(models.EnterpriseClient.assigned_to_user_id == int(assigned_to))
    if q and q.strip():
        q_norm = q.strip().lower()
        like = f"%{q_norm}%"
        matching_stage_keys = [
            stage["key"]
            for stage in catalog.CLIENT_STAGES
            if q_norm in stage["key"].replace("_", " ").lower()
            or q_norm in stage["label"].lower()
        ]
        matching_country_codes = [
            country_item["code"]
            for country_item in catalog.COUNTRIES
            if q_norm in country_item["code"].lower()
            or q_norm in country_item["name"].lower()
            or q_norm in country_item.get("landmark", "").lower()
        ]
        matching_category_keys = [
            item["key"]
            for item in catalog.VISA_CATEGORIES
            if q_norm in item["key"].lower()
            or q_norm in item["label"].lower()
            or q_norm in item.get("short_label", "").lower()
        ]
        matching_assignee_ids = [
            int(row[0])
            for row in (
                db.query(models.EnterpriseOrganizationMember.user_id)
                .join(models.User, models.User.id == models.EnterpriseOrganizationMember.user_id)
                .filter(
                    models.EnterpriseOrganizationMember.organization_id == organization.id,
                    models.EnterpriseOrganizationMember.is_active.is_(True),
                    or_(
                        func.lower(models.User.full_name).like(like),
                        func.lower(models.User.email).like(like),
                    ),
                )
                .all()
            )
        ]

        search_clauses = [
            func.lower(models.EnterpriseClient.full_name).like(like),
            func.lower(models.EnterpriseClient.email).like(like),
            func.lower(models.EnterpriseClient.phone).like(like),
            func.lower(models.EnterpriseClient.nationality).like(like),
            # passport_number is encrypted at rest and therefore not substring-searchable.
            func.lower(models.EnterpriseClient.application_reference).like(like),
            # Intake record — the plain-text bits a counselor would actually type in.
            func.lower(models.EnterpriseClient.whatsapp_number).like(like),
            func.lower(models.EnterpriseClient.current_city).like(like),
            func.lower(models.EnterpriseClient.field_of_study).like(like),
            func.lower(models.EnterpriseClient.branch_name).like(like),
            func.lower(models.EnterpriseClient.lead_source_detail).like(like),
            func.lower(models.EnterpriseClient.guardian_name).like(like),
            func.lower(models.EnterpriseClient.visa_type).like(like),
            func.lower(models.EnterpriseClient.intake).like(like),
            func.lower(models.EnterpriseClient.destination_country_code).like(like),
            func.lower(models.EnterpriseClient.destination_country_name).like(like),
            func.lower(models.EnterpriseClient.status).like(like),
            func.lower(models.EnterpriseClient.priority).like(like),
        ]
        if matching_stage_keys:
            search_clauses.append(models.EnterpriseClient.status.in_(matching_stage_keys))
        if matching_country_codes:
            search_clauses.append(models.EnterpriseClient.destination_country_code.in_(matching_country_codes))
        if matching_category_keys:
            search_clauses.append(models.EnterpriseClient.visa_category.in_(matching_category_keys))
        if matching_assignee_ids:
            search_clauses.append(models.EnterpriseClient.assigned_to_user_id.in_(matching_assignee_ids))

        # Intake selects store an option key, so match on the LABEL the counselor saw
        # ("walk-in", "education loan", "master's") and translate it back to keys.
        for choice_field in (
            "lead_source", "study_level", "admission_stage", "prior_refusal_history",
            "english_test_type", "funding_source", "highest_qualification", "budget_band",
        ):
            matching_option_keys = [
                option["key"]
                for option in client_fields.CLIENT_PROFILE_OPTIONS.get(choice_field, [])
                if q_norm in option["label"].lower() or q_norm in option["key"].replace("_", " ")
            ]
            if matching_option_keys:
                search_clauses.append(getattr(models.EnterpriseClient, choice_field).in_(matching_option_keys))

        query = query.filter(or_(*search_clauses))

    clients = query.order_by(
        models.EnterpriseClient.created_at.desc(), models.EnterpriseClient.id.desc()
    ).all()

    member_names = _org_member_name_map(db, organization.id)
    _sensitive = role.ctx.has("clients.view_sensitive")

    # Separate query from `base`, so it needs the scope filter of its own — otherwise the stage
    # tallies above the list would quietly report the whole workspace's pipeline to someone who
    # can only see one office's clients.
    status_counts = {stage["key"]: 0 for stage in catalog.CLIENT_STAGES}
    for status_key, count in (
        scope_client_query(
            db.query(models.EnterpriseClient.status, func.count(models.EnterpriseClient.id))
            .filter(models.EnterpriseClient.organization_id == organization.id),
            role.ctx,
        )
        .group_by(models.EnterpriseClient.status)
        .all()
    ):
        if status_key in status_counts:
            status_counts[status_key] = int(count)

    total_clients = sum(status_counts.values())

    return {
        "organization_id": organization.id,
        "permissions": _enterprise_permissions_for_role(role),
        "total_clients": total_clients,
        "filtered_count": len(clients),
        "status_counts": status_counts,
        # List payloads never render the passport number, so it is omitted for anyone without
        # the capability rather than shipped to every screen that shows a client table.
        "clients": [
            _serialize_client(c, member_names, include_sensitive=_sensitive)
            for c in clients
        ],
    }


def _apply_client_case_fields(target: models.EnterpriseClient, *, category, country_code, visa_type, intake):
    try:
        resolved = catalog.resolve_visa_case(
            category=category,
            country_code=country_code,
            visa_type=visa_type,
            intake=intake,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    target.visa_category = resolved["category"]
    target.destination_country_code = resolved["country_code"]
    target.destination_country_name = resolved["country_name"]
    target.visa_type = resolved["visa_type"]
    target.intake = resolved["intake"]



def _apply_client_branch(
    db: Session,
    target: models.EnterpriseClient,
    *,
    organization,
    ctx,
    data: dict,
    creating: bool,
) -> None:
    """Attach a client to an office, and keep the denormalized name snapshot in step.

    Rules, in order:
      * key absent on update  -> leave the office alone.
      * key absent on create  -> file it under the member's own office, else the workspace default,
        so nothing lands unfiled (an unfiled client is invisible to office-scoped members).
      * moving an already-filed client needs `clients.set_branch`.
      * an office-scoped member may only target one of their own offices.
    """
    from app import enterprise_team as team

    has_key = "branch_id" in data
    raw = data.get("branch_id")

    if not has_key or raw in ("", None):
        if not creating:
            return
        fallback = ctx.primary_branch_id if ctx is not None else None
        branch = None
        if fallback:
            branch = (
                db.query(models.EnterpriseBranch)
                .filter(
                    models.EnterpriseBranch.id == int(fallback),
                    models.EnterpriseBranch.organization_id == organization.id,
                    models.EnterpriseBranch.is_active.is_(True),
                )
                .first()
            )
        if branch is None:
            branch = team.ensure_default_branch(
                db, organization.id,
                actor_user_id=getattr(target, "created_by_user_id", None),
                company_name=organization.company_name,
            )
        if branch is not None:
            target.branch_id = branch.id
            target.branch_name = branch.name
        return

    branch = team.get_org_branch_or_404(db, organization.id, raw)
    if not branch.is_active:
        raise HTTPException(status_code=422, detail="That office has been archived.")

    moving = (not creating) and target.branch_id and int(target.branch_id) != int(branch.id)
    if moving and ctx is not None and not ctx.has("clients.set_branch"):
        raise HTTPException(status_code=403, detail=access.denied_detail("clients.set_branch"))
    if ctx is not None and ctx.scope_kind == "branch" and int(branch.id) not in ctx.branch_ids:
        raise HTTPException(
            status_code=422,
            detail="You can only file a client under an office you work in.",
        )

    target.branch_id = branch.id
    target.branch_name = branch.name


def _assert_can_assign(db: Session, organization, ctx, assigned_to_user_id, *, current_value=None) -> None:
    """Guard `assigned_to_user_id`: who may hand a case to someone else."""
    if assigned_to_user_id is None or int(assigned_to_user_id or 0) == int(current_value or 0):
        return
    if not _is_active_org_member(db, organization.id, assigned_to_user_id):
        raise HTTPException(status_code=400, detail="Assigned team member is not part of this organization.")
    if ctx is None:
        return
    # Someone limited to their own caseload may take a case, but not push one onto a colleague —
    # and crucially not assign one AWAY, which would make it disappear from their own scope.
    if ctx.scope_kind == "assigned" and int(assigned_to_user_id) != ctx.user_id:
        raise HTTPException(
            status_code=403,
            detail="You can only assign clients to yourself. Ask a manager to reassign a case.",
        )
    if not ctx.has("clients.assign") and int(assigned_to_user_id) != ctx.user_id:
        raise HTTPException(status_code=403, detail=access.denied_detail("clients.assign"))


def _apply_client_intake_fields(target: models.EnterpriseClient, data: dict) -> None:
    """Write the lead-intake record onto a client.

    `data` holds only the keys the caller sent (PATCH semantics on update, everything on
    create), so an absent key is left alone and an explicitly empty one clears the column.
    Select values are validated against the shared catalog, so a stale or tampered payload
    can't write an option that doesn't exist.
    """
    for field in _CLIENT_INTAKE_TEXT_FIELDS:
        # `branch_name` is in this tuple because it is also the READ loop (see
        # _serialize_client_intake) — removing it there would blank the office on every client
        # payload. It is skipped only on the way IN: offices are real records now, so the name is
        # written from the chosen branch. Letting a caller retype it would let a branch-scoped
        # member move a client out of their own scope with a text edit.
        if field == "branch_name":
            continue
        if field in data:
            setattr(target, field, (data[field] or "").strip() or None)

    for field in _CLIENT_INTAKE_CHOICE_FIELDS:
        if field not in data:
            continue
        try:
            setattr(target, field, client_fields.normalize_choice(field, data[field]))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

    for field, label in _CLIENT_INTAKE_DATE_FIELDS:
        if field in data:
            setattr(target, field, _parse_iso_date_or_400(data[field], label))

    for field, label, low, high in _CLIENT_INTAKE_INT_FIELDS:
        if field not in data:
            continue
        raw = ("" if data[field] is None else str(data[field])).strip()
        if not raw:
            setattr(target, field, None)
            continue
        try:
            # OverflowError (not a ValueError) is what "inf"/"1e999" raises here.
            number = int(float(raw))
        except (TypeError, ValueError, OverflowError):
            raise HTTPException(status_code=400, detail=f"{label} must be a number.")
        if not low <= number <= high:
            raise HTTPException(status_code=400, detail=f"{label} must be between {low} and {high}.")
        setattr(target, field, number)

    # Consent is evidence, so each purpose carries its own timestamp: stamped when it is
    # first given, cleared the moment it is withdrawn.
    if "marketing_consent_channels" in data:
        try:
            channels = client_fields.normalize_marketing_channels(data["marketing_consent_channels"])
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
        target.marketing_consent_channels = channels
        if channels:
            if not target.marketing_consent_at:
                target.marketing_consent_at = datetime.utcnow()
        else:
            target.marketing_consent_at = None

    if "institution_share_consent" in data and data["institution_share_consent"] is not None:
        if data["institution_share_consent"]:
            if not target.institution_share_consent_at:
                target.institution_share_consent_at = datetime.utcnow()
        else:
            target.institution_share_consent_at = None


@router.post("/clients")
def enterprise_create_client(
    payload: EnterpriseClientCreateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="clients.create"
    )
    # The active-client cap is the plan's (sandbox 10 / Starter 100 / Growth 500 /
    # Scale 2,000). The infra-fee gate this used to sit beside was retired with the fee.
    billing.enforce_client_limit_or_402(db, organization.id)

    full_name = (payload.full_name or "").strip()
    if len(full_name) < 2:
        raise HTTPException(status_code=400, detail="Client name must be at least 2 characters.")

    # The org is the data controller and must attest the client consented to having
    # their data processed through Rilono. Enforce it server-side (not just in the UI)
    # so the DPA proof-of-consent trail can't have gaps.
    if not payload.client_consent_confirmed:
        raise HTTPException(
            status_code=400,
            detail="Please confirm this client has consented to having their data processed through Rilono before adding them.",
        )

    _assert_can_assign(db, organization, role.ctx, payload.assigned_to_user_id)

    client = models.EnterpriseClient(
        organization_id=organization.id,
        full_name=full_name,
        email=(str(payload.email).strip().lower() if payload.email else None),
        phone=(payload.phone or "").strip() or None,
        nationality=(payload.nationality or "").strip() or None,
        date_of_birth=_parse_iso_date_or_400(
            payload.date_of_birth, "Date of birth", direction=ent_dates.NOT_FUTURE
        ),
        passport_number=(payload.passport_number or "").strip() or None,
        passport_expiry=_parse_iso_date_or_400(payload.passport_expiry, "Passport expiry"),
        application_reference=(payload.application_reference or "").strip() or None,
        status=_stage_key_or_400(payload.status) if payload.status else catalog.DEFAULT_CLIENT_STAGE,
        priority=catalog.normalize_priority(payload.priority),
        target_date=_parse_iso_date_or_400(payload.target_date, "Target date"),
        assigned_to_user_id=payload.assigned_to_user_id,
        created_by_user_id=current_user.id,
        client_consent_confirmed_at=(datetime.utcnow() if payload.client_consent_confirmed else None),
        client_consent_confirmed_by_user_id=(current_user.id if payload.client_consent_confirmed else None),
    )
    _apply_client_case_fields(
        client,
        category=payload.visa_category,
        country_code=payload.destination_country_code,
        visa_type=payload.visa_type,
        intake=payload.intake,
    )
    _apply_client_intake_fields(client, payload.model_dump())
    # Normalizes held_from_status for a client opened straight onto a stage. The opening
    # stage is not a move, so whatever admissions milestone the intake form captured — the
    # walk-in who already holds an admit — is what stands. `creating` says the same thing to
    # the stage record: a walk-in opened at Collecting Documents never worked the stages
    # before it, so they are skipped rather than quietly assumed complete.
    _apply_status_change(client, client.status, db, creating=True)
    _apply_client_branch(
        db, client, organization=organization, ctx=role.ctx,
        data=payload.model_dump(exclude_unset=True), creating=True,
    )
    # Last gate before the row exists. It runs on the fully-built but still-unsaved object,
    # so it matches on exactly the values that are about to be stored — and a 409 here
    # leaves nothing behind.
    duplicates = _assert_not_duplicate_client(
        db, organization=organization, ctx=role.ctx, subject=client,
        allow=payload.allow_duplicate,
    )
    db.add(client)
    db.flush()

    initial_note = (payload.initial_note or "").strip()
    if initial_note:
        db.add(models.EnterpriseClientNote(
            organization_id=organization.id,
            client_id=client.id,
            author_user_id=current_user.id,
            author_name=current_user.full_name or current_user.email,
            body=initial_note,
        ))
    if duplicates:
        # Only reachable with allow_duplicate — otherwise the check above raised.
        db.add(models.EnterpriseClientNote(
            organization_id=organization.id,
            client_id=client.id,
            author_user_id=current_user.id,
            author_name=current_user.full_name or current_user.email,
            body=_duplicate_override_note(duplicates),
        ))

    db.commit()
    db.refresh(client)

    notif.notify_org(
        db, organization.id, type="client_added",
        title=f"{current_user.full_name or current_user.email} added client {client.full_name}",
        body=f"{client.destination_country_name or client.destination_country_code or ''} · {client.visa_type or 'student visa'}".strip(" ·"),
        actor_user_id=current_user.id, reference_type="client", reference_id=client.id, commit=True,
    )

    member_names = _org_member_name_map(db, organization.id)
    return {
        "message": "Client added successfully.",
        "permissions": _enterprise_permissions_for_role(role),
        "client": _serialize_client(
            client, member_names,
            include_sensitive=role.ctx.has("clients.view_sensitive"),
        ),
        "subscription": _serialize_subscription_state(billing.build_subscription_state(db, organization.id)),
    }


@router.get("/clients/{client_id}")
def enterprise_get_client(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="clients.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    member_names = _org_member_name_map(db, organization.id)

    notes = (
        db.query(models.EnterpriseClientNote)
        .filter(models.EnterpriseClientNote.client_id == client.id)
        .order_by(models.EnterpriseClientNote.created_at.desc(), models.EnterpriseClientNote.id.desc())
        .all()
    )
    emails = (
        db.query(models.EnterpriseClientEmail)
        .filter(models.EnterpriseClientEmail.client_id == client.id)
        .order_by(models.EnterpriseClientEmail.created_at.desc(), models.EnterpriseClientEmail.id.desc())
        .all()
    )
    documents = (
        db.query(models.EnterpriseClientDocument)
        .filter(models.EnterpriseClientDocument.client_id == client.id)
        .order_by(models.EnterpriseClientDocument.created_at.desc(), models.EnterpriseClientDocument.id.desc())
        .all()
    )
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "client": _serialize_client(
            client, member_names,
            include_sensitive=role.ctx.has("clients.view_sensitive"),
        ),
        "notes": [_serialize_note(n) for n in notes],
        "emails": [_serialize_client_email(e) for e in emails],
        "documents": [
            _serialize_client_document(d, include_sensitive=role.ctx.has("clients.view_sensitive"))
            for d in documents
        ],
    }


@router.patch("/clients/{client_id}")
def enterprise_update_client(
    client_id: int,
    payload: EnterpriseClientUpdateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="clients.edit"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)

    data = payload.model_dump(exclude_unset=True)
    # Before anything is applied: a caller who cannot READ the passport number may not
    # blind-overwrite it either (the field is omitted from their payload, so their form
    # renders it empty and would otherwise clear it on save).
    _strip_unwritable_sensitive(client, data, role.ctx)
    # ...and the fields they are writing must not have moved since their form was drawn.
    # `force_overwrite` is the second press of the same button, after the conflict banner
    # has shown them exactly which of their fields somebody else changed.
    _assert_client_write_is_current(client, data, force=payload.force_overwrite)
    status_before_edit = client.status
    admission_stage_before_edit = client.admission_stage
    # Snapshot of the matchable identity, taken before anything is applied. The Edit form
    # posts every field on every save, so "email in data" says nothing about whether the
    # email actually changed — comparing the two fingerprints does, and it keeps a plain
    # stage move on a client who shares a sibling's phone number from re-opening a
    # duplicate question the org already settled.
    identity_before_edit = dupes.fingerprint(client)

    if "full_name" in data:
        full_name = (data["full_name"] or "").strip()
        if len(full_name) < 2:
            raise HTTPException(status_code=400, detail="Client name must be at least 2 characters.")
        client.full_name = full_name

    # Visa case fields are validated together when any of them change.
    if any(k in data for k in ("visa_category", "destination_country_code", "visa_type", "intake")):
        _apply_client_case_fields(
            client,
            category=data.get("visa_category", client.visa_category),
            country_code=data.get("destination_country_code", client.destination_country_code),
            visa_type=data.get("visa_type", client.visa_type),
            intake=data.get("intake", client.intake),
        )

    if "email" in data:
        email_val = (data["email"] or "").strip().lower()
        client.email = email_val or None
    if "phone" in data:
        client.phone = (data["phone"] or "").strip() or None
    if "nationality" in data:
        client.nationality = (data["nationality"] or "").strip() or None
    if "passport_number" in data:
        client.passport_number = (data["passport_number"] or "").strip() or None
    if "application_reference" in data:
        client.application_reference = (data["application_reference"] or "").strip() or None
    if "date_of_birth" in data:
        client.date_of_birth = _parse_iso_date_or_400(
            data["date_of_birth"], "Date of birth", direction=ent_dates.NOT_FUTURE
        )
    if "passport_expiry" in data:
        client.passport_expiry = _parse_iso_date_or_400(data["passport_expiry"], "Passport expiry")
    if "target_date" in data:
        client.target_date = _parse_iso_date_or_400(data["target_date"], "Target date")
    if "priority" in data:
        client.priority = catalog.normalize_priority(data["priority"])
    if "assigned_to_user_id" in data:
        new_assignee = data["assigned_to_user_id"]
        _assert_can_assign(
            db, organization, role.ctx, new_assignee,
            current_value=client.assigned_to_user_id,
        )
        client.assigned_to_user_id = new_assignee
    _apply_client_intake_fields(client, data)
    # The Edit-client form sends every field on every save, so one request can carry both a
    # deliberate milestone edit and a stage move — and _apply_status_change cannot tell an
    # edited admission_stage from the one the form echoed back. Decide it here instead: a
    # submitted value that actually changed the record is a staff edit, and it wins over the
    # derivation for this request. A save that only moves the case still re-derives, and
    # derivation still never overwrites a manual-only value (deferred).
    admission_stage_edited = (
        "admission_stage" in data and client.admission_stage != admission_stage_before_edit
    )
    staff_admission_stage = client.admission_stage
    if "status" in data and data["status"]:
        _apply_status_change(client, _stage_key_or_400(data["status"]), db)
        if admission_stage_edited:
            client.admission_stage = staff_admission_stage
    _apply_client_branch(
        db, client, organization=organization, ctx=role.ctx, data=data, creating=False,
    )

    identity_after_edit = dupes.fingerprint(client)
    if identity_after_edit != identity_before_edit:
        _assert_not_duplicate_client(
            db, organization=organization, ctx=role.ctx, subject=identity_after_edit,
            exclude_client_id=client.id, allow=payload.allow_duplicate,
            hint="Check the match before saving these details onto this file.",
        )

    db.commit()
    db.refresh(client)
    if client.status != status_before_edit:
        stage_label = (catalog.CLIENT_STAGE_MAP.get(client.status) or {}).get("label", client.status)
        notif.notify_org(
            db, organization.id, type="status_changed",
            title=f"{current_user.full_name or current_user.email} moved {client.full_name} to {stage_label}",
            actor_user_id=current_user.id, reference_type="client", reference_id=client.id, commit=True,
        )
    member_names = _org_member_name_map(db, organization.id)
    return {
        "message": "Client updated.",
        "permissions": _enterprise_permissions_for_role(role),
        "client": _serialize_client(
            client, member_names,
            include_sensitive=role.ctx.has("clients.view_sensitive"),
        ),
    }


@router.patch("/clients/{client_id}/status")
def enterprise_update_client_status(
    client_id: int,
    payload: EnterpriseClientStatusRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="clients.edit"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    new_status = _stage_key_or_400(payload.status)
    old_status = client.status
    # A stage move from a stale board is the change that writes a false journey entry and
    # announces a move that never happened. Conflict on the STAGE specifically: if the
    # caller's screen agreed with the stored stage, their move is a valid continuation
    # regardless of what else changed on the record meanwhile.
    if payload.expected_status:
        if _norm_for_compare(old_status) != _norm_for_compare(payload.expected_status):
            _raise_stale_write(client, what="client", conflicts=[_CLIENT_FIELD_LABELS["status"]])
    else:
        _assert_fresh_write(client, payload.expected_version, what="client")
    _apply_status_change(client, new_status, db)
    db.commit()
    db.refresh(client)
    if new_status != old_status:
        stage_label = (catalog.CLIENT_STAGE_MAP.get(new_status) or {}).get("label", new_status)
        notif.notify_org(
            db, organization.id, type="status_changed",
            title=f"{current_user.full_name or current_user.email} moved {client.full_name} to {stage_label}",
            actor_user_id=current_user.id, reference_type="client", reference_id=client.id, commit=True,
        )
    member_names = _org_member_name_map(db, organization.id)
    return {
        "message": "Status updated.",
        "permissions": _enterprise_permissions_for_role(role),
        "client": _serialize_client(
            client, member_names,
            include_sensitive=role.ctx.has("clients.view_sensitive"),
        ),
    }


class EnterpriseStageDataUpdateRequest(BaseModel):
    """Save the case record for ONE stage. `values` is {field_key: value}; an empty/omitted
    value clears that field. Unknown keys (e.g. after a catalog change) are ignored."""
    stage_key: str
    values: dict = {}
    expected_version: Optional[int] = None


def _clean_stage_values(allowed: dict[str, dict], incoming) -> dict[str, str]:
    """Coerce a submitted case-record payload to the values that may be stored.

    Keys the catalog doesn't declare for this destination+stage are dropped (a stale or
    tampered payload can't write arbitrary keys). An empty string is kept and means "clear
    this field" — the caller pops it from the record.

    A field the catalog calls a date is stored as one from here on: parsed, bounded and
    normalized to YYYY-MM-DD. NEW WRITES ONLY. The case record is a JSON blob of free text,
    and everything already saved in it stays exactly as it is — unread, unconverted, still
    whatever a counsellor typed. That is the point of validating at this seam instead of
    migrating: no existing value has to be interpreted for today's saves to come out clean.

    No direction is imposed. These fields are a mix of issue dates, appointment dates,
    deadlines and expiries across fifteen destinations; only the sanity window is true of all
    of them (see `enterprise_dates`).
    """
    values = incoming if isinstance(incoming, dict) else {}
    submitted: dict[str, str] = {}
    for key, value in values.items():
        field = allowed.get(key)
        if field is None:
            continue
        text_value = ("" if value is None else str(value)).strip()
        if len(text_value) > 500:
            raise HTTPException(status_code=400, detail=f"'{key}' is too long (max 500 characters).")
        if text_value and field.get("type") == "date":
            parsed = _parse_iso_date_or_400(text_value, field.get("label") or key)
            text_value = parsed.isoformat() if parsed else ""
        submitted[key] = text_value
    return submitted


@router.patch("/clients/{client_id}/stage-data")
def enterprise_update_client_stage_data(
    client_id: int,
    payload: EnterpriseStageDataUpdateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Record the destination-specific case details a counselor captures at a pipeline stage
    (e.g. US: SEVIS ID / DS-160 confirmation; UK: CAS number / IHS reference)."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="clients.edit"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    # The case record is merged key-by-key below, so a stale save only clobbers the fields
    # it actually carries — but those are visa-critical (CAS number, SEVIS ID, funds shown),
    # so a stale writer is still refused rather than merged over a colleague's correction.
    _assert_fresh_write(client, payload.expected_version, what="client")

    stage_key = str(payload.stage_key or "").strip().lower()
    if stage_key not in catalog.CLIENT_STAGE_KEYS:
        raise HTTPException(status_code=400, detail="Unknown stage.")

    # Only fields defined for THIS client's destination + stage are accepted, so a stale or
    # tampered payload can't write arbitrary keys into the record.
    allowed = {f["key"]: f for f in catalog.stage_fields_for(client.destination_country_code, stage_key)}
    if not allowed:
        raise HTTPException(status_code=400, detail="This stage has no record fields for this destination.")

    submitted = _clean_stage_values(allowed, payload.values)

    data, unreadable = _parse_stage_data(client)
    if unreadable:
        # Writing now would rebuild the column from an empty dict and take every other
        # stage's records with it. Refuse: the stored bytes stay on the row exactly as they
        # are, recoverable, and the counselor is told instead of shown a false success.
        raise HTTPException(
            status_code=409,
            detail="This client's saved case records can't be read, so saving would overwrite them. Contact support before editing this stage.",
        )

    # MERGE onto what is stored, never replace it: `allowed` covers only what the current
    # catalog declares, so a replace would delete the keys it no longer names — including
    # RETIRED_STAGE_FIELDS, whose values are still read by the backfill into the client
    # columns. Only the keys the counselor actually submitted change; an empty one clears.
    existing = data.get(stage_key)
    bucket = dict(existing) if isinstance(existing, dict) else {}
    for key, text_value in submitted.items():
        if text_value:
            bucket[key] = text_value
        else:
            bucket.pop(key, None)
    if bucket:
        data[stage_key] = bucket
    else:
        data.pop(stage_key, None)
    client.stage_data = json.dumps(data) if data else None
    db.commit()
    db.refresh(client)

    member_names = _org_member_name_map(db, organization.id)
    return {
        "message": "Case record saved.",
        "permissions": _enterprise_permissions_for_role(role),
        "client": _serialize_client(
            client, member_names,
            include_sensitive=role.ctx.has("clients.view_sensitive"),
        ),
    }


@router.delete("/clients/{client_id}")
def enterprise_delete_client(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="clients.delete"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)

    # Collect the storage keys BEFORE the rows go — the child rows cascade at the DB level
    # (passive_deletes), so no Python-side hook ever runs for them and after the commit
    # there is nothing left to ask. The blobs themselves are purged only AFTER the commit
    # (below): rows first, bytes second, the same invariant every sibling delete handler
    # documents. The old order destroyed the files first, so a commit that failed — dropped
    # connection, timeout, deadlock — left the client alive and still listing passports and
    # bank statements whose bytes were already unrecoverable.
    storage_keys = [
        key
        for (key,) in db.query(models.EnterpriseClientDocument.storage_key)
        .filter(
            models.EnterpriseClientDocument.client_id == client.id,
            models.EnterpriseClientDocument.organization_id == organization.id,
        )
        .all()
        if key
    ]
    storage_keys += [
        key
        for (key,) in db.query(models.EnterpriseClientEmailAttachment.storage_key)
        .filter(
            models.EnterpriseClientEmailAttachment.client_id == client.id,
            models.EnterpriseClientEmailAttachment.organization_id == organization.id,
        )
        .all()
        if key
    ]
    # Reference files on this client's calendar events, reached through the event because the
    # attachment carries no client_id of its own.
    storage_keys += cal_files.keys_for_client(
        db, organization_id=organization.id, client_id=client.id
    )

    # Leads that became this client outlive it (an enquiry is its own record), so detach
    # them explicitly rather than trusting ON DELETE SET NULL — the sqlite sandbox never
    # enforces foreign keys, and a dangling id leaves an "Open client" button that 404s.
    db.query(models.EnterpriseLead).filter(
        models.EnterpriseLead.organization_id == organization.id,
        models.EnterpriseLead.converted_client_id == client.id,
    ).update({"converted_client_id": None}, synchronize_session=False)

    db.delete(client)
    db.commit()

    # Only now, with the rows durably gone, purge the bytes. The DPA's deletion promise is
    # kept by delete_document being best-effort-but-attempted for every key; a blob that
    # fails to delete here is an orphan costing storage, never a live row pointing at
    # nothing — the unrecoverable direction.
    for key in storage_keys:
        enterprise_storage.delete_document(key)

    return {
        "message": "Client deleted.",
        "permissions": _enterprise_permissions_for_role(role),
    }


@router.post("/clients/{client_id}/notes")
def enterprise_add_client_note(
    client_id: int,
    payload: EnterpriseClientNoteRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="notes.write"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    body = (payload.body or "").strip()
    if not body:
        raise HTTPException(status_code=400, detail="Note cannot be empty.")

    note = models.EnterpriseClientNote(
        organization_id=organization.id,
        client_id=client.id,
        author_user_id=current_user.id,
        author_name=current_user.full_name or current_user.email,
        body=body,
    )
    db.add(note)
    db.commit()
    db.refresh(note)
    return {
        "message": "Note added.",
        "permissions": _enterprise_permissions_for_role(role),
        "note": _serialize_note(note),
    }


@router.delete("/clients/{client_id}/notes/{note_id}")
def enterprise_delete_client_note(
    client_id: int,
    note_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Delete a client note. Admins can delete any note (including AI-generated ones);
    editors can only delete notes they authored."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="notes.write"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    note = (
        db.query(models.EnterpriseClientNote)
        .filter(
            models.EnterpriseClientNote.id == int(note_id),
            models.EnterpriseClientNote.organization_id == organization.id,
            models.EnterpriseClientNote.client_id == client.id,
        )
        .first()
    )
    if not note:
        raise HTTPException(status_code=404, detail="Note not found.")
    # Removing a colleague's note is its own permission now, so a branch manager can tidy their
    # office's timeline without being made a workspace admin.
    if not role.ctx.has("notes.moderate") and note.author_user_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="You can only delete your own notes. Ask an organization admin to remove this one.",
        )
    db.delete(note)
    db.commit()
    return {"message": "Note deleted.", "permissions": _enterprise_permissions_for_role(role)}


def _send_and_log_client_email(
    db: Session,
    *,
    organization: models.EnterpriseOrganization,
    client: models.EnterpriseClient,
    subject: str,
    body: str,
    current_user: models.User,
    body_html: Optional[str] = None,
    attachment_rows: Optional[list[models.EnterpriseClientEmailAttachment]] = None,
) -> models.EnterpriseClientEmail:
    # Replies go straight back to the staffer who sent the message, and the two of
    # them carry on in their own mail clients from there. Deliberately unconditional:
    # the tokenized reply+ routing in app/enterprise_inbound_email.py stays dormant,
    # so a stale RESEND_INBOUND_REPLY_DOMAIN can never re-arm a Reply-To that bounces.
    # To revisit portal threading, restore the reply_routing_enabled() branch here.
    reply_to = current_user.email
    payload_attachments = _load_attachment_payloads(attachment_rows or [])
    success, message_id, error = send_enterprise_client_email(
        to_email=client.email,
        subject=subject,
        body=body,
        body_html=body_html,
        organization_name=organization.company_name,
        sender_name=current_user.full_name or current_user.email,
        logo_url=_absolute_enterprise_logo_url(organization),
        reply_to=reply_to,
        # Always true now — Reply-To is a real person's inbox, so telling the client
        # they can reply is accurate rather than the lie it was while it bounced.
        direct_reply_hint=True,
        attachments=payload_attachments,
    )
    row = models.EnterpriseClientEmail(
        organization_id=organization.id,
        client_id=client.id,
        sent_by_user_id=current_user.id,
        sent_by_name=current_user.full_name or current_user.email,
        to_email=client.email,
        subject=subject,
        body=body,
        body_html=body_html,
        status="sent" if success else "failed",
        provider_message_id=message_id,
        error_message=error,
    )
    db.add(row)
    # Bind the (previously draft) attachments to the message that carried them, so the
    # thread shows exactly what the client received. On a failed send they stay drafts
    # instead, which is what lets the composer retry with the same files attached.
    if success:
        for attachment in attachment_rows or []:
            attachment.email = row
    return row


def _load_attachment_payloads(rows: list[models.EnterpriseClientEmailAttachment]) -> list[dict]:
    """Read attachment bytes out of encrypted storage for the mail provider.

    An unreadable file aborts the whole send. Skipping it would deliver a message
    without the document while the thread still showed the file as attached — the
    consultant would believe a passport or offer letter went out when it didn't.
    Failing here leaves the drafts intact so the send can simply be retried.
    """
    payloads: list[dict] = []
    for row in rows:
        try:
            data = enterprise_storage.fetch_document(row.storage_key)
        except Exception:
            logger.exception("Email attachment unreadable (id=%s, key=%s)", row.id, row.storage_key)
            raise HTTPException(
                status_code=502,
                detail=f"“{row.original_filename}” could not be read from storage, so nothing was sent. "
                       "Remove it and attach the file again.",
            )
        if not data:
            raise HTTPException(
                status_code=502,
                detail=f"“{row.original_filename}” is empty in storage, so nothing was sent. "
                       "Remove it and attach the file again.",
            )
        payloads.append({
            "filename": _safe_filename(row.original_filename),
            "content": data,
            "content_type": row.mime_type or None,
        })
    return payloads


@router.post("/clients/{client_id}/email")
def enterprise_email_client(
    client_id: int,
    payload: EnterpriseClientEmailRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="emails.send"
    )
    _enforce_rate_limit_or_429(
        request=request,
        scope="enterprise.client_email",
        limit=ENTERPRISE_CLIENT_EMAIL_RATE_LIMIT,
        window_seconds=ENTERPRISE_CLIENT_EMAIL_RATE_WINDOW_SECONDS,
        extra_key=f"org:{organization.id}:user:{current_user.id}",
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    if not (client.email or "").strip():
        raise HTTPException(status_code=400, detail="This client has no email address on file.")

    subject = payload.subject.strip()
    # Never trust the composer's markup: re-serialize it through the allow-list
    # sanitizer, and derive the plain-text part from the *sanitized* HTML so the two
    # halves of the message can never disagree about what was actually sent.
    body_html = sanitize_email_html(payload.body_html)
    body = (html_to_text(body_html) if body_html else payload.body).strip()
    if not body:
        raise HTTPException(status_code=400, detail="Write a message before sending.")
    body = body[:ENTERPRISE_EMAIL_BODY_MAX]

    attachment_rows = _collect_email_attachments(
        db,
        organization=organization,
        client=client,
        current_user=current_user,
        attachment_ids=payload.attachment_ids,
        document_ids=payload.document_ids,
    )

    row = _send_and_log_client_email(
        db, organization=organization, client=client, subject=subject, body=body,
        current_user=current_user, body_html=body_html, attachment_rows=attachment_rows,
    )
    db.commit()
    db.refresh(row)

    if row.status != "sent":
        raise HTTPException(status_code=502, detail=row.error_message or "Email could not be sent.")

    return {
        "message": f"Email sent to {client.full_name}.",
        "permissions": _enterprise_permissions_for_role(role),
        "email": _serialize_client_email(row),
    }


# ---------------------------------------------------------------------------
# Composer attachments. Files are uploaded while the message is still being
# written (email_id is null = draft) and bound to the email when it sends.
# ---------------------------------------------------------------------------

def _email_attachment_or_404(
    db: Session, *, organization_id: int, client_id: int, attachment_id: int,
    require_uploader_user_id: int | None = None,
) -> models.EnterpriseClientEmailAttachment:
    row = (
        db.query(models.EnterpriseClientEmailAttachment)
        .filter(
            models.EnterpriseClientEmailAttachment.id == int(attachment_id),
            models.EnterpriseClientEmailAttachment.organization_id == int(organization_id),
            models.EnterpriseClientEmailAttachment.client_id == int(client_id),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Attachment not found.")
    # A file attached to a message that hasn't been sent yet is a private draft. The list endpoint
    # already scopes drafts to their uploader; download and delete must apply the same rule, or a
    # colleague can pull or bin a half-written message's attachments.
    if (
        require_uploader_user_id is not None
        and not getattr(row, "email_id", None)
        and int(getattr(row, "uploaded_by_user_id", 0) or 0) != int(require_uploader_user_id)
    ):
        raise HTTPException(status_code=404, detail="Attachment not found.")
    return row


def _sweep_stale_draft_attachments(db: Session, *, organization_id: int, user_id: int) -> list[str]:
    """Mark this user's abandoned draft attachments for deletion (composer closed
    without sending) and return their storage keys.

    Only the DB rows are deleted here. The caller drops the blobs *after* the commit
    succeeds — deleting bytes first would, on a failed commit, leave surviving rows
    pointing at files that no longer exist.
    """
    cutoff = datetime.now(dt_timezone.utc) - timedelta(hours=ENTERPRISE_EMAIL_ATTACH_DRAFT_TTL_HOURS)
    keys: list[str] = []
    try:
        stale = (
            db.query(models.EnterpriseClientEmailAttachment)
            .filter(
                models.EnterpriseClientEmailAttachment.organization_id == organization_id,
                models.EnterpriseClientEmailAttachment.uploaded_by_user_id == user_id,
                models.EnterpriseClientEmailAttachment.email_id.is_(None),
                models.EnterpriseClientEmailAttachment.created_at < cutoff,
            )
            .limit(50)
            .all()
        )
        for row in stale:
            keys.append(row.storage_key)
            db.delete(row)
    except Exception:
        logger.exception("Draft attachment sweep failed (org_id=%s)", organization_id)
    return keys


def _collect_email_attachments(
    db: Session,
    *,
    organization: models.EnterpriseOrganization,
    client: models.EnterpriseClient,
    current_user: models.User,
    attachment_ids: list[int],
    document_ids: list[int],
) -> list[models.EnterpriseClientEmailAttachment]:
    """Resolve the composer's attachment selection into storage-backed rows.

    Draft uploads are looked up (scoped to this org + client + uploader); documents
    already on file are *copied* into the email's own storage key, so deleting the
    document later never rewrites the history of what was sent.

    Everything is resolved and size-checked *before* a single byte is copied. A DB
    rollback cannot un-write an object, so a limit breach discovered after copying
    would strand bytes in the bucket that no row references and no sweep can find.
    """
    rows: list[models.EnterpriseClientEmailAttachment] = []

    for attachment_id in list(dict.fromkeys(attachment_ids or []))[:ENTERPRISE_EMAIL_ATTACH_MAX_FILES]:
        row = (
            db.query(models.EnterpriseClientEmailAttachment)
            .filter(
                models.EnterpriseClientEmailAttachment.id == int(attachment_id),
                models.EnterpriseClientEmailAttachment.organization_id == organization.id,
                models.EnterpriseClientEmailAttachment.client_id == client.id,
                models.EnterpriseClientEmailAttachment.email_id.is_(None),
                models.EnterpriseClientEmailAttachment.uploaded_by_user_id == current_user.id,
            )
            .first()
        )
        if not row:
            raise HTTPException(
                status_code=400,
                detail="One of the attachments is no longer available. Remove it and attach the file again.",
            )
        rows.append(row)

    # Resolve the picked documents (metadata only) and check the ceilings first.
    picked_docs: list[models.EnterpriseClientDocument] = []
    for document_id in list(dict.fromkeys(document_ids or [])):
        if len(rows) + len(picked_docs) >= ENTERPRISE_EMAIL_ATTACH_MAX_FILES:
            break
        doc = (
            db.query(models.EnterpriseClientDocument)
            .filter(
                models.EnterpriseClientDocument.id == int(document_id),
                models.EnterpriseClientDocument.client_id == client.id,
                models.EnterpriseClientDocument.organization_id == organization.id,
            )
            .first()
        )
        if not doc:
            raise HTTPException(status_code=404, detail="One of the selected documents is no longer on file.")
        picked_docs.append(doc)

    projected = (
        sum(int(r.file_size or 0) for r in rows)
        + sum(int(d.file_size or 0) for d in picked_docs)
    )
    if projected > ENTERPRISE_EMAIL_ATTACH_MAX_TOTAL_BYTES:
        raise HTTPException(
            status_code=400,
            detail=(
                "Attachments total "
                f"{projected // (1024 * 1024)} MB — the limit is "
                f"{ENTERPRISE_EMAIL_ATTACH_MAX_TOTAL_BYTES // (1024 * 1024)} MB per email. "
                "Remove a file, or share it as a link instead."
            ),
        )

    written_keys: list[str] = []
    try:
        for doc in picked_docs:
            try:
                data = enterprise_storage.fetch_document(doc.storage_key)
            except Exception:
                logger.exception("Could not read document %s for email attachment", doc.id)
                raise HTTPException(
                    status_code=502,
                    detail=f"Could not read “{doc.original_filename}” from storage. Try attaching it as a file instead.",
                )
            ext = os.path.splitext(doc.original_filename)[1].lower()
            storage_key = f"enterprise/{organization.id}/clients/{client.id}/email/{uuid.uuid4().hex}{ext}"
            try:
                enterprise_storage.store_document(storage_key, data, content_type=doc.mime_type)
            except Exception:
                logger.exception("Could not copy document %s into an email attachment", doc.id)
                raise HTTPException(
                    status_code=502, detail="Could not attach that document right now. Please try again."
                )
            written_keys.append(storage_key)
            row = models.EnterpriseClientEmailAttachment(
                organization_id=organization.id,
                client_id=client.id,
                source_document_id=doc.id,
                original_filename=doc.original_filename,
                storage_key=storage_key,
                file_size=len(data),
                mime_type=doc.mime_type,
                uploaded_by_user_id=current_user.id,
            )
            db.add(row)
            rows.append(row)
    except Exception:
        # The pending DB rows roll back on their own; the objects would not.
        for key in written_keys:
            enterprise_storage.delete_document(key)
        raise

    return rows


@router.get("/clients/{client_id}/email/attachments")
def enterprise_list_draft_email_attachments(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """This user's not-yet-sent attachments for this client, so a refreshed or
    reopened composer picks the files back up instead of orphaning them."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="emails.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    rows = (
        db.query(models.EnterpriseClientEmailAttachment)
        .filter(
            models.EnterpriseClientEmailAttachment.organization_id == organization.id,
            models.EnterpriseClientEmailAttachment.client_id == client.id,
            models.EnterpriseClientEmailAttachment.email_id.is_(None),
            models.EnterpriseClientEmailAttachment.uploaded_by_user_id == current_user.id,
        )
        .order_by(models.EnterpriseClientEmailAttachment.id.asc())
        .all()
    )
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "attachments": [_serialize_email_attachment(r) for r in rows],
    }


@router.post("/clients/{client_id}/email/attachments")
async def enterprise_upload_email_attachment(
    client_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Upload one file for the message currently being composed (a draft attachment)."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="emails.send"
    )
    # Uploads are cheap to trigger and expensive to store, so cap them the same way
    # sends are capped — otherwise a compromised editor account could fill the bucket.
    _enforce_rate_limit_or_429(
        request=request,
        scope="enterprise.email_attachment",
        limit=ENTERPRISE_EMAIL_ATTACH_RATE_LIMIT,
        window_seconds=ENTERPRISE_EMAIL_ATTACH_RATE_WINDOW_SECONDS,
        extra_key=f"org:{organization.id}:user:{current_user.id}",
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)

    if not enterprise_storage.is_configured():
        raise HTTPException(status_code=503, detail="File storage is not configured.")

    original = _safe_filename(file.filename)
    ext = os.path.splitext(original)[1].lower()
    if ext not in ENTERPRISE_DOC_ALLOWED_EXT:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Allowed: PDF, images, Word/Excel, CSV, or text.",
        )

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="That file is empty.")

    pending = (
        db.query(models.EnterpriseClientEmailAttachment)
        .filter(
            models.EnterpriseClientEmailAttachment.organization_id == organization.id,
            models.EnterpriseClientEmailAttachment.client_id == client.id,
            models.EnterpriseClientEmailAttachment.email_id.is_(None),
            models.EnterpriseClientEmailAttachment.uploaded_by_user_id == current_user.id,
        )
        .all()
    )
    if len(pending) >= ENTERPRISE_EMAIL_ATTACH_MAX_FILES:
        raise HTTPException(
            status_code=400,
            detail=f"You can attach up to {ENTERPRISE_EMAIL_ATTACH_MAX_FILES} files to one email.",
        )
    limit_mb = ENTERPRISE_EMAIL_ATTACH_MAX_TOTAL_BYTES // (1024 * 1024)
    if sum(int(p.file_size or 0) for p in pending) + len(data) > ENTERPRISE_EMAIL_ATTACH_MAX_TOTAL_BYTES:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Attachments can total {limit_mb} MB per email. "
                "Send this file on its own, or share it as a link instead."
            ),
        )

    storage_key = f"enterprise/{organization.id}/clients/{client.id}/email/{uuid.uuid4().hex}{ext}"
    try:
        await run_in_threadpool(enterprise_storage.store_document, storage_key, data, content_type=file.content_type)
    except Exception:
        logger.exception("Failed to store email attachment (org_id=%s, client_id=%s)", organization.id, client.id)
        raise HTTPException(status_code=502, detail="Could not upload that file right now. Please try again.")

    row = models.EnterpriseClientEmailAttachment(
        organization_id=organization.id,
        client_id=client.id,
        original_filename=original,
        storage_key=storage_key,
        file_size=len(data),
        mime_type=(file.content_type or None),
        uploaded_by_user_id=current_user.id,
    )
    db.add(row)
    swept_keys = _sweep_stale_draft_attachments(db, organization_id=organization.id, user_id=current_user.id)
    db.commit()
    db.refresh(row)
    for key in swept_keys:
        enterprise_storage.delete_document(key)

    return {
        "message": "File attached.",
        "permissions": _enterprise_permissions_for_role(role),
        "attachment": _serialize_email_attachment(row),
    }


@router.delete("/clients/{client_id}/email/attachments/{attachment_id}")
def enterprise_delete_email_attachment(
    client_id: int,
    attachment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Remove a draft attachment. Files already sent are part of the record and stay."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="emails.send"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    row = _email_attachment_or_404(
        db, organization_id=organization.id, client_id=client.id, attachment_id=attachment_id,
        require_uploader_user_id=current_user.id,
    )
    if row.email_id is not None:
        raise HTTPException(status_code=400, detail="This file was already sent and can't be removed.")
    storage_key = row.storage_key
    db.delete(row)
    # Row first, blob second: a failed commit must never leave a surviving row
    # pointing at bytes that are already gone.
    db.commit()
    enterprise_storage.delete_document(storage_key)
    return {"message": "Attachment removed.", "permissions": _enterprise_permissions_for_role(role)}


@router.get("/clients/{client_id}/email/attachments/{attachment_id}/download")
def enterprise_download_email_attachment(
    client_id: int,
    attachment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="emails.view"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    row = _email_attachment_or_404(
        db, organization_id=organization.id, client_id=client.id, attachment_id=attachment_id,
        require_uploader_user_id=current_user.id,
    )
    try:
        data = enterprise_storage.fetch_document(row.storage_key)
    except Exception:
        logger.exception("Failed to fetch email attachment id=%s", row.id)
        raise HTTPException(status_code=502, detail="Could not retrieve that file right now.")

    # Same hardening as the document download: the served Content-Type comes from the
    # validated extension, never the uploader-supplied mime_type, so a *.pdf holding
    # HTML+<script> can't execute on this origin.
    ext = os.path.splitext(row.original_filename)[1].lower()
    if ext in ENTERPRISE_DOC_INLINE_EXT:
        disposition = "inline"
        media_type = ENTERPRISE_DOC_EXT_MIME.get(ext, "application/octet-stream")
    else:
        disposition = "attachment"
        media_type = "application/octet-stream"
    filename = _safe_filename(row.original_filename)
    return Response(
        content=data,
        media_type=media_type,
        headers={
            "Content-Disposition": f'{disposition}; filename="{filename}"',
            "X-Content-Type-Options": "nosniff",
        },
    )


@router.post("/clients/email/bulk")
def enterprise_email_clients_bulk(
    payload: EnterpriseBulkEmailRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="emails.send_bulk"
    )
    _enforce_rate_limit_or_429(
        request=request,
        scope="enterprise.bulk_client_email",
        limit=ENTERPRISE_BULK_EMAIL_RATE_LIMIT,
        window_seconds=ENTERPRISE_BULK_EMAIL_RATE_WINDOW_SECONDS,
        extra_key=f"org:{organization.id}:user:{current_user.id}",
    )
    client_ids = list(dict.fromkeys(payload.client_ids))[:ENTERPRISE_BULK_EMAIL_MAX_RECIPIENTS]
    subject = payload.subject.strip()
    body = payload.body.strip()

    # The recipient list arrives as caller-supplied ids, which makes this the one place where an
    # unscoped query would let someone email a client they aren't allowed to open. Out-of-scope
    # ids are dropped silently rather than rejected — the same 404-not-403 reasoning: refusing the
    # whole request would confirm those clients exist.
    clients = (
        scope_client_query(
            db.query(models.EnterpriseClient).filter(
                models.EnterpriseClient.organization_id == organization.id,
                models.EnterpriseClient.id.in_(client_ids),
            ),
            role.ctx,
        )
        .all()
    )

    sent = 0
    failed = 0
    skipped = 0
    for client in clients:
        if not (client.email or "").strip():
            skipped += 1
            continue
        row = _send_and_log_client_email(
            db, organization=organization, client=client, subject=subject, body=body, current_user=current_user
        )
        if row.status == "sent":
            sent += 1
        else:
            failed += 1

    db.commit()
    return {
        "message": f"Sent {sent} email(s). {failed} failed, {skipped} skipped (no email).",
        "permissions": _enterprise_permissions_for_role(role),
        "sent": sent,
        "failed": failed,
        "skipped": skipped,
    }


# Dashboard "What's next": how far ahead it looks, and how many rows the card carries before
# it defers to the Calendar. Shorter than the Calendar's own 21-day horizon on purpose — this
# is the "what am I doing today" panel, not the planning view.
DASH_WHATS_NEXT_HORIZON_DAYS = 7
DASH_WHATS_NEXT_LIMIT = 6

# "Needs attention": an open case nobody has touched in this many days. SCAN_LIMIT bounds the
# candidate set the roll-up walks — a workspace with thousands of cold leads must not turn the
# dashboard into a full-table scan, so the count is reported as a floor past that point.
DASH_STALE_DAYS = 14
DASH_STALE_LIMIT = 6
DASH_STALE_SCAN_LIMIT = 200

# A payment request that has been raised but not yet met. Mirrors client_payment_totals() in
# enterprise_payments.py — "created" is the only unpaid-and-still-live status; cancelled and
# failed are dead, and everything at or above "paid" is money in.
DASH_PAYMENT_OPEN_STATUS = "created"
DASH_PAYMENT_COLLECTED_STATUSES = ("paid", "transferred", "settled", "partially_refunded")


def _dash_scope(query, branch_id):
    """Apply the dashboard's optional office filter to an already scope-filtered aggregate."""
    if branch_id:
        return query.filter(models.EnterpriseClient.branch_id == int(branch_id))
    return query


@router.get("/dashboard")
def enterprise_dashboard(
    request: Request,
    branch_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="dashboard.view"
    )
    org_id = organization.id
    ctx = role.ctx

    # Every roll-up on this screen is built from its OWN query, so each one needs the scope filter
    # separately — a single missed aggregate would report whole-workspace numbers to someone who
    # can only see one office.
    def _scoped_clients():
        return scope_client_query(
            db.query(models.EnterpriseClient).filter(
                models.EnterpriseClient.organization_id == org_id
            ),
            ctx,
        )

    def _scoped_agg(*columns):
        return scope_client_query(
            db.query(*columns).filter(models.EnterpriseClient.organization_id == org_id),
            ctx,
        )

    base = _scoped_clients()
    if branch_id:
        base = base.filter(models.EnterpriseClient.branch_id == int(branch_id))
    total_clients = base.count()

    # Counts by status
    status_counts = {stage["key"]: 0 for stage in catalog.CLIENT_STAGES}
    for status_key, count in (
        _dash_scope(_scoped_agg(models.EnterpriseClient.status, func.count(models.EnterpriseClient.id)), branch_id)
        .group_by(models.EnterpriseClient.status)
        .all()
    ):
        if status_key in status_counts:
            status_counts[status_key] = int(count)

    open_stage_keys = {s["key"] for s in catalog.CLIENT_STAGES if s["is_open"]}
    active_clients = sum(v for k, v in status_counts.items() if k in open_stage_keys)
    approved = status_counts.get(catalog.STAGE_APPROVED, 0)
    rejected = status_counts.get(catalog.STAGE_REJECTED, 0)
    decided = approved + rejected
    approval_rate = round((approved / decided) * 100) if decided else None

    # Counts by category
    category_counts = []
    cat_raw = dict(
        _dash_scope(_scoped_agg(models.EnterpriseClient.visa_category, func.count(models.EnterpriseClient.id)), branch_id)
        .group_by(models.EnterpriseClient.visa_category)
        .all()
    )
    for cat in catalog.VISA_CATEGORIES:
        category_counts.append({
            "key": cat["key"],
            "label": cat["label"],
            "accent": cat["accent"],
            "count": int(cat_raw.get(cat["key"], 0)),
        })

    # Counts by specific visa type (top 6)
    visa_type_rows = (
        _dash_scope(_scoped_agg(models.EnterpriseClient.visa_type, func.count(models.EnterpriseClient.id)), branch_id)
        .group_by(models.EnterpriseClient.visa_type)
        .order_by(func.count(models.EnterpriseClient.id).desc())
        .limit(6)
        .all()
    )
    visa_type_counts = [
        {"visa_type": vt or "—", "count": int(count)} for vt, count in visa_type_rows
    ]

    # Top destination countries
    country_rows = (
        _dash_scope(
            _scoped_agg(
                models.EnterpriseClient.destination_country_code,
                func.count(models.EnterpriseClient.id),
            ),
            branch_id,
        )
        .group_by(models.EnterpriseClient.destination_country_code)
        .order_by(func.count(models.EnterpriseClient.id).desc())
        .limit(8)
        .all()
    )
    top_countries = []
    for code, count in country_rows:
        brief = _country_brief(code)
        brief["count"] = int(count)
        top_countries.append(brief)

    # New clients this month
    now = datetime.utcnow()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    new_this_month = (
        base.filter(models.EnterpriseClient.created_at >= month_start).count()
    )

    member_names = _org_member_name_map(db, org_id)
    _sensitive = ctx.has("clients.view_sensitive")

    org_tz, today, org_tz_label = ent_time.org_clock(db, org_id)

    # "What's next" — the Calendar's overdue + upcoming feed, folded into the dashboard so a
    # follow-up that has quietly aged out is visible on the screen people actually land on.
    # Same 30-day look-back as the Calendar and the sidebar badge, so the overdue count agrees
    # across all three; a shorter forward horizon because this is the daily-driver view.
    whats_next = None
    if ctx.has("calendar.view"):
        horizon_days = DASH_WHATS_NEXT_HORIZON_DAYS
        window = _collect_calendar_events(
            db, org_id,
            today - timedelta(days=30), today + timedelta(days=horizon_days),
            include_done=False, ctx=ctx, branch_id=branch_id, today=today,
        )
        wn_overdue = [e for e in window if e.get("overdue")]
        wn_upcoming = [
            e for e in window
            if not e.get("overdue") and (e["date"] or "") >= today.isoformat()
        ]
        whats_next = {
            "today": today.isoformat(),
            "timezone": org_tz,
            "timezone_label": org_tz_label,
            "horizon_days": horizon_days,
            # Totals are the honest counts; the lists are trimmed for the card, which links
            # out to the Calendar for the rest.
            "overdue_total": len(wn_overdue),
            "upcoming_total": len(wn_upcoming),
            "overdue": wn_overdue[:DASH_WHATS_NEXT_LIMIT],
            "upcoming": wn_upcoming[:DASH_WHATS_NEXT_LIMIT],
        }

    # Upcoming deadlines (target_date today or later), nearest first. Only built as the fallback
    # for a role that can see the dashboard but not the calendar — Finance, or a custom role
    # without calendar.view — since "What's next" already carries client key dates.
    upcoming = []
    if whats_next is None:
        upcoming = (
            base
            .filter(
                models.EnterpriseClient.target_date.isnot(None),
                models.EnterpriseClient.target_date >= today,
                models.EnterpriseClient.status.notin_([catalog.STAGE_APPROVED, catalog.STAGE_REJECTED]),
            )
            .order_by(models.EnterpriseClient.target_date.asc())
            .limit(8)
            .all()
        )

    # Recent clients
    recent = (
        base.order_by(models.EnterpriseClient.created_at.desc(), models.EnterpriseClient.id.desc())
        .limit(6)
        .all()
    )

    # ---- Needs attention: open cases that have stopped moving ----------------------------
    # "Moving" is deliberately wider than the client row's own updated_at. A counsellor who logs
    # a note, sends an email or takes in a document is working the case even though the record
    # itself never changed — counting only updated_at would flag the busiest files as stale.
    # The row's own timestamp is the OUTER bound (activity can only make a case look fresher),
    # so filtering on it first gives a superset the activity pass then narrows.
    stale_cutoff = now - timedelta(days=DASH_STALE_DAYS)
    client_touched_at = func.coalesce(
        models.EnterpriseClient.updated_at, models.EnterpriseClient.created_at
    )
    stale_candidates = (
        base
        .filter(
            models.EnterpriseClient.status.in_(open_stage_keys or {""}),
            client_touched_at < stale_cutoff,
        )
        .order_by(client_touched_at.asc())
        .limit(DASH_STALE_SCAN_LIMIT)
        .all()
    )

    def _naive(ts):
        """Postgres hands back tz-aware datetimes and SQLite naive ones; `now` is naive."""
        if ts is None:
            return None
        return ts.replace(tzinfo=None) if getattr(ts, "tzinfo", None) else ts

    stale_rows: list[tuple[models.EnterpriseClient, int]] = []
    if stale_candidates:
        candidate_ids = [c.id for c in stale_candidates]
        last_touch: dict[int, datetime] = {}
        for activity_model in (
            models.EnterpriseClientNote,
            models.EnterpriseClientEmail,
            models.EnterpriseClientDocument,
        ):
            for cid, ts in (
                db.query(activity_model.client_id, func.max(activity_model.created_at))
                .filter(
                    activity_model.organization_id == org_id,
                    activity_model.client_id.in_(candidate_ids),
                )
                .group_by(activity_model.client_id)
                .all()
            ):
                ts = _naive(ts)
                if ts is not None and (cid not in last_touch or ts > last_touch[cid]):
                    last_touch[cid] = ts
        for client in stale_candidates:
            own = _naive(client.updated_at) or _naive(client.created_at)
            latest = max([t for t in (own, last_touch.get(client.id)) if t is not None], default=None)
            if latest is None or latest >= stale_cutoff:
                continue
            stale_rows.append((client, max(0, (now - latest).days)))
        stale_rows.sort(key=lambda pair: pair[1], reverse=True)

    stalled = {
        "days": DASH_STALE_DAYS,
        "total": len(stale_rows),
        # Past the scan limit the total is a floor, not the truth — say so rather than
        # letting the card imply it counted everything.
        "truncated": len(stale_candidates) >= DASH_STALE_SCAN_LIMIT,
        "clients": [
            {
                **_serialize_client(client, member_names, include_sensitive=_sensitive),
                "days_stale": days,
            }
            for client, days in stale_rows[:DASH_STALE_LIMIT]
        ],
    }

    # ---- Money: what has been invoiced and not yet collected ------------------------------
    # Gated on finance.view, so a counsellor without the Finance section never sees org revenue.
    # Totals are grouped by currency and never summed across them — there is no FX rate here,
    # and inventing one is how a report ends up off by an order of magnitude.
    finance_snapshot = None
    if ctx.has("finance.view"):
        payments_q = db.query(models.EnterpriseStudentPayment).filter(
            models.EnterpriseStudentPayment.organization_id == org_id
        )
        # A payment inherits its client's scope. Orphaned rows (client deleted, client_id NULL)
        # stay workspace-scope only — the same rule _get_org_payment_or_404 enforces.
        if branch_id or getattr(ctx, "scope_kind", "all") != "all":
            visible_ids = {int(row[0]) for row in base.with_entities(models.EnterpriseClient.id).all()}
            payments_q = payments_q.filter(
                models.EnterpriseStudentPayment.client_id.in_(visible_ids or {-1})
            )

        def _money_by_currency(query):
            return [
                {
                    "currency": (currency or "INR").upper(),
                    "count": int(count or 0),
                    "amount_minor": int(total or 0),
                }
                for currency, count, total in (
                    query.with_entities(
                        models.EnterpriseStudentPayment.currency,
                        func.count(models.EnterpriseStudentPayment.id),
                        func.sum(models.EnterpriseStudentPayment.amount_paise),
                    )
                    .group_by(models.EnterpriseStudentPayment.currency)
                    .all()
                )
            ]

        open_payments = payments_q.filter(
            models.EnterpriseStudentPayment.status == DASH_PAYMENT_OPEN_STATUS
        )
        finance_snapshot = {
            "outstanding": _money_by_currency(open_payments),
            "overdue_count": open_payments.filter(
                models.EnterpriseStudentPayment.due_date.isnot(None),
                models.EnterpriseStudentPayment.due_date < today,
            ).count(),
            "collected_this_month": _money_by_currency(
                payments_q.filter(
                    models.EnterpriseStudentPayment.status.in_(DASH_PAYMENT_COLLECTED_STATUSES),
                    models.EnterpriseStudentPayment.paid_at >= month_start,
                )
            ),
        }

    pipeline = []
    for stage in catalog.CLIENT_STAGES:
        pipeline.append({
            "key": stage["key"],
            "label": stage["label"],
            "color": stage["color"],
            "count": status_counts.get(stage["key"], 0),
        })

    return {
        "organization": {
            "id": organization.id,
            "company_name": organization.company_name,
            "logo_url": _resolve_enterprise_logo_url(organization),
        },
        "permissions": _enterprise_permissions_for_role(role),
        "kpis": {
            "total_clients": total_clients,
            "active_clients": active_clients,
            "approved": approved,
            "rejected": rejected,
            "approval_rate": approval_rate,
            "new_this_month": new_this_month,
        },
        "pipeline": pipeline,
        "category_counts": category_counts,
        "visa_type_counts": visa_type_counts,
        "top_countries": top_countries,
        "whats_next": whats_next,
        "stalled": stalled,
        "finance_snapshot": finance_snapshot,
        "upcoming_deadlines": [
            _serialize_client(c, member_names, include_sensitive=_sensitive) for c in upcoming
        ],
        "recent_clients": [
            _serialize_client(c, member_names, include_sensitive=_sensitive) for c in recent
        ],
        "subscription": _serialize_subscription_state(billing.build_subscription_state(db, org_id)),
    }


# ===========================================================================
# Calendar — timelines, deadlines & next steps (derived + manual events)
# ===========================================================================

CALENDAR_EVENT_TYPES = {
    "reminder":    {"label": "Reminder",    "color": "#6366f1"},
    "task":        {"label": "Task",        "color": "#0ea5e9"},
    "follow_up":   {"label": "Follow-up",   "color": "#8b5cf6"},
    "appointment": {"label": "Appointment", "color": "#10b981"},
    "deadline":    {"label": "Deadline",    "color": "#f97316"},
    "other":       {"label": "Other",       "color": "#64748b"},
}
DEFAULT_CALENDAR_EVENT_TYPE = "reminder"
CALENDAR_DERIVED_TYPES = {
    "client_deadline": {"label": "Key date / deadline", "color": "#f97316"},
    # Same colour as the manual "Follow-up" event type: to the person reading the calendar these
    # are the same kind of thing, one typed into the intake form and one into a reminder.
    "client_followup": {"label": "Follow-up due",       "color": "#8b5cf6"},
    "passport_expiry": {"label": "Passport expires",    "color": "#f59e0b"},
}
CALENDAR_MAX_RANGE_DAYS = int(os.getenv("ENTERPRISE_CALENDAR_MAX_RANGE_DAYS", "100"))


def _calendar_event_types_payload() -> list[dict]:
    return [{"key": k, "label": v["label"], "color": v["color"]} for k, v in CALENDAR_EVENT_TYPES.items()]


def _normalize_calendar_event_type(raw: str | None) -> str:
    value = str(raw or "").strip().lower()
    return value if value in CALENDAR_EVENT_TYPES else DEFAULT_CALENDAR_EVENT_TYPE


def _parse_calendar_time_or_400(raw: str | None) -> Optional[str]:
    value = str(raw or "").strip()
    if not value:
        return None
    if not re.match(r"^([01]?\d|2[0-3]):[0-5]\d$", value):
        raise HTTPException(status_code=400, detail="Time must be in HH:MM 24-hour format.")
    hh, mm = value.split(":")
    return f"{int(hh):02d}:{mm}"


def _serialize_calendar_manual_event(
    ev: models.EnterpriseCalendarEvent,
    client_name: str | None,
    attachments: list[dict] | None = None,
    *,
    today: date,
) -> dict:
    cfg = CALENDAR_EVENT_TYPES.get(ev.event_type, CALENDAR_EVENT_TYPES[DEFAULT_CALENDAR_EVENT_TYPE])
    ev_date = ev.event_date
    # `today` is the org's date, passed in rather than read here: date.today() is the server
    # process's zone, which for an org west of it flips a tomorrow event to overdue.
    overdue = bool(ev_date and ev_date < today and not ev.is_done)
    # Reference files ride along on the event so the edit modal can render them with no extra
    # round-trip. `attachments` is passed in already batch-loaded — never looked up per event,
    # which on a 100-day month view would be a guaranteed N+1 (see cal_files.lists_by_event).
    files = attachments or []
    return {
        "attachments": files,
        "attachment_count": len(files),
        "id": f"manual-{ev.id}",
        "event_id": ev.id,
        "source": "manual",
        "type": ev.event_type,
        "type_label": cfg["label"],
        "color": cfg["color"],
        "date": ev_date.isoformat() if ev_date else None,
        "time": ev.event_time,
        "title": ev.title,
        "notes": ev.notes,
        "client_id": ev.client_id,
        "client_name": client_name,
        "notify_client": bool(ev.notify_client),
        "is_done": bool(ev.is_done),
        "editable": True,
        "overdue": overdue,
        "created_by_name": ev.created_by_name,
    }


def _serialize_calendar_derived_event(kind: str, client: models.EnterpriseClient, when, *, today: date) -> dict:
    cfg = CALENDAR_DERIVED_TYPES[kind]
    overdue = bool(when and when < today)
    return {
        "id": f"{kind}-{client.id}",
        "event_id": None,
        "source": "client",
        "type": kind,
        "type_label": cfg["label"],
        "color": cfg["color"],
        "date": when.isoformat() if when else None,
        "time": None,
        "title": client.full_name,
        "notes": None,
        "client_id": client.id,
        "client_name": client.full_name,
        "stage": _stage_brief(client.status, client.destination_country_code),
        "is_done": False,
        "editable": False,
        "overdue": overdue,
    }


def _collect_calendar_events(
    db: Session, organization_id: int, start: date, end: date,
    *, include_done: bool = True, ctx=None, branch_id: Optional[int] = None,
    today: Optional[date] = None,
) -> list[dict]:
    # Every "overdue" flag below is decided against the ORG's date, resolved once here so a
    # single response can't grade two events against two different notions of today.
    if today is None:
        today = ent_time.org_today(db, organization_id)
    member_names = _org_member_name_map(db, organization_id)

    # The calendar is three separate sources — manual events, client key dates and passport
    # expiries — so each is scoped in turn. A manual event tied to a client inherits that client's
    # scope; a standalone one belongs to whoever created it once scope is narrowed.
    scoped_ids: set[int] | None = None
    if ctx is not None and ctx.scope_kind != "all":
        scoped_ids = {
            int(row[0])
            for row in scope_client_query(
                db.query(models.EnterpriseClient.id).filter(
                    models.EnterpriseClient.organization_id == organization_id
                ),
                ctx,
            ).all()
        }

    # The dashboard's office filter narrows the same three sources by branch: an event belongs to
    # the office its client sits in. A standalone reminder has no office of its own, so it stays
    # with whoever wrote it rather than disappearing the moment a branch is picked.
    branch_client_ids: set[int] | None = None
    if branch_id:
        branch_client_ids = {
            int(row[0])
            for row in scope_client_query(
                db.query(models.EnterpriseClient.id).filter(
                    models.EnterpriseClient.organization_id == organization_id,
                    models.EnterpriseClient.branch_id == int(branch_id),
                ),
                ctx,
            ).all()
        }

    def _branch_scope(query):
        """Narrow a derived-event query (which selects clients) to the requested office."""
        if branch_id:
            return query.filter(models.EnterpriseClient.branch_id == int(branch_id))
        return query

    manual_q = (
        db.query(models.EnterpriseCalendarEvent)
        .filter(
            models.EnterpriseCalendarEvent.organization_id == organization_id,
            models.EnterpriseCalendarEvent.event_date >= start,
            models.EnterpriseCalendarEvent.event_date <= end,
        )
    )
    if not include_done:
        manual_q = manual_q.filter(models.EnterpriseCalendarEvent.is_done.is_(False))
    if scoped_ids is not None:
        manual_q = manual_q.filter(
            or_(
                models.EnterpriseCalendarEvent.client_id.in_(scoped_ids or {-1}),
                and_(
                    models.EnterpriseCalendarEvent.client_id.is_(None),
                    models.EnterpriseCalendarEvent.created_by_user_id == ctx.user_id,
                ),
            )
        )
    if branch_client_ids is not None:
        own_standalone = models.EnterpriseCalendarEvent.client_id.is_(None)
        if ctx is not None:
            own_standalone = and_(
                own_standalone, models.EnterpriseCalendarEvent.created_by_user_id == ctx.user_id
            )
        manual_q = manual_q.filter(
            or_(
                models.EnterpriseCalendarEvent.client_id.in_(branch_client_ids or {-1}),
                own_standalone,
            )
        )

    client_name_cache: dict[int, str] = {}

    def _client_name(cid: Optional[int]) -> Optional[str]:
        if not cid:
            return None
        if cid not in client_name_cache:
            row = (
                db.query(models.EnterpriseClient.full_name)
                .filter(
                    models.EnterpriseClient.id == cid,
                    models.EnterpriseClient.organization_id == organization_id,
                )
                .first()
            )
            client_name_cache[cid] = row[0] if row else None
        return client_name_cache[cid]

    manual_events = manual_q.all()
    # One query for every event's reference files, before the serializer loop.
    attachments_by_event = cal_files.lists_by_event(
        db, organization_id, [ev.id for ev in manual_events]
    )
    events = [
        _serialize_calendar_manual_event(
            ev, _client_name(ev.client_id), attachments_by_event.get(ev.id), today=today
        )
        for ev in manual_events
    ]

    # Derived: client key dates (target_date) — skip terminal cases
    for client in (
        _branch_scope(scope_client_query(
            db.query(models.EnterpriseClient).filter(
                models.EnterpriseClient.organization_id == organization_id,
                models.EnterpriseClient.target_date.isnot(None),
                models.EnterpriseClient.target_date >= start,
                models.EnterpriseClient.target_date <= end,
                models.EnterpriseClient.status.notin_([catalog.STAGE_APPROVED, catalog.STAGE_REJECTED]),
            ),
            ctx,
        ))
        .all()
    ):
        events.append(_serialize_calendar_derived_event("client_deadline", client, client.target_date, today=today))

    # Derived: the "Next follow-up" date a counsellor sets on the intake record. Same skip of
    # decided cases as the key date — an approved or refused applicant needs no chasing.
    for client in (
        _branch_scope(scope_client_query(
            db.query(models.EnterpriseClient).filter(
                models.EnterpriseClient.organization_id == organization_id,
                models.EnterpriseClient.next_followup_date.isnot(None),
                models.EnterpriseClient.next_followup_date >= start,
                models.EnterpriseClient.next_followup_date <= end,
                models.EnterpriseClient.status.notin_([catalog.STAGE_APPROVED, catalog.STAGE_REJECTED]),
            ),
            ctx,
        ))
        .all()
    ):
        events.append(
            _serialize_calendar_derived_event("client_followup", client, client.next_followup_date, today=today)
        )

    # Derived: passport expiries in range (any active client)
    for client in (
        _branch_scope(scope_client_query(
            db.query(models.EnterpriseClient).filter(
                models.EnterpriseClient.organization_id == organization_id,
                models.EnterpriseClient.passport_expiry.isnot(None),
                models.EnterpriseClient.passport_expiry >= start,
                models.EnterpriseClient.passport_expiry <= end,
            ),
            ctx,
        ))
        .all()
    ):
        events.append(_serialize_calendar_derived_event("passport_expiry", client, client.passport_expiry, today=today))

    events.sort(key=lambda e: (e["date"] or "", e["time"] or "99:99", e["title"] or ""))
    return events


class EnterpriseCalendarEventCreate(BaseModel):
    title: str = Field(..., min_length=1, max_length=200)
    event_date: str = Field(..., min_length=8, max_length=10)
    event_type: str = Field(default=DEFAULT_CALENDAR_EVENT_TYPE, max_length=20)
    event_time: Optional[str] = Field(default=None, max_length=5)
    notes: Optional[str] = Field(default=None, max_length=2000)
    client_id: Optional[int] = None
    notify_client: Optional[bool] = None
    # Reference files uploaded while this reminder was still being composed are held under
    # this per-modal token; creating the event binds them to it.
    attachment_draft_token: Optional[str] = Field(
        default=None, max_length=cal_files.DRAFT_TOKEN_MAX_LEN
    )


class EnterpriseCalendarEventUpdate(BaseModel):
    title: Optional[str] = Field(default=None, min_length=1, max_length=200)
    event_date: Optional[str] = Field(default=None, min_length=8, max_length=10)
    event_type: Optional[str] = Field(default=None, max_length=20)
    event_time: Optional[str] = Field(default=None, max_length=5)
    notes: Optional[str] = Field(default=None, max_length=2000)
    client_id: Optional[int] = None
    notify_client: Optional[bool] = None
    is_done: Optional[bool] = None


def _reject_backdated_client_notice(db: Session, organization_id: int, ev) -> None:
    """Refuse to arm a client notification on a reminder that is already due.

    Logging a past event is legitimate; doing it with client notification on is not. The
    overnight digest sweeps `today - 14 days <= event_date <= today` for events that are
    not done and not yet notified (see services/enterprise_calendar_reminders.py), so a
    back-dated reminder with notify-on doesn't quietly record history — it emails the
    student to say their own appointment is "6 days overdue".

    Checked against the RESULTING row, because the update path can move the date and the
    notify flag independently. Rows the job would skip anyway — already sent, marked done,
    no linked client — are left alone so ordinary edits to old events still save.
    """
    if not (ev.notify_client and ev.client_id and not ev.is_done and ev.client_notified_at is None):
        return
    if ev.event_date and ev.event_date < ent_time.org_today(db, int(organization_id)):
        raise HTTPException(
            status_code=400,
            detail=(
                "This reminder is dated in the past, so notifying the client would email them "
                "that it is already overdue. Switch off the client notification to log it, or "
                "move it to today or later."
            ),
        )


@router.get("/calendar")
def enterprise_calendar(
    request: Request,
    start: Optional[str] = None,
    end: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="calendar.view")

    org_tz, today, org_tz_label = ent_time.org_clock(db, organization.id)
    if start:
        start_date = _parse_iso_date_or_400(start, "start")
    else:
        start_date = today.replace(day=1)
    if end:
        end_date = _parse_iso_date_or_400(end, "end")
    else:
        # default to the end of the start month
        nm = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        end_date = nm - timedelta(days=1)
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="End date must be on or after the start date.")
    if (end_date - start_date).days > CALENDAR_MAX_RANGE_DAYS:
        raise HTTPException(status_code=400, detail=f"Date range too large (max {CALENDAR_MAX_RANGE_DAYS} days).")

    events = _collect_calendar_events(db, organization.id, start_date, end_date, ctx=role.ctx, today=today)
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "start": start_date.isoformat(),
        "end": end_date.isoformat(),
        "today": today.isoformat(),
        # Event times are wall-clock in this zone. The client prints the label next to them —
        # without it "14:30" means a different moment to every reader. Both are null/"" until
        # an office sets a zone, so the UI shows a bare time rather than asserting UTC.
        "timezone": org_tz,
        "timezone_label": org_tz_label,
        "event_types": _calendar_event_types_payload(),
        "events": events,
    }


@router.get("/calendar/upcoming")
def enterprise_calendar_upcoming(
    request: Request,
    days: int = 14,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """The 'what's next' feed: overdue + the next N days of deadlines and reminders."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="calendar.view")
    days = max(1, min(int(days or 14), 90))
    org_tz, today, org_tz_label = ent_time.org_clock(db, organization.id)
    # Look back 30 days so overdue items still surface, forward `days`.
    window = _collect_calendar_events(
        db, organization.id, today - timedelta(days=30), today + timedelta(days=days),
        include_done=False, ctx=role.ctx, today=today,
    )
    overdue = [e for e in window if e.get("overdue")]
    upcoming = [e for e in window if not e.get("overdue") and (e["date"] or "") >= today.isoformat()]
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "today": today.isoformat(),
        "timezone": org_tz,
        "timezone_label": org_tz_label,
        "horizon_days": days,
        "overdue": overdue,
        "upcoming": upcoming[:40],
    }


@router.get("/calendar/clients")
def enterprise_calendar_clients(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Lightweight client list for the @-mention autocomplete in the reminder form.
    Scoped to the clients this member may actually open — an unscoped picker would hand over
    every name and email address in the workspace."""
    _, organization, _role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="calendar.view"
    )
    rows = (
        scope_client_query(
            db.query(
                models.EnterpriseClient.id,
                models.EnterpriseClient.full_name,
                models.EnterpriseClient.email,
            ).filter(models.EnterpriseClient.organization_id == organization.id),
            _role.ctx,
        )
        .order_by(models.EnterpriseClient.full_name.asc())
        .all()
    )
    return {
        "clients": [
            {"id": r[0], "name": r[1], "email": (r[2] or None), "has_email": bool((r[2] or "").strip())}
            for r in rows
        ]
    }


@router.post("/calendar/events")
def enterprise_calendar_create_event(
    payload: EnterpriseCalendarEventCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="calendar.manage"
    )
    title = (payload.title or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="A title is required.")
    event_date = _parse_iso_date_or_400(payload.event_date, "Event date")
    event_time = _parse_calendar_time_or_400(payload.event_time)
    if payload.client_id is not None:
        _get_org_client_or_404(db, organization.id, payload.client_id, ctx=role.ctx)

    ev = models.EnterpriseCalendarEvent(
        organization_id=organization.id,
        client_id=payload.client_id,
        title=title[:200],
        notes=(payload.notes or None),
        event_type=_normalize_calendar_event_type(payload.event_type),
        event_date=event_date,
        event_time=event_time,
        is_done=False,
        # Only meaningful when a client is linked; the client is emailed when due.
        notify_client=bool(payload.notify_client) and payload.client_id is not None,
        created_by_user_id=current_user.id,
        created_by_name=current_user.full_name or current_user.email,
    )
    _reject_backdated_client_notice(db, organization.id, ev)
    db.add(ev)
    # flush, not commit: binding the draft uploads needs ev.id while still inside this
    # transaction, so a failed bind rolls the event back with it rather than leaving a
    # reminder whose files silently went missing.
    db.flush()
    bound = cal_files.bind_draft_to_event(
        db,
        organization_id=organization.id,
        user_id=current_user.id,
        token=payload.attachment_draft_token,
        event_id=ev.id,
    )
    attachments = [cal_files.serialize(row) for row in bound]
    db.commit()
    db.refresh(ev)
    client_name = None
    if ev.client_id:
        c = _get_org_client_or_404(db, organization.id, ev.client_id, ctx=role.ctx)
        client_name = c.full_name
    return {
        "message": "Event added.",
        "permissions": _enterprise_permissions_for_role(role),
        "event": _serialize_calendar_manual_event(
            ev, client_name, attachments, today=ent_time.org_today(db, organization.id)
        ),
    }


def _get_org_calendar_event_or_404(
    db: Session, organization_id: int, event_id: int, *, ctx=None
) -> models.EnterpriseCalendarEvent:
    ev = (
        db.query(models.EnterpriseCalendarEvent)
        .filter(
            models.EnterpriseCalendarEvent.id == int(event_id),
            models.EnterpriseCalendarEvent.organization_id == int(organization_id),
        )
        .first()
    )
    if not ev:
        raise HTTPException(status_code=404, detail="Calendar event not found.")
    # An event that hangs off a client inherits that client's scope — otherwise a scope-restricted
    # member could read, re-point or delete a colleague's appointment for a case they can't see.
    # A standalone event (no client) belongs to whoever created it once scope is narrowed.
    if ctx is not None and ctx.scope_kind != "all":
        if ev.client_id:
            linked = (
                db.query(models.EnterpriseClient)
                .filter(
                    models.EnterpriseClient.id == int(ev.client_id),
                    models.EnterpriseClient.organization_id == int(organization_id),
                )
                .first()
            )
            if not client_in_scope(linked, ctx):
                raise HTTPException(status_code=404, detail="Calendar event not found.")
        elif int(getattr(ev, "created_by_user_id", 0) or 0) != ctx.user_id:
            raise HTTPException(status_code=404, detail="Calendar event not found.")
    return ev


@router.patch("/calendar/events/{event_id}")
def enterprise_calendar_update_event(
    event_id: int,
    payload: EnterpriseCalendarEventUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="calendar.manage"
    )
    ev = _get_org_calendar_event_or_404(db, organization.id, event_id, ctx=role.ctx)

    if payload.title is not None:
        t = payload.title.strip()
        if not t:
            raise HTTPException(status_code=400, detail="A title is required.")
        ev.title = t[:200]
    if payload.event_date is not None:
        ev.event_date = _parse_iso_date_or_400(payload.event_date, "Event date")
    if payload.event_type is not None:
        ev.event_type = _normalize_calendar_event_type(payload.event_type)
    if payload.event_time is not None:
        ev.event_time = _parse_calendar_time_or_400(payload.event_time)
    if payload.notes is not None:
        ev.notes = payload.notes or None
    if payload.client_id is not None:
        new_cid = None if payload.client_id == 0 else payload.client_id
        if new_cid is not None:
            _get_org_client_or_404(db, organization.id, new_cid, ctx=role.ctx)
        if new_cid != ev.client_id:
            # Re-link → allow the new client to be notified afresh.
            ev.client_notified_at = None
        ev.client_id = new_cid
        if new_cid is None:
            ev.notify_client = False
    if payload.notify_client is not None:
        ev.notify_client = bool(payload.notify_client) and ev.client_id is not None
    if payload.is_done is not None:
        ev.is_done = bool(payload.is_done)

    _reject_backdated_client_notice(db, organization.id, ev)
    db.commit()
    db.refresh(ev)
    client_name = None
    if ev.client_id:
        c = _get_org_client_or_404(db, organization.id, ev.client_id, ctx=role.ctx)
        client_name = c.full_name
    return {
        "message": "Event updated.",
        "permissions": _enterprise_permissions_for_role(role),
        # Re-read the files rather than defaulting to []: the client replaces its copy of the
        # event with whatever comes back, so an empty list here would blank the tray.
        "event": _serialize_calendar_manual_event(
            ev, client_name,
            [cal_files.serialize(r) for r in cal_files.for_event(db, organization.id, ev.id)],
            today=ent_time.org_today(db, organization.id),
        ),
    }


@router.delete("/calendar/events/{event_id}")
def enterprise_calendar_delete_event(
    event_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="calendar.manage"
    )
    ev = _get_org_calendar_event_or_404(db, organization.id, event_id, ctx=role.ctx)
    # Collect the blob keys BEFORE the delete: the attachment rows go away with the event
    # (cascade), but nothing tells R2, and an orphaned ciphertext blob is a bill forever.
    storage_keys = cal_files.purge_for_events(
        db, organization_id=organization.id, event_ids=[ev.id]
    )
    db.delete(ev)
    db.commit()
    # Rows first, bytes second — a rolled-back commit must never leave a live row pointing
    # at files that are already gone.
    cal_files.drop_blobs(storage_keys)
    return {"message": "Event deleted.", "permissions": _enterprise_permissions_for_role(role)}


# ---------------------------------------------------------------------------
# Calendar reference files (private R2 storage, authenticated streaming).
#
# Uploads are immediate rather than deferred to the form submit, because the calendar form
# posts JSON — a file input inside it would be silently dropped. An event that already
# exists takes the file straight away; a reminder still being composed parks it under a
# per-modal draft token that POST /calendar/events binds.
# ---------------------------------------------------------------------------

ENTERPRISE_CAL_ATTACH_RATE_LIMIT = int(os.getenv("ENTERPRISE_CAL_ATTACH_RATE_LIMIT", "120"))
ENTERPRISE_CAL_ATTACH_RATE_WINDOW_SECONDS = int(
    os.getenv("ENTERPRISE_CAL_ATTACH_RATE_WINDOW_SECONDS", "3600")
)


def _calendar_attachment_or_404(
    db: Session,
    *,
    organization_id: int,
    attachment_id: int,
    user_id: int,
    ctx=None,
) -> models.EnterpriseCalendarEventAttachment:
    """Resolve one attachment, or 404.

    A bound file inherits its event's record scope — routed through
    _get_org_calendar_event_or_404 so a branch- or assigned-scoped member can never reach a
    colleague's file. An unbound draft has no event to inherit from, so it belongs to its
    uploader alone.
    """
    row = (
        db.query(models.EnterpriseCalendarEventAttachment)
        .filter(
            models.EnterpriseCalendarEventAttachment.id == int(attachment_id),
            models.EnterpriseCalendarEventAttachment.organization_id == int(organization_id),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Attachment not found.")
    if row.event_id:
        _get_org_calendar_event_or_404(db, organization_id, row.event_id, ctx=ctx)
    elif int(row.uploaded_by_user_id or 0) != int(user_id):
        raise HTTPException(status_code=404, detail="Attachment not found.")
    return row


@router.post("/calendar/attachments")
async def enterprise_upload_calendar_attachment(
    request: Request,
    file: UploadFile = File(...),
    event_id: Optional[int] = Form(None),
    draft_token: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="calendar.manage"
    )
    # Uploads are cheap to trigger and expensive to store, and there is no per-org storage
    # quota anywhere in this app — so the rate limit is a real ceiling, not a formality.
    _enforce_rate_limit_or_429(
        request=request,
        scope="enterprise.calendar_attachment",
        limit=ENTERPRISE_CAL_ATTACH_RATE_LIMIT,
        window_seconds=ENTERPRISE_CAL_ATTACH_RATE_WINDOW_SECONDS,
        extra_key=f"org:{organization.id}:user:{current_user.id}",
    )
    if not enterprise_storage.is_configured():
        raise HTTPException(status_code=503, detail="File storage is not configured.")

    token = cal_files.normalize_draft_token(draft_token)
    ev = None
    if event_id:
        # Scope-checked: attaching to an event you cannot see must 404 like everything else.
        ev = _get_org_calendar_event_or_404(db, organization.id, int(event_id), ctx=role.ctx)
        used_count, used_bytes = cal_files.usage_for_event(db, organization.id, ev.id)
    elif token:
        used_count, used_bytes = cal_files.usage_for_draft(db, organization.id, token, current_user.id)
    else:
        raise HTTPException(status_code=400, detail="Nothing to attach this file to.")

    data = await file.read()
    try:
        ext = cal_files.validate_new_file(
            filename=file.filename,
            size=len(data),
            existing_count=used_count,
            existing_bytes=used_bytes,
        )
    except cal_files.CalendarFileError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    try:
        # cal_files.store does a blocking R2 PUT (and a DB write); offload it so this async
        # handler — async only for `await file.read()` — doesn't run it on the event loop.
        # The await suspends until it returns, so `db` is never touched from two threads at once.
        row = await run_in_threadpool(
            cal_files.store,
            db,
            organization_id=organization.id,
            event_id=(ev.id if ev else None),
            draft_token=(None if ev else token),
            filename=file.filename,
            data=data,
            content_type=file.content_type,
            uploaded_by_user_id=current_user.id,
            uploaded_by_name=(current_user.full_name or current_user.email),
            ext=ext,
        )
    except Exception:
        logger.exception("Failed to store calendar attachment (org_id=%s)", organization.id)
        raise HTTPException(status_code=502, detail="Could not upload that file right now. Please try again.")

    # Reclaim uploads abandoned by modals this user closed without saving.
    swept = cal_files.sweep_stale_drafts(db, organization_id=organization.id, user_id=current_user.id)
    try:
        db.commit()
    except Exception:
        # The bytes are already in the bucket but no row will point at them, and nothing else
        # sweeps orphans — so drop them here rather than pay for them forever.
        db.rollback()
        logger.exception("Calendar attachment commit failed (org_id=%s)", organization.id)
        cal_files.drop_blobs([row.storage_key])
        raise HTTPException(status_code=502, detail="Could not attach that file right now. Please try again.")
    db.refresh(row)
    cal_files.drop_blobs(swept)
    return {
        "message": "File attached.",
        "permissions": _enterprise_permissions_for_role(role),
        "attachment": cal_files.serialize(row),
    }


@router.delete("/calendar/attachments/drafts/{draft_token}")
def enterprise_discard_calendar_draft_attachments(
    draft_token: str,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Drop a cancelled modal's uploads now instead of waiting for the TTL sweep.

    A courtesy, never the only cleanup: Escape and overlay-click close the modal with no hook
    to call this at all, which is exactly what the sweep is for.
    """
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="calendar.manage"
    )
    keys = cal_files.discard_draft(
        db, organization_id=organization.id, user_id=current_user.id, token=draft_token
    )
    db.commit()
    cal_files.drop_blobs(keys)
    return {
        "message": "Draft attachments discarded.",
        "discarded": len(keys),
        "permissions": _enterprise_permissions_for_role(role),
    }


@router.delete("/calendar/attachments/{attachment_id}")
def enterprise_delete_calendar_attachment(
    attachment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="calendar.manage"
    )
    row = _calendar_attachment_or_404(
        db,
        organization_id=organization.id,
        attachment_id=attachment_id,
        user_id=current_user.id,
        ctx=role.ctx,
    )
    storage_key = row.storage_key
    db.delete(row)
    # Row first, blob second: a failed commit must never leave a surviving row pointing at
    # bytes that are already gone.
    db.commit()
    cal_files.drop_blobs([storage_key])
    return {"message": "Attachment removed.", "permissions": _enterprise_permissions_for_role(role)}


@router.get("/calendar/attachments/{attachment_id}/download")
def enterprise_download_calendar_attachment(
    attachment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="calendar.view"
    )
    row = _calendar_attachment_or_404(
        db,
        organization_id=organization.id,
        attachment_id=attachment_id,
        user_id=current_user.id,
        ctx=role.ctx,
    )
    try:
        data = enterprise_storage.fetch_document(row.storage_key)
    except Exception:
        logger.exception("Failed to fetch calendar attachment id=%s", row.id)
        raise HTTPException(status_code=502, detail="Could not retrieve that file right now.")

    # The served Content-Type is derived from the validated extension, never from the stored
    # mime_type: a *.pdf can arrive claiming text/html with a <script> body, and served inline
    # under this app's 'unsafe-inline' CSP that would execute on the workspace origin.
    media_type, disposition = cal_files.download_headers_for(row.original_filename)
    filename = cal_files.safe_filename(row.original_filename)
    return Response(
        content=data,
        media_type=media_type,
        headers={
            "Content-Disposition": f'{disposition}; filename="{filename}"',
            "X-Content-Type-Options": "nosniff",
        },
    )


# ===========================================================================
# Help & Support + feature requests
# ===========================================================================

ENTERPRISE_SUPPORT_RATE_LIMIT = int(os.getenv("ENTERPRISE_SUPPORT_RATE_LIMIT", "8"))
ENTERPRISE_SUPPORT_RATE_WINDOW_SECONDS = int(os.getenv("ENTERPRISE_SUPPORT_RATE_WINDOW_SECONDS", "3600"))
SUPPORT_REQUEST_TYPES = {"support", "feature_request"}
# Screenshots and sample files ride along as email attachments to the support inbox — they
# are never written to our bucket, so the only ceiling that matters is what an email can carry.
ENTERPRISE_SUPPORT_ATTACH_MAX_FILES = int(os.getenv("ENTERPRISE_SUPPORT_ATTACH_MAX_FILES", "5"))
ENTERPRISE_SUPPORT_ATTACH_MAX_TOTAL_BYTES = int(
    os.getenv("ENTERPRISE_SUPPORT_ATTACH_MAX_TOTAL_BYTES", str(10 * 1024 * 1024))
)


def _normalize_support_type(raw: str | None) -> str:
    value = str(raw or "").strip().lower()
    if value in {"feature", "feature_request", "feature-request", "idea"}:
        return "feature_request"
    return "support"


def _support_attachment_manifest(r: models.EnterpriseSupportRequest) -> list[dict]:
    """The [{filename, size}] manifest of what the requester attached (best-effort — rows
    written before this column existed simply have none)."""
    raw = getattr(r, "attachments_json", None)
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
    except Exception:
        return []
    out = []
    for item in parsed if isinstance(parsed, list) else []:
        if not isinstance(item, dict):
            continue
        name = str(item.get("filename") or "").strip()
        if not name:
            continue
        try:
            size = int(item.get("size") or 0)
        except Exception:
            size = 0
        out.append({"filename": name[:160], "size": max(0, size)})
    return out


def _serialize_support_request(r: models.EnterpriseSupportRequest) -> dict:
    attachments = _support_attachment_manifest(r)
    return {
        "id": r.id,
        "request_type": r.request_type,
        "type_label": "Feature request" if r.request_type == "feature_request" else "Help & support",
        "subject": r.subject,
        "message": r.message,
        "status": r.status,
        "requester_name": r.requester_name,
        "attachments": attachments,
        "attachment_count": len(attachments),
        "created_at": _iso(r.created_at),
    }


@router.get("/support")
def enterprise_support_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="support.manage")
    rows = (
        db.query(models.EnterpriseSupportRequest)
        .filter(models.EnterpriseSupportRequest.organization_id == organization.id)
        .order_by(models.EnterpriseSupportRequest.created_at.desc(), models.EnterpriseSupportRequest.id.desc())
        .limit(25)
        .all()
    )
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "support_email": os.getenv("ENTERPRISE_SUPPORT_INBOX", "contact@rilono.com").strip() or "contact@rilono.com",
        "attach_max_files": ENTERPRISE_SUPPORT_ATTACH_MAX_FILES,
        "attach_max_total_bytes": ENTERPRISE_SUPPORT_ATTACH_MAX_TOTAL_BYTES,
        "requests": [_serialize_support_request(r) for r in rows],
    }


async def _read_support_attachments(uploads: list[UploadFile]) -> list[dict]:
    """Validate and read the files a requester attached, returning email-attachment dicts of
    {filename, content, content_type}. Extension allow-list + a total-size cap only; nothing
    is stored, so the derived content type is what the inbox sees (never the client's)."""
    picked: list[dict] = []
    total = 0
    for upload in uploads:
        if upload is None or not (upload.filename or "").strip():
            continue
        if len(picked) >= ENTERPRISE_SUPPORT_ATTACH_MAX_FILES:
            raise HTTPException(
                status_code=400,
                detail=f"You can attach up to {ENTERPRISE_SUPPORT_ATTACH_MAX_FILES} files.",
            )
        original = _safe_filename(upload.filename)
        ext = os.path.splitext(original)[1].lower()
        if ext not in ENTERPRISE_DOC_ALLOWED_EXT:
            raise HTTPException(
                status_code=400,
                detail=f"“{original}” isn't a supported file type. Allowed: PDF, images, Word/Excel, CSV, or text.",
            )
        data = await upload.read()
        if not data:
            raise HTTPException(status_code=400, detail=f"“{original}” is empty.")
        total += len(data)
        if total > ENTERPRISE_SUPPORT_ATTACH_MAX_TOTAL_BYTES:
            limit_mb = ENTERPRISE_SUPPORT_ATTACH_MAX_TOTAL_BYTES // (1024 * 1024)
            raise HTTPException(
                status_code=400,
                detail=f"Attachments add up to more than {limit_mb} MB. Send the largest one separately.",
            )
        picked.append({
            "filename": original,
            "content": data,
            "content_type": ENTERPRISE_DOC_EXT_MIME.get(ext, "application/octet-stream"),
        })
    return picked


@router.post("/support")
async def enterprise_support_create(
    request: Request,
    request_type: Optional[str] = Form(None),
    subject: Optional[str] = Form(None),
    message: Optional[str] = Form(None),
    files: Optional[list[UploadFile]] = File(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="support.request")
    _enforce_rate_limit_or_429(
        request=request,
        scope="enterprise.support",
        limit=ENTERPRISE_SUPPORT_RATE_LIMIT,
        window_seconds=ENTERPRISE_SUPPORT_RATE_WINDOW_SECONDS,
        extra_key=str(current_user.id),
    )
    # The portal posts multipart (so files can ride along), but a tab still running the
    # pre-attachments enterprise.js posts JSON — accept both so a mid-session deploy never
    # swallows someone's support request.
    if subject is None and message is None:
        body = {}
        try:
            parsed = await request.json()
            if isinstance(parsed, dict):
                body = parsed
        except Exception:
            body = {}
        request_type = request_type or body.get("request_type")
        subject = body.get("subject")
        message = body.get("message")

    request_type = _normalize_support_type(request_type)
    subject = str(subject or "").strip()
    message = str(message or "").strip()
    if len(subject) < 3:
        raise HTTPException(status_code=400, detail="Please add a short subject.")
    if len(message) < 5:
        raise HTTPException(status_code=400, detail="Please describe your request.")
    attachments = await _read_support_attachments(files or [])

    requester_name = current_user.full_name or current_user.email
    row = models.EnterpriseSupportRequest(
        organization_id=organization.id,
        user_id=current_user.id,
        requester_name=requester_name,
        requester_email=current_user.email,
        request_type=request_type,
        subject=subject[:160],
        message=message[:4000],
        status="open",
        attachments_json=json.dumps(
            [{"filename": a["filename"], "size": len(a["content"])} for a in attachments]
        ) if attachments else None,
    )
    db.add(row)
    db.commit()
    db.refresh(row)

    # Notify the support inbox (best-effort — never fail the request on email error). The
    # Resend call is a blocking HTTPS round trip; offload it so it doesn't sit on the event
    # loop (this handler is async for `await request.json()` / reading the uploads).
    try:
        await run_in_threadpool(
            send_enterprise_support_request_email,
            request_type=request_type,
            subject=subject,
            message=message,
            org_name=organization.company_name or "Unknown organization",
            requester_name=requester_name,
            requester_email=current_user.email or "",
            attachments=attachments,
        )
    except Exception:
        logger.exception("Failed to email enterprise support request (org_id=%s)", organization.id)

    # Feature requests get a warm confirmation back to the requester (best-effort — a
    # confirmation-email failure must never fail the request, mirroring the block above).
    if request_type == "feature_request" and (current_user.email or "").strip():
        try:
            await run_in_threadpool(
                send_feature_request_confirmation,
                to_email=current_user.email,
                full_name=requester_name,
                request_summary=subject,
                product="Rilono Enterprise",
            )
        except Exception:
            logger.exception("Failed to send enterprise feature-request confirmation (org_id=%s)", organization.id)

    friendly = "Thanks! Your feature request is in — we read every one." if request_type == "feature_request" \
        else "Thanks! Our team has your message and will get back to you by email."
    return {
        "message": friendly,
        "permissions": _enterprise_permissions_for_role(role),
        "request": _serialize_support_request(row),
    }


# ===========================================================================
# Billing (per-organization subscriptions) + self-serve signup
# ===========================================================================

def _razorpay_credentials() -> tuple[str, str]:
    return (os.getenv("RAZORPAY_KEY_ID", "").strip(), os.getenv("RAZORPAY_KEY_SECRET", "").strip())


def _razorpay_enabled() -> bool:
    key_id, key_secret = _razorpay_credentials()
    return bool(key_id and key_secret)


def _resolve_charge_currency(hint: Optional[str], organization) -> str:
    """Decide what currency to CHARGE an organization in.

    Precedence: explicit hint from the buyer -> the currency they were last billed in
    (organization.billing_currency, if set) -> the org's country -> INR.

    The hint is validated strictly: an unrecognised or not-yet-enabled code is a 400,
    never a silent coercion to INR. Coercing here would charge someone in a currency
    they did not pick, which is exactly the class of bug this whole module exists to
    prevent. Note this decides only the CURRENCY — the amount always comes from the
    server-side price book.
    """
    raw = (hint or "").strip()
    if raw:
        try:
            return money.normalize_currency(raw, strict=True)
        except money.UnsupportedCurrency:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"We can't charge in {raw.upper()[:8]} yet. Supported: "
                    + ", ".join(money.supported_charge_currencies()) + "."
                ),
            )
    stored = (getattr(organization, "billing_currency", None) or "").strip()
    if stored and money.is_chargeable(stored):
        return money.normalize_currency(stored, strict=True)
    country = (getattr(organization, "country_code", None) or "").strip().upper()
    if country:
        from app.routers import pricing as pricing_fx
        guess = pricing_fx._COUNTRY_TO_CURRENCY.get(country)
        if guess and money.is_chargeable(guess):
            return money.normalize_currency(guess, strict=True)
    return money.DEFAULT_CURRENCY


def _razorpay_request(method: str, path: str, json_payload: dict | None = None) -> dict:
    key_id, key_secret = _razorpay_credentials()
    url = f"{RAZORPAY_API_BASE}/{path.lstrip('/')}"
    try:
        resp = requests.request(method.upper(), url, auth=(key_id, key_secret), json=json_payload, timeout=15)
    except requests.RequestException:
        raise HTTPException(status_code=502, detail="Failed to contact the payment gateway.")
    if resp.status_code >= 400:
        raise HTTPException(status_code=502, detail="Unable to process the payment request right now.")
    try:
        data = resp.json()
    except ValueError:
        raise HTTPException(status_code=502, detail="Invalid response from the payment gateway.")
    if not isinstance(data, dict):
        raise HTTPException(status_code=502, detail="Unexpected response from the payment gateway.")
    return data


# ---------------------------------------------------------------------------
# Finance / marketplace payments (Razorpay Route) — Phase 1: onboarding
#
# The "Finance" portal section: collect payments from students, and (later)
# revenue analytics + usage billing as nested sub-views. Rilono is a technology
# platform, NOT a payment aggregator — student money is collected into Razorpay's
# PA escrow and settled by Razorpay directly to the consultancy's own linked-account
# bank; Rilono never takes custody of it. This phase lets a consultancy connect its
# company bank account (a Route Linked Account). Collection/checkout, webhooks and
# refunds are later phases.
# ---------------------------------------------------------------------------

class EnterpriseLinkedAccountRequest(BaseModel):
    legal_business_name: str | None = Field(None, max_length=200)
    business_type: str | None = Field(None, max_length=60)
    contact_name: str | None = Field(None, max_length=120)
    contact_email: str | None = Field(None, max_length=200)
    contact_phone: str | None = Field(None, max_length=20)
    business_pan: str | None = Field(None, max_length=20)
    gst_number: str | None = Field(None, max_length=20)
    bank_account_number: str | None = Field(None, max_length=40)
    bank_ifsc: str | None = Field(None, max_length=20)
    beneficiary_name: str | None = Field(None, max_length=200)
    attested_service_delivery: bool = False
    attested_turnover_ok: bool = False


@router.get("/finance/summary")
def enterprise_finance_summary(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="finance.view"
    )
    la = enterprise_payments.get_linked_account(db, organization.id)
    # The settlement identity fields (bank IFSC/last 4, beneficiary name, GST, Razorpay account
    # id) are keyed to the capability that can CHANGE them, which is owner-only. Anyone else —
    # including a member who can invoice and refund nothing — sees the status and nothing more.
    is_admin = role.ctx.has("finance.payout_account")
    return {
        "payments_enabled": enterprise_payments.razorpay_enabled(),
        "linked_account": enterprise_payments.serialize_linked_account(la, include_sensitive=is_admin),
        "fee": enterprise_payments.fee_config_public(),
        "permissions": _enterprise_permissions_for_role(role),
    }


@router.post("/finance/linked-account")
def enterprise_finance_connect_bank(
    payload: EnterpriseLinkedAccountRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Connect (or update) the consultancy's Razorpay Route linked account.

    Admin-only. Requires the eligibility attestation (the consultancy itself delivers the
    service to the student and meets the turnover threshold) — RBI/Route only permits a
    split payee that interfaces with the payer for the goods/services. When Razorpay Route
    is enabled, this runs the v2 Accounts onboarding sequence and stores the returned ids +
    activation status; otherwise it saves the local record so activation can begin later.
    """
    _, organization, _ = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="finance.payout_account"
    )
    if not (payload.attested_service_delivery and payload.attested_turnover_ok):
        raise HTTPException(
            status_code=400,
            detail=(
                "To collect payments, please confirm that your organization directly delivers the "
                "service to the student and meets the eligibility criteria."
            ),
        )

    la = enterprise_payments.get_linked_account(db, organization.id)
    if la is None:
        la = models.EnterpriseLinkedAccount(
            organization_id=organization.id,
            created_by_user_id=current_user.id,
        )
        db.add(la)

    # Record the business + settlement details (display-safe only) and the attestation.
    la.legal_business_name = (payload.legal_business_name or "").strip() or la.legal_business_name
    la.business_type = (payload.business_type or "").strip() or la.business_type
    la.contact_name = (payload.contact_name or "").strip() or la.contact_name
    la.contact_email = (payload.contact_email or "").strip() or la.contact_email
    la.contact_phone = (payload.contact_phone or "").strip() or la.contact_phone
    la.business_pan = (payload.business_pan or "").strip().upper() or la.business_pan
    la.gst_number = (payload.gst_number or "").strip().upper() or la.gst_number
    la.beneficiary_name = (payload.beneficiary_name or "").strip() or la.beneficiary_name
    if payload.bank_ifsc:
        la.bank_ifsc = payload.bank_ifsc.strip().upper()
    if payload.bank_account_number:
        digits = re.sub(r"\D", "", payload.bank_account_number)
        la.bank_account_last4 = digits[-4:] if digits else la.bank_account_last4
    la.attested_service_delivery = True
    la.attested_turnover_ok = True
    la.attested_at = datetime.utcnow()
    la.attested_ip = extract_client_ip(request) if request else None
    la.attested_version = FINANCE_ATTESTATION_VERSION

    if not enterprise_payments.razorpay_enabled():
        # Route not switched on yet — keep the details, don't attempt live KYC.
        if la.activation_status in (None, "", "not_started"):
            la.activation_status = "not_started"
        db.commit()
        db.refresh(la)
        return {
            "payments_enabled": False,
            "linked_account": enterprise_payments.serialize_linked_account(la),
            "message": (
                "Your details are saved. Online collection will activate once Rilono enables "
                "Razorpay Route for your account."
            ),
        }

    # Razorpay Route is live: run the v2 onboarding sequence server-side (order enforced).
    la = enterprise_payments.onboard_linked_account(
        db=db,
        linked_account=la,
        payload=payload,
        request_ip=(extract_client_ip(request) if request else None),
    )
    db.commit()
    db.refresh(la)
    return {
        "payments_enabled": True,
        "linked_account": enterprise_payments.serialize_linked_account(la),
        "message": "Bank account submitted to Razorpay for verification.",
    }


@router.post("/finance/linked-account/refresh")
def enterprise_finance_refresh(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Re-sync the linked account's activation status + requirements from Razorpay."""
    _, organization, _ = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="finance.payout_account"
    )
    la = enterprise_payments.get_linked_account(db, organization.id)
    if la is None or not la.razorpay_account_id:
        raise HTTPException(status_code=404, detail="No connected bank account to refresh.")
    la = enterprise_payments.refresh_linked_account_status(db=db, linked_account=la)
    db.commit()
    db.refresh(la)
    return {"linked_account": enterprise_payments.serialize_linked_account(la)}


# ---------------------------------------------------------------------------
# Finance Phase 2 — collect payments from a client (secure emailed pay-link),
# webhook reconciliation, refunds. Money-moving writes are admin-only.
# ---------------------------------------------------------------------------

class EnterprisePaymentRequestCreate(BaseModel):
    amount_paise: int = Field(gt=0)
    description: str = Field(min_length=3, max_length=300)
    due_date: Optional[date] = None


class EnterpriseManualPaymentCreate(BaseModel):
    """Record a payment collected off-platform (cash / bank transfer / UPI / …)."""
    amount_paise: int = Field(gt=0)
    method: str = Field(min_length=2, max_length=30)
    description: Optional[str] = Field(default=None, max_length=300)
    reference: Optional[str] = Field(default=None, max_length=80)
    received_on: Optional[date] = None


class EnterprisePublicPayVerifyRequest(BaseModel):
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str


class EnterprisePaymentRefundRequest(BaseModel):
    reason: Optional[str] = Field(default=None, max_length=500)


def _build_pay_link_url(subdomain_slug, token: str, request: Request | None) -> str:
    """Public pay-page URL for the emailed secure link (mirrors the interview invite URL)."""
    subdomain = str(subdomain_slug or "").strip().lower()
    base = None
    if subdomain:
        host = f"{subdomain}.{ENTERPRISE_ROOT_DOMAIN}"
        port = _request_port_for_local_enterprise_url(request)
        if port:
            host = f"{host}:{port}"
        base = f"{ENTERPRISE_PORTAL_SCHEME}://{host}"
    if not base:
        base = ENTERPRISE_PASSWORD_SETUP_BASE_URL
    return f"{base.rstrip('/')}/pay/{token}"


def _get_org_payment_or_404(
    db: Session, organization_id: int, payment_id: int, *, ctx=None
) -> models.EnterpriseStudentPayment:
    row = (
        db.query(models.EnterpriseStudentPayment)
        .filter(
            models.EnterpriseStudentPayment.id == int(payment_id),
            models.EnterpriseStudentPayment.organization_id == int(organization_id),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Payment not found.")
    # Payment rows are RETAINED when a client is deleted (client_id goes NULL), so an orphaned row
    # has no client to inherit scope from — those stay workspace-scope only rather than falling
    # into everyone's view.
    if ctx is not None and ctx.scope_kind != "all":
        if not row.client_id:
            raise HTTPException(status_code=404, detail="Payment not found.")
        linked = (
            db.query(models.EnterpriseClient)
            .filter(
                models.EnterpriseClient.id == int(row.client_id),
                models.EnterpriseClient.organization_id == int(organization_id),
            )
            .first()
        )
        if not client_in_scope(linked, ctx):
            raise HTTPException(status_code=404, detail="Payment not found.")
    return row


def _send_payment_request_email_safe(payment, organization, client_email, pay_url):
    due = payment.due_date.strftime("%d %b %Y") if payment.due_date else None
    return send_enterprise_payment_request_email(
        to_email=client_email,
        client_name=payment.client_name_snapshot or "there",
        organization_name=organization.company_name,
        # Hand over the stored integer and the row's own currency rather than a
        # pre-divided string: `amount_paise` is minor units of `payment.currency`, and the
        # /100 here was only ever right because Route happens to be INR-only. app.money
        # owns the exponent and the symbol, so this email matches the pay page exactly.
        amount_minor=int(payment.amount_paise or 0),
        currency=(payment.currency or "INR"),
        description=payment.description or "Service payment",
        pay_url=pay_url,
        invoice_number=payment.invoice_number or "",
        due_date_text=due,
        logo_url=_absolute_enterprise_logo_url(organization),
    )


@router.get("/clients/{client_id}/payments")
def enterprise_client_payments(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """The client dossier's Payments tab: totals + the request ledger for this client."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="finance.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    rows = (
        db.query(models.EnterpriseStudentPayment)
        .filter(
            models.EnterpriseStudentPayment.organization_id == organization.id,
            models.EnterpriseStudentPayment.client_id == client.id,
        )
        .order_by(models.EnterpriseStudentPayment.created_at.desc(), models.EnterpriseStudentPayment.id.desc())
        .all()
    )
    la = enterprise_payments.get_linked_account(db, organization.id)
    return {
        "payments": [enterprise_payments.serialize_payment(p) for p in rows],
        "totals": enterprise_payments.client_payment_totals(db, organization.id, client.id),
        "collect_enabled": bool(
            enterprise_payments.razorpay_enabled() and la is not None and la.is_payable
        ),
        "linked_account_status": (la.activation_status if la else "not_started"),
        "fee": enterprise_payments.fee_config_public(),
        "client_email": (client.email or "").strip().lower() or None,
        "permissions": _enterprise_permissions_for_role(role),
    }


@router.post("/clients/{client_id}/payments")
def enterprise_create_client_payment(
    client_id: int,
    payload: EnterprisePaymentRequestCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Raise a payment request for this client and email them a secure pay-link.

    Hard compliance gates: Razorpay live + the org's linked account activated
    (a marketplace must never collect for a non-onboarded payee)."""
    # Collecting a student fee is day-to-day work for a branch manager or a finance member, so
    # this is capability-gated rather than pinned to workspace admins.
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="finance.manage"
    )
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.payment_request",
        limit=30, window_seconds=3600, extra_key=str(organization.id),
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    client_email = (client.email or "").strip().lower()
    if not client_email:
        raise HTTPException(status_code=400, detail="Add an email to this client before requesting a payment.")
    if not enterprise_payments.razorpay_enabled():
        raise HTTPException(
            status_code=409,
            detail="Online collection isn't live yet — you can connect your bank account in Finance meanwhile.",
        )
    la = enterprise_payments.get_linked_account(db, organization.id)

    raw_token = generate_verification_token()
    payment = enterprise_payments.create_payment_request(
        db=db,
        organization=organization,
        linked_account=la,
        client=client,
        amount_paise=int(payload.amount_paise),
        description=payload.description,
        due_date=payload.due_date,
        created_by=current_user,
        pay_token_hash=hash_token(raw_token),
        payer_email=client_email,
        # Sent to Razorpay Checkout as prefill.contact. Required for international cards:
        # the gateway rejects the payment outright if the contact details look like
        # placeholders, so a student abroad cannot pay without a real phone here.
        payer_phone=(client.phone or "").strip() or None,
    )
    pay_url = _build_pay_link_url(organization.subdomain_slug, raw_token, request)
    sent, _mid, err = _send_payment_request_email_safe(payment, organization, client_email, pay_url)
    if sent:
        payment.email_sent_at = datetime.utcnow()
    db.commit()
    db.refresh(payment)
    message = (
        f"Payment request for ₹{payment.amount_paise / 100:,.2f} sent to {client_email}."
        if sent else
        f"Payment request created, but the email could not be sent right now. {err or ''}".strip()
    )
    return {
        "message": message,
        "email_sent": sent,
        "pay_url": pay_url,  # returned once — only the hash is stored
        "payment": enterprise_payments.serialize_payment(payment),
        "totals": enterprise_payments.client_payment_totals(db, organization.id, client.id),
    }


@router.post("/clients/{client_id}/payments/manual")
def enterprise_record_manual_payment(
    client_id: int,
    payload: EnterpriseManualPaymentCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Record a payment the org collected OUTSIDE Rilono (cash / bank transfer / UPI / cheque / …).

    Bookkeeping only — no money moves through the platform, so it needs no linked account and works
    even when online collection isn't live. Lands as a 'paid' row in this client's ledger."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="finance.manage"
    )
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.payment_manual",
        limit=120, window_seconds=3600, extra_key=str(organization.id),
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    payment = enterprise_payments.record_manual_payment(
        db=db,
        organization=organization,
        client=client,
        amount_paise=int(payload.amount_paise),
        method=payload.method,
        description=payload.description,
        reference=payload.reference,
        received_on=payload.received_on,
        created_by=current_user,
    )
    db.commit()
    db.refresh(payment)
    return {
        "message": f"Recorded ₹{payment.amount_paise / 100:,.2f} received from {client.full_name}.",
        "payment": enterprise_payments.serialize_payment(payment),
        "totals": enterprise_payments.client_payment_totals(db, organization.id, client.id),
    }


@router.delete("/finance/payments/{payment_id}/manual")
def enterprise_delete_manual_payment(
    payment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Remove a manually-recorded (off-platform) payment — e.g. to correct a mistake. Only manual
    rows can be removed; real Razorpay payments are immutable financial records."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="finance.manage"
    )
    payment = _get_org_payment_or_404(db, organization.id, payment_id, ctx=role.ctx)
    if (payment.provider or "") != "manual":
        raise HTTPException(status_code=409, detail="Only manually-recorded payments can be removed.")
    client_id = payment.client_id
    db.delete(payment)
    db.commit()
    totals = (
        enterprise_payments.client_payment_totals(db, organization.id, client_id)
        if client_id else None
    )
    return {"message": "Payment record removed.", "totals": totals}


@router.post("/finance/payments/{payment_id}/resend-email")
def enterprise_resend_payment_email(
    payment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Rotate the secure token and re-send the pay-link (also returns it for copying)."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="finance.manage"
    )
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.payment_resend",
        limit=10, window_seconds=3600, extra_key=str(payment_id),
    )
    payment = _get_org_payment_or_404(db, organization.id, payment_id, ctx=role.ctx)
    if payment.status != "created":
        raise HTTPException(status_code=409, detail="Only an unpaid request's link can be re-sent.")
    to_email = payment.payer_email_snapshot
    if not to_email:
        raise HTTPException(status_code=400, detail="This request has no email on file.")

    raw_token = generate_verification_token()
    payment.pay_token_hash = hash_token(raw_token)  # rotates: the old link stops working
    pay_url = _build_pay_link_url(organization.subdomain_slug, raw_token, request)
    sent, _mid, err = _send_payment_request_email_safe(payment, organization, to_email, pay_url)
    if sent:
        payment.email_sent_at = datetime.utcnow()
    db.commit()
    return {
        "message": f"Pay link re-sent to {to_email}." if sent
                   else f"Link rotated, but the email could not be sent. {err or ''}".strip(),
        "email_sent": sent,
        "pay_url": pay_url,
        "payment": enterprise_payments.serialize_payment(payment),
    }


@router.post("/finance/payments/{payment_id}/cancel")
def enterprise_cancel_payment_request(
    payment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="finance.manage"
    )
    payment = _get_org_payment_or_404(db, organization.id, payment_id, ctx=role.ctx)
    if payment.status not in ("created", "failed"):
        raise HTTPException(status_code=409, detail="Only an unpaid request can be cancelled.")
    payment.status = "cancelled"
    payment.cancelled_at = datetime.utcnow()
    db.commit()
    return {"message": "Payment request cancelled.", "payment": enterprise_payments.serialize_payment(payment)}


@router.post("/finance/payments/{payment_id}/refund")
def enterprise_refund_payment(
    payment_id: int,
    payload: EnterprisePaymentRefundRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Refund the student in full (original instrument). Gateway-first: if Razorpay rejects
    it (e.g. the payout already settled), nothing is persisted and the reason is surfaced."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="finance.refund"
    )
    payment = _get_org_payment_or_404(db, organization.id, payment_id, ctx=role.ctx)
    audit = enterprise_payments.issue_full_refund(
        db=db, payment=payment, by_user=current_user, reason=(payload.reason or None),
    )
    db.commit()
    return {
        "message": f"Refund of ₹{audit.amount_paise / 100:,.2f} initiated to the student's original payment method.",
        "payment": enterprise_payments.serialize_payment(payment),
    }


# ---- Phase 3 · Finance books: the consultancy's own income, costs and ROI ----
#
# Everything below reads through app/enterprise_finance.py, which builds ONE ledger per
# request from (a) the org's hand-recorded entries and (b) money the platform already
# knows about (collected payments, the Rilono fee on them, credit top-ups, refunds,
# lost chargebacks). Analytics, the ledger and the export all consume that same list, so
# they cannot disagree. Company-level money is admin-only — an editor manages students,
# not payroll — so every route here gates on require_manage_users.

ENTERPRISE_FINANCE_RATE_LIMIT = int(os.getenv("ENTERPRISE_FINANCE_RATE_LIMIT", "120"))
ENTERPRISE_FINANCE_RATE_WINDOW_SECONDS = int(os.getenv("ENTERPRISE_FINANCE_RATE_WINDOW_SECONDS", "60"))
ENTERPRISE_FINANCE_EXPORT_RATE_LIMIT = int(os.getenv("ENTERPRISE_FINANCE_EXPORT_RATE_LIMIT", "12"))
ENTERPRISE_FINANCE_EXPORT_RATE_WINDOW_SECONDS = int(
    os.getenv("ENTERPRISE_FINANCE_EXPORT_RATE_WINDOW_SECONDS", "300")
)
FINANCE_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class EnterpriseFinanceEntryRequest(BaseModel):
    """One hand-recorded income or expense. `amount_paise` is a positive magnitude —
    `kind` decides which side of the books it lands on (mirrors the admin console)."""
    kind: str = Field("expense", max_length=10)
    category: str | None = Field(None, max_length=60)
    amount_paise: int = Field(..., gt=0)
    tax_paise: int | None = Field(None, ge=0)
    occurred_on: Optional[date] = None
    description: str | None = Field(None, max_length=300)
    counterparty: str | None = Field(None, max_length=160)
    payment_method: str | None = Field(None, max_length=30)
    reference: str | None = Field(None, max_length=120)
    notes: str | None = Field(None, max_length=2000)
    client_id: int | None = None
    repeat_monthly: bool = False
    repeat_until: Optional[date] = None
    # Optimistic-concurrency token on edit (ignored on create). This payload is whole-object
    # by design — "an empty client_id means detach" — so without it the later of two saves
    # silently rewrites every field of the earlier one, including the amount.
    expected_version: Optional[int] = None
    force_overwrite: bool = False


class EnterpriseFinanceSettingsRequest(BaseModel):
    hourly_cost_paise: int | None = Field(None, ge=0)
    opening_balance_paise: int | None = Field(None, ge=0)
    opening_balance_on: Optional[date] = None
    fy_start_month: int | None = Field(None, ge=1, le=12)
    # {"deep_scan": 30, …} — the org's own minutes-per-task baselines.
    savings_minutes: dict[str, int] | None = None


def _require_finance_admin(
    *, db: Session, user: models.User, request: Request, capability: str = "finance.books"
):
    """Gate for the company's own books, with the finance settings row attached.

    Gated in-body rather than via the generic flags so the 403 explains itself — "Only
    organization admins can manage users" makes no sense on a P&L endpoint.

    Two rules here, both deliberate:

      * It is capability-based (`finance.books` / `finance.books_manage`), not admin-only, which is
        the entire point of having a Finance role: the person who keeps the books needs the books,
        and shouldn't have to be handed workspace administration to get them.
      * It additionally requires workspace-wide record scope. Every figure behind these endpoints
        is a whole-company roll-up — profit, receivables, revenue split by office and by staff
        member — so there is no meaningful "branch-scoped P&L" to serve. A member limited to one
        office is refused rather than shown numbers that include offices they can't see.
    """
    _, organization, role = _require_enterprise_membership(
        db=db, user=user, request=request, require_capability=capability
    )
    if not role.ctx.is_org_scope:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=(
                "The company's books cover the whole workspace, so they're only available to "
                "people with workspace-wide access."
            ),
        )
    settings = finance.get_or_create_settings(db, organization.id)
    return organization, role, settings


def _finance_range(range_key: str | None, start: str | None, end: str | None, settings) -> dict:
    return finance.resolve_range(
        range_key,
        start=start,
        end=end,
        fy_start_month=int(getattr(settings, "fy_start_month", 4) or 4),
    )


def _get_org_finance_entry_or_404(db: Session, organization_id: int, entry_id: int):
    row = (
        db.query(models.EnterpriseFinanceEntry)
        .filter(
            models.EnterpriseFinanceEntry.id == int(entry_id),
            models.EnterpriseFinanceEntry.organization_id == int(organization_id),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="That finance entry no longer exists.")
    return row


def _finance_client_for_entry(db: Session, organization_id: int, client_id, *, ctx=None):
    """Resolve an optional client attribution, scoped to the org (never by id alone)."""
    if not client_id:
        return None
    return _get_org_client_or_404(db, organization_id, int(client_id), ctx=ctx)


@router.get("/finance/books")
def enterprise_finance_books(
    request: Request,
    range_key: str = "this_month",
    start: str | None = None,
    end: str | None = None,
    kind: str | None = None,
    category: str | None = None,
    source: str | None = None,
    client_id: int | None = None,
    q: str | None = None,
    limit: int = finance.LEDGER_PAGE_SIZE,
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """The books for a period: filtered ledger page + totals for the whole filtered set."""
    organization, role, settings = _require_finance_admin(db=db, user=current_user, request=request)
    _enforce_rate_limit_or_429(
        request, "enterprise.finance_books",
        ENTERPRISE_FINANCE_RATE_LIMIT, ENTERPRISE_FINANCE_RATE_WINDOW_SECONDS,
        extra_key=str(organization.id),
    )
    rng = _finance_range(range_key, start, end, settings)
    book = finance.build_book(db, organization.id, rng, cache=finance.new_cache())
    rows = finance.filter_rows(
        book["rows"],
        kind=(kind or "").strip().lower() or None,
        category=(category or "").strip() or None,
        source=(source or "").strip() or None,
        client_id=int(client_id) if client_id else None,
        q=(q or "")[:120],
    )
    page = finance.ledger_page(rows, offset=offset, limit=limit)
    return {
        "range": finance.range_payload(rng),
        "ledger": page,
        "totals": finance.summarize(rows),
        "period_totals": finance.summarize(book["rows"]),
        "categories": finance.categories_payload(),
        "settings": finance.settings_payload(settings),
        "truncated": book["truncated"],
        "truncated_note": book["truncated_note"],
        "permissions": _enterprise_permissions_for_role(role),
    }


@router.get("/finance/analytics")
def enterprise_finance_analytics(
    request: Request,
    range_key: str = "this_month",
    start: str | None = None,
    end: str | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """P&L, cash, receivables, collections and per-client/destination profitability."""
    organization, role, settings = _require_finance_admin(db=db, user=current_user, request=request)
    _enforce_rate_limit_or_429(
        request, "enterprise.finance_analytics",
        ENTERPRISE_FINANCE_RATE_LIMIT, ENTERPRISE_FINANCE_RATE_WINDOW_SECONDS,
        extra_key=str(organization.id),
    )
    rng = _finance_range(range_key, start, end, settings)
    payload = finance.analytics(db, organization.id, rng, settings=settings, cache=finance.new_cache())
    payload["settings"] = finance.settings_payload(settings)
    payload["fee"] = enterprise_payments.fee_config_public()
    payload["permissions"] = _enterprise_permissions_for_role(role)
    return payload


@router.get("/finance/savings")
def enterprise_finance_savings(
    request: Request,
    range_key: str = "this_month",
    start: str | None = None,
    end: str | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Staff time and money the platform saved this org, net of what they paid Rilono."""
    organization, role, settings = _require_finance_admin(db=db, user=current_user, request=request)
    _enforce_rate_limit_or_429(
        request, "enterprise.finance_savings",
        ENTERPRISE_FINANCE_RATE_LIMIT, ENTERPRISE_FINANCE_RATE_WINDOW_SECONDS,
        extra_key=str(organization.id),
    )
    rng = _finance_range(range_key, start, end, settings)
    payload = finance.savings(db, organization.id, rng, settings=settings, cache=finance.new_cache())
    payload["settings"] = finance.settings_payload(settings)
    payload["permissions"] = _enterprise_permissions_for_role(role)
    return payload


def _validate_finance_entry_dates(db: Session, organization_id: int, payload) -> None:
    """Bound the dates on a hand-recorded ledger entry.

    `occurred_on` is an ACTUALS date: the books record what has already happened, and the
    period report keys every figure off it, so a future one silently moves money out of the
    month it belongs to.

    `repeat_until` before the start used to be coerced to None by `apply_entry_fields` —
    which the ledger reads as "repeat forever". That fails in the dangerous direction: the
    user asked for the series to END and got an unbounded one, with a success response. The
    coercion stays as a backstop; this turns the case into a 400 they can act on.
    """
    today = ent_time.org_today(db, int(organization_id))
    occurred = ent_dates.validate(
        payload.occurred_on, "The entry date", today=today,
        direction=ent_dates.NOT_FUTURE,
        future_hint=" The books record what has already happened.",
    ) or today  # mirrors apply_entry_fields: no date given means today
    until = ent_dates.validate(getattr(payload, "repeat_until", None), "The repeat-until date", today=today)
    if getattr(payload, "repeat_monthly", False) and until and until < occurred:
        raise HTTPException(
            status_code=400,
            detail="The repeat-until date can't be before the entry date.",
        )


@router.post("/finance/entries")
def enterprise_finance_create_entry(
    payload: EnterpriseFinanceEntryRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Record income or a cost the platform can't see (cash fees, salaries, rent, ads…)."""
    organization, role, _settings = _require_finance_admin(db=db, user=current_user, request=request, capability="finance.books_manage")
    _enforce_rate_limit_or_429(
        request, "enterprise.finance_entry_write",
        ENTERPRISE_FINANCE_RATE_LIMIT, ENTERPRISE_FINANCE_RATE_WINDOW_SECONDS,
        extra_key=str(organization.id),
    )
    if int(payload.amount_paise) > finance.MAX_AMOUNT_PAISE:
        raise HTTPException(
            status_code=400,
            detail=f"Amount exceeds the per-entry limit of ₹{finance.MAX_AMOUNT_PAISE / 100:,.0f}.",
        )
    _validate_finance_entry_dates(db, organization.id, payload)
    client = _finance_client_for_entry(db, organization.id, payload.client_id, ctx=role.ctx)
    entry = models.EnterpriseFinanceEntry(
        organization_id=organization.id,
        created_by_user_id=current_user.id,
        created_by_name=(current_user.full_name or current_user.email or None),
    )
    finance.apply_entry_fields(entry, payload, client=client)
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {
        "message": ("Income recorded." if entry.kind == "income" else "Expense recorded."),
        "entry": finance.serialize_entry(entry),
    }


@router.patch("/finance/entries/{entry_id}")
def enterprise_finance_update_entry(
    entry_id: int,
    payload: EnterpriseFinanceEntryRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Edit a hand-recorded entry. Platform-derived rows have no entry id and can't reach here."""
    organization, role, _settings = _require_finance_admin(db=db, user=current_user, request=request, capability="finance.books_manage")
    _enforce_rate_limit_or_429(
        request, "enterprise.finance_entry_write",
        ENTERPRISE_FINANCE_RATE_LIMIT, ENTERPRISE_FINANCE_RATE_WINDOW_SECONDS,
        extra_key=str(organization.id),
    )
    if int(payload.amount_paise) > finance.MAX_AMOUNT_PAISE:
        raise HTTPException(
            status_code=400,
            detail=f"Amount exceeds the per-entry limit of ₹{finance.MAX_AMOUNT_PAISE / 100:,.0f}.",
        )
    _validate_finance_entry_dates(db, organization.id, payload)
    entry = _get_org_finance_entry_or_404(db, organization.id, entry_id)
    if not payload.force_overwrite:
        _assert_fresh_write(entry, payload.expected_version, what="entry")
    client = _finance_client_for_entry(db, organization.id, payload.client_id, ctx=role.ctx)
    # The form always submits every field, so an empty client_id means "detach".
    finance.apply_entry_fields(entry, payload, client=client, clear_client=(client is None))
    db.commit()
    db.refresh(entry)
    return {"message": "Entry updated.", "entry": finance.serialize_entry(entry)}


@router.delete("/finance/entries/{entry_id}")
def enterprise_finance_delete_entry(
    entry_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    organization, _, _settings = _require_finance_admin(db=db, user=current_user, request=request, capability="finance.books_manage")
    entry = _get_org_finance_entry_or_404(db, organization.id, entry_id)
    db.delete(entry)
    db.commit()
    return {"message": "Entry deleted."}


@router.put("/finance/settings")
def enterprise_finance_update_settings(
    payload: EnterpriseFinanceSettingsRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Update the org's hourly cost, opening balance, financial-year start and the
    minutes-per-task baselines the savings panel values time with."""
    organization, role, settings = _require_finance_admin(db=db, user=current_user, request=request, capability="finance.books_manage")
    _enforce_rate_limit_or_429(
        request, "enterprise.finance_settings",
        ENTERPRISE_FINANCE_RATE_LIMIT, ENTERPRISE_FINANCE_RATE_WINDOW_SECONDS,
        extra_key=str(organization.id),
    )
    if payload.hourly_cost_paise is not None:
        hourly = int(payload.hourly_cost_paise)
        if hourly > finance.MAX_HOURLY_COST_PAISE:
            raise HTTPException(
                status_code=400,
                detail=f"Hourly cost looks too high — the limit is ₹{finance.MAX_HOURLY_COST_PAISE / 100:,.0f}/hour.",
            )
        settings.hourly_cost_paise = hourly or finance.DEFAULT_HOURLY_COST_PAISE
    if payload.opening_balance_paise is not None:
        opening = int(payload.opening_balance_paise)
        if opening > finance.MAX_OPENING_BALANCE_PAISE:
            raise HTTPException(
                status_code=400,
                detail=(
                    "That opening balance looks like a typo — the limit is "
                    f"₹{finance.MAX_OPENING_BALANCE_PAISE // 100:,}."
                ),
            )
        settings.opening_balance_paise = opening
    if payload.opening_balance_on is not None:
        settings.opening_balance_on = payload.opening_balance_on
    if payload.fy_start_month is not None:
        settings.fy_start_month = int(payload.fy_start_month)
    if payload.savings_minutes is not None:
        merged = finance.savings_minutes(settings)
        for key, value in payload.savings_minutes.items():
            if key in merged:
                merged[key] = max(0, min(600, int(value)))
        settings.savings_overrides_json = json.dumps(merged)
    settings.updated_by_user_id = current_user.id
    db.commit()
    db.refresh(settings)
    return {
        "message": "Finance settings saved.",
        "settings": finance.settings_payload(settings),
        "permissions": _enterprise_permissions_for_role(role),
    }


def _finance_csv_safe(value):
    """Defuse spreadsheet formula injection: a cell a user typed (a vendor called
    "=cmd|..." or a description starting with @) must never be evaluated when the export
    is opened. Same guard the credit-ledger export uses."""
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    return value


def _finance_export_csv(sections: list[dict]) -> bytes:
    """CSV fallback (and the explicit ?fmt=csv output): sections stacked into one sheet.
    BOM first so Excel opens it as UTF-8."""
    import csv
    import io as _io

    buffer = _io.StringIO()
    writer = csv.writer(buffer)
    for index, section in enumerate(sections):
        if index:
            writer.writerow([])
        writer.writerow([section["title"].upper()])
        for row in section["rows"]:
            writer.writerow([_finance_csv_safe(value) for value in row])
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def _finance_export_xlsx(sections: list[dict]) -> bytes | None:
    """One worksheet per section. Returns None when openpyxl isn't installed on the
    host so the caller falls back to CSV (never import it at module level)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:  # pragma: no cover - depends on the deployed environment
        logger.warning("openpyxl unavailable; falling back to CSV for the finance export")
        return None

    from io import BytesIO

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4338CA")
    money_format = '"₹"#,##0.00'
    date_format = "yyyy-mm-dd"

    percent_format = '0.0"%"'

    workbook = Workbook()
    workbook.remove(workbook.active)
    for section in sections:
        sheet = workbook.create_sheet(section["title"][:31] or "Sheet")
        rows = section["rows"]
        # The section declares its own shape and column types — inferring a format from
        # the Python type would render a 55.5% share as ₹55.50.
        is_table = bool(section.get("is_table"))
        money_columns = set(section.get("money") or ())
        percent_columns = set(section.get("percent") or ())
        date_columns = set(section.get("date") or ())
        money_rows = set(section.get("money_rows") or ())   # key/value blocks: money by ROW
        widths: dict[int, int] = {}
        for row in rows:
            # Same formula-injection guard as the CSV path: openpyxl writes a leading "="
            # as a live formula.
            sheet.append([_finance_csv_safe(value) for value in row])
            for column_index, value in enumerate(row, start=1):
                widths[column_index] = max(widths.get(column_index, 12), min(46, len(str(value)) + 3))
        if is_table and rows:
            for column_index in range(1, len(rows[0]) + 1):
                cell = sheet.cell(row=1, column=column_index)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.freeze_panes = "A2"
            if sheet.max_row > 1:
                sheet.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{sheet.max_row}"
        for row_index, row_cells in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), start=1):
            for column_index, cell in enumerate(row_cells, start=1):
                if cell.value is None or isinstance(cell.value, str):
                    continue
                if column_index in date_columns or isinstance(cell.value, date):
                    if not isinstance(cell.value, datetime) or column_index in date_columns:
                        cell.number_format = date_format
                elif column_index in percent_columns:
                    cell.number_format = percent_format
                elif column_index in money_columns and (not money_rows or row_index in money_rows):
                    cell.number_format = money_format
        for column_index, width in widths.items():
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


@router.get("/finance/export")
def enterprise_finance_export(
    request: Request,
    range_key: str = "this_month",
    start: str | None = None,
    end: str | None = None,
    fmt: str = "xlsx",
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Download the complete book for a period — ledger, P&L, categories, clients,
    receivables and the Rilono savings summary — as .xlsx (CSV fallback)."""
    organization, _, settings = _require_finance_admin(db=db, user=current_user, request=request, capability="finance.export")
    _enforce_rate_limit_or_429(
        request, "enterprise.finance_export",
        ENTERPRISE_FINANCE_EXPORT_RATE_LIMIT, ENTERPRISE_FINANCE_EXPORT_RATE_WINDOW_SECONDS,
        extra_key=str(organization.id),
    )
    rng = _finance_range(range_key, start, end, settings)
    cache = finance.new_cache()
    sections = finance.export_sections(
        rng=rng,
        book=finance.build_book(db, organization.id, rng, cache=cache),
        ana=finance.analytics(db, organization.id, rng, settings=settings, cache=cache),
        save=finance.savings(db, organization.id, rng, settings=settings, cache=cache),
    )
    stamp = datetime.utcnow().strftime("%Y%m%d-%H%M")
    wants_csv = (fmt or "").strip().lower() == "csv"
    content = None if wants_csv else _finance_export_xlsx(sections)
    if content is None:
        content = _finance_export_csv(sections)
        filename = f"rilono-finance-{stamp}.csv"
        media_type = "text/csv; charset=utf-8"
    else:
        filename = f"rilono-finance-{stamp}.xlsx"
        media_type = FINANCE_XLSX_MEDIA_TYPE
    return Response(
        content=content,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{_safe_filename(filename)}"',
            "Cache-Control": "no-store",
            "X-Content-Type-Options": "nosniff",
        },
    )


# ---- Public pay page (no auth; token = capability) -------------------------

def _public_payment_or_404(db: Session, token: str) -> models.EnterpriseStudentPayment:
    token_hash = hash_token((token or "").strip())
    row = (
        db.query(models.EnterpriseStudentPayment)
        .filter(models.EnterpriseStudentPayment.pay_token_hash == token_hash)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="This payment link is invalid or has been replaced.")
    return row


@router.get("/pay/{token}")
def enterprise_public_pay_info(
    token: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Everything the public pay page needs. The full amount charged is disclosed to the
    student up-front; the consultancy (not Rilono) is named as the payee."""
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.pay_public",
        limit=ENTERPRISE_PUBLIC_RATE_LIMIT, window_seconds=ENTERPRISE_PUBLIC_RATE_WINDOW,
        extra_key=hash_token(token)[:16],
    )
    payment = _public_payment_or_404(db, token)
    organization = db.query(models.EnterpriseOrganization).filter(
        models.EnterpriseOrganization.id == payment.organization_id
    ).first()
    la = enterprise_payments.get_linked_account(db, payment.organization_id)
    payable = (
        payment.status == "created"
        and enterprise_payments.razorpay_enabled()
        and la is not None and bool(la.is_payable)
    )
    info = {
        "organization_name": organization.company_name if organization else "Your consultancy",
        "organization_logo_url": _resolve_enterprise_logo_url(organization) if organization else None,
        "client_name": payment.client_name_snapshot,
        "invoice_number": payment.invoice_number,
        "description": payment.description,
        "amount_paise": payment.amount_paise,
        "currency": payment.currency,
        "status": payment.status,
        "due_date": payment.due_date.isoformat() if payment.due_date else None,
        "paid_at": payment.paid_at.isoformat() if payment.paid_at else None,
        "payable": payable,
    }
    if payable:
        info["razorpay_key_id"] = os.getenv("RAZORPAY_KEY_ID", "").strip()
        info["razorpay_order_id"] = payment.razorpay_order_id
        info["payer_email"] = payment.payer_email_snapshot
        # Prefilled into Checkout. Razorpay rejects international card payments sent with
        # dummy contact details, so this is what makes a foreign card work on an INR invoice.
        info["payer_phone"] = getattr(payment, "payer_phone_snapshot", None)
    return info


@router.post("/pay/{token}/verify")
def enterprise_public_pay_verify(
    token: str,
    payload: EnterprisePublicPayVerifyRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """Checkout callback: verify the Razorpay signature AND confirm capture with the API
    before marking paid (a signature alone proves authenticity, not capture). Webhooks
    remain the reconciliation source of truth for transfer/settlement states."""
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.pay_verify",
        limit=ENTERPRISE_PUBLIC_RATE_LIMIT, window_seconds=ENTERPRISE_PUBLIC_RATE_WINDOW,
        extra_key=hash_token(token)[:16],
    )
    payment = _public_payment_or_404(db, token)
    if payment.status == "cancelled":
        raise HTTPException(status_code=409, detail="This payment request was cancelled.")
    if (payload.razorpay_order_id or "").strip() != (payment.razorpay_order_id or ""):
        raise HTTPException(status_code=400, detail="Payment does not match this request.")
    if not enterprise_payments.verify_checkout_signature(
        payload.razorpay_order_id, payload.razorpay_payment_id, payload.razorpay_signature
    ):
        raise HTTPException(status_code=400, detail="Payment verification failed.")
    enterprise_payments.confirm_captured_and_mark_paid(
        db=db, payment=payment, razorpay_payment_id=payload.razorpay_payment_id.strip()
    )
    db.commit()
    return {"status": "paid", "message": "Payment received. You can close this page."}


# ---- Razorpay Route webhook (reconciliation source of truth) ---------------

@router.post("/webhook/razorpay-route")
async def enterprise_route_webhook(request: Request, db: Session = Depends(get_db)):
    """Idempotent, out-of-order-tolerant reconciliation of Route events. Signature is
    HMAC-SHA256 of the RAW body with a dedicated secret (distinct from the API keys and
    the B2C subscription webhook secret); events dedupe on the unique event id."""
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.route_webhook", limit=300, window_seconds=60,
    )
    secret = os.getenv("RAZORPAY_ROUTE_WEBHOOK_SECRET", "").strip()
    if not secret:
        raise HTTPException(status_code=503, detail="Webhook not configured.")
    body = await request.body()
    signature = (request.headers.get("X-Razorpay-Signature") or "").strip()
    expected = hmac.new(secret.encode(), body, hashlib.sha256).hexdigest()
    if not signature or not hmac.compare_digest(expected, signature):
        raise HTTPException(status_code=400, detail="Invalid webhook signature.")

    try:
        event = json.loads(body.decode("utf-8"))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid webhook payload.")
    event_type = str(event.get("event") or "").strip()
    event_id = (request.headers.get("x-razorpay-event-id") or "").strip() or None

    container = (event.get("payload") or {})
    entity = {}
    for kind in ("payment", "transfer", "refund", "order", "dispute"):
        maybe = (container.get(kind) or {}).get("entity") if isinstance(container.get(kind), dict) else None
        if maybe:
            entity = maybe
            break

    # Peek at the target row so the ledger row carries org/payment links, then dedupe-insert.
    target = enterprise_payments._find_payment_for_event(db, entity) if entity else None
    fresh = enterprise_payments.record_webhook_event(
        db,
        event_id=event_id,
        event_type=event_type,
        entity_type=str(entity.get("entity") or "") or None,
        entity_id=str(entity.get("id") or "") or None,
        amount_paise=int(entity.get("amount") or 0) or None,
        payload_json=body.decode("utf-8", errors="replace"),
        organization_id=target.organization_id if target else None,
        student_payment_id=target.id if target else None,
    )
    if not fresh:
        return {"status": "duplicate"}
    applied = enterprise_payments.apply_webhook_event(db, event_type, event)
    db.commit()

    # Chargebacks demand staff action before the evidence deadline — alert org admins.
    # Best-effort: an email failure must never make the webhook 5xx (Razorpay retries).
    if event_type.startswith("payment.dispute.") and applied is not None:
        try:
            _send_dispute_alert_to_org_admins(db, payment=applied, event_type=event_type, entity=entity)
        except Exception:
            logger.exception("route-webhook: dispute alert email failed (payment_id=%s)", applied.id)

    return {"status": "ok"}


def _reconcile_plan_payment(db: Session, *, plan_row, entity: dict, order_id: str) -> dict:
    """Activate a plan whose browser callback never arrived (abandoned 3DS, closed tab).

    Mirrors /billing/verify exactly, including the same idempotency rule: the period is
    only extended when this payment has not already been honoured, so a webhook racing the
    browser callback activates once. `sync_plan_credits` is idempotent on its own key, so
    the credit grant is safe to call unconditionally.
    """
    # Same guard as the credit path: never honour an order whose amount or currency does
    # not match what we priced.
    stored_currency = money.normalize_currency(plan_row.currency, strict=False)
    if money.normalize_currency(entity.get("currency"), strict=False) != stored_currency:
        logger.error("plan-webhook: currency mismatch order=%s stored=%s got=%s",
                     order_id, stored_currency, entity.get("currency"))
        return {"status": "mismatch"}
    if int(entity.get("amount") or 0) != int(plan_row.amount_paise):
        logger.error("plan-webhook: amount mismatch order=%s stored=%s got=%s",
                     order_id, plan_row.amount_paise, entity.get("amount"))
        return {"status": "mismatch"}

    if plan_row.status == "verified":
        return {"status": "ok"}

    now = datetime.utcnow()
    plan_row.razorpay_payment_id = str(entity.get("id") or "").strip() or None
    plan_row.status = "verified"
    plan_row.verified_at = now
    plan_row.error_message = None

    period_days = 365 if plan_row.billing_cycle == "yearly" else billing.PLAN_PERIOD_DAYS
    sub = billing.get_or_create_org_subscription(db, plan_row.organization_id, commit=False)
    sub.plan = plan_row.plan
    sub.status = "active"
    base = sub.current_period_end
    if base is not None and getattr(base, "tzinfo", None):
        base = base.replace(tzinfo=None)
    start = base if (base and base > now) else now
    sub.current_period_end = start + timedelta(days=period_days)
    db.commit()

    try:
        credits.sync_plan_credits(db, plan_row.organization_id, commit=True)
    except Exception:
        logger.exception("plan-webhook: credit grant failed for org %s", plan_row.organization_id)
    return {"status": "ok"}


def _sub_by_mandate(db: Session, subscription_id: str):
    if not subscription_id:
        return None
    return (
        db.query(models.EnterpriseSubscription)
        .filter(models.EnterpriseSubscription.razorpay_subscription_id == subscription_id)
        .first()
    )


def _handle_mandate_lifecycle(db: Session, *, event_name: str, event: dict) -> dict:
    """Record a mandate state change. Never moves money; only decides whether it will again."""
    entity = (((event.get("payload") or {}).get("subscription") or {}).get("entity")) or {}
    subscription_id = str(entity.get("id") or "").strip()
    sub = _sub_by_mandate(db, subscription_id)
    if not sub:
        return {"status": "ignored", "reason": "mandate_not_found"}

    state = event_name.split(".", 1)[1]           # halted | cancelled | paused | activated | ...
    sub.mandate_status = "active" if state in {"activated", "resumed"} else state
    if state in {"cancelled", "completed", "halted", "paused"}:
        # Stop auto-renewal. Access is NOT revoked here — `current_period_end` still governs
        # that, so the customer keeps the time they already paid for and lapses naturally.
        sub.cancel_at_period_end = True
        if state in {"cancelled", "completed"} and sub.canceled_at is None:
            sub.canceled_at = datetime.utcnow()
    else:
        sub.cancel_at_period_end = False
        sub.canceled_at = None
    db.commit()
    logger.info("mandate: org=%s %s -> %s", sub.organization_id, subscription_id, sub.mandate_status)
    return {"status": "ok", "mandate_status": sub.mandate_status}


def _handle_subscription_charged(db: Session, *, event: dict, payment_entity: dict | None = None) -> dict:
    """A recurring charge succeeded: extend the paid period and grant the month's credits.

    Idempotent on the Razorpay payment id — Razorpay retries webhooks, and applying the same
    renewal twice would hand out a second month of credits for one charge.
    """
    payload = event.get("payload") or {}
    entity = payment_entity or ((payload.get("payment") or {}).get("entity")) or {}
    sub_entity = ((payload.get("subscription") or {}).get("entity")) or {}
    subscription_id = (
        str(entity.get("subscription_id") or "").strip()
        or str(sub_entity.get("id") or "").strip()
    )
    payment_id = str(entity.get("id") or "").strip()
    if not subscription_id or not payment_id:
        return {"status": "ignored", "reason": "missing_ids"}

    sub = _sub_by_mandate(db, subscription_id)
    seed_row = (
        db.query(models.EnterpriseSubscriptionPayment)
        .filter(models.EnterpriseSubscriptionPayment.razorpay_subscription_id == subscription_id)
        .order_by(models.EnterpriseSubscriptionPayment.id.asc())
        .first()
    )
    if sub is None and seed_row is None:
        return {"status": "ignored", "reason": "mandate_not_found"}
    org_id = int(sub.organization_id if sub is not None else seed_row.organization_id)

    # IDEMPOTENCY: this exact charge may only be applied once.
    if db.query(models.EnterpriseSubscriptionPayment).filter(
        models.EnterpriseSubscriptionPayment.razorpay_payment_id == payment_id
    ).first():
        return {"status": "ok", "idempotent": True}

    plan_key = billing.normalize_plan_key(seed_row.plan if seed_row else (sub.plan if sub else None))
    if plan_key not in billing.PAID_PLAN_KEYS:
        return {"status": "ignored", "reason": "not_a_paid_plan"}

    amount = int(entity.get("amount") or 0)
    currency = money.normalize_currency(entity.get("currency"), strict=False)
    # Re-derive the tax split from the amount actually charged rather than copying the seed
    # row: a price change between periods must not be reported with the old period's tax.
    subtotal = money.tax_inclusive_net_minor(amount, currency) if amount else 0
    tax = max(0, amount - subtotal)

    db.add(models.EnterpriseSubscriptionPayment(
        organization_id=org_id,
        created_by_user_id=(seed_row.created_by_user_id if seed_row else None),
        plan=plan_key,
        billing_cycle="monthly",
        amount_paise=amount,
        subtotal_paise=subtotal,
        tax_paise=tax,
        tax_percent=(seed_row.tax_percent if seed_row else None),
        tax_label=(seed_row.tax_label if seed_row else None),
        included_credits=billing.included_credits_for(plan_key),
        currency=currency,
        # Each renewal needs its own unique "order" key; the payment id is unique per charge.
        razorpay_order_id=payment_id,
        razorpay_payment_id=payment_id,
        razorpay_subscription_id=subscription_id,
        status="verified",
        verified_at=datetime.utcnow(),
    ))

    now = datetime.utcnow()
    sub = billing.get_or_create_org_subscription(db, org_id, commit=False)
    sub.plan = plan_key
    sub.status = "active"
    sub.razorpay_subscription_id = subscription_id
    sub.mandate_status = "active"
    # A successful charge means the mandate is healthy again. _handle_mandate_lifecycle sets
    # cancel_at_period_end on "halted" (card declining); leaving it set after Razorpay
    # recovers would tell a paying customer forever that their plan is about to stop.
    # A genuine customer-initiated cancel re-arrives as subscription.cancelled and re-sets it.
    sub.cancel_at_period_end = False
    sub.canceled_at = None
    base = sub.current_period_end
    if base is not None and getattr(base, "tzinfo", None):
        base = base.replace(tzinfo=None)
    # Stack on the period already bought so a charge arriving early never shortens it.
    start = base if (base and base > now) else now
    sub.current_period_end = start + timedelta(days=billing.PLAN_PERIOD_DAYS)
    db.commit()

    try:
        credits.sync_plan_credits(db, org_id, commit=True)
    except Exception:
        logger.exception("subscription-charged: credit grant failed for org %s", org_id)
    logger.info("subscription-charged: org=%s renewed to %s", org_id, sub.current_period_end)
    return {"status": "ok", "renewed_until": str(sub.current_period_end)}


@router.post("/webhook/razorpay-credits")
async def enterprise_credits_webhook(request: Request, db: Session = Depends(get_db)):
    """Reconcile credit / infra-fee payments that the browser never confirmed.

    Credit top-ups and infra fees are credited ONLY from the browser callback
    (/credits/topup/verify). The row lock there prevents DOUBLE-crediting, but nothing
    prevents ZERO-crediting: if the buyer closes the tab, loses connection, or abandons
    the 3DS redirect after paying, the money is captured and the wallet is never topped up.

    That failure is rare on domestic UPI/NetBanking and common on international cards,
    which always round-trip through a 3DS/redirect step. Enabling international payments
    without this endpoint produces a steady trickle of consultancies charged but not
    credited — the worst possible support ticket.

    Configure in the Razorpay dashboard as a SEPARATE webhook (its own secret) subscribed
    to `payment.captured`. Idempotency is the same ledger-row check the verify route uses,
    so a webhook and a browser callback racing each other credit exactly once.
    """
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.credits_webhook", limit=300, window_seconds=60,
    )
    secret = (
        os.getenv("RAZORPAY_CREDITS_WEBHOOK_SECRET", "").strip()
        or os.getenv("RAZORPAY_WEBHOOK_SECRET", "").strip()
    )
    if not secret:
        raise HTTPException(status_code=503, detail="Webhook not configured.")
    body = await request.body()
    signature = (request.headers.get("X-Razorpay-Signature") or "").strip()
    expected = hmac.new(secret.encode(), body, hashlib.sha256).hexdigest()
    if not signature or not hmac.compare_digest(expected, signature):
        raise HTTPException(status_code=400, detail="Invalid webhook signature.")

    try:
        event = json.loads(body.decode("utf-8"))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid webhook payload.")
    event_name = str(event.get("event") or "").strip()

    # A MANDATE LIFECYCLE CHANGE. "halted" means Razorpay gave up retrying a failing card;
    # "cancelled"/"completed"/"paused" mean no further charge is coming. None of these move
    # money, but all of them mean the plan must stop auto-renewing — and, critically, the org
    # must lapse when its paid period runs out instead of keeping the tier forever.
    if event_name in {"subscription.halted", "subscription.cancelled", "subscription.completed",
                      "subscription.paused", "subscription.activated", "subscription.resumed"}:
        return _handle_mandate_lifecycle(db, event_name=event_name, event=event)

    # A RENEWAL CHARGE. This is what actually makes the subscription recurring: Razorpay
    # auto-debits the mandate each month and tells us here. Without handling it the customer
    # is charged and receives nothing — no extended period and no monthly credits.
    if event_name == "subscription.charged":
        return _handle_subscription_charged(db, event=event)

    if event_name != "payment.captured":
        return {"status": "ignored"}

    entity = (((event.get("payload") or {}).get("payment") or {}).get("entity")) or {}
    order_id = str(entity.get("order_id") or "").strip()
    if not order_id:
        # A subscription's FIRST charge arrives as payment.captured with a subscription_id
        # and no order_id. Route it to the same renewal path so the mandate is recorded even
        # when the buyer abandoned the browser callback.
        if str(entity.get("subscription_id") or "").strip():
            return _handle_subscription_charged(db, event=event, payment_entity=entity)
        return {"status": "ignored"}

    query = db.query(models.EnterpriseCreditPayment).filter(
        models.EnterpriseCreditPayment.razorpay_order_id == order_id
    )
    if db.bind and db.bind.dialect.name != "sqlite":
        query = query.with_for_update()
    payment_row = query.first()
    if not payment_row:
        # Plan subscriptions live in their own table, and they need this safety net MORE
        # than top-ups do: an abandoned 3DS redirect on a plan purchase leaves an org
        # charged ₹3,538.82 with no plan and no credits. Handled here rather than in a
        # second webhook so one Razorpay subscription covers every enterprise order.
        plan_query = db.query(models.EnterpriseSubscriptionPayment).filter(
            models.EnterpriseSubscriptionPayment.razorpay_order_id == order_id
        )
        if db.bind and db.bind.dialect.name != "sqlite":
            plan_query = plan_query.with_for_update()
        plan_row = plan_query.first()
        if plan_row:
            return _reconcile_plan_payment(db, plan_row=plan_row, entity=entity, order_id=order_id)
        # Not one of ours (B2C pass, Route collection, …) — ack so Razorpay stops retrying.
        return {"status": "ignored"}

    # Trust the webhook entity for amount/currency only after checking it against what we
    # stored: a mismatch means this order was not the one we priced, and crediting on it
    # would hand out credits for a different (possibly smaller) payment.
    stored_currency = money.normalize_currency(payment_row.currency, strict=False)
    if money.normalize_currency(entity.get("currency"), strict=False) != stored_currency:
        logger.error(
            "credits-webhook: currency mismatch order=%s stored=%s got=%s",
            order_id, stored_currency, entity.get("currency"),
        )
        return {"status": "mismatch"}
    if int(entity.get("amount") or 0) != int(payment_row.amount_paise):
        logger.error(
            "credits-webhook: amount mismatch order=%s stored=%s got=%s",
            order_id, payment_row.amount_paise, entity.get("amount"),
        )
        return {"status": "mismatch"}

    if payment_row.status != "verified":
        payment_row.razorpay_payment_id = str(entity.get("id") or "").strip() or None
        base_minor, fx_rate = money.settled_inr_minor(entity)
        if base_minor is not None:
            payment_row.base_amount_paise = base_minor
            payment_row.fx_rate_used = fx_rate
        payment_row.base_currency = "INR"
        payment_row.is_international = bool(entity.get("international"))
        payment_row.status = "verified"
        payment_row.verified_at = datetime.utcnow()
        payment_row.error_message = None

    if payment_row.kind == "infra_fee":
        wallet = credits.get_or_create_wallet(db, payment_row.organization_id, commit=False)
        paid_until = wallet.infra_fee_paid_until
        # Only extend if the browser callback did not already do it for THIS payment.
        if not paid_until or paid_until < datetime.utcnow():
            credits.mark_infra_fee_paid(db, payment_row.organization_id, commit=False)
    else:
        total_credits = int(payment_row.credits) + int(payment_row.bonus_credits)
        already = (
            db.query(models.EnterpriseCreditTransaction)
            .filter(
                models.EnterpriseCreditTransaction.organization_id == payment_row.organization_id,
                models.EnterpriseCreditTransaction.reference_type == "payment",
                models.EnterpriseCreditTransaction.reference_id == payment_row.id,
            )
            .first()
        )
        if not already and total_credits > 0:
            pkg = credits.get_package(payment_row.package_key)
            label = pkg["label"] if pkg else "Credit top-up"
            credits.add_credits(
                db, payment_row.organization_id, total_credits,
                txn_type="topup",
                description=f"{label} (+{total_credits} credits)",
                reference_type="payment", reference_id=payment_row.id,
                commit=False,
            )
    db.commit()
    return {"status": "ok"}


def _send_dispute_alert_to_org_admins(
    db: Session, *, payment: "models.EnterpriseStudentPayment", event_type: str, entity: dict
) -> None:
    """Email every active org admin when a dispute opens or needs action. Won/lost/closed
    updates are also sent so staff see the outcome without polling the dashboard."""
    admins = (
        db.query(models.User)
        .join(
            models.EnterpriseOrganizationMember,
            models.EnterpriseOrganizationMember.user_id == models.User.id,
        )
        .filter(
            models.EnterpriseOrganizationMember.organization_id == payment.organization_id,
            models.EnterpriseOrganizationMember.is_active.is_(True),
            models.EnterpriseOrganizationMember.role == ENTERPRISE_ROLE_ADMIN,
            models.User.is_active.is_(True),
        )
        .all()
    )
    organization = db.query(models.EnterpriseOrganization).filter(
        models.EnterpriseOrganization.id == payment.organization_id
    ).first()
    if not admins or organization is None:
        return

    suffix = event_type.rsplit(".", 1)[-1].replace("_", " ")
    respond_by_text = None
    try:
        raw = entity.get("respond_by")
        if raw:
            respond_by_text = datetime.utcfromtimestamp(int(raw)).strftime("%d %b %Y, %H:%M UTC")
    except (TypeError, ValueError, OSError, OverflowError):
        respond_by_text = None

    for admin in admins:
        if getattr(admin, "email_notifications_enabled", True) is False:
            continue
        send_enterprise_payment_dispute_alert_email(
            to_email=admin.email,
            organization_name=organization.company_name,
            client_name=payment.client_name_snapshot or "a client",
            # The dispute entity's `amount` is in the DISPUTED PAYMENT's own presentment
            # currency, which is not necessarily INR — dividing it by 100 and letting the
            # template prefix "₹" mislabelled every foreign chargeback in the one email an
            # org uses to decide whether to fight it. Hand over the integer minor amount
            # plus the currency it is denominated in and let app.money render both.
            amount_minor=int(entity.get("amount") or payment.amount_paise or 0),
            currency=(
                str(entity.get("currency") or "").strip().upper()
                or (payment.currency or "INR")
            ),
            invoice_number=payment.invoice_number or "",
            dispute_state=suffix,
            reason_code=str(entity.get("reason_code") or "") or None,
            respond_by_text=respond_by_text,
        )


# ---- Resend inbound-email webhook (client replies -> Emails thread) ---------

@router.post("/webhooks/inbound-email")
async def enterprise_inbound_email_webhook(request: Request, db: Session = Depends(get_db)):
    """Receive Resend Inbound `email.received` events (Svix-signed over the RAW body)
    and thread client replies into the CRM. Replies arrive on the tokenized
    reply+c{id}-{sig}@{inbound domain} address set as Reply-To on outbound client
    emails; the HMAC token resolves the client. Unmatched/duplicate mail returns
    200 so the provider doesn't retry forever."""
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.inbound_email_webhook", limit=300, window_seconds=60,
    )
    secret = inbound_email.webhook_secret()
    if not secret:
        raise HTTPException(status_code=503, detail="Webhook not configured.")
    try:
        declared_length = int(request.headers.get("content-length") or 0)
    except ValueError:
        declared_length = 0
    if declared_length > 1_000_000:
        raise HTTPException(status_code=413, detail="Payload too large.")
    body = await request.body()
    if len(body) > 1_000_000:
        raise HTTPException(status_code=413, detail="Payload too large.")
    if not inbound_email.verify_svix_signature(secret=secret, headers=request.headers, body=body):
        raise HTTPException(status_code=400, detail="Invalid webhook signature.")

    try:
        event = json.loads(body.decode("utf-8"))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid webhook payload.")
    if str(event.get("type") or "").strip() != "email.received":
        return {"status": "ignored"}
    data = event.get("data") or {}
    if not isinstance(data, dict):
        return {"status": "ignored"}

    # Resolve the client from our tokenized reply+ recipient address.
    client_id = None
    matched_address = None
    recipients = inbound_email.recipient_addresses(data)
    for addr in recipients:
        client_id = inbound_email.parse_reply_address(addr)
        if client_id is not None:
            matched_address = addr
            break
    if client_id is None:
        # 200 so the provider doesn't retry, but leave an operator trace —
        # a real reply landing here means a stale/foreign token.
        logger.info("inbound-email webhook: no reply-token match (recipients=%s)", recipients[:10])
        return {"status": "no_match"}
    client = (
        db.query(models.EnterpriseClient)
        .filter(models.EnterpriseClient.id == int(client_id))
        .first()
    )
    if client is None:
        return {"status": "no_match"}

    provider_message_id = str(data.get("email_id") or data.get("id") or "").strip() or None
    if provider_message_id:
        duplicate = (
            db.query(models.EnterpriseClientEmail)
            .filter(
                models.EnterpriseClientEmail.provider_message_id == provider_message_id,
                models.EnterpriseClientEmail.direction == "inbound",
            )
            .first()
        )
        if duplicate:
            return {"status": "duplicate"}

    from_email = inbound_email.sender_address(data)
    subject = str(data.get("subject") or "").strip()[:300] or "(no subject)"
    # The body fetch is a blocking HTTP call (the webhook carries metadata only) —
    # run it off the event loop. On fetch failure, 5xx so Svix retries with the
    # same email_id instead of us committing an empty reply forever.
    reply_text, fetch_failed = await run_in_threadpool(inbound_email.extract_reply_text, data)
    if fetch_failed and not reply_text:
        logger.warning("inbound-email webhook: body fetch failed for %s; asking provider to retry",
                       provider_message_id)
        raise HTTPException(status_code=500, detail="Inbound body fetch failed — retry.")
    if not reply_text:
        reply_text = "(Empty message — the reply may only contain an attachment.)"

    # From: is spoofable and the reply+ address is forwardable — never assert the
    # client wrote it unless the sender matches the email we have on file.
    sender_matches = bool(
        from_email and (client.email or "").strip() and from_email == client.email.strip().lower()
    )
    row = models.EnterpriseClientEmail(
        organization_id=client.organization_id,
        client_id=client.id,
        sent_by_user_id=None,
        sent_by_name=client.full_name,
        to_email=matched_address or (inbound_email.reply_address_for_client(client.id) or "inbound"),
        subject=subject,
        body=reply_text[:20000],
        status="received",
        provider_message_id=provider_message_id,
        direction="inbound",
        from_email=from_email,
    )
    db.add(row)
    notif.notify_org(
        db,
        client.organization_id,
        type="client_email_reply",
        # Subject in the title keeps distinct replies from tripping notify_org's
        # 10-minute identical-title dedupe.
        title=(f"✉️ {client.full_name} replied: {subject}" if sender_matches
               else f"✉️ Reply from {from_email or 'unknown sender'} on {client.full_name}'s thread"),
        body=reply_text[:140] or None,
        reference_type="client",
        reference_id=client.id,
    )
    try:
        db.commit()
    except IntegrityError:
        # Unique index on inbound provider_message_id: a concurrent Svix
        # redelivery beat us to the insert.
        db.rollback()
        return {"status": "duplicate"}

    # Email heads-up so staff don't have to be watching the portal. Best-effort:
    # a mail failure must never make the webhook 5xx (the reply is already saved).
    try:
        await _alert_staff_of_inbound_reply(
            db, organization_id=client.organization_id, client=client,
            reply_subject=subject, reply_body=reply_text,
        )
    except Exception:
        logger.exception("inbound-email: staff heads-up email failed (client_id=%s)", client.id)
    return {"status": "ok"}


def _inbound_reply_recipients(
    db: Session, *, organization_id: int, client: models.EnterpriseClient
) -> list[models.User]:
    """Who gets the email nudge for an inbound reply: the staffer who sent the most
    recent outbound message in this thread (the person waiting on the reply); fall
    back to active org admins. Honors each user's email-notification preference."""
    def _wants_email(u: models.User) -> bool:
        return (
            u is not None and u.is_active
            and getattr(u, "email_notifications_enabled", True) is not False
            and bool((u.email or "").strip())
        )

    recipients: list[models.User] = []
    last_out = (
        db.query(models.EnterpriseClientEmail)
        .filter(
            models.EnterpriseClientEmail.client_id == client.id,
            models.EnterpriseClientEmail.direction == "outbound",
            models.EnterpriseClientEmail.sent_by_user_id.isnot(None),
        )
        .order_by(models.EnterpriseClientEmail.created_at.desc())
        .first()
    )
    if last_out and last_out.sent_by_user_id:
        sender = db.query(models.User).filter(models.User.id == last_out.sent_by_user_id).first()
        if _wants_email(sender):
            recipients.append(sender)
    if not recipients:
        admins = (
            db.query(models.User)
            .join(
                models.EnterpriseOrganizationMember,
                models.EnterpriseOrganizationMember.user_id == models.User.id,
            )
            .filter(
                models.EnterpriseOrganizationMember.organization_id == organization_id,
                models.EnterpriseOrganizationMember.is_active.is_(True),
                models.EnterpriseOrganizationMember.role == ENTERPRISE_ROLE_ADMIN,
                models.User.is_active.is_(True),
            )
            .all()
        )
        recipients.extend(a for a in admins if _wants_email(a))

    seen: set[str] = set()
    deduped: list[models.User] = []
    for u in recipients:
        key = (u.email or "").strip().lower()
        if key and key not in seen:
            seen.add(key)
            deduped.append(u)
    return deduped


async def _alert_staff_of_inbound_reply(
    db: Session, *, organization_id: int, client: models.EnterpriseClient,
    reply_subject: str, reply_body: str,
) -> None:
    """Send the 'client replied' email, carrying the full reply. Staff can answer
    from their own inbox (Reply-To = the client on file, untracked) or from the
    portal (tracked). Resend calls run off the event loop."""
    recipients = _inbound_reply_recipients(db, organization_id=organization_id, client=client)
    if not recipients:
        return
    organization = (
        db.query(models.EnterpriseOrganization)
        .filter(models.EnterpriseOrganization.id == organization_id)
        .first()
    )
    org_name = organization.company_name if organization else "your consultancy"
    logo_url = _absolute_enterprise_logo_url(organization) if organization else None
    for user in recipients:
        try:
            await run_in_threadpool(
                send_enterprise_inbound_reply_alert_email,
                to_email=user.email,
                staff_name=user.full_name or user.email,
                organization_name=org_name,
                client_name=client.full_name,
                client_id=client.id,
                reply_subject=reply_subject,
                reply_body=reply_body,
                # The address on file, not the inbound From: — that one is spoofable.
                client_reply_to=(client.email or "").strip() or None,
                logo_url=logo_url,
            )
        except Exception:
            logger.exception("inbound-email: alert email to %s failed", user.email)


@router.get("/billing/plans")
def enterprise_billing_plans(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request)
    return {
        "plans": billing.public_plans_payload(),
        "currency": billing.CURRENCY,
        "checkout_enabled": _razorpay_enabled(),
        "razorpay_key_id": os.getenv("RAZORPAY_KEY_ID", "").strip() or None,
        "permissions": _enterprise_permissions_for_role(role),
        "subscription": _serialize_subscription_state(billing.build_subscription_state(db, organization.id)),
    }


@router.get("/billing/subscription")
def enterprise_billing_subscription(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request)
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "subscription": _serialize_subscription_state(billing.build_subscription_state(db, organization.id)),
        "plans": billing.public_plans_payload(),
    }


AUTO_RENEW = os.getenv("ENTERPRISE_AUTO_RENEW", "true").strip().lower() in {"1", "true", "yes", "on"}

# How many billing cycles a mandate is authorised for. Razorpay requires a finite count;
# ~10 years of monthly charges is effectively "until cancelled" while still bounded.
RECURRING_TOTAL_COUNT = int(os.getenv("ENTERPRISE_RECURRING_TOTAL_COUNT", "120") or "120")


def _ensure_razorpay_plan_id(db: Session, *, plan_key: str, currency: str, amount_minor: int) -> str:
    """Get-or-create the Razorpay Plan backing this tier at this exact charged amount.

    Keyed on (tier, currency, amount, period) so a coupon that changes the amount gets its
    own plan automatically. Razorpay does not dedupe plans by name, so without this cache
    every checkout would mint a duplicate plan and reconciliation would become guesswork.
    """
    period = "monthly"
    existing = (
        db.query(models.EnterpriseRazorpayPlan)
        .filter(
            models.EnterpriseRazorpayPlan.plan_key == plan_key,
            models.EnterpriseRazorpayPlan.currency == currency,
            models.EnterpriseRazorpayPlan.amount_minor == int(amount_minor),
            models.EnterpriseRazorpayPlan.period == period,
        )
        .first()
    )
    if existing:
        return existing.razorpay_plan_id

    label = (billing.get_plan(plan_key) or {}).get("label", plan_key)
    data = _razorpay_request("POST", "/plans", {
        "period": period,
        "interval": 1,
        "item": {
            "name": f"Rilono Enterprise — {label}",
            # The customer-visible amount INCLUDES GST: this is what gets auto-charged.
            "description": f"{label} plan, billed monthly (incl. tax)",
            "amount": int(amount_minor),
            "currency": currency,
        },
        "notes": {"plan_key": plan_key, "price_book_version": money.PRICE_BOOK_VERSION},
    })
    plan_id = str(data.get("id") or "").strip()
    if not plan_id.startswith("plan_"):
        raise HTTPException(status_code=502, detail="Could not set up the recurring plan.")
    db.add(models.EnterpriseRazorpayPlan(
        plan_key=plan_key, currency=currency, amount_minor=int(amount_minor),
        period=period, razorpay_plan_id=plan_id,
    ))
    try:
        db.commit()
    except Exception:
        # Another request created the same row concurrently — reuse theirs and drop ours.
        db.rollback()
        again = (
            db.query(models.EnterpriseRazorpayPlan)
            .filter(
                models.EnterpriseRazorpayPlan.plan_key == plan_key,
                models.EnterpriseRazorpayPlan.currency == currency,
                models.EnterpriseRazorpayPlan.amount_minor == int(amount_minor),
                models.EnterpriseRazorpayPlan.period == period,
            )
            .first()
        )
        if again:
            return again.razorpay_plan_id
        raise
    return plan_id


def _activate_free_plan(
    db: Session, *, organization, user, plan: dict, cycle: str,
    subtotal: int, tax: int, quote: dict, currency: str,
    coupon_code: str | None, coupon_percent, list_amount: int,
) -> None:
    """Activate a paid tier that a 100%-off coupon has fully covered.

    Writes the same rows a real payment would (a zero-amount, verified payment row for the
    audit trail, the subscription period, and the credit grant) so a comped account is
    indistinguishable downstream from a paying one — including lapsing on schedule.
    """
    now = datetime.utcnow()
    period_days = 365 if cycle == "yearly" else billing.PLAN_PERIOD_DAYS
    row = models.EnterpriseSubscriptionPayment(
        organization_id=organization.id,
        created_by_user_id=user.id,
        plan=plan["key"],
        billing_cycle=cycle,
        amount_paise=0,
        subtotal_paise=subtotal,
        tax_paise=tax,
        tax_percent=quote.get("tax_percent"),
        tax_label=quote.get("tax_label"),
        included_credits=int(plan.get("included_credits") or 0),
        currency=currency,
        razorpay_order_id=f"free_{organization.id}_{secrets.token_hex(8)}"[:40],
        status="verified",
        coupon_code=coupon_code,
        coupon_percent_off=coupon_percent,
        original_amount_paise=list_amount,
        verified_at=now,
    )
    db.add(row)
    sub = billing.get_or_create_org_subscription(db, organization.id, commit=False)
    sub.plan = plan["key"]
    sub.status = "active"
    base = sub.current_period_end
    if base is not None and getattr(base, "tzinfo", None):
        base = base.replace(tzinfo=None)
    start = base if (base and base > now) else now
    sub.current_period_end = start + timedelta(days=period_days)
    db.commit()
    try:
        credits.sync_plan_credits(db, organization.id, commit=True)
    except Exception:
        logger.exception("free-plan: credit grant failed for org %s", organization.id)


@router.post("/billing/checkout")
def enterprise_billing_checkout(
    payload: EnterpriseBillingCheckoutRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, _ = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="billing.manage"
    )
    if billing.ENTERPRISE_FREE:
        return {"action": "free", "message": "Rilono Enterprise is free — no billing required."}
    plan = billing.get_plan(payload.plan)
    if not plan or plan["key"] not in billing.PAID_PLAN_KEYS:
        raise HTTPException(status_code=400, detail="Please choose a valid paid plan.")
    cycle = billing.normalize_billing_cycle(payload.billing_cycle)
    # Price and currency resolve together. Reading billing.CURRENCY here instead would let
    # a misconfigured ENTERPRISE_PLAN_CURRENCY create a USD order for a rupee amount.
    plan_currency = billing.resolve_plan_currency(plan["key"], billing.CURRENCY)
    list_amount = billing.plan_price_minor(plan["key"], plan_currency)
    if list_amount <= 0:
        raise HTTPException(status_code=400, detail="This plan is not available for online checkout.")

    # Per-account discount code (admin-managed). It reduces the TAXABLE subtotal, so it is
    # resolved before the GST quote below — never applied to a tax-inclusive figure.
    coupon_code = None
    coupon_percent = None
    raw_coupon = (payload.coupon_code or "").strip()
    if raw_coupon:
        coupon = enterprise_coupons.resolve_active_coupon_or_400(
            db, organization.id, raw_coupon, context="billing"
        )
        coupon_percent = enterprise_coupons.parse_percent_off(coupon.percent_off)
        coupon_code = enterprise_coupons.normalize_code(coupon.code)

    # list → discount → GST → total, in that order and in exactly one place.
    quote = billing.checkout_quote(plan["key"], plan_currency, discount_percent=coupon_percent)
    subtotal = quote["subtotal_minor"]
    tax = quote["tax_minor"]
    # The ONLY figure that may reach Razorpay. Sending `subtotal` here would undercharge
    # every Indian customer by 18% and produce an invoice the pricing page contradicts.
    amount = quote["total_minor"]

    # A fully-covering discount leaves nothing to charge. /coupons/validate already previews
    # this as "free" and the UI renders ₹0/mo, so rejecting it here (the old
    # apply_to_amount_or_400 floor check) meant the customer was shown a free plan and then
    # hard-failed with a 400 at the moment they clicked pay. Activate it directly instead —
    # this mirrors what the credits top-up path already does for a 100%-off coupon.
    if amount < money.min_charge_minor(plan_currency):
        _activate_free_plan(
            db, organization=organization, user=current_user, plan=plan, cycle=cycle,
            subtotal=subtotal, tax=tax, quote=quote, currency=plan_currency,
            coupon_code=coupon_code, coupon_percent=coupon_percent, list_amount=list_amount,
        )
        return {
            "action": "activated",
            "free": True,
            "message": f"Your {plan['label']} plan is now active.",
            "subscription": _serialize_subscription_state(
                billing.build_subscription_state(db, organization.id)
            ),
            "wallet": credits.wallet_state(db, organization.id),
        }

    if not _razorpay_enabled():
        return {
            "action": "contact_sales",
            "message": "Online checkout is being enabled. Please contact sales to activate your plan.",
        }

    notes = {
        "organization_id": str(organization.id),
        "plan": plan["key"],
        "billing_cycle": cycle,
        "user_id": str(current_user.id),
        "coupon_code": coupon_code or "",
        "subtotal_paise": str(subtotal),
        "tax_paise": str(tax),
        "tax_label": quote["tax_label"] or "",
        "price_book_version": money.PRICE_BOOK_VERSION,
    }

    # RECURRING PATH (the default). A Razorpay Subscription is a MANDATE: standing permission
    # to auto-charge every month. The legacy /orders path below charges once and leaves the
    # customer to remember to come back — which for a product sold as "₹2,999/month" means it
    # silently stops billing after one month. Set ENTERPRISE_AUTO_RENEW=false to fall back.
    if AUTO_RENEW:
        # Any mandate this org already holds is superseded ONLY once the replacement is paid
        # for — see the cancellation in /billing/verify. Razorpay charges every mandate it
        # holds, so an org upgrading Starter -> Growth must not keep both; but opening a
        # checkout is not a commitment to buy, and cancelling here would mean that merely
        # opening the modal and closing it silently kills auto-renewal on the plan they are
        # still paying for, with nothing to re-arm it.
        rzp_plan_id = _ensure_razorpay_plan_id(
            db, plan_key=plan["key"], currency=plan_currency, amount_minor=amount,
        )
        sub_data = _razorpay_request("POST", "/subscriptions", {
            "plan_id": rzp_plan_id,
            "total_count": RECURRING_TOTAL_COUNT,
            "customer_notify": 1,
            "quantity": 1,
            "notes": notes,
        })
        subscription_id = str(sub_data.get("id") or "").strip()
        if not subscription_id.startswith("sub_"):
            raise HTTPException(status_code=502, detail="Could not start the subscription.")
        # The payment row keys on razorpay_order_id (unique), and for the recurring flow the
        # subscription id plays that role until the first charge arrives.
        db.add(models.EnterpriseSubscriptionPayment(
            organization_id=organization.id,
            created_by_user_id=current_user.id,
            plan=plan["key"],
            billing_cycle=cycle,
            amount_paise=amount,
            subtotal_paise=subtotal,
            tax_paise=tax,
            tax_percent=quote.get("tax_percent"),
            tax_label=quote.get("tax_label"),
            included_credits=billing.included_credits_for(plan["key"]),
            currency=plan_currency,
            razorpay_order_id=subscription_id,
            razorpay_subscription_id=subscription_id,
            coupon_code=coupon_code,
            coupon_percent_off=coupon_percent,
            original_amount_paise=list_amount,
            status="created",
        ))
        db.commit()
        return {
            "action": "checkout",
            "checkout_mode": "subscription",
            "razorpay_key_id": os.getenv("RAZORPAY_KEY_ID", "").strip(),
            "subscription_id": subscription_id,
            "amount": amount,
            "original_amount": list_amount,
            "subtotal": subtotal,
            "discount_paise": list_amount - subtotal,
            "tax_paise": tax,
            "tax_label": quote["tax_label"],
            "tax_percent": quote["tax_percent"],
            "total_display": quote["total_display"],
            "subtotal_display": quote["subtotal_display"],
            "tax_display": quote["tax_display"],
            "coupon_code": coupon_code,
            "coupon_percent_off": float(coupon_percent) if coupon_percent is not None else None,
            "currency": plan_currency,
            "plan": plan["key"],
            "plan_label": plan["label"],
            "included_credits": billing.included_credits_for(plan["key"]),
            "billing_cycle": cycle,
            "auto_renew": True,
            "organization_name": organization.company_name,
            "prefill": {
                "name": current_user.full_name or "",
                "email": current_user.email or "",
            },
        }

    receipt = f"reln_{organization.id}_{secrets.token_hex(6)}"[:40]
    order = _razorpay_request("POST", "/orders", {
        "amount": amount,
        "currency": plan_currency,
        "receipt": receipt,
        "notes": notes,
    })
    order_id = str(order.get("id") or "").strip()
    if not order_id:
        raise HTTPException(status_code=502, detail="Could not create the payment order.")

    db.add(models.EnterpriseSubscriptionPayment(
        organization_id=organization.id,
        created_by_user_id=current_user.id,
        provider="razorpay",
        plan=plan["key"],
        billing_cycle=cycle,
        amount_paise=amount,          # charged (incl. tax)
        subtotal_paise=subtotal,      # taxable value
        tax_paise=tax,
        tax_percent=quote["tax_percent"],
        tax_label=quote["tax_label"],
        # Snapshotted so re-pricing the tier later cannot change what this order bought.
        included_credits=billing.included_credits_for(plan["key"]),
        original_amount_paise=list_amount,   # ex-tax list price, pre-discount
        coupon_code=coupon_code,
        coupon_percent_off=coupon_percent,
        currency=plan_currency,
        razorpay_order_id=order_id,
        status="created",
    ))
    db.commit()

    return {
        "action": "checkout",
        "checkout_mode": "order",
        "auto_renew": False,
        "razorpay_key_id": os.getenv("RAZORPAY_KEY_ID", "").strip(),
        "order_id": order_id,
        # `amount` is the charged total — the Razorpay modal must open on this.
        "amount": amount,
        "original_amount": list_amount,
        "subtotal": subtotal,
        "discount_paise": list_amount - subtotal,
        "tax_paise": tax,
        "tax_label": quote["tax_label"],
        "tax_percent": quote["tax_percent"],
        "total_display": quote["total_display"],
        "subtotal_display": quote["subtotal_display"],
        "tax_display": quote["tax_display"],
        "coupon_code": coupon_code,
        "coupon_percent_off": float(coupon_percent) if coupon_percent is not None else None,
        "currency": plan_currency,
        "plan": plan["key"],
        "plan_label": plan["label"],
        "included_credits": billing.included_credits_for(plan["key"]),
        "billing_cycle": cycle,
        "organization_name": organization.company_name,
        "prefill": {
            "name": current_user.full_name or "",
            "email": current_user.email or "",
        },
    }


@router.post("/billing/cancel")
def enterprise_billing_cancel(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Stop auto-renewal at the end of the period already paid for.

    Deliberately NOT immediate: the customer bought this month, so they keep it. Only the
    NEXT charge is cancelled. Access ends when `current_period_end` passes, exactly as it
    would if they had simply not renewed.
    """
    _, organization, _ = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="billing.manage"
    )
    sub = billing.get_or_create_org_subscription(db, organization.id, commit=False)
    mandate_id = (sub.razorpay_subscription_id or "").strip()
    if not mandate_id:
        raise HTTPException(status_code=400, detail="This plan does not auto-renew, so there is nothing to cancel.")
    if sub.cancel_at_period_end:
        return {
            "message": "Auto-renewal is already off.",
            "subscription": _serialize_subscription_state(billing.build_subscription_state(db, organization.id)),
        }
    try:
        _razorpay_request("POST", f"/subscriptions/{mandate_id}/cancel", {"cancel_at_cycle_end": 1})
    except HTTPException:
        raise
    except Exception:
        logger.exception("billing-cancel: Razorpay cancel failed for org %s", organization.id)
        raise HTTPException(status_code=502, detail="Could not cancel the subscription right now.")

    sub.cancel_at_period_end = True
    sub.canceled_at = datetime.utcnow()
    db.commit()
    return {
        "message": "Auto-renewal is off. Your plan stays active until the end of the current period.",
        "subscription": _serialize_subscription_state(billing.build_subscription_state(db, organization.id)),
    }


@router.post("/billing/verify")
def enterprise_billing_verify(
    payload: EnterpriseBillingVerifyRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, _ = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="billing.manage"
    )
    key_id, key_secret = _razorpay_credentials()
    if not key_id or not key_secret:
        raise HTTPException(status_code=503, detail="Payment verification is not configured.")

    payment_row = (
        db.query(models.EnterpriseSubscriptionPayment)
        .filter(
            models.EnterpriseSubscriptionPayment.razorpay_order_id == payload.razorpay_order_id.strip(),
            models.EnterpriseSubscriptionPayment.organization_id == organization.id,
        )
        .first()
    )
    if not payment_row:
        raise HTTPException(status_code=404, detail="Payment order not found for this organization.")

    # Idempotency: a payment may be redeemed only once. Replaying a valid (order, payment,
    # signature) triple must NOT re-extend the plan (that would allow renewal-without-payment),
    # matching the credit/infra verify paths. Return the current state instead of re-activating.
    if payment_row.status == "verified":
        return {
            "message": "This payment has already been verified.",
            "subscription": _serialize_subscription_state(billing.build_subscription_state(db, organization.id)),
        }

    # THE TWO FLOWS SIGN DIFFERENT STRINGS, IN DIFFERENT ORDERS. A one-off order signs
    # "<order_id>|<payment_id>"; a recurring mandate signs "<payment_id>|<subscription_id>".
    # Applying the order formula to a mandate rejects every legitimate recurring payment and
    # marks its row failed — the customer is charged and the plan never activates. The flow is
    # read from the STORED row, never from the request, so a caller cannot choose the formula.
    is_recurring = bool(payment_row.razorpay_subscription_id)
    signed = (
        f"{payload.razorpay_payment_id}|{payment_row.razorpay_subscription_id}"
        if is_recurring
        else f"{payload.razorpay_order_id}|{payload.razorpay_payment_id}"
    )
    expected_signature = hmac.new(
        key_secret.encode("utf-8"),
        signed.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    if not hmac.compare_digest(expected_signature, payload.razorpay_signature):
        payment_row.status = "failed"
        payment_row.error_message = "Invalid payment signature."
        db.commit()
        raise HTTPException(status_code=400, detail="Invalid payment signature.")

    # IDEMPOTENCY ACROSS BOTH PATHS, not just this one. The renewal webhook may have already
    # applied THIS charge — it arrives first whenever the buyer closes the tab or the 3DS
    # redirect is slow — and it records the charge as its own row keyed on the payment id,
    # leaving this seed row still "created". The status check above therefore does not fire,
    # and re-applying here would stack a second 30 days onto a period the webhook already
    # extended and mint a second month of credits for one payment.
    charge_id = payload.razorpay_payment_id.strip()
    already = (
        db.query(models.EnterpriseSubscriptionPayment)
        .filter(
            models.EnterpriseSubscriptionPayment.organization_id == organization.id,
            models.EnterpriseSubscriptionPayment.razorpay_payment_id == charge_id,
        )
        .first()
    )
    if already is not None:
        payment_row.status = "verified"
        payment_row.verified_at = payment_row.verified_at or datetime.utcnow()
        db.commit()
        return {
            "message": "This payment has already been applied.",
            "subscription": _serialize_subscription_state(billing.build_subscription_state(db, organization.id)),
            "wallet": credits.wallet_state(db, organization.id),
        }

    now = datetime.utcnow()
    payment_row.razorpay_payment_id = charge_id
    payment_row.status = "verified"
    payment_row.verified_at = now
    payment_row.error_message = None

    # Monthly only (billing.normalize_billing_cycle collapses everything to "monthly"), but
    # an in-flight order created before that change may still carry "yearly" — honour it.
    period_days = 365 if payment_row.billing_cycle == "yearly" else billing.PLAN_PERIOD_DAYS
    sub = billing.get_or_create_org_subscription(db, organization.id, commit=False)
    sub.plan = payment_row.plan
    sub.status = "active"
    if is_recurring:
        # NOW cancel the mandate this one replaces — the replacement is paid for, so there is
        # no window in which the org has neither. Razorpay auto-debits EVERY mandate it holds,
        # so leaving the old one live would charge an upgrading customer for both tiers every
        # month. Immediate (not cycle-end) because access is governed by `current_period_end`,
        # which this payment just extended.
        superseded = (sub.razorpay_subscription_id or "").strip()
        if superseded and superseded != payment_row.razorpay_subscription_id:
            try:
                _razorpay_request("POST", f"/subscriptions/{superseded}/cancel", {"cancel_at_cycle_end": 0})
            except Exception:
                # Already cancelled is fine. Anything else must not fail the sale the customer
                # just paid for, but it MUST be loud: an uncancelled mandate is a double charge.
                logger.exception(
                    "billing: could not cancel superseded mandate %s for org %s — CHECK FOR DOUBLE BILLING",
                    superseded, organization.id,
                )
        # Record the mandate so a renewal webhook can find this org, and clear any earlier
        # cancellation — re-subscribing must not stay flagged to stop at period end.
        sub.razorpay_subscription_id = payment_row.razorpay_subscription_id
        sub.mandate_status = "active"
        sub.cancel_at_period_end = False
        sub.canceled_at = None
    # Stack rather than reset, so paying early never shortens the period already bought.
    base = sub.current_period_end
    if base is not None and getattr(base, "tzinfo", None):
        base = base.replace(tzinfo=None)
    start = base if (base and base > now) else now
    sub.current_period_end = start + timedelta(days=period_days)

    db.commit()

    # Grant this period's included AI credits. Runs AFTER the commit above so it reads the
    # new period end and derives the right idempotency key; it is itself idempotent, and
    # the lazy sync on every wallet read is the safety net if this ever fails.
    granted = 0
    try:
        txn = credits.sync_plan_credits(db, organization.id, commit=True)
        granted = int(txn.credits) if txn is not None else 0
    except Exception:
        logger.exception("plan-credits: grant after checkout failed for org %s", organization.id)

    plan_label = (billing.get_plan(payment_row.plan) or {}).get("label", "Your")
    message = f"Your {plan_label} plan is now active."
    if granted > 0:
        message += f" {granted:,} AI credits have been added to your wallet."
    return {
        "message": message,
        "credits_granted": granted,
        "subscription": _serialize_subscription_state(billing.build_subscription_state(db, organization.id)),
        "wallet": credits.wallet_state(db, organization.id),
    }


# ===========================================================================
# Rilono Credits — prepaid wallet for premium AI features (the revenue model)
# ===========================================================================

ENTERPRISE_CREDIT_TXN_PAGE_SIZE = int(os.getenv("ENTERPRISE_CREDIT_TXN_PAGE_SIZE", "25"))


def _serialize_credit_txn(
    txn: models.EnterpriseCreditTransaction,
    client_names: Optional[dict] = None,
) -> dict:
    # Resolve the client name when this entry was a per-client action (Deep Scan,
    # mock interview) so the ledger shows *where* the credits were used.
    client_name = None
    if txn.reference_type == "client" and txn.reference_id is not None and client_names:
        client_name = client_names.get(txn.reference_id)
    return {
        "id": txn.id,
        "type": txn.type,
        "action_key": txn.action_key,
        "credits": int(txn.credits),
        "balance_after": int(txn.balance_after),
        "description": txn.description,
        "created_by_name": txn.created_by_name,
        "reference_type": txn.reference_type,
        "reference_id": txn.reference_id,
        "client_name": client_name,
        "created_at": _iso(txn.created_at),
    }


@router.get("/credits/wallet")
def enterprise_credits_wallet(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="credits.view")
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "wallet": credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization)),
        "usage": credits.usage_breakdown(
            db, organization.id, include_by_member=role.ctx.is_org_scope,
        ),
        # Priced in the org's billing currency, with the full ladder on each package so
        # the top-up modal can offer a currency selector without a second round-trip.
        "packages": credits.packages_payload(_resolve_charge_currency(None, organization)),
        "charge_currencies": list(money.supported_charge_currencies()),
        "checkout_enabled": _razorpay_enabled(),
        "razorpay_key_id": os.getenv("RAZORPAY_KEY_ID", "").strip() or None,
    }


@router.get("/credits/transactions")
def enterprise_credits_transactions(
    request: Request,
    limit: int = 25,
    offset: int = 0,
    kind: str = "",
    action: str = "",
    member_id: Optional[int] = None,
    client_id: Optional[int] = None,
    days: int = 0,
    start: str = "",
    end: str = "",
    q: str = "",
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """The credit ledger, filterable — this is the 'who / when / what / why' log
    behind the Credits → Analytics tab. Every filter is optional, so the plain
    `?limit=N` call the wallet page has always made keeps working unchanged.

    `days` selects a preset window and `start`/`end` an explicit one, resolved by
    the same helper the analytics endpoint uses so the ledger under a chart always
    covers exactly the period the chart is drawing."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="credits.view")
    # The ceiling is high because the Analytics tab's CSV export pulls the whole
    # filtered ledger in one request rather than paging through it.
    limit = max(1, min(int(limit or 25), 2000))
    offset = max(0, int(offset or 0))

    T = models.EnterpriseCreditTransaction
    query = db.query(T).filter(T.organization_id == organization.id)
    # Every per-client charge bakes the client's full name into `description` ("Deep Scan —
    # Priya Menon"), so scoping only the name-resolution map below would still hand a
    # scope-limited member the names of clients in other offices. Applied here, before the
    # caller's `client_id` and `q` filters, so neither can widen it.
    if not role.ctx.is_org_scope:
        query = query.filter(
            or_(
                T.reference_type != "client",
                T.reference_id.is_(None),
                T.reference_id.in_(scoped_client_ids_subq(db, organization.id, role.ctx)),
            )
        )

    kind = (kind or "").strip().lower()
    if kind in {"debit", "topup", "bonus", "adjustment"}:
        query = query.filter(T.type == kind)

    action = (action or "").strip().lower()
    if action:
        if action == "other":
            # Legacy / unmapped debits: anything whose action key we no longer price.
            query = query.filter(
                T.type == "debit",
                or_(T.action_key.is_(None), T.action_key.notin_(list(credits.ACTIONS.keys()))),
            )
        elif action in credits.ACTIONS:
            query = query.filter(T.action_key == action)

    if member_id is not None:
        query = query.filter(T.created_by_user_id == int(member_id))
    if client_id is not None:
        query = query.filter(T.reference_type == "client", T.reference_id == int(client_id))

    window = credits.resolve_analytics_range(days=days, start=start, end=end)
    if window["since"] is not None:
        query = query.filter(T.created_at >= window["since"])
    if window["until"] is not None:
        query = query.filter(T.created_at <= window["until"])

    q = (q or "").strip()
    if q:
        # Free-text search spans the ledger's own text AND the client the action
        # was run on, because "show me everything we spent on Priya" is the
        # question people actually type into this box.
        needle = f"%{q}%"
        matching_clients = [
            row[0] for row in scope_client_query(
                db.query(models.EnterpriseClient.id).filter(
                    models.EnterpriseClient.organization_id == organization.id,
                    models.EnterpriseClient.full_name.ilike(needle),
                ),
                role.ctx,
            ).limit(500).all()
        ]
        clauses = [T.description.ilike(needle), T.created_by_name.ilike(needle)]
        if matching_clients:
            clauses.append(and_(T.reference_type == "client", T.reference_id.in_(matching_clients)))
        query = query.filter(or_(*clauses))

    total = int(query.count())
    rows = (
        query.order_by(T.created_at.desc(), T.id.desc())
        .offset(offset).limit(limit).all()
    )
    # Batch-resolve client names for per-client actions so each ledger row can
    # show which client the credits were spent on.
    client_ids = {
        t.reference_id for t in rows
        if t.reference_type == "client" and t.reference_id is not None
    }
    client_names: dict = {}
    if client_ids:
        # Scoped: the ledger row itself is a workspace record, but the client NAME attached to it
        # must not become a way to learn who another office is working with.
        for cid, name in (
            scope_client_query(
                db.query(models.EnterpriseClient.id, models.EnterpriseClient.full_name).filter(
                    models.EnterpriseClient.organization_id == organization.id,
                    models.EnterpriseClient.id.in_(client_ids),
                ),
                role.ctx,
            )
            .all()
        ):
            client_names[cid] = name
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "transactions": [_serialize_credit_txn(t, client_names) for t in rows],
        "total": total,
        "offset": offset,
        "limit": limit,
        "has_more": (offset + len(rows)) < total,
        "range": {
            "key": window["key"],
            "custom": window["custom"],
            "days": window["days"],
            "label": window["label"],
            "since": (window["since"].isoformat() if window["since"] else None),
            "until": (window["until"].isoformat() if window["until"] else None),
        },
    }


@router.get("/credits/analytics")
def enterprise_credits_analytics(
    request: Request,
    days: int = credits.DEFAULT_ANALYTICS_DAYS,
    start: str = "",
    end: str = "",
    bucket: str = "",
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Spend analytics for the org's own wallet: where credits went, on which
    clients, by whom, over time — plus burn rate, runway and free allowances.

    The window is either the `days` preset or an explicit `start`/`end` pair of
    YYYY-MM-DD dates; `bucket` (day/week/month) overrides how the timeline and the
    period-by-period table are grouped. A malformed or out-of-range value is
    clamped by the resolver rather than rejected — a date picker mid-edit should
    not 400 the whole tab."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="credits.view")
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "wallet": credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization)),
        "analytics": credits.spend_analytics(
            db, organization.id, days=days, start=start, end=end, bucket=bucket,
            include_by_member=role.ctx.is_org_scope,
            ctx=role.ctx,
        ),
    }


@router.get("/credits/payments")
def enterprise_credits_payments(
    request: Request,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Purchase history: every top-up and infra-fee charge this org has made,
    with what was bought, what was actually paid and what was refunded.

    Admin-only — these rows carry payment amounts and processor references,
    which the rest of the team has no reason to see."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="credits.purchase")
    perms = _enterprise_permissions_for_role(role)
    if not perms["can_manage_users"]:
        raise HTTPException(
            status_code=403,
            detail="Only an organization admin can view purchase history and receipts.",
        )
    limit = max(1, min(int(limit or 50), 200))

    P = models.EnterpriseCreditPayment
    rows = (
        db.query(P)
        .filter(
            P.organization_id == organization.id,
            # 'created' rows are abandoned checkouts — an order was opened and
            # never paid. Showing them as "history" would be misleading.
            P.status != "created",
        )
        .order_by(P.verified_at.desc().nullslast(), P.id.desc())
        .limit(limit)
        .all()
    )

    buyer_names: dict = {}
    buyer_ids = {p.created_by_user_id for p in rows if p.created_by_user_id}
    if buyer_ids:
        for uid, full_name, email in (
            db.query(models.User.id, models.User.full_name, models.User.email)
            .filter(models.User.id.in_(buyer_ids)).all()
        ):
            buyer_names[uid] = full_name or email

    payments = []
    # Every *_paise column on these rows is minor units of THAT ROW's currency, so a
    # running total can only be kept per currency — adding a $39 top-up's 3900 cents to a
    # ₹2,999 row's 299900 paise yields a number that is neither. Keyed by currency, then
    # rendered as one string per currency the org has actually paid in.
    collected_by_currency: dict[str, int] = {}
    credits_bought = 0
    for p in rows:
        package = credits.get_package(p.package_key) if p.kind == "credits" else None
        pay_currency = money.normalize_currency(p.currency, strict=False)
        refunded = int(p.refunded_amount_paise or 0)
        net = max(0, int(p.amount_paise or 0) - refunded)
        total_credits = int(p.credits or 0) + int(p.bonus_credits or 0)
        if p.status in credits.REVENUE_PAYMENT_STATUSES:
            collected_by_currency[pay_currency] = collected_by_currency.get(pay_currency, 0) + net
            credits_bought += total_credits
        payments.append({
            "id": p.id,
            "kind": p.kind,
            "package_key": p.package_key,
            "label": (
                (package or {}).get("label")
                or ("Infrastructure server fee" if p.kind == "infra_fee" else (p.package_key or "Top-up"))
            ),
            "credits": int(p.credits or 0),
            "bonus_credits": int(p.bonus_credits or 0),
            "total_credits": total_credits,
            # Legacy field name; minor units of `currency`, which now travels with it so
            # the client never has to infer rupees from the key.
            "amount_paise": int(p.amount_paise or 0),
            "currency": pay_currency,
            "amount_display": money.format_money(p.amount_paise, pay_currency),
            "original_amount_paise": p.original_amount_paise,
            "coupon_code": p.coupon_code,
            "coupon_percent_off": (float(p.coupon_percent_off) if p.coupon_percent_off is not None else None),
            "status": p.status,
            "refunded_amount_paise": refunded,
            "refunded_display": money.format_money(refunded, pay_currency),
            "net_amount_paise": net,
            "net_amount_display": money.format_money(net, pay_currency),
            "payment_reference": p.razorpay_payment_id or p.razorpay_order_id,
            "created_by_name": buyer_names.get(p.created_by_user_id),
            "created_at": _iso(p.created_at),
            "verified_at": _iso(p.verified_at),
        })

    # Plan subscription charges. They live in their own table, so before the tiered plans
    # launched this receipt list showed only top-ups — an org would have been charged
    # ₹3,538.82 every month with no receipt for it anywhere in the product.
    SP = models.EnterpriseSubscriptionPayment
    plan_rows = (
        db.query(SP)
        .filter(SP.organization_id == organization.id, SP.status != "created")
        .order_by(SP.verified_at.desc().nullslast(), SP.id.desc())
        .limit(limit)
        .all()
    )
    plan_buyer_ids = {p.created_by_user_id for p in plan_rows if p.created_by_user_id} - set(buyer_names)
    if plan_buyer_ids:
        for uid, full_name, email in (
            db.query(models.User.id, models.User.full_name, models.User.email)
            .filter(models.User.id.in_(plan_buyer_ids)).all()
        ):
            buyer_names[uid] = full_name or email

    for p in plan_rows:
        pay_currency = money.normalize_currency(p.currency, strict=False)
        amount = int(p.amount_paise or 0)
        tax = int(p.tax_paise or 0)
        if p.status in credits.REVENUE_PAYMENT_STATUSES:
            collected_by_currency[pay_currency] = collected_by_currency.get(pay_currency, 0) + amount
        plan_meta = billing.get_plan(p.plan) or {}
        payments.append({
            "id": p.id,
            "kind": "plan",
            "package_key": p.plan,
            "label": f"{plan_meta.get('label', p.plan)} plan · monthly",
            "credits": 0,
            "bonus_credits": 0,
            # Credits the plan period granted — shown in the same column as a top-up's, so
            # the receipt list answers "what did this buy" the same way for both.
            "total_credits": int(p.included_credits or 0),
            "amount_paise": amount,
            "currency": pay_currency,
            "amount_display": money.format_money(amount, pay_currency),
            "subtotal_paise": int(p.subtotal_paise or 0) or None,
            "tax_paise": tax,
            "tax_label": p.tax_label,
            "tax_display": money.format_money(tax, pay_currency) if tax else None,
            "original_amount_paise": p.original_amount_paise,
            "coupon_code": p.coupon_code,
            "coupon_percent_off": (float(p.coupon_percent_off) if p.coupon_percent_off is not None else None),
            "status": p.status,
            "refunded_amount_paise": 0,
            "refunded_display": money.format_money(0, pay_currency),
            "net_amount_paise": amount,
            "net_amount_display": money.format_money(amount, pay_currency),
            "payment_reference": p.razorpay_payment_id or p.razorpay_order_id,
            "created_by_name": buyer_names.get(p.created_by_user_id),
            "created_at": _iso(p.created_at),
            "verified_at": _iso(p.verified_at),
        })

    # One list, newest first, regardless of which table a charge came from.
    payments.sort(key=lambda r: (r.get("verified_at") or r.get("created_at") or ""), reverse=True)

    return {
        "permissions": perms,
        "payments": payments,
        "summary": {
            "count": len(payments),
            # Kept for the deployed SPA, but it is ONLY the INR subtotal now — it is not a
            # grand total and must never be presented as one on a mixed-currency account.
            "collected_paise": int(collected_by_currency.get("INR", 0)),
            "collected_by_currency": [
                {"currency": code, "amount_minor": amt, "display": money.format_money(amt, code)}
                for code, amt in sorted(collected_by_currency.items())
            ],
            "collected_display": (
                " + ".join(
                    money.format_money(amt, code)
                    for code, amt in sorted(collected_by_currency.items())
                )
                or money.format_money(0, "INR")
            ),
            "credits_purchased": credits_bought,
        },
    }


@router.post("/coupons/validate")
def enterprise_coupon_validate(
    payload: EnterpriseCouponValidateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Preview a per-account discount code for a given purchase before checkout."""
    _, organization, _ = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="credits.purchase")
    context = (payload.context or "credits").strip().lower()
    if context not in ("credits", "billing"):
        raise HTTPException(status_code=400, detail="Unknown checkout context.")

    if context == "credits":
        package = credits.get_package(payload.package)
        if not package:
            raise HTTPException(status_code=400, detail="Please choose a valid credit package.")
        # This is a PREVIEW of a specific checkout, so it must price the item exactly the
        # way that checkout will. Reading PACKAGES[...]["amount_paise"] quoted the INR list
        # price to every org: a USD buyer saw "₹999 → ₹499" and was then charged
        # $12.99 → $6.49. Same resolver, same price book, same currency as the order.
        currency = _resolve_charge_currency(None, organization)
        base_amount = credits.package_price_minor(package["key"], currency)
    else:
        plan = billing.get_plan(payload.plan)
        if not plan or plan["key"] not in billing.PAID_PLAN_KEYS:
            raise HTTPException(status_code=400, detail="Please choose a valid paid plan.")
        cycle = billing.normalize_billing_cycle(payload.billing_cycle)
        # The SaaS plan book is still INR-only (billing.plan_amount_paise has no currency
        # dimension), so this branch is genuinely rupees. Say so rather than inheriting it.
        currency = "INR"
        base_amount = billing.plan_amount_paise(plan["key"], cycle)
    if base_amount <= 0:
        raise HTTPException(status_code=400, detail="This item is not available for checkout.")

    coupon = enterprise_coupons.resolve_active_coupon_or_400(
        db, organization.id, payload.code, context=context
    )
    percent = enterprise_coupons.parse_percent_off(coupon.percent_off)
    # A fully-covering discount (e.g. 100% off) leaves a zero amount. Don't block — report
    # it as free; checkout will grant the credits without Razorpay. The "is it free?"
    # threshold is per-currency for the same reason the checkout floor is: the shared
    # constant means ₹1, and reusing it on a USD order would mean $1.00 and call a
    # 93%-off order free. Must stay identical to the test in the checkout handler.
    amount = enterprise_coupons.compute_discounted_amount_paise(base_amount, percent)
    is_free = amount < money.min_charge_minor(currency)
    return {
        "valid": True,
        "free": is_free,
        "code": enterprise_coupons.normalize_code(coupon.code),
        "percent_off": float(percent),
        "percent_display": enterprise_coupons.format_percent_off(percent) + "%",
        # Minor units of `currency` below — paise only when that says INR.
        "currency": currency,
        "base_amount_paise": base_amount,
        "amount_paise": amount,
        "discount_paise": base_amount - amount,
        "base_amount_display": money.format_money(base_amount, currency),
        "amount_display": money.format_money(amount, currency),
    }


def _grant_free_credit_topup(
    db, organization, current_user, package, base_amount, amount, coupon_code, coupon_percent,
    currency=None,
):
    """A discount covered the full amount (e.g. 100% off) → add the credits without
    Razorpay. Records a verified 'free' payment so the redemption is counted and the
    purchase shows in history, then credits the wallet (idempotent ledger reference).

    `currency` is the currency the checkout was priced in — `amount` and `base_amount` are
    minor units OF THAT CURRENCY. Stamping credits.CURRENCY here instead wrote a USD-priced
    row (original_amount_paise=1299) labelled INR, which every downstream reader — purchase
    history, the admin console, the revenue sums — would then render and add as rupees."""
    total_credits = int(package["credits"]) + int(package["bonus_credits"])
    payment = models.EnterpriseCreditPayment(
        organization_id=organization.id,
        created_by_user_id=current_user.id,
        provider="free",
        kind="credits",
        package_key=package["key"],
        credits=int(package["credits"]),
        bonus_credits=int(package["bonus_credits"]),
        amount_paise=int(amount),
        original_amount_paise=int(base_amount),
        coupon_code=coupon_code,
        coupon_percent_off=coupon_percent,
        currency=(currency or credits.CURRENCY),
        # A fully-discounted order never reached a gateway, so there is no settlement to
        # look up. It is worth 0 either way, and stating that explicitly keeps the revenue
        # sums (which read base_amount_paise) from falling back to a non-INR amount_paise.
        base_amount_paise=0,
        base_currency="INR",
        razorpay_order_id=f"free_{secrets.token_hex(8)}",  # NOT NULL + unique column
        status="verified",
        verified_at=datetime.utcnow(),
    )
    db.add(payment)
    db.flush()  # assign payment.id for the ledger reference
    pct = enterprise_coupons.format_percent_off(coupon_percent) if coupon_percent is not None else ""
    credits.add_credits(
        db, organization.id, total_credits,
        txn_type="topup",
        description=f"{package['label']} (+{total_credits} credits · {coupon_code} {pct}% off)",
        reference_type="payment", reference_id=payment.id,
        user=current_user, commit=False,
    )
    db.commit()
    return {
        "action": "granted",
        "message": f"{coupon_code} covered the full amount — {total_credits} credits added to your wallet.",
        "total_credits": total_credits,
        "coupon_code": coupon_code,
        "wallet": credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization)),
    }


@router.post("/credits/topup/checkout")
def enterprise_credits_topup_checkout(
    payload: EnterpriseCreditTopupRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, _ = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="credits.purchase"
    )
    package = credits.get_package(payload.package)
    if not package:
        raise HTTPException(status_code=400, detail="Please choose a valid credit package.")
    # Currency is a client HINT; the price is always resolved server-side from the shared
    # price book. An unsupported code is a 400 rather than a silent fall back to INR.
    currency = _resolve_charge_currency(payload.currency, organization)
    base_amount = credits.package_price_minor(package["key"], currency)
    if base_amount <= 0:
        raise HTTPException(status_code=400, detail="This package is not available for checkout.")

    # Per-account discount code (admin-managed). Reduces the payable amount.
    amount = base_amount
    coupon_code = None
    coupon_percent = None
    raw_coupon = (payload.coupon_code or "").strip()
    if raw_coupon:
        coupon = enterprise_coupons.resolve_active_coupon_or_400(
            db, organization.id, raw_coupon, context="credits"
        )
        coupon_percent = enterprise_coupons.parse_percent_off(coupon.percent_off)
        coupon_code = enterprise_coupons.normalize_code(coupon.code)
        amount = enterprise_coupons.compute_discounted_amount_paise(base_amount, coupon_percent)

    # Fully covered by the discount → grant the credits for free (no Razorpay order).
    #
    # The threshold has to be currency-aware for the same reason the floor below does.
    # enterprise_coupons.is_free_checkout() hardcodes MIN_CHECKOUT_PAISE = 100, which means
    # "₹1" — but `amount` is now minor units of `currency`, so on a USD order the same 100
    # means "$1.00". A 93%-off coupon leaves $0.91 on a $12.99 package, which that helper
    # would call "free" and hand the credits over for nothing, while the identical coupon
    # on the INR list price still charges ₹69.93. Compare against the same per-currency
    # floor the charge path uses so a discount can never buy the package outright.
    checkout_floor = money.min_charge_minor(currency)
    if coupon_code and amount < checkout_floor:
        return _grant_free_credit_topup(
            db, organization, current_user, package, base_amount, amount, coupon_code,
            coupon_percent, currency,
        )
    # Currency-aware floor. The old constant was 100 = "₹1"; the same integer in USD
    # would be $1.00, a 87× stricter floor applied by accident.
    if amount < checkout_floor:
        raise HTTPException(status_code=400, detail="This amount is too low for online checkout.")

    if not _razorpay_enabled():
        return {
            "action": "contact_sales",
            "message": "Online top-up is being enabled. Please contact us to add credits.",
        }

    receipt = f"reln_cr_{organization.id}_{currency.lower()}_{secrets.token_hex(4)}"[:40]
    order = _razorpay_request("POST", "/orders", {
        "amount": amount,
        "currency": currency,
        "receipt": receipt,
        "notes": {
            "organization_id": str(organization.id),
            "kind": "credits",
            "package": package["key"],
            "user_id": str(current_user.id),
            "coupon_code": coupon_code or "",
            "currency": currency,
            "price_book_version": money.PRICE_BOOK_VERSION,
        },
    })
    order_id = str(order.get("id") or "").strip()
    if not order_id:
        raise HTTPException(status_code=502, detail="Could not create the payment order.")

    db.add(models.EnterpriseCreditPayment(
        organization_id=organization.id,
        created_by_user_id=current_user.id,
        provider="razorpay",
        kind="credits",
        package_key=package["key"],
        credits=int(package["credits"]),
        bonus_credits=int(package["bonus_credits"]),
        amount_paise=amount,
        original_amount_paise=base_amount,
        coupon_code=coupon_code,
        coupon_percent_off=coupon_percent,
        currency=currency,
        price_book_version=money.PRICE_BOOK_VERSION,
        razorpay_order_id=order_id,
        status="created",
    ))
    # Make the currency choice sticky so the next top-up doesn't re-guess from country.
    if getattr(organization, "billing_currency", None) != currency:
        organization.billing_currency = currency
    db.commit()

    return {
        "action": "checkout",
        "razorpay_key_id": os.getenv("RAZORPAY_KEY_ID", "").strip(),
        "order_id": order_id,
        "amount": amount,
        "original_amount": base_amount,
        "discount_paise": base_amount - amount,
        "coupon_code": coupon_code,
        "coupon_percent_off": float(coupon_percent) if coupon_percent is not None else None,
        "currency": currency,
        "amount_display": money.format_money(amount, currency),
        "package": package["key"],
        "package_label": package["label"],
        "total_credits": int(package["credits"]) + int(package["bonus_credits"]),
        "organization_name": organization.company_name,
        # International payments fail on placeholder contact data, so send the real
        # values we hold and omit rather than invent.
        # https://razorpay.com/docs/payments/international-payments/?preferred-country=IN
        "prefill": {
            "name": current_user.full_name or "",
            "email": current_user.email or "",
            "contact": (getattr(current_user, "phone", None) or ""),
        },
    }


@router.post("/credits/topup/verify")
def enterprise_credits_topup_verify(
    payload: EnterpriseCreditVerifyRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, _ = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="credits.purchase"
    )
    payment_row = _verify_credit_payment_or_402(
        db=db, organization=organization, payload=payload, expected_kind="credits"
    )

    total_credits = int(payment_row.credits) + int(payment_row.bonus_credits)
    # Idempotency: credit the wallet only once (skip if a ledger row already exists).
    already = (
        db.query(models.EnterpriseCreditTransaction)
        .filter(
            models.EnterpriseCreditTransaction.organization_id == organization.id,
            models.EnterpriseCreditTransaction.reference_type == "payment",
            models.EnterpriseCreditTransaction.reference_id == payment_row.id,
        )
        .first()
    )
    if not already and total_credits > 0:
        pkg = credits.get_package(payment_row.package_key)
        label = pkg["label"] if pkg else "Credit top-up"
        credits.add_credits(
            db, organization.id, total_credits,
            txn_type="topup",
            description=f"{label} (+{total_credits} credits)",
            reference_type="payment", reference_id=payment_row.id,
            user=current_user, commit=False,
        )
    db.commit()

    return {
        "message": f"{total_credits} credits added to your wallet.",
        "wallet": credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization)),
    }


@router.post("/credits/infra/checkout")
def enterprise_infra_fee_checkout(
    request: Request,
    payload: Optional[EnterpriseInfraCheckoutRequest] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """RETIRED 2026-08-02 — the ₹999/month infrastructure server fee no longer exists.

    Kept as an explicit 410 rather than deleted: a browser running a cached SPA bundle
    still has the "Activate ₹999/mo" button, and a 404 there would read as a bug. The
    verify sibling below stays FUNCTIONAL so an order created moments before the cutover
    can still be honoured — only new orders are refused.
    """
    _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="billing.manage"
    )
    raise HTTPException(
        status_code=410,
        detail=(
            "The infrastructure server fee has been replaced by Rilono plans. "
            "Open Plans & Billing to choose a plan — it includes your monthly AI credits."
        ),
    )


def _retired_infra_fee_checkout_body(organization, current_user, payload, db):
    # Dead code retained for one release so the retirement is a one-line revert if the
    # rollout is halted. Nothing calls it.
    currency = _resolve_charge_currency(payload.currency if payload else None, organization)
    amount = money.price_minor("infra_fee", currency)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="The infrastructure fee is not configured.")
    if not _razorpay_enabled():
        return {
            "action": "contact_sales",
            "message": "Online payment is being enabled. Please contact us to activate the infrastructure fee.",
        }

    receipt = f"reln_infra_{organization.id}_{currency.lower()}_{secrets.token_hex(4)}"[:40]
    order = _razorpay_request("POST", "/orders", {
        "amount": amount,
        "currency": currency,
        "receipt": receipt,
        "notes": {
            "organization_id": str(organization.id),
            "kind": "infra_fee",
            "user_id": str(current_user.id),
            "currency": currency,
            "price_book_version": money.PRICE_BOOK_VERSION,
        },
    })
    order_id = str(order.get("id") or "").strip()
    if not order_id:
        raise HTTPException(status_code=502, detail="Could not create the payment order.")

    db.add(models.EnterpriseCreditPayment(
        organization_id=organization.id,
        created_by_user_id=current_user.id,
        provider="razorpay",
        kind="infra_fee",
        package_key="infra",
        credits=0,
        bonus_credits=0,
        amount_paise=amount,
        currency=currency,
        price_book_version=money.PRICE_BOOK_VERSION,
        razorpay_order_id=order_id,
        status="created",
    ))
    if getattr(organization, "billing_currency", None) != currency:
        organization.billing_currency = currency
    db.commit()

    return {
        "action": "checkout",
        "razorpay_key_id": os.getenv("RAZORPAY_KEY_ID", "").strip(),
        "order_id": order_id,
        "amount": amount,
        "currency": currency,
        "amount_display": money.format_money(amount, currency),
        "organization_name": organization.company_name,
        "prefill": {
            "name": current_user.full_name or "",
            "email": current_user.email or "",
            "contact": (getattr(current_user, "phone", None) or ""),
        },
    }


@router.post("/credits/infra/verify")
def enterprise_infra_fee_verify(
    payload: EnterpriseCreditVerifyRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, _ = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="billing.manage"
    )
    payment_row = _verify_credit_payment_or_402(
        db=db, organization=organization, payload=payload, expected_kind="infra_fee"
    )
    # Idempotency: only extend the paid-until window once per payment.
    already = (
        db.query(models.EnterpriseCreditTransaction)
        .filter(
            models.EnterpriseCreditTransaction.organization_id == organization.id,
            models.EnterpriseCreditTransaction.reference_type == "infra_payment",
            models.EnterpriseCreditTransaction.reference_id == payment_row.id,
        )
        .first()
    )
    if not already:
        credits.mark_infra_fee_paid(db, organization.id, commit=False)
        # Record a zero-credit ledger marker so the credit ledger shows the charge.
        wallet = credits.get_or_create_wallet(db, organization.id, commit=False)
        db.add(models.EnterpriseCreditTransaction(
            organization_id=organization.id,
            type="infra_fee",
            credits=0,
            balance_after=int(wallet.balance_credits),
            # Describe what THIS org was actually charged, from the payment row that just
            # verified — the INR list constant printed "₹999/mo" into the ledger of an org
            # that had paid $12.99.
            description=(
                "Infrastructure server fee ("
                + money.format_money(
                    payment_row.amount_paise,
                    money.normalize_currency(payment_row.currency, strict=False),
                )
                + "/mo)"
            ),
            reference_type="infra_payment",
            reference_id=payment_row.id,
            created_by_user_id=current_user.id,
            created_by_name=current_user.full_name or current_user.email,
        ))
    db.commit()

    return {
        "message": "Infrastructure server fee activated.",
        "wallet": credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization)),
    }


def _verify_credit_payment_or_402(
    *,
    db: Session,
    organization: models.EnterpriseOrganization,
    payload: EnterpriseCreditVerifyRequest,
    expected_kind: str,
) -> models.EnterpriseCreditPayment:
    """Validate a Razorpay signature for a credit/infra payment and mark it verified."""
    key_id, key_secret = _razorpay_credentials()
    if not key_id or not key_secret:
        raise HTTPException(status_code=503, detail="Payment verification is not configured.")

    payment_row = (
        db.query(models.EnterpriseCreditPayment)
        .filter(
            models.EnterpriseCreditPayment.razorpay_order_id == payload.razorpay_order_id.strip(),
            models.EnterpriseCreditPayment.organization_id == organization.id,
            models.EnterpriseCreditPayment.kind == expected_kind,
        )
        # Lock the payment row (Postgres) so two concurrent verifies for the SAME payment
        # serialize: the second blocks here until the first commits, then the caller's
        # "already credited?" ledger check sees the committed row and skips — closing the
        # non-atomic check-then-insert race that could otherwise double-credit the wallet.
        .with_for_update()
        .first()
    )
    if not payment_row:
        raise HTTPException(status_code=404, detail="Payment order not found for this organization.")

    expected_signature = hmac.new(
        key_secret.encode("utf-8"),
        f"{payload.razorpay_order_id}|{payload.razorpay_payment_id}".encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    if not hmac.compare_digest(expected_signature, payload.razorpay_signature):
        payment_row.status = "failed"
        payment_row.error_message = "Invalid payment signature."
        db.commit()
        raise HTTPException(status_code=400, detail="Invalid payment signature.")

    # The signature proves only that "order_id|payment_id" came from Razorpay — it binds
    # neither the amount nor the currency, and says nothing about whether the money was
    # captured. Re-fetch both entities and assert against what we stored at checkout.
    # These run INSIDE the with_for_update() lock above so the double-credit guard holds.
    order_data = _razorpay_request("GET", f"/orders/{payload.razorpay_order_id.strip()}")
    payment_data = _razorpay_request("GET", f"/payments/{payload.razorpay_payment_id.strip()}")

    def _reject(detail: str):
        payment_row.status = "failed"
        payment_row.error_message = detail
        db.commit()
        raise HTTPException(status_code=400, detail=detail)

    expected_currency = money.normalize_currency(payment_row.currency, strict=False)
    if str(order_data.get("id") or "") != payment_row.razorpay_order_id:
        _reject("Razorpay order mismatch.")
    if str(order_data.get("status", "")).lower() != "paid":
        _reject("This payment has not completed yet.")
    if int(order_data.get("amount", 0) or 0) != int(payment_row.amount_paise):
        _reject("Payment amount mismatch.")
    if money.normalize_currency(order_data.get("currency"), strict=False) != expected_currency:
        _reject("Payment currency mismatch.")
    if str(payment_data.get("order_id") or "") != payment_row.razorpay_order_id:
        _reject("Payment does not belong to this order.")
    if int(payment_data.get("amount", 0) or 0) != int(payment_row.amount_paise):
        _reject("Payment amount mismatch.")
    if money.normalize_currency(payment_data.get("currency"), strict=False) != expected_currency:
        _reject("Payment currency mismatch.")
    # Authorized-but-uncaptured still produces a valid checkout signature, and capture
    # failure is materially more common on international cards. Crediting a wallet here
    # would hand out credits for money that never arrives.
    if not bool(payment_data.get("captured")):
        _reject("This payment has not been captured yet.")

    if payment_row.status != "verified":
        payment_row.razorpay_payment_id = payload.razorpay_payment_id.strip()
        # Razorpay's own INR settlement figure — the only amount that may be summed for
        # revenue once rows carry mixed currencies.
        base_minor, fx_rate = money.settled_inr_minor(payment_data)
        if base_minor is not None:
            payment_row.base_amount_paise = base_minor
            payment_row.fx_rate_used = fx_rate
        payment_row.base_currency = "INR"
        payment_row.is_international = bool(payment_data.get("international"))
        payment_row.status = "verified"
        payment_row.verified_at = datetime.utcnow()
        payment_row.error_message = None
    return payment_row


ENTERPRISE_SIGNUP_OTP_EXPIRES_MINUTES = int(os.getenv("ENTERPRISE_SIGNUP_OTP_EXPIRES_MINUTES", "10") or "10")
ENTERPRISE_SIGNUP_OTP_MAX_ATTEMPTS = int(os.getenv("ENTERPRISE_SIGNUP_OTP_MAX_ATTEMPTS", "5") or "5")


class EnterpriseSignupSendCodeRequest(BaseModel):
    email: EmailStr
    cf_turnstile_token: Optional[str] = None


@router.post("/signup/send-code")
def enterprise_signup_send_code(
    payload: EnterpriseSignupSendCodeRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """Step 1 of workspace signup: email the owner a 6-digit code. The workspace itself
    is only created by /signup once the code is verified — so unverified emails never
    produce junk workspaces or claim subdomains."""
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.signup_send_code", limit=6, window_seconds=3600,
    )

    # Bot gate lives on this step (the final /signup is gated by the code itself).
    turnstile_token = (payload.cf_turnstile_token or "").strip()
    if is_turnstile_enabled() and not is_request_ip_whitelisted(request):
        if turnstile_token:
            client_ip = extract_client_ip(request) if request else None
            if not verify_turnstile_token(turnstile_token, client_ip):
                raise HTTPException(status_code=400, detail="Security verification failed. Please try again.")
        elif _is_turnstile_required():
            raise HTTPException(status_code=400, detail="Security verification is required.")

    email = payload.email.lower().strip()
    if db.query(models.User).filter(models.User.email == email).first():
        raise HTTPException(
            status_code=409,
            detail=ENTERPRISE_SIGNUP_EMAIL_TAKEN_DETAIL,
        )

    code = f"{secrets.randbelow(1_000_000):06d}"
    row = (
        db.query(models.EnterpriseSignupOtp)
        .filter(models.EnterpriseSignupOtp.email == email)
        .first()
    )
    expires_at = datetime.utcnow() + timedelta(minutes=ENTERPRISE_SIGNUP_OTP_EXPIRES_MINUTES)
    if row:
        row.code_hash = hash_token(code)
        row.expires_at = expires_at
        row.attempts = 0
    else:
        db.add(models.EnterpriseSignupOtp(email=email, code_hash=hash_token(code), expires_at=expires_at))
    db.commit()

    sent = send_email_otp(email, code, expires_in_minutes=ENTERPRISE_SIGNUP_OTP_EXPIRES_MINUTES)
    result = {
        "message": f"We've emailed a 6-digit code to {email}.",
        "expires_in_minutes": ENTERPRISE_SIGNUP_OTP_EXPIRES_MINUTES,
    }
    if not sent:
        if _is_development_env():
            # Local sandbox without an email provider — surface the code for testing.
            result["dev_code"] = code
        else:
            raise HTTPException(
                status_code=502,
                detail="We couldn't send the verification email right now. Please try again in a moment.",
            )
    return result


def _verify_signup_otp_or_400(db: Session, email: str, code: Optional[str]) -> models.EnterpriseSignupOtp:
    """Validate the signup email code; raises 400 with a friendly message otherwise."""
    code_clean = "".join(ch for ch in str(code or "") if ch.isdigit())
    if len(code_clean) != 6:
        raise HTTPException(status_code=400, detail="Enter the 6-digit code we emailed you.")
    row = (
        db.query(models.EnterpriseSignupOtp)
        .filter(models.EnterpriseSignupOtp.email == email)
        .first()
    )
    if not row:
        raise HTTPException(status_code=400, detail="No verification code found for this email — please request a new one.")
    expires = row.expires_at.replace(tzinfo=None) if getattr(row.expires_at, "tzinfo", None) else row.expires_at
    if expires < datetime.utcnow():
        raise HTTPException(status_code=400, detail="That code has expired — please request a new one.")
    if int(row.attempts or 0) >= ENTERPRISE_SIGNUP_OTP_MAX_ATTEMPTS:
        raise HTTPException(status_code=400, detail="Too many incorrect attempts — please request a new code.")
    if not token_matches(code_clean, row.code_hash):
        row.attempts = int(row.attempts or 0) + 1
        db.commit()
        raise HTTPException(status_code=400, detail="That code isn't right — please check the email and try again.")
    return row


@router.post("/signup")
def enterprise_signup(
    payload: EnterpriseSignupRequest,
    request: Request,
    response: Response,
    db: Session = Depends(get_db),
):
    # Plain `def` for the same reason as enterprise_login: it never awaits, but it runs a
    # bcrypt hash (~170ms CPU), Turnstile verification and sync DB writes. Sync handlers run
    # in FastAPI's threadpool, keeping this off the shared event loop.
    _enforce_rate_limit_or_429(
        request=request,
        scope="enterprise.signup",
        limit=ENTERPRISE_SIGNUP_RATE_LIMIT,
        window_seconds=ENTERPRISE_SIGNUP_RATE_WINDOW_SECONDS,
    )

    # Bot-gating (Turnstile) happens at /signup/send-code; this step is gated by the
    # emailed code itself, which is a stronger proof (a verified, reachable inbox).

    if not payload.accepted_terms_privacy:
        raise HTTPException(
            status_code=400,
            detail="You must accept the Terms & Conditions and Privacy Policy to create your workspace.",
        )

    if not payload.accepted_dpa:
        raise HTTPException(
            status_code=400,
            detail="You must accept the Data Processing Agreement to manage client data in your workspace.",
        )

    company_name = (payload.company_name or "").strip()
    if len(company_name) < 2:
        raise HTTPException(status_code=400, detail="Company name must be at least 2 characters.")
    full_name = (payload.full_name or "").strip()
    if len(full_name) < 2:
        raise HTTPException(status_code=400, detail="Your name must be at least 2 characters.")
    email = payload.email.lower().strip()
    subdomain_slug = _parse_enterprise_subdomain_or_400(payload.subdomain_slug)

    password_error = validate_password_strength(payload.password, email)
    if password_error:
        raise HTTPException(status_code=400, detail=password_error)

    existing_user = db.query(models.User).filter(models.User.email == email).first()
    if existing_user:
        raise HTTPException(
            status_code=409,
            detail=ENTERPRISE_SIGNUP_EMAIL_TAKEN_DETAIL,
        )

    # The owner must prove the inbox before anything is created (or a subdomain claimed).
    otp_row = _verify_signup_otp_or_400(db, email, payload.email_otp)

    _assert_subdomain_available(db=db, subdomain_slug=subdomain_slug)

    password_hash = get_password_hash(payload.password)

    # Capture proof-of-consent: when, from where, and which version of the legal
    # documents the owner agreed to. Stored on the user row for audit/compliance.
    consent_ip = extract_client_ip(request) if request else None
    consent_user_agent = (request.headers.get("user-agent") if request else None) or None
    if consent_user_agent:
        consent_user_agent = consent_user_agent[:1000]

    # Create the owner user. Enterprise access is granted via their organization
    # membership below (see _has_enterprise_access) — NOT via an EnterpriseCredential.
    # Credentials are reserved for platform-admin-granted accounts and would otherwise
    # be promoted to is_admin by seed_enterprise_user() on the next restart.
    user = models.User(
        email=email,
        username=None,
        hashed_password=password_hash,
        full_name=full_name,
        university=company_name,
        is_active=True,
        email_verified=True,
        # Enterprise-created account (workspace owner) → blocked from the B2C consumer app.
        is_enterprise_account=True,
        accepted_terms_privacy_at=datetime.utcnow(),
        accepted_terms_privacy_ip=consent_ip,
        accepted_terms_privacy_user_agent=consent_user_agent,
        accepted_terms_privacy_version=LEGAL_TERMS_PRIVACY_VERSION,
        age_confirmed_at=datetime.utcnow(),
        marketing_emails_consent=bool(payload.marketing_emails_consent),
        marketing_emails_consent_at=(datetime.utcnow() if payload.marketing_emails_consent else None),
    )
    db.add(user)
    db.flush()

    organization = models.EnterpriseOrganization(
        company_name=company_name,
        subdomain_slug=subdomain_slug,
        logo_url=_build_default_enterprise_logo_url(
            organization_id=None,
            company_name=company_name,
            subdomain_slug=subdomain_slug,
            randomize=True,
        ),
        created_by_user_id=user.id,
        # Proof the organization accepted the Data Processing Agreement (it is the
        # controller of its clients' data; Rilono is the processor).
        dpa_accepted_at=datetime.utcnow(),
        dpa_accepted_version=LEGAL_DPA_VERSION,
        dpa_accepted_by_user_id=user.id,
    )
    db.add(organization)
    db.flush()

    db.add(models.EnterpriseOrganizationMember(
        organization_id=organization.id,
        user_id=user.id,
        # role_key is the source of truth — `role` below is only its legacy mirror
        # (legacy_role_for(owner) == "admin"). It MUST be set explicitly: the column
        # defaults to "viewer", and normalize_role_key() trusts a present role_key over
        # the mirror, so omitting it makes the person who just created the workspace a
        # viewer of it — locked out of settings, billing and invites. Owner (not admin)
        # because access.resolve_access_context derives is_owner from this exact value.
        role_key=access.ROLE_OWNER,
        role=ENTERPRISE_ROLE_ADMIN,
        is_active=True,
        invited_by_user_id=user.id,
    ))
    organization.logo_url = _build_default_enterprise_logo_url(
        organization_id=organization.id,
        company_name=company_name,
        subdomain_slug=subdomain_slug,
        randomize=True,
    )
    billing.get_or_create_org_subscription(db, organization.id, commit=False)

    user.last_login_at = datetime.utcnow()
    user.first_login_at = datetime.utcnow()
    # The signup code is single-use: burn it with the same commit that creates the workspace.
    db.delete(otp_row)
    db.commit()
    db.refresh(user)

    portal_url = _build_enterprise_portal_url(subdomain_slug, request)

    # Warm welcome email to the new workspace owner — best-effort, never block signup.
    try:
        send_enterprise_welcome_email(
            to_email=user.email,
            full_name=user.full_name or "",
            company=company_name,
            portal_url=portal_url,
        )
    except Exception:
        logger.exception("Enterprise welcome email failed for %s", user.email)

    access_token = create_access_token(
        data={"sub": user.email},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
    )
    set_auth_cookie(request, response, access_token)

    context = _build_enterprise_context(db, user, request)
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": user.id,
            "email": user.email,
            "full_name": user.full_name,
            "is_admin": user.is_admin,
        },
        "portal_url": portal_url,
        **context,
    }


# ===========================================================================
# AI copilot (Gemini function-calling agent, read-only, org-scoped)
# ===========================================================================

ENTERPRISE_AI_RATE_LIMIT = int(os.getenv("ENTERPRISE_AI_RATE_LIMIT", "30"))
ENTERPRISE_AI_RATE_WINDOW_SECONDS = int(os.getenv("ENTERPRISE_AI_RATE_WINDOW_SECONDS", "300"))

ENTERPRISE_AI_SUGGESTIONS = [
    "How many clients got approved this week?",
    "Which clients need my attention right now?",
    "How many clients are at the interview stage?",
    "Show me clients applying to Canada.",
    "What's happened in the portal in the last 7 days?",
    "How is the team's workload split?",
]


class EnterpriseAIChatTurn(BaseModel):
    role: str = Field(..., max_length=16)
    content: str = Field(..., max_length=6000)


class EnterpriseAIChatRequest(BaseModel):
    message: str = Field(..., min_length=1, max_length=4000)
    # Legacy stateless clients only (SPA loaded before threads shipped). When
    # `conversation_id` is present the transcript is replayed from the DB and this
    # field is IGNORED — client-supplied history was a context-poisoning surface
    # (a fabricated "earlier you confirmed…" turn fed straight to the tool agent).
    history: Optional[list[EnterpriseAIChatTurn]] = None
    conversation_id: Optional[int] = None
    # Sent (true) by clients that understand saved threads. A legacy SPA loaded before
    # threads shipped never sends it — and must NOT get a thread-per-message dribble of
    # single-exchange rows, so persistence only starts once the client opts in.
    client_threads: bool = False


# Saved assistant threads: per-member scratchpad, hard-deleted on retention/offboarding.
ENTERPRISE_AI_CHAT_RETENTION_DAYS = int(os.getenv("ENTERPRISE_AI_CHAT_RETENTION_DAYS", "90"))
ENTERPRISE_AI_CONVERSATION_LIST_LIMIT = 50
ENTERPRISE_AI_CONVERSATION_MESSAGES_LIMIT = 200
ENTERPRISE_AI_TITLE_MAX = 80
# Parity with the old client-supplied history cap (EnterpriseAIChatTurn.content) —
# persistence must not silently grow the prompt the credit meter was tuned against.
ENTERPRISE_AI_HISTORY_TURN_CHARS = 6000


def _serialize_ai_conversation(conv: models.EnterpriseAiConversation) -> dict:
    return {
        "id": conv.id,
        "title": conv.title or "New conversation",
        "message_count": int(conv.message_count or 0),
        "created_at": conv.created_at.isoformat() if conv.created_at else None,
        "last_message_at": conv.last_message_at.isoformat() if conv.last_message_at else None,
    }


def _get_own_ai_conversation_or_404(
    db: Session, *, organization_id: int, user_id: int, conversation_id: int
) -> models.EnterpriseAiConversation:
    """Fetch a thread scoped to BOTH the org and the member — threads are private, so
    another member's id (or another org's) is indistinguishable from a missing one."""
    conv = (
        db.query(models.EnterpriseAiConversation)
        .filter(
            models.EnterpriseAiConversation.id == int(conversation_id),
            models.EnterpriseAiConversation.organization_id == int(organization_id),
            models.EnterpriseAiConversation.user_id == int(user_id),
        )
        .first()
    )
    if not conv:
        raise HTTPException(status_code=404, detail="Conversation not found.")
    return conv


def _refetch_ai_conversation(
    db: Session, conv: models.EnterpriseAiConversation
) -> Optional[models.EnterpriseAiConversation]:
    """Fresh SELECT of a thread row (None if it no longer exists). Used to re-check a
    thread after a long model call before appending to it — appending to a row deleted
    mid-request raises StaleDataError at commit."""
    return (
        db.query(models.EnterpriseAiConversation)
        .filter(
            models.EnterpriseAiConversation.id == conv.id,
            models.EnterpriseAiConversation.organization_id == conv.organization_id,
            models.EnterpriseAiConversation.user_id == conv.user_id,
        )
        .first()
    )


def _delete_ai_conversations(db: Session, conversation_ids: list[int]) -> None:
    """Messages first, then threads — never relies on DB-level FK cascade (SQLite dev
    DBs don't enforce it). Caller commits."""
    if not conversation_ids:
        return
    db.query(models.EnterpriseAiMessage).filter(
        models.EnterpriseAiMessage.conversation_id.in_(conversation_ids)
    ).delete(synchronize_session=False)
    db.query(models.EnterpriseAiConversation).filter(
        models.EnterpriseAiConversation.id.in_(conversation_ids)
    ).delete(synchronize_session=False)


def _purge_stale_ai_conversations(db: Session, *, organization_id: int, user_id: int) -> None:
    """Rolling retention for one member's threads, run opportunistically when they list
    their history — these rows carry client PII into the prod DB, so old transcripts
    age out instead of accumulating forever. Caller commits."""
    if ENTERPRISE_AI_CHAT_RETENTION_DAYS <= 0:
        return
    cutoff = datetime.now(dt_timezone.utc) - timedelta(days=ENTERPRISE_AI_CHAT_RETENTION_DAYS)
    stale_ids = [
        cid for (cid,) in db.query(models.EnterpriseAiConversation.id).filter(
            models.EnterpriseAiConversation.organization_id == int(organization_id),
            models.EnterpriseAiConversation.user_id == int(user_id),
            models.EnterpriseAiConversation.last_message_at < cutoff,
        ).all()
    ]
    _delete_ai_conversations(db, stale_ids)


def _load_ai_conversation_history(db: Session, conv: models.EnterpriseAiConversation) -> list[dict]:
    """Rebuild the model-facing history from the stored transcript: the most recent
    MAX_HISTORY_TURNS messages, oldest first, each capped at the same per-turn size the
    old client-supplied history had."""
    rows = (
        db.query(models.EnterpriseAiMessage.role, models.EnterpriseAiMessage.content)
        .filter(models.EnterpriseAiMessage.conversation_id == conv.id)
        .order_by(models.EnterpriseAiMessage.id.desc())
        .limit(enterprise_ai.MAX_HISTORY_TURNS)
        .all()
    )
    return [
        {"role": role, "content": (content or "")[:ENTERPRISE_AI_HISTORY_TURN_CHARS]}
        for role, content in reversed(rows)
    ]


def _append_ai_turn(
    db: Session,
    conv: models.EnterpriseAiConversation,
    *,
    user_message: str,
    model_answer: str,
) -> None:
    """Persist one exchange onto a thread. Caller commits (alongside the meter, so a
    metered answer and its transcript land atomically)."""
    now = datetime.now(dt_timezone.utc)
    db.add(models.EnterpriseAiMessage(
        conversation_id=conv.id, organization_id=conv.organization_id,
        role="user", content=user_message, created_at=now,
    ))
    db.add(models.EnterpriseAiMessage(
        conversation_id=conv.id, organization_id=conv.organization_id,
        role="model", content=model_answer, created_at=now,
    ))
    # SQL-side increment, not read-modify-write — two tabs answering on the same thread
    # concurrently must not lose one tab's count.
    conv.message_count = models.EnterpriseAiConversation.message_count + 2
    conv.last_message_at = now


@router.get("/ai/meta")
def enterprise_ai_meta(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _require_enterprise_membership(db=db, user=current_user, request=request)
    return {
        "enabled": enterprise_ai.is_ai_configured(),
        "suggestions": ENTERPRISE_AI_SUGGESTIONS,
    }


@router.post("/ai/chat")
def enterprise_ai_chat(
    payload: EnterpriseAIChatRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability=("ai.assistant", "credits.spend"))
    _enforce_rate_limit_or_429(
        request=request,
        scope="enterprise.ai",
        limit=ENTERPRISE_AI_RATE_LIMIT,
        window_seconds=ENTERPRISE_AI_RATE_WINDOW_SECONDS,
        extra_key=str(current_user.id),
    )

    if not enterprise_ai.is_ai_configured():
        raise HTTPException(
            status_code=503,
            detail="The AI assistant isn't available right now. Please try again later.",
        )

    # Resolve the thread FIRST: an id the member doesn't own must 404 before any
    # metering or model work happens.
    conv = None
    if payload.conversation_id is not None:
        conv = _get_own_ai_conversation_or_404(
            db, organization_id=organization.id, user_id=current_user.id,
            conversation_id=payload.conversation_id,
        )

    # Cost guardrail: reject obviously off-topic prompts before spending model tokens.
    # (A refused off-topic message is free — it doesn't touch the copilot meter.)
    if ai_guardrails.is_off_topic(payload.message):
        ai_guardrails.record_block(source="enterprise_copilot", detail="enterprise")
        # An EXISTING thread keeps the exchange so its transcript reads coherently on
        # reload; a refusal never creates a new thread. Re-fetch before appending: the
        # thread may have been deleted since the request started (another tab, the
        # retention sweep) and persisting onto the dead row raises StaleDataError.
        if conv is not None:
            conv = _refetch_ai_conversation(db, conv)
        if conv is not None:
            _append_ai_turn(db, conv, user_message=payload.message.strip()[:4000],
                            model_answer=ai_guardrails.OFF_TOPIC_REFUSAL)
            db.commit()
        response = {"answer": ai_guardrails.OFF_TOPIC_REFUSAL, "permissions": _enterprise_permissions_for_role(role)}
        if conv is not None:
            response["conversation_id"] = conv.id
        return response

    # Meter the copilot: free daily allowance, then 1 credit per bundle of messages.
    # Block a paid message the wallet can't cover BEFORE spending any Gemini tokens.
    credits.copilot_precheck_or_402(db, organization.id)

    # A known thread replays its own server-side transcript; the request's history is
    # only honoured for legacy stateless clients that predate saved threads.
    if conv is not None:
        history = _load_ai_conversation_history(db, conv)
    else:
        history = [turn.model_dump() for turn in (payload.history or [])]
    try:
        turn = enterprise_ai.run_enterprise_ai_chat(
            db=db,
            organization=organization,
            user=current_user,
            role=role,
            message=payload.message,
            history=history,
            ctx=role.ctx,
        )
    except Exception:
        logger.exception("Enterprise AI chat failed (org_id=%s)", organization.id)
        raise HTTPException(
            status_code=502,
            detail="The AI assistant ran into a problem answering that. Please try again.",
        )
    answer = turn.answer

    # Answered successfully → persist the exchange, then record the turn against the
    # meter (may debit a credit); record_copilot_message commits, so the transcript and
    # its metering land together. A failed answer persists nothing.
    prompt_text = payload.message.strip()[:4000]
    persist = True
    if conv is not None:
        # The model call took many seconds — the thread may have been deleted meanwhile
        # (another tab, the retention sweep, member offboarding). Persisting onto the
        # dead row would raise StaleDataError at commit, 500ing a successful answer AND
        # rolling back its metering. Re-fetch; if it's gone, fall through to a fresh
        # thread carrying this exchange (the response's conversation_id tells the
        # client where the conversation now lives).
        conv = _refetch_ai_conversation(db, conv)
    else:
        # No thread requested: only clients that understand saved threads get one.
        # A legacy stateless SPA (no client_threads flag) would otherwise dribble one
        # single-exchange thread per message into History.
        persist = bool(payload.client_threads)
    if persist:
        if conv is None:
            conv = models.EnterpriseAiConversation(
                organization_id=organization.id,
                user_id=current_user.id,
                # Title = first message truncated. Never a model call — a flash call per
                # new thread is real money against the meter for what a substring does.
                title=prompt_text[:ENTERPRISE_AI_TITLE_MAX],
            )
            db.add(conv)
            db.flush()  # assign conv.id before the messages reference it
        _append_ai_turn(db, conv, user_message=prompt_text, model_answer=answer)

    # The turn's REAL cost — summed across every tool round-trip, not just the last one —
    # decides its message weight, so a six-tool answer can't be sold at a one-call price.
    meter = credits.record_copilot_message(
        db, organization.id, user=current_user, turn_cost_usd=turn.usage.cost_usd,
    )
    response = {
        "answer": answer,
        "permissions": _enterprise_permissions_for_role(role),
        "credits_meter": meter,
    }
    if conv is not None:
        response["conversation_id"] = conv.id
        response["conversation"] = _serialize_ai_conversation(conv)
    if meter.get("credits_charged"):
        response["wallet"] = credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization))
    return response


@router.get("/ai/conversations")
def enterprise_ai_conversation_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """The member's OWN saved assistant threads, newest first. Listing is also when the
    rolling retention sweep runs — threads past the retention window are hard-deleted."""
    _, organization, _role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability=("ai.assistant", "credits.spend"),
    )
    _purge_stale_ai_conversations(db, organization_id=organization.id, user_id=current_user.id)
    db.commit()
    rows = (
        db.query(models.EnterpriseAiConversation)
        .filter(
            models.EnterpriseAiConversation.organization_id == organization.id,
            models.EnterpriseAiConversation.user_id == current_user.id,
        )
        .order_by(models.EnterpriseAiConversation.last_message_at.desc())
        .limit(ENTERPRISE_AI_CONVERSATION_LIST_LIMIT)
        .all()
    )
    return {
        "conversations": [_serialize_ai_conversation(c) for c in rows],
        "retention_days": ENTERPRISE_AI_CHAT_RETENTION_DAYS,
    }


@router.get("/ai/conversations/{conversation_id}")
def enterprise_ai_conversation_detail(
    conversation_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, _role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability=("ai.assistant", "credits.spend"),
    )
    conv = _get_own_ai_conversation_or_404(
        db, organization_id=organization.id, user_id=current_user.id, conversation_id=conversation_id,
    )
    rows = (
        db.query(models.EnterpriseAiMessage)
        .filter(models.EnterpriseAiMessage.conversation_id == conv.id)
        .order_by(models.EnterpriseAiMessage.id.desc())
        .limit(ENTERPRISE_AI_CONVERSATION_MESSAGES_LIMIT)
        .all()
    )
    return {
        "conversation": _serialize_ai_conversation(conv),
        "messages": [
            {"role": m.role, "content": m.content,
             "created_at": m.created_at.isoformat() if m.created_at else None}
            for m in reversed(rows)
        ],
    }


@router.delete("/ai/conversations/{conversation_id}")
def enterprise_ai_conversation_delete(
    conversation_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, _role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability=("ai.assistant", "credits.spend"),
    )
    conv = _get_own_ai_conversation_or_404(
        db, organization_id=organization.id, user_id=current_user.id, conversation_id=conversation_id,
    )
    _delete_ai_conversations(db, [conv.id])
    db.commit()
    return {"message": "Conversation deleted."}


# ===========================================================================
# Rilono Copilot (Chrome extension) — enterprise staff mode
#
# Staff authenticate as themselves and pick a CLIENT (EnterpriseClient CRM row)
# to work on behalf of; the client only shapes the AI context, so every message
# stays attributable to the staff user. Metered on the same org copilot meter
# as the dashboard assistant (free daily allowance, then credits).
# ===========================================================================

ENTERPRISE_COPILOT_CLIENT_LIMIT = 300
ENTERPRISE_COPILOT_MAX_ATTACHMENTS = 8


class EnterpriseCopilotChatRequest(BaseModel):
    client_id: int = Field(..., ge=1)
    message: str = Field(..., min_length=1, max_length=8000)
    conversation_history: Optional[list[dict]] = None
    session_attachments: Optional[list[ChatSessionAttachment]] = None


@router.get("/copilot/context")
def enterprise_copilot_extension_context(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Session probe for the Chrome extension: is this signed-in user enterprise
    staff? Returns 200 with enterprise=False for regular students (no error), so
    the extension can fall back to the B2C copilot silently."""
    membership, organization = _get_active_enterprise_membership(db, current_user.id)
    if not membership or not organization or not (organization.subdomain_slug or "").strip():
        return {"enterprise": False}
    _enforce_request_subdomain_matches_org(request, organization)
    # Resolve the real access context rather than reading the legacy `role` string: answering
    # `permissions` from the mirror made this the one endpoint that could disagree with the rest
    # of the app (a branch manager reported as "viewer"), and the count below has to be scoped
    # like every other client count.
    ctx = access.resolve_access_context(db, membership, organization)
    role = EnterpriseRoleContext(access.legacy_role_for(ctx.role_key, ctx.capabilities), ctx)
    client_count = (
        scope_client_query(
            db.query(models.EnterpriseClient).filter(
                models.EnterpriseClient.organization_id == organization.id
            ),
            ctx,
        ).count()
    )
    return {
        "enterprise": True,
        "organization": {
            "id": organization.id,
            "company_name": organization.company_name,
            "subdomain_slug": (organization.subdomain_slug or "").strip().lower() or None,
            # Absolute: this one renders inside the extension, not on our origin.
            "logo_url": _absolute_enterprise_logo_url(organization),
        },
        "role": role,
        "permissions": _enterprise_permissions_for_role(role),
        "access": access.access_payload(ctx),
        "copilot_enabled": enterprise_copilot.is_provider_available(),
        "client_count": int(client_count or 0),
    }


@router.get("/copilot/clients")
def enterprise_copilot_extension_clients(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Compact client list for the extension's on-behalf-of picker (view-level).
    Deliberately excludes sensitive fields like passport numbers — the chat
    context is assembled server-side, so the extension never needs them."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="clients.view"
    )
    base = scope_client_query(
        db.query(models.EnterpriseClient).filter(
            models.EnterpriseClient.organization_id == organization.id
        ),
        role.ctx,
    )
    total = base.count()
    rows = (
        base.order_by(models.EnterpriseClient.updated_at.desc())
        .limit(ENTERPRISE_COPILOT_CLIENT_LIMIT)
        .all()
    )
    member_names = _org_member_name_map(db, organization.id)
    clients = []
    for client in rows:
        assigned_name = None
        if client.assigned_to_user_id:
            assigned_name = member_names.get(int(client.assigned_to_user_id))
        clients.append({
            "id": client.id,
            "full_name": client.full_name,
            "email": client.email,
            "visa_category": client.visa_category,
            "visa_category_label": _category_label(client.visa_category),
            "destination_country_code": client.destination_country_code,
            "destination_country_name": client.destination_country_name,
            "visa_type": client.visa_type,
            "intake": client.intake,
            "status": client.status,
            "stage": _stage_brief(client.status, client.destination_country_code),
            "priority": client.priority,
            "assigned_to_name": assigned_name,
            "updated_at": _iso(client.updated_at),
        })
    return {
        "clients": clients,
        "total": int(total or 0),
        "capped": bool(total and total > ENTERPRISE_COPILOT_CLIENT_LIMIT),
    }


@router.post("/copilot/chat")
def enterprise_copilot_extension_chat(
    payload: EnterpriseCopilotChatRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """One staff-mode Copilot turn about a specific client. Mirrors /ai/chat's
    guardrail → precheck → generate → meter flow; the response field is named
    `response` to match the extension's B2C chat contract."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability=("ai.assistant", "credits.spend"))
    _enforce_rate_limit_or_429(
        request=request,
        scope="enterprise.copilot",
        limit=ENTERPRISE_AI_RATE_LIMIT,
        window_seconds=ENTERPRISE_AI_RATE_WINDOW_SECONDS,
        extra_key=str(current_user.id),
    )

    if not enterprise_copilot.is_provider_available():
        raise HTTPException(
            status_code=503,
            detail="Rilono Copilot isn't available right now. Please try again later.",
        )

    client = _get_org_client_or_404(db, organization.id, payload.client_id, ctx=role.ctx)

    if len(payload.session_attachments or []) > ENTERPRISE_COPILOT_MAX_ATTACHMENTS:
        raise HTTPException(
            status_code=400,
            detail=f"At most {ENTERPRISE_COPILOT_MAX_ATTACHMENTS} attachments are allowed per message.",
        )

    # Cost guardrail: refuse obviously off-topic prompts before spending tokens.
    # Also screen the newest user turn in the supplied history — otherwise a
    # short "continue" message with the real (off-topic) request smuggled into
    # conversation_history bypasses this free pre-model check entirely.
    latest_history_turn = ""
    for turn in reversed(payload.conversation_history or []):
        if isinstance(turn, dict) and (turn.get("role") or "user") != "assistant":
            latest_history_turn = str(turn.get("content") or "")
            break
    if ai_guardrails.is_off_topic(payload.message) or (
        len(payload.message) < 200
        and latest_history_turn
        and ai_guardrails.is_off_topic(latest_history_turn)
    ):
        ai_guardrails.record_block(source=enterprise_copilot.USAGE_SOURCE, detail="enterprise_extension")
        return {"response": ai_guardrails.OFF_TOPIC_REFUSAL}

    # Meter: free daily allowance, then credits — block unaffordable paid
    # messages BEFORE any Gemini tokens are spent.
    credits.copilot_precheck_or_402(db, organization.id)

    try:
        copilot_turn = enterprise_copilot.run_enterprise_copilot_chat(
            db,
            organization=organization,
            staff_user=current_user,
            role=role,
            client=client,
            message=payload.message,
            conversation_history=payload.conversation_history,
            session_attachments=payload.session_attachments,
            # Record scope picked the client; capabilities decide what the prompt may
            # disclose about them — the extension must not be a way around the
            # clients.view_sensitive / documents.* gates the dashboard enforces.
            ctx=role.ctx,
        )
        answer = copilot_turn.answer
    except Exception:
        logger.exception(
            "Enterprise extension copilot failed (org_id=%s client_id=%s)",
            organization.id, client.id,
        )
        raise HTTPException(
            status_code=502,
            detail="Rilono Copilot ran into a problem answering that. Please try again.",
        )

    # Answered successfully → record the message against the meter (may debit a
    # credit). Metering failures must never destroy the already-generated (and
    # already-paid-for) answer: deliver it unmetered and log loudly instead —
    # a 500 here would just make the org re-spend the full Gemini cost on retry.
    response = {
        "response": answer,
        "client": {"id": client.id, "full_name": client.full_name},
    }
    try:
        meter = credits.record_copilot_message(
            db, organization.id, user=current_user,
            turn_cost_usd=copilot_turn.usage.cost_usd,
        )
        response["credits_meter"] = meter
        if meter.get("credits_charged"):
            response["wallet"] = credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization))
    except Exception:
        logger.exception(
            "Copilot metering failed after successful generation (org_id=%s) — answer delivered unmetered",
            organization.id,
        )
    return response


# ===========================================================================
# Per-client documents (private R2 storage, authenticated streaming)
# ===========================================================================

ENTERPRISE_DOC_MAX_BYTES = int(os.getenv("ENTERPRISE_DOC_MAX_BYTES", str(25 * 1024 * 1024)))
ENTERPRISE_DOC_ALLOWED_EXT = {
    ".pdf", ".jpg", ".jpeg", ".png", ".webp", ".gif", ".heic",
    ".doc", ".docx", ".xls", ".xlsx", ".csv", ".txt",
}
ENTERPRISE_DOC_INLINE_EXT = {".pdf", ".jpg", ".jpeg", ".png", ".webp", ".gif"}
# Safe served Content-Type per (validated) extension. The download endpoint uses THIS, never
# the uploader-supplied mime_type — a file named *.pdf can arrive with Content-Type text/html
# and a <script> body, which served inline under our 'unsafe-inline' CSP would execute in the
# portal origin (account takeover). Deriving from the extension makes that impossible.
ENTERPRISE_DOC_EXT_MIME = {
    ".pdf": "application/pdf",
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
    ".webp": "image/webp", ".gif": "image/gif", ".heic": "image/heic",
    ".doc": "application/msword",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xls": "application/vnd.ms-excel",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".csv": "text/csv", ".txt": "text/plain",
}


def _document_scan_pricing(db: Session, organization_id: int) -> dict:
    """What one document scan costs this org right now, for the upload card and the
    per-document Scan button (so the price is shown before the credit is spent)."""
    return {
        "action_key": ENTERPRISE_DOC_SCAN_ACTION_KEY,
        "cost_credits": credits.action_cost(ENTERPRISE_DOC_SCAN_ACTION_KEY),
        "can_afford": credits.can_afford(db, organization_id, ENTERPRISE_DOC_SCAN_ACTION_KEY),
        "ai_available": gemini_service.is_ai_configured(),
    }


def _scrub_sensitive_autofill(extracted, include_sensitive: bool):
    """Mask sensitive identity values inside a document's stored autofill record.

    `_ent_autofill_profile` records what it wrote and what disagreed, and that block is
    persisted into `extracted_fields` and served back on the client-detail and documents
    responses. Left alone it hands the passport number to precisely the roles
    `_serialize_client` withholds it from — Viewer holds `clients.view` and `documents.view`
    but not `clients.view_sensitive` — so the masking has to happen HERE, on the read path,
    not only where the note is composed. Read-path masking is also what covers the rows
    written before any of this existed.

    Matched on `attr` where present and on the human label otherwise, because rows stored
    before `attr` was recorded carry only the label.
    """
    if include_sensitive or not isinstance(extracted, dict):
        return extracted
    autofill = extracted.get("autofill")
    if not isinstance(autofill, dict):
        return extracted

    def _is_sensitive(entry):
        if not isinstance(entry, dict):
            return False
        attr = entry.get("attr")
        if attr:
            return attr in _CLIENT_SENSITIVE_FIELDS
        return str(entry.get("field") or "") in _CLIENT_SENSITIVE_LABELS

    def _mask(entries, keys):
        out = []
        for entry in (entries or []):
            if _is_sensitive(entry):
                entry = {**entry, **{k: "••••••••" for k in keys if k in entry}}
            out.append(entry)
        return out

    scrubbed = {**autofill}
    if isinstance(autofill.get("filled"), list):
        scrubbed["filled"] = _mask(autofill["filled"], ("value",))
    if isinstance(autofill.get("conflicts"), list):
        scrubbed["conflicts"] = _mask(autofill["conflicts"], ("existing", "document"))
    return {**extracted, "autofill": scrubbed}


def _serialize_client_document(
    doc: models.EnterpriseClientDocument, *, include_sensitive: bool = False
) -> dict:
    extracted = None
    if doc.extracted_fields:
        try:
            import json
            extracted = json.loads(doc.extracted_fields)
        except Exception:
            extracted = None
    extracted = _scrub_sensitive_autofill(extracted, include_sensitive)
    # A staff override reads as "valid" everywhere downstream, but it was NOT the AI that
    # cleared it — surface the provenance so the UI can label it "Manually approved".
    # Rows accepted before these columns existed fall back to the audit keys the accept
    # endpoint has always written into extracted_fields.
    payload = extracted if isinstance(extracted, dict) else {}
    accepted_by = (doc.manually_accepted_by or payload.get("accepted_by") or "").strip() or None
    accepted_at = _iso(doc.manually_accepted_at) or (payload.get("accepted_at") or None)
    manually_accepted = bool(accepted_by or doc.manually_accepted_at or payload.get("accepted_at"))
    return {
        "id": doc.id,
        "client_id": doc.client_id,
        "document_type": doc.document_type,
        "original_filename": doc.original_filename,
        "file_size": doc.file_size,
        "mime_type": doc.mime_type,
        "uploaded_by_name": doc.uploaded_by_name,
        "created_at": _iso(doc.created_at),
        "download_url": f"/api/enterprise/clients/{doc.client_id}/documents/{doc.id}/download",
        # AI validation. null = a scan is running right now; "not_scanned" = never asked
        # for (the document is stored, which is free, and can be scanned on demand later).
        "validation_status": doc.validation_status,
        "validation_message": doc.validation_message,
        "validated_at": _iso(doc.validated_at),
        "validation_credits_charged": int(doc.validation_credits_charged or 0),
        # Human-in-the-loop override (staff accepted a document Rilono AI flagged).
        "manually_accepted": manually_accepted,
        "manually_accepted_by": accepted_by,
        "manually_accepted_at": accepted_at,
        "ai_flag_before_accept": (payload.get("ai_flag_before_accept") or None),
        "extracted": extracted,
    }


def _safe_filename(name: str) -> str:
    base = os.path.basename(str(name or "").strip()) or "document"
    base = re.sub(r"[\r\n\"]+", "", base)
    return base[:160]


@router.get("/clients/{client_id}/documents")
def enterprise_list_client_documents(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="documents.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    rows = (
        db.query(models.EnterpriseClientDocument)
        .filter(models.EnterpriseClientDocument.client_id == client.id)
        .order_by(models.EnterpriseClientDocument.created_at.desc(), models.EnterpriseClientDocument.id.desc())
        .all()
    )
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "document_types": list(catalog.STUDENT_DOCUMENT_TYPES),
        "documents": [
            _serialize_client_document(d, include_sensitive=role.ctx.has("clients.view_sensitive"))
            for d in rows
        ],
        # So the upload card and every Scan button can price themselves without a second call.
        "scan_pricing": _document_scan_pricing(db, organization.id),
        # A scan is debited by the background worker, AFTER the upload/scan response has
        # already been sent — so this list (which the UI polls for the verdict) is where the
        # post-charge balance actually becomes visible. Without it the sidebar badge would
        # keep showing the pre-scan balance until the next full page load.
        "wallet": (
            credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization))
            if role.ctx.has("credits.view") else None
        ),
    }


def _ent_parse_bool_form(value, default: bool = False) -> bool:
    """Parse a checkbox value out of multipart form data. FormData carries strings, so an
    unparsed truthy check would read "false" as True and silently bill an opted-out scan."""
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


@router.post("/clients/{client_id}/documents")
async def enterprise_upload_client_document(
    client_id: int,
    request: Request,
    document_type: str = Form("Other"),
    scan: str = Form("true"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Store a document (free) and, when `scan` is set, run the billed Rilono AI
    scan & validate on it (credits action "document_scan")."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="documents.upload"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)

    if not enterprise_storage.is_configured():
        raise HTTPException(status_code=503, detail="Document storage is not configured.")

    # Scanning debits the wallet, so it needs credits.spend on top of documents.upload —
    # a member who may file documents but not spend the org's money simply stores the file.
    # Silently degrading (rather than 403ing) is deliberate: the upload itself is allowed,
    # and the response reports scan_requested so the UI can say why nothing was scanned.
    want_scan = _ent_parse_bool_form(scan, default=True)
    if want_scan and not role.ctx.has("credits.spend"):
        want_scan = False
    if want_scan and not gemini_service.is_ai_configured():
        want_scan = False
    # Block an unaffordable scan BEFORE storing anything, so the member gets a clean 402
    # with the file still on their machine rather than a stored-but-unscanned surprise.
    if want_scan:
        credits.enforce_action_or_402(db, organization.id, ENTERPRISE_DOC_SCAN_ACTION_KEY)

    original = _safe_filename(file.filename)
    ext = os.path.splitext(original)[1].lower()
    if ext not in ENTERPRISE_DOC_ALLOWED_EXT:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Allowed: PDF, images, Word/Excel, CSV, or text.",
        )

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="The file is empty.")
    if len(data) > ENTERPRISE_DOC_MAX_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"File is too large. Maximum size is {ENTERPRISE_DOC_MAX_BYTES // (1024 * 1024)} MB.",
        )

    storage_key = f"enterprise/{organization.id}/clients/{client.id}/{uuid.uuid4().hex}{ext}"
    try:
        await run_in_threadpool(enterprise_storage.store_document, storage_key, data, content_type=file.content_type)
    except Exception:
        logger.exception("Failed to store client document (org_id=%s, client_id=%s)", organization.id, client.id)
        raise HTTPException(status_code=502, detail="Could not store the document right now. Please try again.")

    doc = models.EnterpriseClientDocument(
        organization_id=organization.id,
        client_id=client.id,
        document_type=catalog.normalize_document_type(document_type),
        original_filename=original,
        storage_key=storage_key,
        file_size=len(data),
        mime_type=(file.content_type or None),
        uploaded_by_user_id=current_user.id,
        uploaded_by_name=current_user.full_name or current_user.email,
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)

    # Text extraction (free) always runs so the copilot can read the document; the billed
    # scan only when asked for. Both happen in the background so neither slows the upload.
    _start_document_processing(
        doc.id, data, original, file.content_type,
        validate=want_scan, charge_user_id=current_user.id,
    )

    return {
        "message": "Document uploaded." + (" Rilono AI is scanning it." if want_scan else ""),
        "permissions": _enterprise_permissions_for_role(role),
        "document": _serialize_client_document(
            doc, include_sensitive=role.ctx.has("clients.view_sensitive")
        ),
        "scan_requested": want_scan,
        "scan_pricing": _document_scan_pricing(db, organization.id),
    }


def _ent_flexible_parse_date(value):
    """Best-effort parse of a human-readable date (passport/ID dates come in many formats)
    into a date. Prefers day-first (most passports). Returns None if unparseable.

    An implausible result is treated as UNPARSEABLE rather than raising: this reads machine
    output, and the caller's contract for None is "skip this field". The `fuzzy=True`
    fallback below is what makes the check necessary — it will pull a date out of prose and
    fill whatever components the text didn't supply, so a bad OCR read yields a confident,
    wrong date rather than nothing.
    """
    s = str(value or "").strip()
    if not s or s.lower() in {"null", "none", "n/a", "na", "not available", "unknown"}:
        return None
    parsed = None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%m/%d/%Y",
                "%d %b %Y", "%d %B %Y", "%d-%b-%Y", "%d %b, %Y", "%b %d, %Y", "%B %d, %Y", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(s, fmt).date()
            break
        except ValueError:
            continue
    if parsed is None:
        try:
            from dateutil import parser as _dateparser
            parsed = _dateparser.parse(s, dayfirst=True, fuzzy=True).date()
        except Exception:
            return None
    return parsed if ent_dates.is_sane(parsed) else None


def _ent_clean_extracted(validation: dict) -> dict:
    """Pull identity fields out of validate_and_extract_document()'s response, dropping
    empty/placeholder values."""
    def val(key):
        raw = validation.get(key)
        if raw is None:
            return None
        s = str(raw).strip()
        if not s or s.lower() in {"null", "none", "n/a", "na", "not available", "unknown", "-"}:
            return None
        return s
    return {
        "name": val("Name"),
        "date_of_birth": val("Date of Birth"),
        "document_number": val("Document Number"),
        "expiration_date": val("Expiration Date"),
        "issue_date": val("Issue Date"),
        "country": val("Country"),
        "other": val("Other Information"),
    }


def _ent_profile_field_plan(document_type: str, fields: dict):
    """Which extracted fields map onto the client profile, per document type. Only identity
    documents (passport) populate the profile today; extend this for more types later."""
    dt = (document_type or "").lower()
    is_passport = ("passport" in dt) and ("photo" not in dt)
    if not is_passport:
        return []
    return [  # (client attribute, human label, extracted value, kind)
        ("full_name", "Name", fields.get("name"), "text"),
        # "past_date", not "date": a passport EXPIRY is legitimately in the future, a date of
        # birth never is — and a fuzzy parse that loses the year defaults it to the current
        # one, which is exactly how a plausible-looking future DOB gets written.
        ("date_of_birth", "Date of birth", fields.get("date_of_birth"), "past_date"),
        ("nationality", "Nationality", fields.get("country"), "text"),
        ("passport_number", "Passport number", fields.get("document_number"), "text"),
        ("passport_expiry", "Passport expiry", fields.get("expiration_date"), "date"),
    ]


# Upload-time AI cross-validation context. No hardcoded identity/consistency rules live
# in this codebase: we hand the AI the client's profile and their other documents, and the
# AI decides — per document type and destination — what to check (identity, dates, funds,
# study plan, …). A material conflict makes the AI fail the validation itself.
_ENT_VALIDATION_PROFILE_CHARS = 4000
_ENT_VALIDATION_DOCS_CHARS = 12000
_ENT_VALIDATION_MAX_RELATED_DOCS = 6
_ENT_VALIDATION_TEXT_PER_DOC_CHARS = 1800

# The per-document scan is a billed premium action. The usage source is deliberately NOT
# the shared "document_ai" bucket the free text extraction uses, so the admin margin report
# can weigh this action's real Gemini cost against the credits it actually earned.
ENTERPRISE_DOC_SCAN_ACTION_KEY = "document_scan"
ENTERPRISE_DOC_SCAN_USAGE_SOURCE = "enterprise_document_scan"
# A re-scan re-reads the file from storage and re-bills, so it gets its own limit — without
# one, one member holding down "Re-scan" is an unbounded Gemini bill.
ENTERPRISE_DOC_SCAN_RATE_LIMIT = int(os.getenv("ENTERPRISE_DOC_SCAN_RATE_LIMIT", "30") or "30")
ENTERPRISE_DOC_SCAN_RATE_WINDOW_SECONDS = int(
    os.getenv("ENTERPRISE_DOC_SCAN_RATE_WINDOW_SECONDS", "3600") or "3600"
)


def _ent_client_profile_context(client) -> str:
    """Compact snapshot of the client profile for AI cross-validation at upload time."""
    if client is None:
        return ""

    def _d(v):
        return v.isoformat() if hasattr(v, "isoformat") else ("" if v is None else str(v).strip())

    dest_name = (getattr(client, "destination_country_name", "") or "").strip()
    dest_code = (getattr(client, "destination_country_code", "") or "").strip()
    destination = f"{dest_name} ({dest_code})" if dest_name and dest_code else (dest_name or dest_code)
    rows = [
        ("Full name", getattr(client, "full_name", None)),
        ("Date of birth", _d(getattr(client, "date_of_birth", None))),
        ("Nationality", getattr(client, "nationality", None)),
        ("Passport number", getattr(client, "passport_number", None)),
        ("Passport expiry", _d(getattr(client, "passport_expiry", None))),
        ("Destination", destination),
        ("Visa type", getattr(client, "visa_type", None)),
        ("Intake", getattr(client, "intake", None)),
        ("Key date / deadline", _d(getattr(client, "target_date", None))),
        ("Email", getattr(client, "email", None)),
    ]
    lines = [f"{label}: {str(value).strip()}" for label, value in rows if value and str(value).strip()]
    if not lines:
        return ""
    text = (
        "This is the client this document was uploaded for. Cross-check the "
        "document against this profile — including that the document actually belongs to this "
        "person:\n" + "\n".join(lines)
    )
    return text[:_ENT_VALIDATION_PROFILE_CHARS]


def _ent_related_documents_context(db: Session, client, exclude_document_id) -> str:
    """Bounded snapshots of the client's already-VALIDATED documents, so the AI can
    cross-validate names, dates, numbers, universities and timelines across everything
    that has itself been checked.

    Only "valid" rows are used as the reference set. A document that Rilono AI red-flagged
    (wrong person, expired, forged) or that was never scanned is not evidence — feeding it
    in lets one bad document silently shape the verdict on every document uploaded after
    it. Staff-accepted rows stay in (a human vouched for them) but keep the explicit
    "MANUALLY APPROVED" label below so the model never reads them as AI-verified.
    """
    if client is None:
        return ""
    try:
        rows = (
            db.query(models.EnterpriseClientDocument)
            .filter(
                models.EnterpriseClientDocument.client_id == client.id,
                models.EnterpriseClientDocument.id != int(exclude_document_id or 0),
                models.EnterpriseClientDocument.validation_status == "valid",
            )
            .order_by(models.EnterpriseClientDocument.created_at.desc())
            .limit(_ENT_VALIDATION_MAX_RELATED_DOCS)
            .all()
        )
    except Exception:
        logger.exception("Failed to load related documents for validation context (client_id=%s)", client.id)
        return ""
    blocks, used = [], 0
    for index, doc in enumerate(rows, start=1):
        # Every row here is "valid" (the query filters on it). A staff override is ALSO
        # stored as "valid" — say so explicitly, so the model does not read a human-accepted
        # document as one Rilono AI itself cleared.
        status_label = "VALIDATED BY RILONO AI"
        if doc.manually_accepted_by or doc.manually_accepted_at:
            status_label = "MANUALLY APPROVED BY STAFF, NOT AI-VALIDATED"
        header = (
            f"\n--- PRIOR DOCUMENT {index}: {(doc.document_type or 'document').upper()} "
            f"({doc.original_filename}) [{status_label}] ---\n"
        )
        body = ""
        if doc.extracted_fields:
            try:
                fields = (json.loads(doc.extracted_fields) or {}).get("fields") or {}
                if fields:
                    body += "Extracted fields: " + json.dumps(fields, default=str) + "\n"
            except Exception:
                pass
        if doc.extracted_text:
            body += str(doc.extracted_text)[:_ENT_VALIDATION_TEXT_PER_DOC_CHARS] + "\n"
        remaining = _ENT_VALIDATION_DOCS_CHARS - used - len(header)
        if remaining <= 0:
            break
        block = header + body[: max(0, remaining)]
        blocks.append(block)
        used += len(block)
    if not blocks:
        return ""
    return "Previously uploaded documents for this client (for cross-validation):" + "".join(blocks)


def _ent_autofill_profile(db: Session, client: models.EnterpriseClient, document_type: str, fields: dict) -> dict:
    """Fill EMPTY client profile fields from a validated document's details. Never overwrites an
    existing value — a differing value is returned as a 'conflict' for staff to review. Adds an
    audit note. Returns {'filled': [...], 'conflicts': [...]}."""
    filled, conflicts = [], []

    def _disp(v):
        return v.isoformat() if hasattr(v, "isoformat") else str(v)

    is_date_kind = {"date", "past_date"}.__contains__
    for attr, label, raw_value, kind in _ent_profile_field_plan(document_type, fields):
        if not raw_value:
            continue
        if is_date_kind(kind):
            new_value = _ent_flexible_parse_date(raw_value)
            if new_value is None:
                continue
            # Skip rather than 400: a counsellor can't fix an OCR misread by re-uploading,
            # and blocking the upload would cost them the document scan they paid for. The
            # field is simply left for them to type.
            if kind == "past_date" and new_value > date.today() + timedelta(days=ent_dates.GRACE_DAYS):
                continue
        else:
            new_value = str(raw_value).strip()
        current = getattr(client, attr, None)
        current_empty = current is None or (isinstance(current, str) and not current.strip())
        if current_empty:
            setattr(client, attr, new_value)
            # `attr` travels with the record so the read path can tell which entries hold a
            # value governed by `clients.view_sensitive` (see _scrub_sensitive_autofill).
            filled.append({"field": label, "attr": attr, "value": _disp(new_value)})
        else:
            if is_date_kind(kind):
                same = (current == new_value)
            else:
                same = (str(current).strip().lower() == new_value.strip().lower())
            if not same:
                conflicts.append({
                    "field": label, "attr": attr,
                    "existing": _disp(current), "document": _disp(new_value),
                })

    if filled or conflicts:
        lines = []
        if filled:
            lines.append("Auto-filled from validated " + str(document_type) + ": "
                         + ", ".join(f["field"] for f in filled) + ".")
        if conflicts:
            # The note is readable by anyone with `notes.view` — which Viewer holds and
            # `clients.view_sensitive` is separate from. Spelling a mismatched passport
            # number out here would hand it to exactly the roles the serializer withholds
            # it from, so a sensitive field's note says only THAT it differs.
            def _conflict_phrase(c):
                if c.get("attr") in _CLIENT_SENSITIVE_FIELDS:
                    return f'{c["field"]} (differs from the profile — open the document to compare)'
                return f'{c["field"]} (profile "{c["existing"]}" vs document "{c["document"]}")'

            lines.append("Needs review — document differs from existing profile: "
                         + "; ".join(_conflict_phrase(c) for c in conflicts) + ".")
        try:
            db.add(models.EnterpriseClientNote(
                organization_id=client.organization_id,
                client_id=client.id,
                author_user_id=None,
                author_name="Rilono AI",
                body=" ".join(lines),
            ))
        except Exception:
            logger.exception("Failed to add autofill audit note (client_id=%s)", client.id)
    return {"filled": filled, "conflicts": conflicts}


def _start_document_processing(
    document_id: int,
    data: bytes,
    filename: str,
    mime_type: str | None,
    *,
    validate: bool,
    charge_user_id: int | None = None,
) -> None:
    """Background work for a newly stored document, in two halves.

    ALWAYS (free): extract the document's text, so the AI copilot can read it and Deep
    Scan has something to map over. Storing a document must never cost credits — that is
    the CRM, not a premium AI action.

    ONLY when `validate` (billed, credits action "document_scan"): run Rilono AI's
    validation + structured extraction, cross-checked against the client profile and their
    already-validated documents, and auto-fill empty profile fields from a passed identity
    document. The caller must already have run credits.enforce_action_or_402 — this worker
    debits the wallet *after* a successful scan, so a failed scan is never billed.

    Used by the staff upload, the client secure-link upload (never validates — no
    authenticated member to bill) and the on-demand re-scan endpoint.
    """
    def _worker():
        import json
        from datetime import timezone
        db2 = SessionLocal()
        try:
            row = (
                db2.query(models.EnterpriseClientDocument)
                .filter(models.EnterpriseClientDocument.id == int(document_id))
                .first()
            )
            if row is None:
                return
            # Fresh thread = fresh contextvars: attribute this worker's Gemini calls
            # (extraction + validation) to the org so their cost isn't unattributed.
            ai_usage.set_usage_account(organization_id=row.organization_id)
            client = (
                db2.query(models.EnterpriseClient)
                .filter(models.EnterpriseClient.id == row.client_id)
                .first()
            )

            # 1) Full-text extraction for the copilot (best-effort, never billed).
            try:
                extracted = gemini_service.extract_text_from_document(
                    data, filename, mime_type or "application/octet-stream"
                )
                if extracted:
                    row.extracted_text = extracted[:200000]
            except Exception:
                logger.exception("Enterprise doc text extraction failed (document_id=%s)", document_id)

            # The scan was not asked for: keep the document, skip the paid half. An explicit
            # "not_scanned" (rather than NULL) lets the UI tell "never scanned" apart from
            # "scan still running" and offer the Scan button on exactly the former.
            if not validate:
                row.validation_status = "not_scanned"
                row.validation_message = None
                db2.commit()
                return

            # 2) Validate the document + extract structured identity fields.
            destination_code = client.destination_country_code if client is not None else None
            destination_summary = (
                f"{client.destination_country_name} — {client.visa_type}" if client is not None else None
            )
            validation = None
            try:
                # Hand the AI the client's profile + their VALIDATED documents: the AI decides,
                # per document type and destination, what to cross-check (identity, dates,
                # funds, study plan, …) and FAILS the validation itself on any material
                # conflict — e.g. a passport that belongs to a different person.
                validation = gemini_service.validate_and_extract_document(
                    data, filename, mime_type or "application/octet-stream",
                    document_type=row.document_type,
                    current_date_for_evaluation=datetime.now(timezone.utc).isoformat(),
                    student_profile_context=_ent_client_profile_context(client),
                    related_documents_context=_ent_related_documents_context(db2, client, row.id),
                    destination_country_code=destination_code,
                    destination_summary=destination_summary,
                    usage_source=ENTERPRISE_DOC_SCAN_USAGE_SOURCE,
                )
            except Exception:
                logger.exception("Enterprise doc validation failed (document_id=%s)", document_id)

            row.validated_at = datetime.now(timezone.utc)
            payload: dict = {}
            if isinstance(validation, dict):
                verdict = str(validation.get("Document Validation", "")).strip().lower()
                row.validation_status = "valid" if verdict == "yes" else ("invalid" if verdict == "no" else "error")
                row.validation_message = (str(validation.get("Message") or "").strip() or None)
                fields = _ent_clean_extracted(validation)
                payload["fields"] = fields
                cvf = validation.get("Cross Validation Flags")
                if cvf:
                    payload["cross_validation_flags"] = cvf
                # Auto-fill ONLY runs when the AI passed the document — a red-flagged
                # document (wrong person, expired, inconsistent) never touches the profile.
                if row.validation_status == "valid" and client is not None:
                    autofill = _ent_autofill_profile(db2, client, row.document_type, fields)
                    payload["autofill"] = autofill
                    filled, conflicts = autofill.get("filled") or [], autofill.get("conflicts") or []
                    if filled or conflicts:
                        bits = []
                        if filled:
                            bits.append("auto-filled " + ", ".join(f["field"] for f in filled))
                        if conflicts:
                            bits.append(f"{len(conflicts)} field(s) to review")
                        try:
                            notif.notify_org(
                                db2, row.organization_id, type="document_validated",
                                title=f"Rilono AI validated {row.document_type} for {client.full_name} — " + "; ".join(bits),
                                reference_type="client", reference_id=client.id, commit=False,
                            )
                        except Exception:
                            pass
                elif row.validation_status == "invalid" and client is not None:
                    # Surface the AI's red flag to the whole org, not just the uploader.
                    # The filename keeps re-uploads distinct (notify_org dedupes on title).
                    try:
                        notif.notify_org(
                            db2, row.organization_id, type="document_flagged",
                            title=(f"⚠ Rilono AI flagged {row.document_type} "
                                   f"({row.original_filename}) for {client.full_name} — needs review"),
                            reference_type="client", reference_id=client.id, commit=False,
                        )
                    except Exception:
                        pass
            else:
                row.validation_status = "error"
                row.validation_message = "Rilono AI could not validate this document automatically."

            if payload:
                row.extracted_fields = json.dumps(payload)[:100000]

            # Bill only a scan that actually returned a verdict. "error" means Rilono AI
            # could not read the document — the agency gets that for free, same invariant
            # every other AI action here holds (a failed action is never charged).
            if row.validation_status in {"valid", "invalid"}:
                charging_user = (
                    db2.query(models.User).filter(models.User.id == int(charge_user_id)).first()
                    if charge_user_id else None
                )
                try:
                    credits.charge_action(
                        db2, row.organization_id, ENTERPRISE_DOC_SCAN_ACTION_KEY,
                        user=charging_user, reference_type="client", reference_id=row.client_id,
                        description=(
                            f"Document scan — {row.document_type}"
                            + (f" · {client.full_name}" if client is not None else "")
                        ),
                        commit=False,
                    )
                    row.validation_credits_charged = credits.action_cost(ENTERPRISE_DOC_SCAN_ACTION_KEY)
                except Exception:
                    # The wallet was drained by a concurrent action between this scan's
                    # pre-check and here. The Gemini call is already paid for, so keep the
                    # result rather than discarding work — but record 0 charged so the
                    # ledger and this row never disagree about what was actually billed.
                    logger.warning(
                        "Document scan completed but could not be charged (document_id=%s, org_id=%s)",
                        document_id, row.organization_id, exc_info=True,
                    )
                    row.validation_credits_charged = 0

            db2.commit()
        except Exception:
            logger.exception("Background document processing failed (document_id=%s)", document_id)
        finally:
            db2.close()

    threading.Thread(target=_worker, daemon=True).start()


@router.post("/clients/{client_id}/documents/{document_id}/scan")
def enterprise_scan_client_document(
    client_id: int,
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Billed: run Rilono AI's scan & validate on ONE already-stored document.

    Serves both halves of the opt-in — a document uploaded without a scan, and a re-scan of
    one already checked (useful once more of the client's dossier has been validated, since
    the cross-validation reference set grows as documents pass). Priced per run either way.
    """
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request,
        require_capability=("documents.upload", "credits.spend"),
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    _enforce_rate_limit_or_429(
        request, scope="enterprise.document_scan",
        limit=ENTERPRISE_DOC_SCAN_RATE_LIMIT,
        window_seconds=ENTERPRISE_DOC_SCAN_RATE_WINDOW_SECONDS,
        extra_key=str(current_user.id),
    )

    doc = (
        db.query(models.EnterpriseClientDocument)
        .filter(
            models.EnterpriseClientDocument.id == int(document_id),
            models.EnterpriseClientDocument.client_id == client.id,
            models.EnterpriseClientDocument.organization_id == organization.id,
        )
        .first()
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found.")
    if doc.validation_status is None:
        # NULL means a scan is already in flight; a second one would bill twice for the
        # same document and race the first worker's write.
        raise HTTPException(status_code=409, detail="Rilono AI is already scanning this document.")
    if not gemini_service.is_ai_configured():
        raise HTTPException(status_code=503, detail="Document scanning isn't available right now.")

    # Block before spending anything — same order as every other billed action here.
    credits.enforce_action_or_402(db, organization.id, ENTERPRISE_DOC_SCAN_ACTION_KEY)

    try:
        data = enterprise_storage.fetch_document(doc.storage_key)
    except enterprise_storage.DocumentNotFound:
        logger.warning("Document blob missing for scan (document_id=%s)", doc.id)
        raise HTTPException(
            status_code=410,
            detail="This document's file is no longer available and can't be scanned. Please re-upload it.",
        )
    except Exception:
        logger.exception("Failed to fetch document for scan (document_id=%s)", doc.id)
        raise HTTPException(status_code=502, detail="Could not retrieve the document right now. Please try again.")

    # Hand the worker a clean slate: a re-scan that fails must not leave the previous
    # verdict on screen looking like the new one. NULL marks the scan as in flight.
    doc.validation_status = None
    doc.validation_message = None
    db.commit()
    db.refresh(doc)

    _start_document_processing(
        doc.id, data, doc.original_filename, doc.mime_type,
        validate=True, charge_user_id=current_user.id,
    )

    return {
        "message": "Rilono AI is scanning this document.",
        "permissions": _enterprise_permissions_for_role(role),
        "document": _serialize_client_document(
            doc, include_sensitive=role.ctx.has("clients.view_sensitive")
        ),
        "scan_pricing": _document_scan_pricing(db, organization.id),
    }


@router.get("/clients/{client_id}/documents/{document_id}/download")
def enterprise_download_client_document(
    client_id: int,
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    # Downloading the raw file means a passport scan, a bank letter or a transcript, so it needs
    # its own capability — read-only members deliberately don't get it. Resolving the CLIENT first
    # is what applies record scope: filtering the document by (id, client_id, organization_id)
    # alone let anyone walk sequential document ids against any client id in the workspace.
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="documents.download"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    doc = (
        db.query(models.EnterpriseClientDocument)
        .filter(
            models.EnterpriseClientDocument.id == int(document_id),
            models.EnterpriseClientDocument.client_id == client.id,
            models.EnterpriseClientDocument.organization_id == organization.id,
        )
        .first()
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found.")

    try:
        data = enterprise_storage.fetch_document(doc.storage_key)
    except enterprise_storage.DocumentNotFound:
        logger.warning("Document blob missing on download (document_id=%s)", doc.id)
        raise HTTPException(
            status_code=410,
            detail="This file is no longer available. It may have been removed; please re-upload it.",
        )
    except Exception:
        logger.exception("Failed to fetch client document id=%s", doc.id)
        raise HTTPException(status_code=502, detail="Could not retrieve the document right now. Please try again.")

    ext = os.path.splitext(doc.original_filename)[1].lower()
    # Serve a Content-Type derived from the validated extension — NEVER the uploader-supplied
    # doc.mime_type. Only known-safe visual types (PDF/images) render inline; everything else is
    # forced to an octet-stream attachment. This prevents a *.pdf uploaded as text/html+<script>
    # from executing in the portal origin (stored XSS / account takeover).
    if ext in ENTERPRISE_DOC_INLINE_EXT:
        disposition = "inline"
        media_type = ENTERPRISE_DOC_EXT_MIME.get(ext, "application/octet-stream")
    else:
        disposition = "attachment"
        media_type = "application/octet-stream"
    filename = _safe_filename(doc.original_filename)
    return Response(
        content=data,
        media_type=media_type,
        headers={
            "Content-Disposition": f'{disposition}; filename="{filename}"',
            "X-Content-Type-Options": "nosniff",
        },
    )


@router.delete("/clients/{client_id}/documents/{document_id}")
def enterprise_delete_client_document(
    client_id: int,
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="documents.delete"
    )
    # Same reasoning as the download route: resolve the client so record scope applies before the
    # document is touched at all.
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    doc = (
        db.query(models.EnterpriseClientDocument)
        .filter(
            models.EnterpriseClientDocument.id == int(document_id),
            models.EnterpriseClientDocument.client_id == client.id,
            models.EnterpriseClientDocument.organization_id == organization.id,
        )
        .first()
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found.")
    enterprise_storage.delete_document(doc.storage_key)
    db.delete(doc)
    db.commit()
    return {"message": "Document deleted.", "permissions": _enterprise_permissions_for_role(role)}


@router.post("/clients/{client_id}/documents/{document_id}/accept")
def enterprise_accept_client_document(
    client_id: int,
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Human-in-the-loop override: staff accepts a document that Rilono AI red-flagged
    (after checking it themselves). Flips it to valid — with an audit trail — and runs
    the normal profile auto-fill. The AI stays the default gatekeeper; this is the escape
    hatch for its false alarms."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="documents.accept"
    )
    # Resolve the client FIRST, like the sibling document routes do. Resolving it after the
    # document lookup turned this endpoint into an existence oracle: against a client outside the
    # caller's scope, a real document id answered 400 ("only flagged documents…") and a made-up
    # one answered 404 — which is enough to enumerate another office's documents and their
    # validation state without ever being allowed to open the client.
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    doc = (
        db.query(models.EnterpriseClientDocument)
        .filter(
            models.EnterpriseClientDocument.id == int(document_id),
            models.EnterpriseClientDocument.client_id == int(client_id),
            models.EnterpriseClientDocument.organization_id == organization.id,
        )
        .first()
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found.")
    if doc.validation_status not in ("invalid", "error"):
        raise HTTPException(status_code=400, detail="Only documents Rilono AI flagged can be accepted manually.")

    staff_name = (current_user.full_name or current_user.email or "staff").strip()
    prior_message = (doc.validation_message or "").strip()
    try:
        payload = json.loads(doc.extracted_fields) if doc.extracted_fields else {}
        if not isinstance(payload, dict):
            payload = {}
    except Exception:
        payload = {}

    doc.validation_status = "valid"
    doc.validation_message = (
        f"Accepted by {staff_name} after manual review."
        + (f" Rilono AI had flagged: {prior_message}" if prior_message else "")
    )[:2000]
    accepted_at = datetime.now(dt_timezone.utc)
    # Provenance columns: the document is valid, but a human — not the AI — cleared it.
    doc.manually_accepted_at = accepted_at
    doc.manually_accepted_by = staff_name[:255]
    payload["accepted_by"] = staff_name
    payload["accepted_at"] = accepted_at.isoformat()
    if prior_message:
        payload["ai_flag_before_accept"] = prior_message[:2000]

    fields = payload.get("fields") or {}
    if client is not None and isinstance(fields, dict) and fields:
        payload["autofill"] = _ent_autofill_profile(db, client, doc.document_type, fields)
    doc.extracted_fields = json.dumps(payload)[:100000]

    try:
        db.add(models.EnterpriseClientNote(
            organization_id=organization.id,
            client_id=int(client_id),
            author_user_id=current_user.id,
            author_name=staff_name,
            body=(f"Manually accepted the {doc.document_type} ({doc.original_filename}) that "
                  f"Rilono AI had flagged" + (f': "{prior_message[:300]}"' if prior_message else ".")),
        ))
    except Exception:
        logger.exception("Failed to add accept audit note (document_id=%s)", document_id)
    db.commit()
    return {
        "document": _serialize_client_document(
            doc, include_sensitive=role.ctx.has("clients.view_sensitive")
        ),
        "permissions": _enterprise_permissions_for_role(role),
    }


ENTERPRISE_DEEP_SCAN_RATE_LIMIT = int(os.getenv("ENTERPRISE_DEEP_SCAN_RATE_LIMIT", "10"))
ENTERPRISE_DEEP_SCAN_RATE_WINDOW_SECONDS = int(os.getenv("ENTERPRISE_DEEP_SCAN_RATE_WINDOW_SECONDS", "600"))


def _serialize_deep_scan(scan: models.EnterpriseClientDeepScan, *, include_findings: bool = True) -> dict:
    def _json_list(raw):
        try:
            value = json.loads(raw) if raw else []
        except Exception:
            value = []
        return value if isinstance(value, list) else []

    def _json_dict(raw):
        try:
            value = json.loads(raw) if raw else {}
        except Exception:
            value = {}
        return value if isinstance(value, dict) else {}

    out = {
        "id": scan.id,
        "client_id": scan.client_id,
        "risk_level": scan.risk_level,
        "summary": scan.summary,
        "stats": _json_dict(scan.stats),
        "credits_charged": scan.credits_charged or 0,
        "triggered_by_name": scan.triggered_by_name,
        "created_at": _iso(scan.created_at),
        # model_used stays server-side only (provider details are internal).
    }
    if include_findings:
        out["findings"] = _json_list(scan.findings)
        out["checks_passed"] = _json_list(scan.checks_passed)
    return out


def _deep_scan_count(db: Session, organization_id: int, client_id: int) -> int:
    return (
        db.query(func.count(models.EnterpriseClientDeepScan.id))
        .filter(
            models.EnterpriseClientDeepScan.organization_id == int(organization_id),
            models.EnterpriseClientDeepScan.client_id == int(client_id),
        )
        .scalar()
    ) or 0


def _deep_scan_pricing(db: Session, organization_id: int, client_id: int) -> dict:
    """What the NEXT scan of this client will cost. Each client's first scan is free,
    bounded by the org's monthly free-scan budget (anti-farming — see enterprise_credits)."""
    scans_run = _deep_scan_count(db, organization_id, client_id)
    free_remaining = max(0, credits.DEEP_SCAN_FREE_SCANS_PER_CLIENT - scans_run)
    if free_remaining and credits.deep_scan_free_budget_left(db, organization_id) <= 0:
        free_remaining = 0
    return {
        "cost_credits": credits.action_cost("deep_scan"),
        "free_remaining": free_remaining,
        "next_scan_free": free_remaining > 0,
    }


@router.get("/clients/{client_id}/deep-scans")
def enterprise_client_deep_scan_history(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Stored Deep Scan history for a client (newest first, summaries only) plus the
    pricing state for the next run. Viewers can read history; running one needs edit."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="clients.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    scans = (
        db.query(models.EnterpriseClientDeepScan)
        .filter(
            models.EnterpriseClientDeepScan.organization_id == organization.id,
            models.EnterpriseClientDeepScan.client_id == client.id,
        )
        .order_by(models.EnterpriseClientDeepScan.created_at.desc(), models.EnterpriseClientDeepScan.id.desc())
        .limit(50)
        .all()
    )
    return {
        "scans": [_serialize_deep_scan(s, include_findings=False) for s in scans],
        "pricing": _deep_scan_pricing(db, organization.id, client.id),
        "ai_available": enterprise_ai.is_ai_configured(),
        "permissions": _enterprise_permissions_for_role(role),
    }


@router.get("/clients/{client_id}/deep-scans/{scan_id}")
def enterprise_client_deep_scan_detail(
    client_id: int,
    scan_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="clients.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    scan = (
        db.query(models.EnterpriseClientDeepScan)
        .filter(
            models.EnterpriseClientDeepScan.id == int(scan_id),
            models.EnterpriseClientDeepScan.organization_id == organization.id,
            models.EnterpriseClientDeepScan.client_id == client.id,
        )
        .first()
    )
    if not scan:
        raise HTTPException(status_code=404, detail="Deep Scan not found.")
    return {
        "scan": _serialize_deep_scan(scan),
        "permissions": _enterprise_permissions_for_role(role),
    }


@router.post("/clients/{client_id}/deep-scan")
def enterprise_client_deep_scan(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Premium AI action: Rilono AI strictly audits the client's ENTIRE dossier —
    profile, stage case records, document contents, notes, emails, universities,
    interview results and payments — and stores the structured result as history.
    Each client's first scan is free; after that it's credit-billed."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability=("ai.deepscan", "credits.spend")
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    _enforce_rate_limit_or_429(
        request, scope="enterprise.deep_scan",
        limit=ENTERPRISE_DEEP_SCAN_RATE_LIMIT,
        window_seconds=ENTERPRISE_DEEP_SCAN_RATE_WINDOW_SECONDS,
        extra_key=str(current_user.id),
    )

    if not enterprise_ai.is_ai_configured():
        raise HTTPException(status_code=503, detail="Deep Scan isn't available right now.")

    # Each client's first scan is free (counted from STORED scans — failures store
    # nothing, so they never burn the freebie), bounded by the org's monthly free-scan
    # budget so client create→scan→delete churn can't farm unlimited free audits.
    # Paid runs are hard-blocked before spending any Gemini tokens.
    is_free = (
        _deep_scan_count(db, organization.id, client.id) < credits.DEEP_SCAN_FREE_SCANS_PER_CLIENT
        and credits.deep_scan_free_budget_left(db, organization.id) > 0
    )
    if not is_free:
        credits.enforce_action_or_402(db, organization.id, "deep_scan")

    documents = (
        db.query(models.EnterpriseClientDocument)
        .filter(models.EnterpriseClientDocument.client_id == client.id)
        .order_by(models.EnterpriseClientDocument.created_at.asc(), models.EnterpriseClientDocument.id.asc())
        .all()
    )

    # Attribute every Gemini call in this scan (map + reduce) to the org so the
    # admin margin analytics see the real per-scan token cost.
    usage_token = ai_usage.set_usage_account(user_id=current_user.id, organization_id=organization.id)
    try:
        result = enterprise_ai.run_deep_scan_audit(
            db=db,
            client=client,
            organization=organization,
            documents=documents,
            current_date=datetime.utcnow().date().isoformat(),
        )
        # The audit bills Gemini per document before it ever returns, so a None here
        # means we have already spent real money: fail loudly rather than let the
        # unpacking below raise a bare TypeError outside this handler.
        if not isinstance(result, dict) or "risk_level" not in result:
            raise RuntimeError(f"Deep Scan returned no audit payload (got {type(result).__name__})")
    except Exception:
        # No charge on any failure (the debit only happens after success below).
        logger.exception("Deep Scan failed (org_id=%s, client_id=%s)", organization.id, client.id)
        raise HTTPException(status_code=502, detail="The Deep Scan ran into a problem. Please try again.")
    finally:
        ai_usage.reset_usage_account(usage_token)

    def _json_capped(items: list, cap_chars: int) -> str:
        """JSON-encode a list, dropping tail items (never slicing mid-string —
        a raw [:N] cut would store invalid JSON that parses back as empty)."""
        items = list(items)
        while True:
            encoded = json.dumps(items, ensure_ascii=False)
            if len(encoded) <= cap_chars or not items:
                return encoded
            items = items[:-1]

    staff_name = current_user.full_name or current_user.email
    charged = 0 if is_free else credits.action_cost("deep_scan")
    scan = models.EnterpriseClientDeepScan(
        organization_id=organization.id,
        client_id=client.id,
        risk_level=result["risk_level"],
        summary=result.get("summary") or None,
        findings=_json_capped(result.get("findings") or [], 400000),
        checks_passed=_json_capped(result.get("checks_passed") or [], 100000),
        stats=json.dumps(result.get("stats") or {}, ensure_ascii=False),
        model_used=result.get("model_used"),
        credits_charged=charged,
        triggered_by_user_id=current_user.id,
        triggered_by_name=staff_name,
    )
    db.add(scan)

    # Charge (or consume the monthly free budget) only after a successful audit —
    # the scan row and the wallet change commit together.
    if is_free:
        credits.consume_deep_scan_free(db, organization.id)
    else:
        credits.charge_action(
            db, organization.id, "deep_scan",
            user=current_user, reference_type="client", reference_id=client.id,
            description=f"Deep Scan — {client.full_name}", commit=False,
        )
    db.commit()
    db.refresh(scan)

    return {
        "permissions": _enterprise_permissions_for_role(role),
        "scan": _serialize_deep_scan(scan),
        "credits_charged": charged,
        "was_free": is_free,
        "pricing": _deep_scan_pricing(db, organization.id, client.id),
        "wallet": credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization)),
    }


# ===========================================================================
# Writing Studio — AI-drafted SOPs and Letters of Recommendation (credit-billed)
#
# Two things make this different from the B2C SOP generator: the draft is grounded in a
# CONSULTANCY's dossier for the client (case records, documents, shortlist, notes), and
# the deliverable is a formatted Word file the office actually hands over. Generation and
# refinement each cost one credit; re-downloading a stored draft is free (no model call).
# ===========================================================================

WRITING_STUDIO_ACTION_KEY = "writing_studio"
ENTERPRISE_WRITING_RATE_LIMIT = int(os.getenv("ENTERPRISE_WRITING_RATE_LIMIT", "20"))
ENTERPRISE_WRITING_RATE_WINDOW_SECONDS = int(os.getenv("ENTERPRISE_WRITING_RATE_WINDOW_SECONDS", "600"))

DOCX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"


class EnterpriseWritingGenerateRequest(BaseModel):
    doc_type: str = Field("sop", max_length=8)
    university: Optional[str] = Field(None, max_length=200)
    program: Optional[str] = Field(None, max_length=200)
    study_level: Optional[str] = Field(None, max_length=60)
    intake: Optional[str] = Field(None, max_length=60)
    brief: Optional[str] = Field(None, max_length=4000)
    # LOR only — whose voice the letter is written in.
    recommender_type: Optional[str] = Field(None, max_length=32)
    recommender_name: Optional[str] = Field(None, max_length=160)
    recommender_title: Optional[str] = Field(None, max_length=160)
    recommender_org: Optional[str] = Field(None, max_length=200)
    recommender_email: Optional[str] = Field(None, max_length=254)
    relationship_context: Optional[str] = Field(None, max_length=2000)


class EnterpriseWritingRefineRequest(BaseModel):
    instruction: str = Field(..., min_length=3, max_length=1000)


def _writing_pricing(db: Session, organization_id: int) -> dict:
    cost = credits.action_cost(WRITING_STUDIO_ACTION_KEY)
    return {
        "cost_credits": cost,
        "can_afford": credits.can_afford(db, organization_id, WRITING_STUDIO_ACTION_KEY),
    }


def _writing_drafts_for_client(db: Session, organization_id: int, client_id: int) -> list:
    """Latest version of each document chain, newest chain first."""
    rows = (
        db.query(models.EnterpriseClientWritingDraft)
        .filter(
            models.EnterpriseClientWritingDraft.organization_id == int(organization_id),
            models.EnterpriseClientWritingDraft.client_id == int(client_id),
        )
        .order_by(models.EnterpriseClientWritingDraft.id.desc())
        .all()
    )
    latest: dict[int, models.EnterpriseClientWritingDraft] = {}
    for row in rows:
        key = row.root_id or row.id
        if key not in latest or int(row.version or 1) > int(latest[key].version or 1):
            latest[key] = row
    return sorted(latest.values(), key=lambda r: (r.root_id or r.id), reverse=True)


def _writing_latest_of_root(
    db: Session, organization_id: int, client_id: int, root_id: int
) -> Optional[models.EnterpriseClientWritingDraft]:
    return (
        db.query(models.EnterpriseClientWritingDraft)
        .filter(
            models.EnterpriseClientWritingDraft.organization_id == int(organization_id),
            models.EnterpriseClientWritingDraft.client_id == int(client_id),
            models.EnterpriseClientWritingDraft.root_id == int(root_id),
        )
        .order_by(models.EnterpriseClientWritingDraft.version.desc())
        .first()
    )


def _writing_defaults(db: Session, client: models.EnterpriseClient) -> dict:
    """Prefill for the composer: the shortlist entries a counselor would actually write
    for, with any admitted/applied university offered first."""
    rows = (
        db.query(models.EnterpriseClientUniversity)
        .filter(models.EnterpriseClientUniversity.client_id == client.id)
        .order_by(models.EnterpriseClientUniversity.created_at.desc())
        .limit(30)
        .all()
    )
    rank = {"admitted": 0, "applied": 1, "considering": 2, "rejected": 3}
    rows.sort(key=lambda u: rank.get((u.status or "").lower(), 2))
    return {
        "intake": client.intake,
        "country_code": client.destination_country_code,
        "universities": [
            {"id": u.id, "university_name": u.university_name, "program": u.program, "status": u.status}
            for u in rows
        ],
    }


@router.get("/clients/{client_id}/writing")
def enterprise_client_writing_list(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Everything the Writing Studio tab needs on open: stored drafts (latest version of
    each chain), the composer's reference data and prefill, and the credit price."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="clients.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    drafts = _writing_drafts_for_client(db, organization.id, client.id)
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "ai_available": enterprise_writing.is_ai_configured(),
        "catalog": enterprise_writing.catalog_payload(),
        "defaults": _writing_defaults(db, client),
        "pricing": _writing_pricing(db, organization.id),
        "drafts": [enterprise_writing.serialize_draft(d, include_content=False) for d in drafts],
        "active": enterprise_writing.serialize_draft(drafts[0]) if drafts else None,
    }


@router.get("/clients/{client_id}/writing/{root_id}/versions")
def enterprise_client_writing_versions(
    client_id: int,
    root_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Every stored version of one document chain, oldest first."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="clients.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    rows = (
        db.query(models.EnterpriseClientWritingDraft)
        .filter(
            models.EnterpriseClientWritingDraft.organization_id == organization.id,
            models.EnterpriseClientWritingDraft.client_id == client.id,
            models.EnterpriseClientWritingDraft.root_id == int(root_id),
        )
        .order_by(models.EnterpriseClientWritingDraft.version.asc())
        .all()
    )
    if not rows:
        raise HTTPException(status_code=404, detail="That document could not be found.")
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "versions": [enterprise_writing.serialize_draft(r) for r in rows],
    }


def _writing_generate_or_502(runner, *, organization_id: int, client_id: int, what: str) -> dict:
    """Run a Writing Studio model call, attributing token cost to the org and turning any
    failure into a 502 BEFORE a credit is charged (the debit only happens on success)."""
    usage_token = ai_usage.set_usage_account(organization_id=organization_id)
    try:
        return runner()
    except Exception:
        logger.exception("Writing Studio %s failed (org_id=%s, client_id=%s)",
                         what, organization_id, client_id)
        raise HTTPException(
            status_code=502,
            detail="Rilono AI couldn't finish this draft. No credits were used — please try again.",
        )
    finally:
        ai_usage.reset_usage_account(usage_token)


@router.post("/clients/{client_id}/writing/generate")
def enterprise_client_writing_generate(
    client_id: int,
    payload: EnterpriseWritingGenerateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Premium AI action: draft a new SOP or LOR for this client from their real dossier."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability=("ai.writing", "credits.spend")
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    _enforce_rate_limit_or_429(
        request, scope="enterprise.writing_studio",
        limit=ENTERPRISE_WRITING_RATE_LIMIT,
        window_seconds=ENTERPRISE_WRITING_RATE_WINDOW_SECONDS,
        extra_key=str(current_user.id),
    )
    # Validate the brief BEFORE checking service availability, so a missing recommender
    # name always reports the fixable problem rather than being masked by a 503.
    doc_type = enterprise_writing.normalize_doc_type(payload.doc_type)
    university = (payload.university or "").strip() or None
    program = (payload.program or "").strip() or None
    recommender_name = (payload.recommender_name or "").strip() or None
    if doc_type == "lor" and not recommender_name:
        # Without a named recommender there is no letterhead and no signature block —
        # the output would be an unusable letter, so this is rejected up front.
        raise HTTPException(
            status_code=400,
            detail="Add the recommender's name — a letter of recommendation is written in their voice and signed by them.",
        )
    if doc_type == "sop" and not (university or program):
        raise HTTPException(
            status_code=400,
            detail="Add the target university or program so the statement can be written for a specific course.",
        )

    if not enterprise_writing.is_ai_configured():
        raise HTTPException(status_code=503, detail="The Writing Studio isn't available right now.")

    # Hard-block an unaffordable run before spending any Gemini tokens.
    credits.enforce_action_or_402(db, organization.id, WRITING_STUDIO_ACTION_KEY)

    result = _writing_generate_or_502(
        lambda: enterprise_writing.generate_draft(
            db,
            client=client,
            doc_type=doc_type,
            university=university,
            program=program,
            study_level=(payload.study_level or "").strip() or None,
            intake=(payload.intake or "").strip() or None,
            brief=(payload.brief or "").strip() or None,
            recommender_type=payload.recommender_type,
            recommender_name=recommender_name,
            recommender_title=(payload.recommender_title or "").strip() or None,
            recommender_org=(payload.recommender_org or "").strip() or None,
            relationship_context=(payload.relationship_context or "").strip() or None,
        ),
        organization_id=organization.id, client_id=client.id, what="generation",
    )

    staff_name = current_user.full_name or current_user.email
    charged = credits.action_cost(WRITING_STUDIO_ACTION_KEY)
    draft = models.EnterpriseClientWritingDraft(
        organization_id=organization.id,
        client_id=client.id,
        doc_type=doc_type,
        version=1,
        country_code=client.destination_country_code,
        university=university,
        program=program,
        study_level=(payload.study_level or "").strip() or None,
        intake=(payload.intake or "").strip() or client.intake,
        recommender_type=(enterprise_writing.normalize_recommender_type(payload.recommender_type)
                          if doc_type == "lor" else None),
        recommender_name=recommender_name if doc_type == "lor" else None,
        recommender_title=((payload.recommender_title or "").strip() or None) if doc_type == "lor" else None,
        recommender_org=((payload.recommender_org or "").strip() or None) if doc_type == "lor" else None,
        recommender_email=((payload.recommender_email or "").strip() or None) if doc_type == "lor" else None,
        relationship_context=((payload.relationship_context or "").strip() or None) if doc_type == "lor" else None,
        brief=(payload.brief or "").strip() or None,
        instruction=None,
        title=result["title"],
        content_md=result["content_md"],
        notes_md=result["notes_md"] or None,
        word_count=result["word_count"],
        model_used=result["model_used"],
        credits_charged=charged,
        created_by_user_id=current_user.id,
        created_by_name=staff_name,
    )
    db.add(draft)
    db.flush()
    draft.root_id = draft.id  # version 1 roots its own chain
    # Charge only after a successful draft — the row and the wallet commit together.
    credits.charge_action(
        db, organization.id, WRITING_STUDIO_ACTION_KEY,
        user=current_user, reference_type="client", reference_id=client.id,
        description=f"{'LOR' if doc_type == 'lor' else 'SOP'} draft — {client.full_name}",
        commit=False,
    )
    db.commit()
    db.refresh(draft)

    return {
        "permissions": _enterprise_permissions_for_role(role),
        "draft": enterprise_writing.serialize_draft(draft),
        "credits_charged": charged,
        "coverage": result.get("coverage") or {},
        "pricing": _writing_pricing(db, organization.id),
        "wallet": credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization)),
    }


@router.post("/clients/{client_id}/writing/{root_id}/refine")
def enterprise_client_writing_refine(
    client_id: int,
    root_id: int,
    payload: EnterpriseWritingRefineRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Premium AI action: revise a stored draft per the counselor's instruction, storing
    the result as the next immutable version of the same chain."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability=("ai.writing", "credits.spend")
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    _enforce_rate_limit_or_429(
        request, scope="enterprise.writing_studio",
        limit=ENTERPRISE_WRITING_RATE_LIMIT,
        window_seconds=ENTERPRISE_WRITING_RATE_WINDOW_SECONDS,
        extra_key=str(current_user.id),
    )
    if not enterprise_writing.is_ai_configured():
        raise HTTPException(status_code=503, detail="The Writing Studio isn't available right now.")

    latest = _writing_latest_of_root(db, organization.id, client.id, root_id)
    if not latest:
        raise HTTPException(status_code=404, detail="That document could not be found.")

    instruction = payload.instruction.strip()
    credits.enforce_action_or_402(db, organization.id, WRITING_STUDIO_ACTION_KEY)

    result = _writing_generate_or_502(
        lambda: enterprise_writing.refine_draft(
            db, client=client, latest=latest, instruction=instruction),
        organization_id=organization.id, client_id=client.id, what="refinement",
    )

    charged = credits.action_cost(WRITING_STUDIO_ACTION_KEY)
    draft = models.EnterpriseClientWritingDraft(
        organization_id=organization.id,
        client_id=client.id,
        doc_type=latest.doc_type,
        root_id=latest.root_id or latest.id,
        version=int(latest.version or 1) + 1,
        country_code=latest.country_code,
        university=latest.university,
        program=latest.program,
        study_level=latest.study_level,
        intake=latest.intake,
        recommender_type=latest.recommender_type,
        recommender_name=latest.recommender_name,
        recommender_title=latest.recommender_title,
        recommender_org=latest.recommender_org,
        recommender_email=latest.recommender_email,
        relationship_context=latest.relationship_context,
        brief=latest.brief,
        instruction=instruction,
        title=result["title"],
        content_md=result["content_md"],
        notes_md=result["notes_md"] or None,
        word_count=result["word_count"],
        model_used=result["model_used"],
        credits_charged=charged,
        created_by_user_id=current_user.id,
        created_by_name=current_user.full_name or current_user.email,
    )
    db.add(draft)
    credits.charge_action(
        db, organization.id, WRITING_STUDIO_ACTION_KEY,
        user=current_user, reference_type="client", reference_id=client.id,
        description=f"{'LOR' if latest.doc_type == 'lor' else 'SOP'} revision — {client.full_name}",
        commit=False,
    )
    db.commit()
    db.refresh(draft)

    return {
        "permissions": _enterprise_permissions_for_role(role),
        "draft": enterprise_writing.serialize_draft(draft),
        "credits_charged": charged,
        "pricing": _writing_pricing(db, organization.id),
        "wallet": credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization)),
    }


@router.get("/clients/{client_id}/writing/{draft_id}/docx")
def enterprise_client_writing_docx(
    client_id: int,
    draft_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Download one stored version as a formatted Word document. Free — the draft is
    already paid for, so re-exporting it never calls the model or charges again."""
    _, organization, _role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="ai.writing")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=_role.ctx)
    draft = (
        db.query(models.EnterpriseClientWritingDraft)
        .filter(
            models.EnterpriseClientWritingDraft.id == int(draft_id),
            models.EnterpriseClientWritingDraft.organization_id == organization.id,
            models.EnterpriseClientWritingDraft.client_id == client.id,
        )
        .first()
    )
    if not draft:
        raise HTTPException(status_code=404, detail="That document could not be found.")

    try:
        data = enterprise_writing.build_docx(
            draft, client=client, organization_name=(organization.company_name or "").strip() or "Rilono")
    except Exception:
        logger.exception("Writing Studio Word export failed (draft_id=%s)", draft.id)
        raise HTTPException(status_code=502, detail="Could not build the Word file. Please try again.")

    filename = _safe_filename(enterprise_writing.docx_filename(draft, client))
    return Response(
        content=data,
        media_type=DOCX_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Content-Type-Options": "nosniff",
        },
    )


@router.delete("/clients/{client_id}/writing/{root_id}")
def enterprise_client_writing_delete(
    client_id: int,
    root_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Delete a document and all its versions. Credits already spent are not refunded."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="ai.writing"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    deleted = (
        db.query(models.EnterpriseClientWritingDraft)
        .filter(
            models.EnterpriseClientWritingDraft.organization_id == organization.id,
            models.EnterpriseClientWritingDraft.client_id == client.id,
            models.EnterpriseClientWritingDraft.root_id == int(root_id),
        )
        .delete(synchronize_session=False)
    )
    db.commit()
    if not deleted:
        raise HTTPException(status_code=404, detail="That document could not be found.")
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "deleted_versions": deleted,
    }


# ===========================================================================
# Mock visa interview (Gemini role-plays the visa officer)
# ===========================================================================

import json as _json

ENTERPRISE_INTERVIEW_RATE_LIMIT = int(os.getenv("ENTERPRISE_INTERVIEW_RATE_LIMIT", "60"))
ENTERPRISE_INTERVIEW_RATE_WINDOW_SECONDS = int(os.getenv("ENTERPRISE_INTERVIEW_RATE_WINDOW_SECONDS", "300"))


class EnterpriseInterviewTurn(BaseModel):
    role: str = Field(..., max_length=16)
    content: str = Field(..., max_length=8000)


class EnterpriseInterviewChatRequest(BaseModel):
    message: Optional[str] = Field(default=None, max_length=4000)
    history: Optional[list[EnterpriseInterviewTurn]] = None
    start: bool = False


class EnterpriseInterviewFeedbackRequest(BaseModel):
    history: list[EnterpriseInterviewTurn] = Field(..., min_length=1)
    mode: str = Field(default="chat", max_length=12)


def _recent_client_notes(db: Session, client_id: int, limit: int = 6):
    return (
        db.query(models.EnterpriseClientNote)
        .filter(models.EnterpriseClientNote.client_id == int(client_id))
        .order_by(models.EnterpriseClientNote.created_at.desc())
        .limit(limit)
        .all()
    )


def _recent_client_documents(db: Session, client_id: int, limit: int = 12):
    """This client's uploaded documents (newest first) so the mock interview can be
    grounded in — and cross-examine against — the applicant's real evidence."""
    return (
        db.query(models.EnterpriseClientDocument)
        .filter(models.EnterpriseClientDocument.client_id == int(client_id))
        .order_by(
            models.EnterpriseClientDocument.created_at.desc(),
            models.EnterpriseClientDocument.id.desc(),
        )
        .limit(limit)
        .all()
    )


def _serialize_interview_session(s: models.EnterpriseInterviewSession, include_detail: bool = False) -> dict:
    data = {
        "id": s.id,
        "conducted_by_name": s.conducted_by_name,
        "mode": s.mode,
        "verdict": s.verdict,
        "created_at": _iso(s.created_at),
    }
    if include_detail:
        try:
            data["transcript"] = _json.loads(s.transcript) if s.transcript else []
        except Exception:
            data["transcript"] = []
        data["feedback"] = s.feedback
    return data


@router.post("/clients/{client_id}/interview/chat")
def enterprise_interview_chat(
    client_id: int,
    payload: EnterpriseInterviewChatRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability=("interviews.run", "credits.spend"))
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.interview", limit=ENTERPRISE_INTERVIEW_RATE_LIMIT,
        window_seconds=ENTERPRISE_INTERVIEW_RATE_WINDOW_SECONDS, extra_key=str(current_user.id),
    )
    if not enterprise_interview.is_ai_configured():
        raise HTTPException(status_code=503, detail="The mock interview isn't available right now.")

    is_start = bool(payload.start)
    # A STAFF-run interview is a preview/test tool — the self-serve link a student takes
    # is the real product. Staff get a few free previews per org, then it costs the normal
    # mock_interview price. Hard-block before any tokens are spent if unaffordable.
    if is_start:
        credits.enforce_staff_interview_or_402(db, organization.id)

    history = [t.model_dump() for t in (payload.history or [])]
    try:
        turn = enterprise_interview.run_interview_turn(
            client=client,
            organization=organization,
            recent_notes=_recent_client_notes(db, client.id),
            documents=_recent_client_documents(db, client.id),
            history=history,
            message=payload.message or "",
            is_start=is_start,
        )
    except Exception:
        logger.exception("Mock interview turn failed (org_id=%s, client_id=%s)", organization.id, client.id)
        raise HTTPException(status_code=502, detail="The interviewer ran into a problem. Please try again.")

    response_payload = {
        "reply": turn["reply"],
        "finished": turn["finished"],
        "decision": turn["decision"],
        "permissions": _enterprise_permissions_for_role(role),
    }
    if is_start:
        meter = credits.consume_staff_interview(
            db, organization.id,
            user=current_user, reference_id=client.id,
            description=f"Mock interview (staff preview) — {client.full_name}",
        )
        response_payload["credits_charged"] = meter["charged"]
        response_payload["was_preview"] = meter["was_preview"]
        response_payload["previews_remaining"] = meter["previews_remaining"]
        response_payload["wallet"] = credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization))
    return response_payload


@router.post("/clients/{client_id}/interview/feedback")
def enterprise_interview_feedback(
    client_id: int,
    payload: EnterpriseInterviewFeedbackRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability=("interviews.run", "credits.spend"))
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.interview", limit=ENTERPRISE_INTERVIEW_RATE_LIMIT,
        window_seconds=ENTERPRISE_INTERVIEW_RATE_WINDOW_SECONDS, extra_key=str(current_user.id),
    )
    if not enterprise_interview.is_ai_configured():
        raise HTTPException(status_code=503, detail="Feedback isn't available right now.")

    history = [t.model_dump() for t in payload.history]
    try:
        feedback = enterprise_interview.generate_interview_feedback(
            client=client, organization=organization, history=history,
            documents=_recent_client_documents(db, client.id),
        )
    except Exception:
        logger.exception("Mock interview feedback failed (org_id=%s, client_id=%s)", organization.id, client.id)
        raise HTTPException(status_code=502, detail="Could not generate feedback. Please try again.")

    verdict = enterprise_interview.extract_verdict(feedback)
    session = models.EnterpriseInterviewSession(
        organization_id=organization.id,
        client_id=client.id,
        conducted_by_user_id=current_user.id,
        conducted_by_name=current_user.full_name or current_user.email,
        mode=("voice" if payload.mode == "voice" else "chat"),
        transcript=_json.dumps(history)[:400000],
        feedback=feedback,
        verdict=verdict,
    )
    db.add(session)
    db.commit()
    db.refresh(session)

    return {
        "feedback": feedback,
        "verdict": verdict,
        "session": _serialize_interview_session(session),
        "permissions": _enterprise_permissions_for_role(role),
    }


@router.get("/clients/{client_id}/interview/sessions")
def enterprise_interview_sessions(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="interviews.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    rows = (
        db.query(models.EnterpriseInterviewSession)
        .filter(models.EnterpriseInterviewSession.client_id == client.id)
        .order_by(models.EnterpriseInterviewSession.created_at.desc(), models.EnterpriseInterviewSession.id.desc())
        .all()
    )
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "ai_enabled": enterprise_interview.is_ai_configured(),
        "sessions": [_serialize_interview_session(s) for s in rows],
    }


@router.get("/clients/{client_id}/interview/sessions/{session_id}")
def enterprise_interview_session_detail(
    client_id: int,
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="interviews.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    s = (
        db.query(models.EnterpriseInterviewSession)
        .filter(
            models.EnterpriseInterviewSession.id == int(session_id),
            models.EnterpriseInterviewSession.client_id == client.id,
            models.EnterpriseInterviewSession.organization_id == organization.id,
        )
        .first()
    )
    if not s:
        raise HTTPException(status_code=404, detail="Interview session not found.")
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "session": _serialize_interview_session(s, include_detail=True),
    }


# ===========================================================================
# Send mock interview to client via secure email link (self-serve)
# ===========================================================================

import secrets as _secrets

ENTERPRISE_INTERVIEW_INVITE_EXPIRES_DAYS = int(os.getenv("ENTERPRISE_INTERVIEW_INVITE_EXPIRES_DAYS", "30"))
ENTERPRISE_INTERVIEW_SESSION_HOURS = int(os.getenv("ENTERPRISE_INTERVIEW_SESSION_HOURS", "3"))
ENTERPRISE_INTERVIEW_CODE_EXPIRES_MIN = int(os.getenv("ENTERPRISE_INTERVIEW_CODE_EXPIRES_MIN", "15"))
ENTERPRISE_INTERVIEW_CODE_MAX_ATTEMPTS = int(os.getenv("ENTERPRISE_INTERVIEW_CODE_MAX_ATTEMPTS", "6"))
ENTERPRISE_INTERVIEW_INVITE_MAX_COUNT = int(os.getenv("ENTERPRISE_INTERVIEW_INVITE_MAX_COUNT", "20"))
ENTERPRISE_PUBLIC_RATE_LIMIT = int(os.getenv("ENTERPRISE_PUBLIC_INTERVIEW_RATE_LIMIT", "30"))
ENTERPRISE_PUBLIC_RATE_WINDOW = int(os.getenv("ENTERPRISE_PUBLIC_INTERVIEW_RATE_WINDOW", "300"))
ENTERPRISE_CODE_RATE_LIMIT = int(os.getenv("ENTERPRISE_INTERVIEW_CODE_RATE_LIMIT", "6"))
ENTERPRISE_CODE_RATE_WINDOW = int(os.getenv("ENTERPRISE_INTERVIEW_CODE_RATE_WINDOW", "900"))


class EnterpriseInterviewInviteRequest(BaseModel):
    allowed_count: int = Field(default=3, ge=1, le=ENTERPRISE_INTERVIEW_INVITE_MAX_COUNT)


class PublicInterviewVerifyRequest(BaseModel):
    code: str = Field(..., min_length=4, max_length=10)


class PublicInterviewChatRequest(BaseModel):
    session_token: str = Field(..., min_length=10, max_length=4000)
    message: Optional[str] = Field(default=None, max_length=4000)
    history: Optional[list[EnterpriseInterviewTurn]] = None
    start: bool = False


class PublicInterviewFeedbackRequest(BaseModel):
    session_token: str = Field(..., min_length=10, max_length=4000)
    history: list[EnterpriseInterviewTurn] = Field(..., min_length=1)
    mode: str = Field(default="chat", max_length=12)
    decision: Optional[str] = Field(default=None, max_length=20)


def _mask_email(email: str) -> str:
    e = (email or "").strip()
    if "@" not in e:
        return e
    local, dom = e.split("@", 1)
    ml = (local[0] + "*") if len(local) <= 2 else (local[0] + "*" * (len(local) - 2) + local[-1])
    return f"{ml}@{dom}"


def _build_interview_invite_url(subdomain_slug, token: str, request: Request | None) -> str:
    subdomain = str(subdomain_slug or "").strip().lower()
    base = None
    if subdomain:
        host = f"{subdomain}.{ENTERPRISE_ROOT_DOMAIN}"
        port = _request_port_for_local_enterprise_url(request)
        if port:
            host = f"{host}:{port}"
        base = f"{ENTERPRISE_PORTAL_SCHEME}://{host}"
    if not base:
        base = ENTERPRISE_PASSWORD_SETUP_BASE_URL
    return f"{base.rstrip('/')}/interview/{token}"


def _interview_invite_remaining(invite: models.EnterpriseInterviewInvite) -> int:
    return max(0, int(invite.allowed_count or 0) - int(invite.used_count or 0))


def _interview_invite_is_live(invite: models.EnterpriseInterviewInvite) -> bool:
    if invite.revoked:
        return False
    if invite.expires_at:
        exp = invite.expires_at.replace(tzinfo=None) if getattr(invite.expires_at, "tzinfo", None) else invite.expires_at
        if exp < datetime.utcnow():
            return False
    return True


def _serialize_invite_status(invite: models.EnterpriseInterviewInvite | None) -> Optional[dict]:
    if not invite:
        return None
    started = int(invite.used_count or 0)
    completed = int(getattr(invite, "completed_count", 0) or 0)
    return {
        "id": invite.id,
        "email": invite.email,
        "allowed_count": invite.allowed_count,
        "used_count": invite.used_count,
        "started_count": started,
        "completed_count": completed,
        "last_completed_at": _iso(getattr(invite, "last_completed_at", None)),
        "remaining": _interview_invite_remaining(invite),
        "revoked": bool(invite.revoked),
        "live": _interview_invite_is_live(invite),
        "created_by_name": invite.created_by_name,
        "created_at": _iso(invite.created_at),
        "expires_at": _iso(invite.expires_at),
    }


def _latest_client_invite(db: Session, organization_id: int, client_id: int):
    return (
        db.query(models.EnterpriseInterviewInvite)
        .filter(
            models.EnterpriseInterviewInvite.organization_id == int(organization_id),
            models.EnterpriseInterviewInvite.client_id == int(client_id),
        )
        .order_by(models.EnterpriseInterviewInvite.created_at.desc(), models.EnterpriseInterviewInvite.id.desc())
        .first()
    )


def _issue_interview_session_token(invite_id: int) -> str:
    return create_access_token(
        data={"sub": f"entiv:{int(invite_id)}", "scope": "ent_interview", "inv": int(invite_id)},
        expires_delta=timedelta(hours=ENTERPRISE_INTERVIEW_SESSION_HOURS),
    )


def _decode_interview_session_token(token: str) -> int:
    try:
        payload = jose_jwt.decode(token, AUTH_SECRET_KEY, algorithms=[AUTH_ALGORITHM])
    except JWTError:
        raise HTTPException(status_code=401, detail="Your interview session has expired. Please verify your email again.")
    if payload.get("scope") != "ent_interview" or not payload.get("inv"):
        raise HTTPException(status_code=401, detail="Invalid interview session.")
    return int(payload["inv"])


def _public_invite_or_404(db: Session, token: str) -> models.EnterpriseInterviewInvite:
    token_hash = hash_token((token or "").strip())
    invite = (
        db.query(models.EnterpriseInterviewInvite)
        .filter(models.EnterpriseInterviewInvite.token_hash == token_hash)
        .first()
    )
    if not invite or not _interview_invite_is_live(invite):
        raise HTTPException(status_code=404, detail="This interview link is invalid or has expired.")
    return invite


# ---- Staff: create / view / revoke the invite -----------------------------

@router.post("/clients/{client_id}/interview/invite")
def enterprise_create_interview_invite(
    client_id: int,
    payload: EnterpriseInterviewInviteRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="interviews.invite"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    email = (client.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Add an email to this client before sending an interview link.")

    # Don't send a link the wallet can't fund — otherwise the client receives the
    # email, tries to start, and hits a confusing "contact your consultancy" wall.
    # Require enough credits up-front to cover every interview this link offers.
    credits.enforce_units_or_402(db, organization.id, "mock_interview", int(payload.allowed_count))

    # Supersede any prior invites for this client.
    db.query(models.EnterpriseInterviewInvite).filter(
        models.EnterpriseInterviewInvite.client_id == client.id,
        models.EnterpriseInterviewInvite.revoked.is_(False),
    ).update({"revoked": True})

    raw_token = generate_verification_token()
    invite = models.EnterpriseInterviewInvite(
        organization_id=organization.id,
        client_id=client.id,
        token_hash=hash_token(raw_token),
        email=email,
        allowed_count=int(payload.allowed_count),
        used_count=0,
        expires_at=datetime.utcnow() + timedelta(days=ENTERPRISE_INTERVIEW_INVITE_EXPIRES_DAYS),
        created_by_user_id=current_user.id,
        created_by_name=current_user.full_name or current_user.email,
    )
    db.add(invite)
    db.commit()
    db.refresh(invite)

    link = _build_interview_invite_url(organization.subdomain_slug, raw_token, request)
    sent, _mid, err = send_enterprise_interview_invite_email(
        to_email=email,
        client_name=client.full_name,
        organization_name=organization.company_name,
        interview_url=link,
        allowed_count=invite.allowed_count,
        destination_country=client.destination_country_name,
        visa_type=client.visa_type,
        logo_url=_absolute_enterprise_logo_url(organization),
    )
    message = (f"Sent {invite.allowed_count} mock interview(s) to {email}."
               if sent else f"Invite created but the email could not be sent right now. {err or ''}".strip())
    return {
        "message": message,
        "email_sent": sent,
        "invite": _serialize_invite_status(invite),
        "permissions": _enterprise_permissions_for_role(role),
    }


@router.get("/clients/{client_id}/interview/invite")
def enterprise_get_interview_invite(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="interviews.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    invite = _latest_client_invite(db, organization.id, client.id)
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "invite": _serialize_invite_status(invite),
    }


@router.post("/clients/{client_id}/interview/invite/revoke")
def enterprise_revoke_interview_invite(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="interviews.invite"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    db.query(models.EnterpriseInterviewInvite).filter(
        models.EnterpriseInterviewInvite.client_id == client.id,
        models.EnterpriseInterviewInvite.revoked.is_(False),
    ).update({"revoked": True})
    db.commit()
    return {"message": "Interview link revoked.", "permissions": _enterprise_permissions_for_role(role)}


# ---- Public (client-facing, token-scoped, no staff auth) ------------------

@router.get("/public/interview/{token}")
def public_interview_info(token: str, db: Session = Depends(get_db)):
    invite = _public_invite_or_404(db, token)
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == invite.client_id).first()
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == invite.organization_id).first()
    if not client or not org:
        raise HTTPException(status_code=404, detail="This interview link is no longer available.")
    remaining = _interview_invite_remaining(invite)
    return {
        "organization_name": org.company_name,
        "logo_url": _resolve_enterprise_logo_url(org),
        "client_first_name": (client.full_name or "there").split(" ")[0],
        "destination_country": client.destination_country_name,
        "visa_type": client.visa_type,
        "masked_email": _mask_email(invite.email),
        "remaining": remaining,
        "exhausted": remaining <= 0,
    }


@router.post("/public/interview/{token}/send-code")
def public_interview_send_code(token: str, request: Request, db: Session = Depends(get_db)):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.interview_code",
        limit=ENTERPRISE_CODE_RATE_LIMIT, window_seconds=ENTERPRISE_CODE_RATE_WINDOW, extra_key=hash_token(token)[:16],
    )
    invite = _public_invite_or_404(db, token)
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == invite.organization_id).first()
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == invite.client_id).first()
    if not org or not client:
        raise HTTPException(status_code=404, detail="This interview link is no longer available.")

    code = f"{_secrets.randbelow(900000) + 100000:06d}"
    invite.code_hash = hash_token(code)
    invite.code_expires_at = datetime.utcnow() + timedelta(minutes=ENTERPRISE_INTERVIEW_CODE_EXPIRES_MIN)
    invite.code_attempts = 0
    db.commit()

    sent, _mid, err = send_enterprise_interview_code_email(
        to_email=invite.email, client_name=client.full_name, organization_name=org.company_name, code=code,
    )
    if not sent:
        logger.warning("Interview code email failed for invite %s: %s", invite.id, err)
    return {"sent": bool(sent), "masked_email": _mask_email(invite.email)}


@router.post("/public/interview/{token}/verify")
def public_interview_verify(token: str, payload: PublicInterviewVerifyRequest, request: Request, db: Session = Depends(get_db)):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.interview_verify",
        limit=ENTERPRISE_PUBLIC_RATE_LIMIT, window_seconds=ENTERPRISE_PUBLIC_RATE_WINDOW, extra_key=hash_token(token)[:16],
    )
    invite = _public_invite_or_404(db, token)
    if not invite.code_hash or not invite.code_expires_at:
        raise HTTPException(status_code=400, detail="Please request a verification code first.")
    code_exp = invite.code_expires_at.replace(tzinfo=None) if getattr(invite.code_expires_at, "tzinfo", None) else invite.code_expires_at
    if code_exp < datetime.utcnow():
        raise HTTPException(status_code=400, detail="That code has expired. Please request a new one.")
    if int(invite.code_attempts or 0) >= ENTERPRISE_INTERVIEW_CODE_MAX_ATTEMPTS:
        raise HTTPException(status_code=429, detail="Too many attempts. Please request a new code.")

    invite.code_attempts = int(invite.code_attempts or 0) + 1
    if hash_token((payload.code or "").strip()) != invite.code_hash:
        db.commit()
        raise HTTPException(status_code=400, detail="That code is incorrect. Please try again.")

    # Verified — consume the code and issue a short-lived interview session token.
    invite.code_hash = None
    invite.code_expires_at = None
    db.commit()
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == invite.client_id).first()
    return {
        "session_token": _issue_interview_session_token(invite.id),
        "client_first_name": (client.full_name or "there").split(" ")[0] if client else "there",
        "destination_country": client.destination_country_name if client else "",
        "visa_type": client.visa_type if client else "",
        "remaining": _interview_invite_remaining(invite),
        "session_hours": ENTERPRISE_INTERVIEW_SESSION_HOURS,
    }


def _public_load_invite_context(db: Session, session_token: str):
    invite_id = _decode_interview_session_token(session_token)
    invite = db.query(models.EnterpriseInterviewInvite).filter(models.EnterpriseInterviewInvite.id == invite_id).first()
    if not invite or not _interview_invite_is_live(invite):
        raise HTTPException(status_code=401, detail="This interview link is no longer active.")
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == invite.client_id).first()
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == invite.organization_id).first()
    if not client or not org:
        raise HTTPException(status_code=404, detail="This interview is no longer available.")
    return invite, client, org


@router.post("/public/interview/chat")
def public_interview_chat(payload: PublicInterviewChatRequest, request: Request, db: Session = Depends(get_db)):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.interview_public",
        limit=ENTERPRISE_PUBLIC_RATE_LIMIT, window_seconds=ENTERPRISE_PUBLIC_RATE_WINDOW,
    )
    if not enterprise_interview.is_ai_configured():
        raise HTTPException(status_code=503, detail="The mock interview isn't available right now.")
    invite, client, org = _public_load_invite_context(db, payload.session_token)

    if payload.start:
        if _interview_invite_remaining(invite) <= 0:
            raise HTTPException(status_code=403, detail="You've used all your mock interviews for this link.")
        # Meter the org's wallet — a self-serve interview costs the same 20 credits as a
        # staff-run one, so the "send to student" links can't be used to run Gemini for free.
        if not credits.can_afford(db, org.id, "mock_interview"):
            raise HTTPException(
                status_code=402,
                detail="This mock interview isn't available right now. Please contact your consultancy.",
            )
        invite.used_count = int(invite.used_count or 0) + 1  # a start consumes one interview
        db.commit()
        credits.charge_action(
            db, org.id, "mock_interview",
            reference_type="client", reference_id=client.id,
            description=f"Mock interview (self-serve) — {client.full_name}", commit=True,
        )

    history = [t.model_dump() for t in (payload.history or [])]
    try:
        turn = enterprise_interview.run_interview_turn(
            client=client, organization=org, recent_notes=[],  # never leak staff notes to the client
            documents=_recent_client_documents(db, client.id),  # the applicant's own documents
            history=history, message=payload.message or "", is_start=bool(payload.start),
        )
    except Exception:
        logger.exception("Public interview turn failed (invite=%s)", invite.id)
        raise HTTPException(status_code=502, detail="The interviewer ran into a problem. Please try again.")
    return {
        "reply": turn["reply"],
        "finished": turn["finished"],
        "decision": turn["decision"],
        "remaining": _interview_invite_remaining(invite),
    }


@router.post("/public/interview/feedback")
def public_interview_feedback(payload: PublicInterviewFeedbackRequest, request: Request, db: Session = Depends(get_db)):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.interview_public",
        limit=ENTERPRISE_PUBLIC_RATE_LIMIT, window_seconds=ENTERPRISE_PUBLIC_RATE_WINDOW,
    )
    if not enterprise_interview.is_ai_configured():
        raise HTTPException(status_code=503, detail="Feedback isn't available right now.")
    invite, client, org = _public_load_invite_context(db, payload.session_token)

    history = [t.model_dump() for t in payload.history]
    raw_decision = (payload.decision or "").strip().lower()
    officer_decision = raw_decision if raw_decision in ("approved", "refused") else None
    try:
        feedback = enterprise_interview.generate_interview_feedback(
            client=client, organization=org, history=history, officer_decision=officer_decision,
            documents=_recent_client_documents(db, client.id),  # the applicant's own documents
        )
    except Exception:
        logger.exception("Public interview feedback failed (invite=%s)", invite.id)
        raise HTTPException(status_code=502, detail="Could not generate feedback. Please try again.")

    decision_label = {"approved": "Approved", "refused": "Refused"}.get(officer_decision)
    verdict = enterprise_interview.extract_verdict(feedback)
    # Keep the officer's decision separate from the coaching feedback so it isn't
    # shown twice in the UI/email; fold it into the *stored* copy for staff.
    stored_feedback = (
        f"**Officer's decision (simulated): Visa {decision_label}**\n\n" + feedback
        if decision_label else feedback
    )
    session = models.EnterpriseInterviewSession(
        organization_id=org.id, client_id=client.id, conducted_by_user_id=None,
        conducted_by_name=f"{client.full_name} (self · via link)",
        mode=("voice" if payload.mode == "voice" else "chat"),
        transcript=_json.dumps(history)[:400000], feedback=stored_feedback, verdict=verdict,
    )
    db.add(session)
    # Mark the invite as having a completed interview so staff can see the student
    # actually finished it (used_count only tells us they *started*).
    invite.completed_count = int(invite.completed_count or 0) + 1
    invite.last_completed_at = datetime.utcnow()
    db.commit()

    # Tell the whole team (actor is the external student, so nobody is excluded).
    notif.notify_org(
        db, org.id, type="interview_completed",
        title=f"🎤 {client.full_name} completed their mock interview",
        body=(f"Verdict: {verdict}" if verdict else None),
        reference_type="client", reference_id=client.id, commit=True,
    )

    # Email the report to the applicant who took the interview (best-effort).
    emailed = False
    try:
        emailed, _mid, err = send_enterprise_interview_report_email(
            to_email=invite.email,
            client_name=client.full_name,
            organization_name=org.company_name,
            destination_country=client.destination_country_name,
            visa_type=client.visa_type,
            decision_label=decision_label,
            feedback_markdown=feedback,
            logo_url=_absolute_enterprise_logo_url(org),
        )
        if not emailed:
            logger.warning("Interview report email failed for invite %s: %s", invite.id, err)
    except Exception:
        logger.exception("Interview report email crashed (invite=%s)", invite.id)

    return {
        "feedback": feedback,
        "verdict": verdict,
        "decision": decision_label,
        "remaining": _interview_invite_remaining(invite),
        "emailed": emailed,
        "masked_email": _mask_email(invite.email),
    }


# ===========================================================================
# Secure document requests — email a client a tokenized, OTP-verified link so
# they can upload the specific documents staff asked for. Files land in the same
# private encrypted storage as staff uploads. Mirrors the interview-invite model.
# ===========================================================================

ENTERPRISE_DOCREQ_EXPIRES_DAYS = int(os.getenv("ENTERPRISE_DOCREQ_EXPIRES_DAYS", "30"))
ENTERPRISE_DOCREQ_SESSION_HOURS = int(os.getenv("ENTERPRISE_DOCREQ_SESSION_HOURS", "6"))
ENTERPRISE_DOCREQ_MAX_TYPES = int(os.getenv("ENTERPRISE_DOCREQ_MAX_TYPES", "20"))


class EnterpriseDocumentRequestCreate(BaseModel):
    document_types: list[str] = Field(..., min_length=1, max_length=ENTERPRISE_DOCREQ_MAX_TYPES)
    message: Optional[str] = Field(default=None, max_length=2000)


class PublicDocRequestVerifyRequest(BaseModel):
    code: str = Field(..., min_length=4, max_length=10)


def _build_document_request_url(subdomain_slug, token: str, request: Request | None) -> str:
    subdomain = str(subdomain_slug or "").strip().lower()
    base = None
    if subdomain:
        host = f"{subdomain}.{ENTERPRISE_ROOT_DOMAIN}"
        port = _request_port_for_local_enterprise_url(request)
        if port:
            host = f"{host}:{port}"
        base = f"{ENTERPRISE_PORTAL_SCHEME}://{host}"
    if not base:
        base = ENTERPRISE_PASSWORD_SETUP_BASE_URL
    return f"{base.rstrip('/')}/upload/{token}"


def _docreq_is_live(req: models.EnterpriseDocumentRequest) -> bool:
    if req.revoked:
        return False
    if req.expires_at:
        exp = req.expires_at.replace(tzinfo=None) if getattr(req.expires_at, "tzinfo", None) else req.expires_at
        if exp < datetime.utcnow():
            return False
    return True


def _recompute_docreq_status(req: models.EnterpriseDocumentRequest) -> None:
    items = req.items or []
    received = sum(1 for i in items if i.status == "received")
    if received == 0:
        req.status = "pending"
        req.completed_at = None
    elif received >= len(items):
        req.status = "completed"
        if not req.completed_at:
            req.completed_at = datetime.utcnow()
    else:
        req.status = "partial"
        req.completed_at = None


def _serialize_docreq_item(item: models.EnterpriseDocumentRequestItem) -> dict:
    return {
        "id": item.id,
        "document_type": item.document_type,
        "status": item.status,
        "received": item.status == "received",
        "received_at": _iso(item.received_at),
        "document_id": item.document_id,
    }


def _serialize_docreq(req: models.EnterpriseDocumentRequest | None) -> Optional[dict]:
    if not req:
        return None
    items = list(req.items or [])
    received = sum(1 for i in items if i.status == "received")
    return {
        "id": req.id,
        "email": req.email,
        "message": req.message,
        "status": req.status,
        "revoked": bool(req.revoked),
        "live": _docreq_is_live(req),
        "total": len(items),
        "received": received,
        "pending": len(items) - received,
        "items": [_serialize_docreq_item(i) for i in items],
        "created_by_name": req.created_by_name,
        "created_at": _iso(req.created_at),
        "expires_at": _iso(req.expires_at),
        "completed_at": _iso(req.completed_at),
    }


def _latest_client_docreq(db: Session, organization_id: int, client_id: int):
    return (
        db.query(models.EnterpriseDocumentRequest)
        .filter(
            models.EnterpriseDocumentRequest.organization_id == int(organization_id),
            models.EnterpriseDocumentRequest.client_id == int(client_id),
        )
        .order_by(models.EnterpriseDocumentRequest.created_at.desc(), models.EnterpriseDocumentRequest.id.desc())
        .first()
    )


def _public_docreq_or_404(db: Session, token: str) -> models.EnterpriseDocumentRequest:
    token_hash = hash_token((token or "").strip())
    req = (
        db.query(models.EnterpriseDocumentRequest)
        .filter(models.EnterpriseDocumentRequest.token_hash == token_hash)
        .first()
    )
    if not req or not _docreq_is_live(req):
        raise HTTPException(status_code=404, detail="This upload link is invalid or has expired.")
    return req


def _issue_docreq_session_token(request_id: int) -> str:
    return create_access_token(
        data={"sub": f"entdr:{int(request_id)}", "scope": "ent_docreq", "dr": int(request_id)},
        expires_delta=timedelta(hours=ENTERPRISE_DOCREQ_SESSION_HOURS),
    )


def _decode_docreq_session_token(token: str) -> int:
    try:
        payload = jose_jwt.decode(token, AUTH_SECRET_KEY, algorithms=[AUTH_ALGORITHM])
    except JWTError:
        raise HTTPException(status_code=401, detail="Your upload session has expired. Please verify your email again.")
    if payload.get("scope") != "ent_docreq" or not payload.get("dr"):
        raise HTTPException(status_code=401, detail="Invalid upload session.")
    return int(payload["dr"])


def _public_load_docreq_context(db: Session, session_token: str):
    request_id = _decode_docreq_session_token(session_token)
    req = (
        db.query(models.EnterpriseDocumentRequest)
        .filter(models.EnterpriseDocumentRequest.id == request_id)
        .first()
    )
    if not req or not _docreq_is_live(req):
        raise HTTPException(status_code=401, detail="This upload link is no longer active.")
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == req.client_id).first()
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == req.organization_id).first()
    if not client or not org:
        raise HTTPException(status_code=404, detail="This upload link is no longer available.")
    return req, client, org


# ---- Staff: create / view / revoke the document request -------------------

@router.post("/clients/{client_id}/document-requests")
def enterprise_create_document_request(
    client_id: int,
    payload: EnterpriseDocumentRequestCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="documents.request"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    email = (client.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Add an email to this client before requesting documents.")

    # Normalize + de-duplicate the requested document types (preserve order).
    seen: set[str] = set()
    doc_types: list[str] = []
    for raw in payload.document_types:
        normalized = catalog.normalize_document_type(raw)
        if normalized and normalized not in seen:
            seen.add(normalized)
            doc_types.append(normalized)
    if not doc_types:
        raise HTTPException(status_code=400, detail="Choose at least one document to request.")

    # Supersede any prior active requests for this client.
    db.query(models.EnterpriseDocumentRequest).filter(
        models.EnterpriseDocumentRequest.client_id == client.id,
        models.EnterpriseDocumentRequest.revoked.is_(False),
    ).update({"revoked": True})

    raw_token = generate_verification_token()
    req = models.EnterpriseDocumentRequest(
        organization_id=organization.id,
        client_id=client.id,
        token_hash=hash_token(raw_token),
        email=email,
        message=(payload.message or "").strip() or None,
        status="pending",
        expires_at=datetime.utcnow() + timedelta(days=ENTERPRISE_DOCREQ_EXPIRES_DAYS),
        created_by_user_id=current_user.id,
        created_by_name=current_user.full_name or current_user.email,
    )
    db.add(req)
    db.flush()
    for dt in doc_types:
        db.add(models.EnterpriseDocumentRequestItem(
            request_id=req.id,
            organization_id=organization.id,
            document_type=dt,
            status="pending",
        ))
    db.commit()
    db.refresh(req)

    link = _build_document_request_url(organization.subdomain_slug, raw_token, request)
    sent, _mid, err = send_enterprise_document_request_email(
        to_email=email,
        client_name=client.full_name,
        organization_name=organization.company_name,
        upload_url=link,
        document_types=doc_types,
        message=req.message,
        logo_url=_absolute_enterprise_logo_url(organization),
    )
    message = (f"Document request sent to {email}."
               if sent else f"Request created but the email could not be sent right now. {err or ''}".strip())
    return {
        "message": message,
        "email_sent": sent,
        "request": _serialize_docreq(req),
        "permissions": _enterprise_permissions_for_role(role),
    }


@router.get("/clients/{client_id}/document-requests")
def enterprise_get_document_request(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="documents.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    req = _latest_client_docreq(db, organization.id, client.id)
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "document_types": list(catalog.STUDENT_DOCUMENT_TYPES),
        "request": _serialize_docreq(req),
    }


@router.post("/clients/{client_id}/document-requests/revoke")
def enterprise_revoke_document_request(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="documents.request"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    db.query(models.EnterpriseDocumentRequest).filter(
        models.EnterpriseDocumentRequest.client_id == client.id,
        models.EnterpriseDocumentRequest.revoked.is_(False),
    ).update({"revoked": True})
    db.commit()
    return {"message": "Document request revoked.", "permissions": _enterprise_permissions_for_role(role)}


# ---- Public (client-facing, token-scoped, no staff auth) ------------------

@router.get("/public/document-request/{token}")
def public_document_request_info(token: str, db: Session = Depends(get_db)):
    req = _public_docreq_or_404(db, token)
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == req.client_id).first()
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == req.organization_id).first()
    if not client or not org:
        raise HTTPException(status_code=404, detail="This upload link is no longer available.")
    return {
        "organization_name": org.company_name,
        "logo_url": _resolve_enterprise_logo_url(org),
        "client_first_name": (client.full_name or "there").split(" ")[0],
        "masked_email": _mask_email(req.email),
        "message": req.message,
        "status": req.status,
        "items": [_serialize_docreq_item(i) for i in (req.items or [])],
        "expires_at": _iso(req.expires_at),
    }


@router.post("/public/document-request/{token}/send-code")
def public_document_request_send_code(token: str, request: Request, db: Session = Depends(get_db)):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.docreq_code",
        limit=ENTERPRISE_CODE_RATE_LIMIT, window_seconds=ENTERPRISE_CODE_RATE_WINDOW, extra_key=hash_token(token)[:16],
    )
    req = _public_docreq_or_404(db, token)
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == req.organization_id).first()
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == req.client_id).first()
    if not org or not client:
        raise HTTPException(status_code=404, detail="This upload link is no longer available.")

    code = f"{_secrets.randbelow(900000) + 100000:06d}"
    req.code_hash = hash_token(code)
    req.code_expires_at = datetime.utcnow() + timedelta(minutes=ENTERPRISE_INTERVIEW_CODE_EXPIRES_MIN)
    req.code_attempts = 0
    db.commit()

    sent, _mid, err = send_enterprise_document_request_code_email(
        to_email=req.email, client_name=client.full_name, organization_name=org.company_name, code=code,
    )
    if not sent:
        logger.warning("Document request code email failed for request %s: %s", req.id, err)
    return {"sent": bool(sent), "masked_email": _mask_email(req.email)}


@router.post("/public/document-request/{token}/verify")
def public_document_request_verify(token: str, payload: PublicDocRequestVerifyRequest, request: Request, db: Session = Depends(get_db)):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.docreq_verify",
        limit=ENTERPRISE_PUBLIC_RATE_LIMIT, window_seconds=ENTERPRISE_PUBLIC_RATE_WINDOW, extra_key=hash_token(token)[:16],
    )
    req = _public_docreq_or_404(db, token)
    if not req.code_hash or not req.code_expires_at:
        raise HTTPException(status_code=400, detail="Please request a verification code first.")
    code_exp = req.code_expires_at.replace(tzinfo=None) if getattr(req.code_expires_at, "tzinfo", None) else req.code_expires_at
    if code_exp < datetime.utcnow():
        raise HTTPException(status_code=400, detail="That code has expired. Please request a new one.")
    if int(req.code_attempts or 0) >= ENTERPRISE_INTERVIEW_CODE_MAX_ATTEMPTS:
        raise HTTPException(status_code=429, detail="Too many attempts. Please request a new code.")

    req.code_attempts = int(req.code_attempts or 0) + 1
    if hash_token((payload.code or "").strip()) != req.code_hash:
        db.commit()
        raise HTTPException(status_code=400, detail="That code is incorrect. Please try again.")

    # Verified — consume the code and issue a short-lived upload session token.
    req.code_hash = None
    req.code_expires_at = None
    db.commit()
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == req.client_id).first()
    return {
        "session_token": _issue_docreq_session_token(req.id),
        "client_first_name": (client.full_name or "there").split(" ")[0] if client else "there",
        "message": req.message,
        "items": [_serialize_docreq_item(i) for i in (req.items or [])],
        "session_hours": ENTERPRISE_DOCREQ_SESSION_HOURS,
    }


@router.post("/public/document-request/upload")
async def public_document_request_upload(
    request: Request,
    session_token: str = Form(...),
    item_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.docreq_upload",
        limit=ENTERPRISE_PUBLIC_RATE_LIMIT, window_seconds=ENTERPRISE_PUBLIC_RATE_WINDOW,
    )
    req, client, org = _public_load_docreq_context(db, session_token)

    item = (
        db.query(models.EnterpriseDocumentRequestItem)
        .filter(
            models.EnterpriseDocumentRequestItem.id == int(item_id),
            models.EnterpriseDocumentRequestItem.request_id == req.id,
        )
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="That document is not part of this request.")

    if not enterprise_storage.is_configured():
        raise HTTPException(status_code=503, detail="Document upload is not available right now. Please contact your consultancy.")

    original = _safe_filename(file.filename)
    ext = os.path.splitext(original)[1].lower()
    if ext not in ENTERPRISE_DOC_ALLOWED_EXT:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Allowed: PDF, images, Word/Excel, CSV, or text.",
        )

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="The file is empty.")
    if len(data) > ENTERPRISE_DOC_MAX_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"File is too large. Maximum size is {ENTERPRISE_DOC_MAX_BYTES // (1024 * 1024)} MB.",
        )

    storage_key = f"enterprise/{org.id}/clients/{client.id}/{uuid.uuid4().hex}{ext}"
    try:
        await run_in_threadpool(enterprise_storage.store_document, storage_key, data, content_type=file.content_type)
    except Exception:
        logger.exception("Failed to store client-uploaded document (org_id=%s, client_id=%s)", org.id, client.id)
        raise HTTPException(status_code=502, detail="Could not store the document right now. Please try again.")

    doc = models.EnterpriseClientDocument(
        organization_id=org.id,
        client_id=client.id,
        document_type=item.document_type,
        original_filename=original,
        storage_key=storage_key,
        file_size=len(data),
        mime_type=(file.content_type or None),
        uploaded_by_user_id=None,
        uploaded_by_name=f"{client.full_name} (uploaded via secure link)",
    )
    db.add(doc)
    db.flush()

    item.document_id = doc.id
    item.status = "received"
    item.received_at = datetime.utcnow()
    was_completed = bool(req.completed_at)
    _recompute_docreq_status(req)
    db.commit()
    db.refresh(req)

    # Notify the team when the request BECOMES complete (not on every file — limited comms).
    if req.completed_at and not was_completed:
        notif.notify_org(
            db, org.id, type="docs_submitted",
            title=f"📁 {client.full_name} submitted all requested documents",
            reference_type="client", reference_id=client.id, commit=True,
        )

    # Extract text in the background so the AI copilot can read the new document. The
    # billed scan deliberately does NOT run here: this endpoint is reached by a student
    # holding an emailed link, with no authenticated member and no capability check, so
    # auto-scanning would let an outsider spend the agency's credits at will (and, by
    # re-uploading, without limit). Staff scan these from the document card instead.
    _start_document_processing(doc.id, data, original, file.content_type, validate=False)

    return {
        "message": "Uploaded.",
        "items": [_serialize_docreq_item(i) for i in (req.items or [])],
        "status": req.status,
    }


# ===========================================================================
# Per-client university shortlisting (B2B)
#
# Each consultancy client gets their own shortlist, tailored to THAT student's
# destination country. Reuses the proven B2C recommendation engine
# (app/university_shortlist.py) — it is a pure function over a country name — but
# every row here is org+client scoped, staff-attributed, and the AI action is
# metered against the organization's Rilono Credits wallet.
# ===========================================================================

UNIVERSITY_ACTION_KEY = "university_match"
_UNIVERSITY_DIFFICULTY = {"reach", "match", "safety"}


def _serialize_client_university(row: models.EnterpriseClientUniversity) -> dict:
    try:
        requirements = json.loads(row.key_requirements) if row.key_requirements else []
        if not isinstance(requirements, list):
            requirements = []
    except Exception:
        requirements = []
    return {
        "id": int(row.id),
        "university_name": row.university_name,
        "program": row.program,
        "location": row.location,
        "country_code": row.country_code,
        "status": row.status or "considering",
        "source": row.source or "manual",
        "est_tuition": row.est_tuition,
        "rationale": row.rationale,
        "notes": row.notes,
        "qs_world_rank": row.qs_world_rank,
        "country_rank": row.country_rank,
        "admission_difficulty": row.admission_difficulty,
        "application_fee": row.application_fee,
        "website_url": row.website_url,
        "admissions_url": row.admissions_url,
        "key_requirements": [str(x)[:140] for x in requirements][:6],
        "added_by_name": row.added_by_name,
        "created_at": row.created_at.isoformat() if row.created_at else None,
    }


def _safe_external_url(value) -> Optional[str]:
    """Only absolute http(s) URLs are stored — these are rendered as hrefs, so a
    javascript:/data: value from the model (or a crafted API call) must never persist."""
    s = str(value or "").strip()
    if not s or not re.match(r"^https?://[^\s/$.?#].[^\s]*$", s, re.I):
        return None
    return s[:400]


def _normalize_university_status(value) -> str:
    from app.university_shortlist import VALID_STATUSES, DEFAULT_STATUS
    normalized = str(value or "").strip().lower()
    return normalized if normalized in VALID_STATUSES else DEFAULT_STATUS


def _client_universities_query(db: Session, organization_id: int, client_id: int):
    return (
        db.query(models.EnterpriseClientUniversity)
        .filter(
            models.EnterpriseClientUniversity.organization_id == organization_id,
            models.EnterpriseClientUniversity.client_id == client_id,
        )
        .order_by(
            models.EnterpriseClientUniversity.created_at.desc(),
            models.EnterpriseClientUniversity.id.desc(),
        )
    )


class EnterpriseUniversityCreate(BaseModel):
    university_name: str = Field(..., min_length=1, max_length=200)
    program: Optional[str] = Field(None, max_length=200)
    location: Optional[str] = Field(None, max_length=160)
    status: Optional[str] = None
    source: Optional[str] = "manual"
    est_tuition: Optional[str] = Field(None, max_length=80)
    rationale: Optional[str] = Field(None, max_length=600)
    notes: Optional[str] = Field(None, max_length=1000)
    qs_world_rank: Optional[str] = Field(None, max_length=20)
    country_rank: Optional[str] = Field(None, max_length=20)
    admission_difficulty: Optional[str] = Field(None, max_length=20)
    application_fee: Optional[str] = Field(None, max_length=60)
    website_url: Optional[str] = Field(None, max_length=400)
    admissions_url: Optional[str] = Field(None, max_length=400)
    key_requirements: Optional[list[str]] = None


class EnterpriseUniversityUpdate(BaseModel):
    status: Optional[str] = None
    notes: Optional[str] = Field(None, max_length=1000)


class EnterpriseUniversityRecommend(BaseModel):
    field_of_study: str = Field(..., min_length=1, max_length=120)
    level: Optional[str] = Field(None, max_length=60)
    budget: Optional[str] = Field(None, max_length=60)
    gpa: Optional[str] = Field(None, max_length=60)
    test_scores: Optional[str] = Field(None, max_length=160)
    preferences: Optional[str] = Field(None, max_length=300)
    max_results: int = Field(6, ge=1, le=8)


@router.get("/clients/{client_id}/universities")
def enterprise_client_universities_list(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """The client's shortlist + the destination context the UI tailors itself to."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="universities.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    rows = _client_universities_query(db, organization.id, client.id).all()
    from app import university_shortlist

    return {
        "permissions": _enterprise_permissions_for_role(role),
        "entries": [_serialize_client_university(r) for r in rows],
        "destination_country_code": client.destination_country_code,
        "destination_country": client.destination_country_name,
        "client_name": client.full_name,
        "recommend_available": university_shortlist.ai_available(),
        "recommend_cost": credits.action_cost(UNIVERSITY_ACTION_KEY),
    }


@router.get("/clients/{client_id}/universities/search")
def enterprise_client_universities_search(
    client_id: int,
    request: Request,
    q: str = "",
    limit: int = 8,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Typeahead over the shared registry, scoped to THIS client's destination country,
    so a UK applicant never sees US-only schools."""
    _, organization, _role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="universities.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=_role.ctx)

    term = (q or "").strip()
    if len(term) < 2:
        return {"country_code": client.destination_country_code, "results": []}
    take = max(1, min(int(limit or 8), 15))
    code = (client.destination_country_code or "").strip().upper()

    rows = (
        db.query(models.USUniversity)
        .filter(
            models.USUniversity.country_code == code,
            models.USUniversity.university_name.ilike(f"%{term}%"),
        )
        .limit(take * 6)
        .all()
    )
    # One university can hold several rows (the registry PK is the email domain), so
    # dedupe by name, then surface prefix matches before mid-string matches.
    seen, prefix, contains = set(), [], []
    lowered = term.lower()
    for row in rows:
        name = (row.university_name or "").strip()
        key = name.lower()
        if not name or key in seen:
            continue
        seen.add(key)
        item = {"name": name, "location": row.location}
        (prefix if key.startswith(lowered) else contains).append(item)
    return {"country_code": code, "results": (prefix + contains)[:take]}


@router.post("/clients/{client_id}/universities", status_code=status.HTTP_201_CREATED)
def enterprise_client_university_add(
    client_id: int,
    payload: EnterpriseUniversityCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="universities.manage"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)

    name = (payload.university_name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="University name is required.")

    difficulty = str(payload.admission_difficulty or "").strip().lower()
    requirements = [str(x).strip()[:140] for x in (payload.key_requirements or []) if str(x).strip()][:6]

    row = models.EnterpriseClientUniversity(
        organization_id=organization.id,
        client_id=client.id,
        # Snapshot the destination the shortlist was built for (server-decided, never client-supplied).
        country_code=client.destination_country_code,
        university_name=name[:200],
        program=(payload.program or "").strip()[:200] or None,
        location=(payload.location or "").strip()[:160] or None,
        status=_normalize_university_status(payload.status),
        source="ai" if str(payload.source or "").strip().lower() == "ai" else "manual",
        est_tuition=(payload.est_tuition or "").strip()[:80] or None,
        rationale=(payload.rationale or "").strip()[:600] or None,
        notes=(payload.notes or "").strip()[:1000] or None,
        qs_world_rank=(payload.qs_world_rank or "").strip()[:20] or None,
        country_rank=(payload.country_rank or "").strip()[:20] or None,
        admission_difficulty=difficulty if difficulty in _UNIVERSITY_DIFFICULTY else None,
        application_fee=(payload.application_fee or "").strip()[:60] or None,
        # Re-validate server-side: these render as clickable links, so only absolute
        # http(s) URLs are stored (a javascript:/data: href would be a stored-XSS vector).
        website_url=_safe_external_url(payload.website_url),
        admissions_url=_safe_external_url(payload.admissions_url),
        key_requirements=json.dumps(requirements) if requirements else None,
        added_by_user_id=current_user.id,
        added_by_name=(current_user.full_name or current_user.email or "")[:120] or None,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "entry": _serialize_client_university(row),
    }


@router.patch("/clients/{client_id}/universities/{entry_id}")
def enterprise_client_university_update(
    client_id: int,
    entry_id: int,
    payload: EnterpriseUniversityUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="universities.manage"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    row = (
        db.query(models.EnterpriseClientUniversity)
        .filter(
            models.EnterpriseClientUniversity.id == entry_id,
            models.EnterpriseClientUniversity.organization_id == organization.id,
            models.EnterpriseClientUniversity.client_id == client.id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="University not found.")

    fields = payload.model_dump(exclude_unset=True)
    if "status" in fields:
        row.status = _normalize_university_status(fields.get("status"))
    if "notes" in fields:
        note = (fields.get("notes") or "").strip()
        row.notes = note[:1000] or None
    db.commit()
    db.refresh(row)
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "entry": _serialize_client_university(row),
    }


@router.delete("/clients/{client_id}/universities/{entry_id}")
def enterprise_client_university_delete(
    client_id: int,
    entry_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="universities.manage"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    deleted = (
        db.query(models.EnterpriseClientUniversity)
        .filter(
            models.EnterpriseClientUniversity.id == entry_id,
            models.EnterpriseClientUniversity.organization_id == organization.id,
            models.EnterpriseClientUniversity.client_id == client.id,
        )
        .delete(synchronize_session=False)
    )
    if not deleted:
        raise HTTPException(status_code=404, detail="University not found.")
    db.commit()
    return {"deleted": True, "id": entry_id, "permissions": _enterprise_permissions_for_role(role)}


@router.post("/clients/{client_id}/universities/recommend")
def enterprise_client_university_recommend(
    client_id: int,
    payload: EnterpriseUniversityRecommend,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """AI university matches for this client, tailored to their destination country.

    Choreography mirrors the proven B2C/Deep-Scan ordering: rate-limit → wallet
    pre-check → generate → fail without charging → charge ONLY on a usable result.
    """
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability=("ai.shortlist", "credits.spend")
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)

    _enforce_rate_limit_or_429(
        request=request,
        scope="enterprise.university_recommend",
        limit=20,
        window_seconds=600,
        extra_key=str(current_user.id),
    )

    from app import university_shortlist

    if not university_shortlist.ai_available():
        raise HTTPException(status_code=503, detail="AI recommendations are not configured.")

    destination = (client.destination_country_name or "").strip()
    if not destination:
        raise HTTPException(status_code=400, detail="Set this client's destination country first.")

    # Hard-block a broke wallet BEFORE spending any Gemini tokens.
    credits.enforce_action_or_402(db, organization.id, UNIVERSITY_ACTION_KEY)

    ai_usage.set_usage_account(organization_id=organization.id)
    result = university_shortlist.recommend_universities(
        destination_country=destination,
        field_of_study=payload.field_of_study,
        level=payload.level,
        budget=payload.budget,
        gpa=payload.gpa,
        test_scores=payload.test_scores,
        home_country=client.nationality,
        preferences=payload.preferences,
        max_results=payload.max_results,
        usage_source="enterprise_university_shortlist",
    )

    if not result.get("available"):
        raise HTTPException(status_code=503, detail=result.get("message") or "Recommendations are unavailable right now.")
    universities = result.get("universities") or []
    if not universities:
        # Nothing usable came back — never bill the consultancy for an empty result.
        raise HTTPException(
            status_code=502,
            detail="Couldn't generate recommendations. Refine the field of study or preferences and retry.",
        )

    txn = credits.charge_action(
        db, organization.id, UNIVERSITY_ACTION_KEY,
        user=current_user, reference_type="client", reference_id=client.id,
        description=f"University shortlist — {client.full_name}", commit=True,
    )
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "universities": universities,
        "destination_country": destination,
        "grounded": bool(result.get("grounded")),
        "credits_charged": credits.action_cost(UNIVERSITY_ACTION_KEY) if txn else 0,
        "wallet": credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization)),
    }


# ===========================================================================
# Client portal share — read-only case tracking for the client
#
# Staff share a client's case as a secure emailed link. The client verifies an
# OTP sent to their own email, then sees a VIEW-ONLY portal: journey stages with
# the per-stage case record, profile details, documents, universities and their
# payment history. Security model mirrors interview invites / document requests
# (hashed capability token + OTP + short-lived signed session token). There is
# deliberately NO write path, and staff notes / internal financials (commission
# split, settlement state) are never included in the payload.
# ===========================================================================

ENTERPRISE_PORTAL_SHARE_EXPIRES_DAYS = int(os.getenv("ENTERPRISE_PORTAL_SHARE_EXPIRES_DAYS", "180"))
ENTERPRISE_PORTAL_SESSION_HOURS = int(os.getenv("ENTERPRISE_PORTAL_SESSION_HOURS", "24"))


class PublicPortalVerifyRequest(BaseModel):
    code: str = Field(..., min_length=4, max_length=10)


class PublicPortalDataRequest(BaseModel):
    session_token: str = Field(..., min_length=10, max_length=4000)


def _build_portal_share_url(subdomain_slug, token: str, request: Request | None) -> str:
    subdomain = str(subdomain_slug or "").strip().lower()
    base = None
    if subdomain:
        host = f"{subdomain}.{ENTERPRISE_ROOT_DOMAIN}"
        port = _request_port_for_local_enterprise_url(request)
        if port:
            host = f"{host}:{port}"
        base = f"{ENTERPRISE_PORTAL_SCHEME}://{host}"
    if not base:
        base = ENTERPRISE_PASSWORD_SETUP_BASE_URL
    return f"{base.rstrip('/')}/portal/{token}"


def _portal_share_is_live(share: models.EnterpriseClientPortalShare) -> bool:
    if share.revoked:
        return False
    if share.expires_at:
        exp = share.expires_at.replace(tzinfo=None) if getattr(share.expires_at, "tzinfo", None) else share.expires_at
        if exp < datetime.utcnow():
            return False
    return True


def _serialize_portal_share_status(share: models.EnterpriseClientPortalShare | None) -> Optional[dict]:
    if not share:
        return None
    return {
        "id": share.id,
        "email": share.email,
        "revoked": bool(share.revoked),
        "live": _portal_share_is_live(share),
        "last_opened_at": _iso(share.last_opened_at),
        "open_count": int(share.open_count or 0),
        "created_by_name": share.created_by_name,
        "created_at": _iso(share.created_at),
        "expires_at": _iso(share.expires_at),
    }


def _latest_client_portal_share(db: Session, organization_id: int, client_id: int):
    return (
        db.query(models.EnterpriseClientPortalShare)
        .filter(
            models.EnterpriseClientPortalShare.organization_id == int(organization_id),
            models.EnterpriseClientPortalShare.client_id == int(client_id),
        )
        .order_by(models.EnterpriseClientPortalShare.created_at.desc(), models.EnterpriseClientPortalShare.id.desc())
        .first()
    )


def _issue_portal_session_token(share_id: int) -> str:
    return create_access_token(
        data={"sub": f"entps:{int(share_id)}", "scope": "ent_portal", "psh": int(share_id)},
        expires_delta=timedelta(hours=ENTERPRISE_PORTAL_SESSION_HOURS),
    )


def _decode_portal_session_token(token: str) -> int:
    try:
        payload = jose_jwt.decode(token, AUTH_SECRET_KEY, algorithms=[AUTH_ALGORITHM])
    except JWTError:
        raise HTTPException(status_code=401, detail="Your portal session has expired. Please verify your email again.")
    if payload.get("scope") != "ent_portal" or not payload.get("psh"):
        raise HTTPException(status_code=401, detail="Invalid portal session.")
    return int(payload["psh"])


def _public_portal_share_or_404(db: Session, token: str) -> models.EnterpriseClientPortalShare:
    token_hash = hash_token((token or "").strip())
    share = (
        db.query(models.EnterpriseClientPortalShare)
        .filter(models.EnterpriseClientPortalShare.token_hash == token_hash)
        .first()
    )
    if not share or not _portal_share_is_live(share):
        raise HTTPException(status_code=404, detail="This portal link is invalid or has expired.")
    return share


def _public_load_portal_context(db: Session, session_token: str):
    share_id = _decode_portal_session_token(session_token)
    share = db.query(models.EnterpriseClientPortalShare).filter(models.EnterpriseClientPortalShare.id == share_id).first()
    if not share or not _portal_share_is_live(share):
        raise HTTPException(status_code=401, detail="This portal link is no longer active.")
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == share.client_id).first()
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == share.organization_id).first()
    if not client or not org:
        raise HTTPException(status_code=404, detail="This portal is no longer available.")
    return share, client, org


# Payment states a student should see. Route plumbing states (transferred /
# settled / on_hold) are all just "paid" from the payer's side; the fee split
# (commission/payout) is consultancy-internal and never serialized here.
_PORTAL_PAYMENT_STATUS = {
    "created": ("pending", "Payment requested"),
    "paid": ("paid", "Paid"),
    "transferred": ("paid", "Paid"),
    "on_hold": ("paid", "Paid"),
    "settled": ("paid", "Paid"),
    "failed": ("failed", "Failed"),
    "refunded": ("refunded", "Refunded"),
    "partially_refunded": ("partially_refunded", "Partially refunded"),
    "cancelled": ("cancelled", "Cancelled"),
}


def _mask_passport_number(value) -> Optional[str]:
    """'M1234567' -> '•••• 567'. Short values mask fully; None stays None."""
    p = str(value or "").strip()
    if not p:
        return None
    if len(p) < 6:
        return "••••"
    return f"•••• {p[-3:]}"


def _portal_stage_records(client: models.EnterpriseClient) -> dict:
    """Per-stage recorded fields, resolved against the destination-aware catalog.

    Returns {stage_key: [{label, value, type}, …]} containing only fields with a
    recorded value, in this destination's stage order — the client sees exactly what
    staff filled in."""
    data = _load_stage_data(client)
    out: dict[str, list] = {}
    for stage in catalog.stages_for(client.destination_country_code):
        stage_key = stage["key"]
        recorded = data.get(stage_key) or {}
        if not recorded:
            continue
        fields = []
        for field in catalog.stage_fields_for(client.destination_country_code, stage_key):
            # Textarea fields are counselor free-text (hold_notes, refusal_notes,
            # shortfall notes, "rebuttal plan" debriefs …) written before any
            # client-visible surface existed. The share modal promises staff that
            # internal notes are never shown — only structured facts go out.
            if (field.get("type") or "").strip().lower() == "textarea":
                continue
            value = recorded.get(field["key"])
            if value is None or str(value).strip() == "":
                continue
            fields.append({"label": field.get("label") or field["key"], "value": str(value)[:2000], "type": field.get("type") or "text"})
        if fields:
            out[stage_key] = fields
    return out


def _build_client_portal_payload(db: Session, share: models.EnterpriseClientPortalShare,
                                 client: models.EnterpriseClient, org: models.EnterpriseOrganization) -> dict:
    documents = (
        db.query(models.EnterpriseClientDocument)
        .filter(models.EnterpriseClientDocument.client_id == client.id)
        .order_by(models.EnterpriseClientDocument.created_at.desc())
        .all()
    )
    universities = _client_universities_query(db, org.id, client.id).all()
    payments = (
        db.query(models.EnterpriseStudentPayment)
        .filter(
            models.EnterpriseStudentPayment.organization_id == org.id,
            models.EnterpriseStudentPayment.client_id == client.id,
        )
        .order_by(models.EnterpriseStudentPayment.created_at.desc())
        .all()
    )
    interviews_done = (
        db.query(func.count(models.EnterpriseInterviewSession.id))
        .filter(models.EnterpriseInterviewSession.client_id == client.id)
        .scalar()
    ) or 0

    doc_status = {"valid": "verified", "invalid": "needs_review", "error": None, None: None}
    pay_rows = []
    for p in payments:
        state, state_label = _PORTAL_PAYMENT_STATUS.get(p.status, ("pending", "Payment requested"))
        pay_rows.append({
            "invoice_number": p.invoice_number,
            "description": p.description or "Consultancy fee",
            "amount": round((p.amount_paise or 0) / 100, 2),
            "refunded_amount": round((p.refunded_amount_paise or 0) / 100, 2),
            "currency": p.currency or "INR",
            "status": state,
            "status_label": state_label,
            "due_date": _iso(p.due_date),
            "paid_at": _iso(p.paid_at),
            "created_at": _iso(p.created_at),
        })

    return {
        "organization": {
            "name": org.company_name,
            "logo_url": _resolve_enterprise_logo_url(org),
        },
        "client": {
            "full_name": client.full_name,
            "email": client.email,
            "phone": client.phone,
            "nationality": client.nationality,
            "date_of_birth": _iso(client.date_of_birth),
            # Masked: the portal's only auth anchor is the client's inbox (link + OTP),
            # so a compromised inbox must not yield a full identity kit. Last 3 chars
            # are enough for the client to confirm which passport is on file.
            "passport_number": _mask_passport_number(client.passport_number),
            "passport_expiry": _iso(client.passport_expiry),
            "visa_category_label": _category_label(client.visa_category),
            "destination_country_code": client.destination_country_code,
            "destination_country_name": client.destination_country_name,
            "country": _country_brief(client.destination_country_code),
            "visa_type": client.visa_type,
            "intake": client.intake,
            "application_reference": client.application_reference,
            "target_date": _iso(client.target_date),
            "created_at": _iso(client.created_at),
            "updated_at": _iso(client.updated_at or client.created_at),
        },
        # Normalized to match the stage brief beside it: shipping the raw key would let
        # the student's status pill and their journey track disagree about the same case.
        "status": catalog.normalize_stage(client.status),
        "stage": _stage_brief(client.status, client.destination_country_code),
        "held_from_status": catalog.normalize_stage(client.held_from_status) if getattr(client, "held_from_status", None) else None,
        "held_from_stage": _stage_brief(client.held_from_status, client.destination_country_code) if getattr(client, "held_from_status", None) else None,
        # This portal is scoped to ONE client, so the stepper is worded and sequenced for that
        # client's destination — the same resolution behind the status pill above, which
        # otherwise reads differently from the step it is meant to be pointing at.
        "stages": [
            {k: s[k] for k in ("key", "label", "description", "order", "color", "is_open", "is_terminal")}
            for s in catalog.stages_for(client.destination_country_code)
        ],
        "stage_records": _portal_stage_records(client),
        # The stages this case went past without going through, so the student's track shows a
        # skipped step as skipped instead of ticking it complete — the same list the staff
        # journey renders, or the two views tell contradictory stories. Just the keys: when each
        # internal move happened is not the student's business.
        "stages_skipped": _skipped_stage_keys(client),
        "documents": [
            {
                "document_type": d.document_type,
                "original_filename": d.original_filename,
                "file_size": d.file_size,
                "status": doc_status.get(d.validation_status),
                "uploaded_at": _iso(d.created_at),
            }
            for d in documents
        ],
        "universities": [
            {
                "university_name": u.university_name,
                "program": u.program,
                "location": u.location,
                "status": u.status or "considering",
                "est_tuition": u.est_tuition,
                "application_fee": u.application_fee,
                "qs_world_rank": u.qs_world_rank,
                "country_rank": u.country_rank,
                "admission_difficulty": u.admission_difficulty,
                "key_requirements": _serialize_client_university(u)["key_requirements"],
                "website_url": _safe_external_url(u.website_url),
                "admissions_url": _safe_external_url(u.admissions_url),
                "rationale": u.rationale,
            }
            for u in universities
        ],
        "payments": pay_rows,
        "mock_interviews_completed": int(interviews_done),
    }


# ---- Staff: create / view / revoke the share ------------------------------

@router.post("/clients/{client_id}/portal-share")
def enterprise_create_portal_share(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="portal.share"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    email = (client.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Add an email to this client before sharing their portal.")

    # Supersede any prior shares for this client — one live link at a time.
    db.query(models.EnterpriseClientPortalShare).filter(
        models.EnterpriseClientPortalShare.client_id == client.id,
        models.EnterpriseClientPortalShare.revoked.is_(False),
    ).update({"revoked": True})

    raw_token = generate_verification_token()
    share = models.EnterpriseClientPortalShare(
        organization_id=organization.id,
        client_id=client.id,
        token_hash=hash_token(raw_token),
        email=email,
        expires_at=datetime.utcnow() + timedelta(days=ENTERPRISE_PORTAL_SHARE_EXPIRES_DAYS),
        created_by_user_id=current_user.id,
        created_by_name=current_user.full_name or current_user.email,
    )
    db.add(share)
    db.commit()
    db.refresh(share)

    link = _build_portal_share_url(organization.subdomain_slug, raw_token, request)
    sent, _mid, err = send_enterprise_portal_share_email(
        to_email=email,
        client_name=client.full_name,
        organization_name=organization.company_name,
        portal_url=link,
        destination_country=client.destination_country_name,
        visa_type=client.visa_type,
        logo_url=_absolute_enterprise_logo_url(organization),
    )
    message = (f"Portal access sent to {email}."
               if sent else f"Share created but the email could not be sent right now. {err or ''}".strip())
    return {
        "message": message,
        "email_sent": sent,
        # Returned once for copy/WhatsApp convenience. Opening it still requires the
        # OTP sent to the client's own email, so the link alone grants nothing.
        "link": link,
        "share": _serialize_portal_share_status(share),
        "permissions": _enterprise_permissions_for_role(role),
    }


@router.get("/clients/{client_id}/portal-share")
def enterprise_get_portal_share(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="clients.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    share = _latest_client_portal_share(db, organization.id, client.id)
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "share": _serialize_portal_share_status(share),
    }


@router.post("/clients/{client_id}/portal-share/revoke")
def enterprise_revoke_portal_share(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="portal.share"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    db.query(models.EnterpriseClientPortalShare).filter(
        models.EnterpriseClientPortalShare.client_id == client.id,
        models.EnterpriseClientPortalShare.revoked.is_(False),
    ).update({"revoked": True})
    db.commit()
    return {"message": "Portal access revoked.", "permissions": _enterprise_permissions_for_role(role)}


# ---- Public (client-facing, token-scoped, no staff auth) ------------------

@router.get("/public/portal/{token}")
def public_portal_info(token: str, db: Session = Depends(get_db)):
    share = _public_portal_share_or_404(db, token)
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == share.client_id).first()
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == share.organization_id).first()
    if not client or not org:
        raise HTTPException(status_code=404, detail="This portal link is no longer available.")
    return {
        "organization_name": org.company_name,
        "logo_url": _resolve_enterprise_logo_url(org),
        "client_first_name": (client.full_name or "there").split(" ")[0],
        "destination_country": client.destination_country_name,
        "visa_type": client.visa_type,
        "masked_email": _mask_email(share.email),
    }


@router.post("/public/portal/{token}/send-code")
def public_portal_send_code(token: str, request: Request, db: Session = Depends(get_db)):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.portal_code",
        limit=ENTERPRISE_CODE_RATE_LIMIT, window_seconds=ENTERPRISE_CODE_RATE_WINDOW, extra_key=hash_token(token)[:16],
    )
    share = _public_portal_share_or_404(db, token)
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == share.organization_id).first()
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == share.client_id).first()
    if not org or not client:
        raise HTTPException(status_code=404, detail="This portal link is no longer available.")

    code = f"{_secrets.randbelow(900000) + 100000:06d}"
    share.code_hash = hash_token(code)
    share.code_expires_at = datetime.utcnow() + timedelta(minutes=ENTERPRISE_INTERVIEW_CODE_EXPIRES_MIN)
    share.code_attempts = 0
    db.commit()

    sent, _mid, err = send_enterprise_portal_code_email(
        to_email=share.email, client_name=client.full_name, organization_name=org.company_name, code=code,
    )
    if not sent:
        logger.warning("Portal code email failed for share %s: %s", share.id, err)
    return {"sent": bool(sent), "masked_email": _mask_email(share.email)}


@router.post("/public/portal/{token}/verify")
def public_portal_verify(token: str, payload: PublicPortalVerifyRequest, request: Request, db: Session = Depends(get_db)):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.portal_verify",
        limit=ENTERPRISE_PUBLIC_RATE_LIMIT, window_seconds=ENTERPRISE_PUBLIC_RATE_WINDOW, extra_key=hash_token(token)[:16],
    )
    share = _public_portal_share_or_404(db, token)
    if not share.code_hash or not share.code_expires_at:
        raise HTTPException(status_code=400, detail="Please request a verification code first.")
    code_exp = share.code_expires_at.replace(tzinfo=None) if getattr(share.code_expires_at, "tzinfo", None) else share.code_expires_at
    if code_exp < datetime.utcnow():
        raise HTTPException(status_code=400, detail="That code has expired. Please request a new one.")
    if int(share.code_attempts or 0) >= ENTERPRISE_INTERVIEW_CODE_MAX_ATTEMPTS:
        raise HTTPException(status_code=429, detail="Too many attempts. Please request a new code.")

    share.code_attempts = int(share.code_attempts or 0) + 1
    if hash_token((payload.code or "").strip()) != share.code_hash:
        db.commit()
        raise HTTPException(status_code=400, detail="That code is incorrect. Please try again.")

    # Verified — consume the code and issue a short-lived read-only session token.
    share.code_hash = None
    share.code_expires_at = None
    db.commit()
    return {
        "session_token": _issue_portal_session_token(share.id),
        "session_hours": ENTERPRISE_PORTAL_SESSION_HOURS,
    }


@router.post("/public/portal/data")
def public_portal_data(payload: PublicPortalDataRequest, request: Request, db: Session = Depends(get_db)):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.portal_data",
        limit=ENTERPRISE_PUBLIC_RATE_LIMIT, window_seconds=ENTERPRISE_PUBLIC_RATE_WINDOW,
    )
    share, client, org = _public_load_portal_context(db, payload.session_token)
    # Atomic SQL increment — concurrent portal loads must not lose updates.
    db.query(models.EnterpriseClientPortalShare).filter(
        models.EnterpriseClientPortalShare.id == share.id
    ).update({
        "last_opened_at": datetime.utcnow(),
        "open_count": models.EnterpriseClientPortalShare.open_count + 1,
    }, synchronize_session=False)
    db.commit()
    return _build_client_portal_payload(db, share, client, org)


# ===========================================================================
# Copilot client invite — the client's own Copilot chat about their case
#
# Staff share the org's AI copilot with a client as a secure emailed link
# (page served at /assist/{token}). The client verifies an OTP sent to their
# own email, then chats about THEIR case only. Security model mirrors
# interview invites / portal shares (hashed capability token + OTP + short-
# lived signed session token, scope "ent_copilot", re-checked per request).
# Billing is a flat per-client unlock: the org wallet is charged once when
# the client first verifies (never for links that are ignored), and the
# invite's message counters cap total usage — the staff copilot's per-message
# meter is deliberately NOT shared with this surface. Context never includes
# staff notes; the client only sees their own CRM profile and documents.
# ===========================================================================

ENTERPRISE_COPILOT_INVITE_EXPIRES_DAYS = int(os.getenv("ENTERPRISE_COPILOT_INVITE_EXPIRES_DAYS", "30"))
ENTERPRISE_COPILOT_SESSION_HOURS = int(os.getenv("ENTERPRISE_COPILOT_SESSION_HOURS", "24"))
ENTERPRISE_COPILOT_INVITE_MESSAGES = int(os.getenv("ENTERPRISE_COPILOT_INVITE_MESSAGES", "100"))
# Wrong guesses allowed across the WHOLE life of one link, however many codes are
# sent. The per-code cap resets on every resend and the rate limits key on the
# caller's IP, so without this a leaked link is a 30-day brute-force budget that
# only grows with the attacker's address pool. 30 guesses against a 900k code
# space is a ~0.003% chance; a client who genuinely fumbles asks staff to resend.
ENTERPRISE_COPILOT_MAX_TOTAL_CODE_ATTEMPTS = int(os.getenv("ENTERPRISE_COPILOT_MAX_TOTAL_CODE_ATTEMPTS", "30"))
COPILOT_CLIENT_ACTION_KEY = "copilot_client"


class PublicCopilotVerifyRequest(BaseModel):
    code: str = Field(..., min_length=4, max_length=10)


class PublicCopilotSessionRequest(BaseModel):
    session_token: str = Field(..., min_length=10, max_length=4000)


class PublicCopilotTurn(BaseModel):
    role: str = Field(..., max_length=16)
    # Generous: our own replies land back here and are not length-capped at
    # generation time. Rejecting a stored reply would 422 every later turn
    # (a bricked session). The engine budgets history to 4k/turn anyway.
    content: str = Field(..., max_length=20000)


class PublicCopilotChatRequest(BaseModel):
    session_token: str = Field(..., min_length=10, max_length=4000)
    message: str = Field(..., min_length=1, max_length=8000)
    # Capped at the same 200 turns the page keeps and the engine budgets. Without a
    # bound, an unauthenticated caller could post a huge array: validation runs
    # before the endpoint body, so it lands before the session check or rate limit.
    history: Optional[List[PublicCopilotTurn]] = Field(None, max_length=200)


def _build_copilot_invite_url(subdomain_slug, token: str, request: Request | None) -> str:
    subdomain = str(subdomain_slug or "").strip().lower()
    base = None
    if subdomain:
        host = f"{subdomain}.{ENTERPRISE_ROOT_DOMAIN}"
        port = _request_port_for_local_enterprise_url(request)
        if port:
            host = f"{host}:{port}"
        base = f"{ENTERPRISE_PORTAL_SCHEME}://{host}"
    if not base:
        base = ENTERPRISE_PASSWORD_SETUP_BASE_URL
    # /assist because bare /copilot already serves the B2C SPA (main.py).
    return f"{base.rstrip('/')}/assist/{token}"


def _copilot_invite_is_live(invite: models.EnterpriseCopilotInvite) -> bool:
    if invite.revoked:
        return False
    if invite.expires_at:
        exp = invite.expires_at.replace(tzinfo=None) if getattr(invite.expires_at, "tzinfo", None) else invite.expires_at
        if exp < datetime.utcnow():
            return False
    return True


def _copilot_invite_remaining(invite: models.EnterpriseCopilotInvite) -> int:
    return max(0, int(invite.allowed_messages or 0) - int(invite.used_messages or 0))


def _serialize_copilot_invite_status(invite: models.EnterpriseCopilotInvite | None) -> Optional[dict]:
    if not invite:
        return None
    return {
        "id": invite.id,
        "email": invite.email,
        "allowed_messages": int(invite.allowed_messages or 0),
        "used_messages": int(invite.used_messages or 0),
        "remaining_messages": _copilot_invite_remaining(invite),
        "unlocked": bool(invite.unlocked_at),
        "unlocked_at": _iso(invite.unlocked_at),
        "last_message_at": _iso(invite.last_message_at),
        "revoked": bool(invite.revoked),
        "live": _copilot_invite_is_live(invite),
        "created_by_name": invite.created_by_name,
        "created_at": _iso(invite.created_at),
        "expires_at": _iso(invite.expires_at),
    }


def _copilot_invite_config() -> dict:
    """What a NEW invite would cost/allow — shown in the staff send modal."""
    return {
        "allowed_messages": ENTERPRISE_COPILOT_INVITE_MESSAGES,
        "expires_days": ENTERPRISE_COPILOT_INVITE_EXPIRES_DAYS,
        "cost_credits": credits.action_cost(COPILOT_CLIENT_ACTION_KEY),
    }


def _latest_client_copilot_invite(db: Session, organization_id: int, client_id: int):
    return (
        db.query(models.EnterpriseCopilotInvite)
        .filter(
            models.EnterpriseCopilotInvite.organization_id == int(organization_id),
            models.EnterpriseCopilotInvite.client_id == int(client_id),
        )
        .order_by(models.EnterpriseCopilotInvite.created_at.desc(), models.EnterpriseCopilotInvite.id.desc())
        .first()
    )


def _issue_copilot_session_token(invite_id: int) -> str:
    return create_access_token(
        data={"sub": f"entcp:{int(invite_id)}", "scope": "ent_copilot", "cp": int(invite_id)},
        expires_delta=timedelta(hours=ENTERPRISE_COPILOT_SESSION_HOURS),
    )


def _decode_copilot_session_token(token: str) -> int:
    try:
        payload = jose_jwt.decode(token, AUTH_SECRET_KEY, algorithms=[AUTH_ALGORITHM])
    except JWTError:
        raise HTTPException(status_code=401, detail="Your copilot session has expired. Please verify your email again.")
    if payload.get("scope") != "ent_copilot" or not payload.get("cp"):
        raise HTTPException(status_code=401, detail="Invalid copilot session.")
    return int(payload["cp"])


def _public_copilot_invite_or_404(db: Session, token: str) -> models.EnterpriseCopilotInvite:
    token_hash = hash_token((token or "").strip())
    invite = (
        db.query(models.EnterpriseCopilotInvite)
        .filter(models.EnterpriseCopilotInvite.token_hash == token_hash)
        .first()
    )
    if not invite or not _copilot_invite_is_live(invite):
        raise HTTPException(status_code=404, detail="This copilot link is invalid or has expired.")
    return invite


def _public_load_copilot_context(db: Session, session_token: str):
    invite_id = _decode_copilot_session_token(session_token)
    invite = db.query(models.EnterpriseCopilotInvite).filter(models.EnterpriseCopilotInvite.id == invite_id).first()
    if not invite or not _copilot_invite_is_live(invite):
        # 410, not 401: the session token is fine — the LINK is dead (staff revoked
        # it, or it passed its expiry). Answering 401 made the page offer a re-verify
        # that could never succeed; the client needs to be told to ask for a new link.
        raise HTTPException(status_code=410, detail="This copilot link is no longer active.")
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == invite.client_id).first()
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == invite.organization_id).first()
    if not client or not org:
        raise HTTPException(status_code=404, detail="This copilot is no longer available.")
    return invite, client, org


# ---- Staff endpoints ------------------------------------------------------

@router.post("/clients/{client_id}/copilot/invite")
def enterprise_create_copilot_invite(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="copilot.invite"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    email = (client.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Add an email to this client before sharing their copilot.")

    # Serialize sends for this client before reading anything. Two of them at once
    # (a double-click, or two staff on the same case) each read "no live invite",
    # each revoke the other's row and each insert — leaving TWO live links, both
    # chargeable at verify. The client row is the only row that always exists to
    # lock, so it is the gate; locking the invite rows alone can't help when there
    # are none yet. (No-op on SQLite, which serializes writes anyway.)
    db.query(models.EnterpriseClient).filter(
        models.EnterpriseClient.id == client.id
    ).with_for_update().first()
    # Then lock this client's live invites, which orders us against a client's
    # in-flight FIRST verify: its unlock UPDATE waits on these locks, so carry_over
    # below is decided on a settled row instead of a stale "not yet paid" read that
    # would revoke the just-paid invite and charge the org a second time.
    db.query(models.EnterpriseCopilotInvite).filter(
        models.EnterpriseCopilotInvite.client_id == client.id,
        models.EnterpriseCopilotInvite.revoked.is_(False),
    ).with_for_update().all()

    # "Charged once per client" is the promise (price-list copy + send modal):
    # a resend while a PAID window is still live must carry the entitlement
    # over — same unlock, same counters, same expiry — not restart a fresh
    # 30-day window that would charge the org a second time at verify.
    prior = _latest_client_copilot_invite(db, organization.id, client.id)
    carry_over = bool(prior and prior.unlocked_at and _copilot_invite_is_live(prior))

    if not carry_over:
        # Affordability gate only — the flat unlock is charged when the client
        # first verifies, so an ignored link never costs the org anything.
        credits.enforce_units_or_402(db, organization.id, COPILOT_CLIENT_ACTION_KEY, 1)

    # Supersede any prior invites for this client — one live link at a time.
    db.query(models.EnterpriseCopilotInvite).filter(
        models.EnterpriseCopilotInvite.client_id == client.id,
        models.EnterpriseCopilotInvite.revoked.is_(False),
    ).update({"revoked": True})

    raw_token = generate_verification_token()
    invite = models.EnterpriseCopilotInvite(
        organization_id=organization.id,
        client_id=client.id,
        token_hash=hash_token(raw_token),
        email=email,
        allowed_messages=(prior.allowed_messages if carry_over else ENTERPRISE_COPILOT_INVITE_MESSAGES),
        used_messages=(prior.used_messages if carry_over else 0),
        unlocked_at=(prior.unlocked_at if carry_over else None),
        credits_charged=(prior.credits_charged if carry_over else 0),
        expires_at=(prior.expires_at if carry_over
                    else datetime.utcnow() + timedelta(days=ENTERPRISE_COPILOT_INVITE_EXPIRES_DAYS)),
        created_by_user_id=current_user.id,
        created_by_name=current_user.full_name or current_user.email,
    )
    db.add(invite)
    db.commit()
    db.refresh(invite)

    link = _build_copilot_invite_url(organization.subdomain_slug, raw_token, request)
    sent, _mid, err = send_enterprise_copilot_invite_email(
        to_email=email,
        client_name=client.full_name,
        organization_name=organization.company_name,
        copilot_url=link,
        destination_country=client.destination_country_name,
        visa_type=client.visa_type,
        logo_url=_absolute_enterprise_logo_url(organization),
    )
    message = (f"Copilot access sent to {email}."
               if sent else f"Invite created but the email could not be sent right now. {err or ''}".strip())
    return {
        "message": message,
        "email_sent": sent,
        # Returned once for copy/WhatsApp convenience. Chatting still requires the OTP
        # sent to the client's own email, so the link alone grants no case access — it
        # does render the greeting card (consultancy, first name, destination), which
        # is why it goes to the client and nobody else.
        "link": link,
        "invite": _serialize_copilot_invite_status(invite),
        "config": _copilot_invite_config(),
        "permissions": _enterprise_permissions_for_role(role),
    }


@router.get("/clients/{client_id}/copilot/invite")
def enterprise_get_copilot_invite(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="clients.view")
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    invite = _latest_client_copilot_invite(db, organization.id, client.id)
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "invite": _serialize_copilot_invite_status(invite),
        "config": _copilot_invite_config(),
    }


@router.post("/clients/{client_id}/copilot/invite/revoke")
def enterprise_revoke_copilot_invite(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="copilot.invite"
    )
    client = _get_org_client_or_404(db, organization.id, client_id, ctx=role.ctx)
    # No refund: an un-activated invite was never charged, and an activated one
    # was consumed (mirrors interview invites — nothing is escrowed).
    db.query(models.EnterpriseCopilotInvite).filter(
        models.EnterpriseCopilotInvite.client_id == client.id,
        models.EnterpriseCopilotInvite.revoked.is_(False),
    ).update({"revoked": True})
    db.commit()
    return {"message": "Copilot access revoked.", "permissions": _enterprise_permissions_for_role(role)}


# ---- Public (client-facing, token-scoped, no staff auth) ------------------

@router.get("/public/copilot/{token}")
def public_copilot_info(token: str, db: Session = Depends(get_db)):
    invite = _public_copilot_invite_or_404(db, token)
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == invite.client_id).first()
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == invite.organization_id).first()
    if not client or not org:
        raise HTTPException(status_code=404, detail="This copilot link is no longer available.")
    remaining = _copilot_invite_remaining(invite)
    return {
        "organization_name": org.company_name,
        "logo_url": _resolve_enterprise_logo_url(org),
        "client_first_name": (client.full_name or "there").split(" ")[0],
        "destination_country": client.destination_country_name,
        "visa_type": client.visa_type,
        "masked_email": _mask_email(invite.email),
        "remaining_messages": remaining,
        "exhausted": remaining <= 0,
        "unlocked": bool(invite.unlocked_at),
        "expires_at": _iso(invite.expires_at),
    }


@router.post("/public/copilot/{token}/send-code")
def public_copilot_send_code(token: str, request: Request, db: Session = Depends(get_db)):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.copilot_code",
        limit=ENTERPRISE_CODE_RATE_LIMIT, window_seconds=ENTERPRISE_CODE_RATE_WINDOW, extra_key=hash_token(token)[:16],
    )
    invite = _public_copilot_invite_or_404(db, token)
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == invite.organization_id).first()
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == invite.client_id).first()
    if not org or not client:
        raise HTTPException(status_code=404, detail="This copilot link is no longer available.")

    # The lifetime budget is checked HERE too, not just at verify: once it's spent
    # there is no code worth emailing, and refusing at send tells the client (and
    # the consultancy) to reissue the link rather than silently mailing dead codes.
    if int(invite.code_attempts_total or 0) >= ENTERPRISE_COPILOT_MAX_TOTAL_CODE_ATTEMPTS:
        raise HTTPException(
            status_code=429,
            detail="Too many verification attempts on this link. Please ask your consultancy to send you a new one.",
        )

    code = f"{_secrets.randbelow(900000) + 100000:06d}"
    invite.code_hash = hash_token(code)
    invite.code_expires_at = datetime.utcnow() + timedelta(minutes=ENTERPRISE_INTERVIEW_CODE_EXPIRES_MIN)
    invite.code_attempts = 0
    db.commit()

    sent, _mid, err = send_enterprise_copilot_code_email(
        to_email=invite.email, client_name=client.full_name, organization_name=org.company_name, code=code,
    )
    if not sent:
        logger.warning("Copilot code email failed for invite %s: %s", invite.id, err)
    return {"sent": bool(sent), "masked_email": _mask_email(invite.email)}


@router.post("/public/copilot/{token}/verify")
def public_copilot_verify(token: str, payload: PublicCopilotVerifyRequest, request: Request, db: Session = Depends(get_db)):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.copilot_verify",
        limit=ENTERPRISE_PUBLIC_RATE_LIMIT, window_seconds=ENTERPRISE_PUBLIC_RATE_WINDOW, extra_key=hash_token(token)[:16],
    )
    invite = _public_copilot_invite_or_404(db, token)
    client = db.query(models.EnterpriseClient).filter(models.EnterpriseClient.id == invite.client_id).first()
    org = db.query(models.EnterpriseOrganization).filter(models.EnterpriseOrganization.id == invite.organization_id).first()
    if not client or not org:
        raise HTTPException(status_code=404, detail="This copilot link is no longer available.")
    if not invite.code_hash or not invite.code_expires_at:
        raise HTTPException(status_code=400, detail="Please request a verification code first.")
    code_exp = invite.code_expires_at.replace(tzinfo=None) if getattr(invite.code_expires_at, "tzinfo", None) else invite.code_expires_at
    if code_exp < datetime.utcnow():
        raise HTTPException(status_code=400, detail="That code has expired. Please request a new one.")
    # Burn the attempt ATOMICALLY, and only while budget remains. Read-then-write let
    # a parallel burst all read the same count, all pass the cap and each take a free
    # guess (the rate limits key on the caller's IP, so an address pool multiplies
    # it). The conditional UPDATE is the real cap — same shape as the message
    # reservation below. Both counters move together: the per-code one, and the
    # lifetime one a resend cannot reset.
    attempts_col = models.EnterpriseCopilotInvite.code_attempts
    total_col = models.EnterpriseCopilotInvite.code_attempts_total
    bumped = db.query(models.EnterpriseCopilotInvite).filter(
        models.EnterpriseCopilotInvite.id == invite.id,
        func.coalesce(attempts_col, 0) < ENTERPRISE_INTERVIEW_CODE_MAX_ATTEMPTS,
        func.coalesce(total_col, 0) < ENTERPRISE_COPILOT_MAX_TOTAL_CODE_ATTEMPTS,
    ).update(
        {
            "code_attempts": func.coalesce(attempts_col, 0) + 1,
            "code_attempts_total": func.coalesce(total_col, 0) + 1,
        },
        synchronize_session=False,
    )
    db.commit()
    if not bumped:
        db.expire(invite)
        if int(invite.code_attempts_total or 0) >= ENTERPRISE_COPILOT_MAX_TOTAL_CODE_ATTEMPTS:
            raise HTTPException(
                status_code=429,
                detail="Too many verification attempts on this link. Please ask your consultancy to send you a new one.",
            )
        raise HTTPException(status_code=429, detail="Too many attempts. Please request a new code.")

    # Constant-time compare via the project's helper, not `!=` on the digests.
    if not token_matches((payload.code or "").strip(), invite.code_hash):
        raise HTTPException(status_code=400, detail="That code is incorrect. Please try again.")

    # First successful verify = activation: charge the flat unlock. The code is
    # NOT consumed on a broke wallet (only the attempt is), so the client can
    # retry the same code once the consultancy tops up. Client-safe wording —
    # never mention credits or the org's wallet to the client.
    if not invite.unlocked_at and not credits.can_afford(db, invite.organization_id, COPILOT_CLIENT_ACTION_KEY):
        # The attempt is already committed above; nothing else to persist here.
        raise HTTPException(status_code=402, detail="Your copilot isn't available right now. Please contact your consultancy.")

    # Verified — consume the code (single-use) and issue the session token.
    invite.code_hash = None
    invite.code_expires_at = None
    if not invite.unlocked_at:
        # Claim the unlock ATOMICALLY: a conditional UPDATE means concurrent
        # verifies of the same code (two tabs, a scripted burst) or a verify
        # racing a staff revoke can never charge twice — charge_action's row
        # lock serializes wallet math but does not deduplicate unlocks.
        claimed = db.query(models.EnterpriseCopilotInvite).filter(
            models.EnterpriseCopilotInvite.id == invite.id,
            models.EnterpriseCopilotInvite.unlocked_at.is_(None),
            models.EnterpriseCopilotInvite.revoked.is_(False),
        ).update({
            "unlocked_at": datetime.utcnow(),
            "credits_charged": int(credits.action_cost(COPILOT_CLIENT_ACTION_KEY) or 0),
        }, synchronize_session=False)
        if claimed:
            try:
                # commit=False: the debit, ledger row and unlock land in ONE commit.
                credits.charge_action(
                    db, invite.organization_id, COPILOT_CLIENT_ACTION_KEY,
                    reference_type="client", reference_id=invite.client_id,
                    description=f"Client copilot access — {client.full_name}",
                    commit=False,
                )
            except HTTPException:
                # Wallet drained between can_afford and the locked debit: roll the
                # claim back (invite stays locked-but-unpaid otherwise) and keep
                # the staff-facing top-up copy away from the client.
                db.rollback()
                raise HTTPException(status_code=402, detail="Your copilot isn't available right now. Please contact your consultancy.")
        else:
            # Another request claimed the unlock, or staff revoked mid-verify.
            db.expire(invite)
            if not _copilot_invite_is_live(invite):
                db.rollback()
                raise HTTPException(status_code=401, detail="This copilot link is no longer active.")
            # Already unlocked by the concurrent verify — proceed without charging,
            # but re-consume the code in case the expire reloaded it.
            invite.code_hash = None
            invite.code_expires_at = None
    db.commit()
    return {
        "session_token": _issue_copilot_session_token(invite.id),
        "session_hours": ENTERPRISE_COPILOT_SESSION_HOURS,
        "client_first_name": (client.full_name or "there").split(" ")[0],
        "destination_country": client.destination_country_name,
        "visa_type": client.visa_type,
        "remaining_messages": _copilot_invite_remaining(invite),
    }


@router.post("/public/copilot/session")
def public_copilot_session(payload: PublicCopilotSessionRequest, request: Request, db: Session = Depends(get_db)):
    """Session probe: lets the page restore a stored session without re-verifying.
    Re-checks invite liveness (revoke/expiry take effect immediately)."""
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.copilot_public",
        limit=ENTERPRISE_PUBLIC_RATE_LIMIT, window_seconds=ENTERPRISE_PUBLIC_RATE_WINDOW,
    )
    invite, client, org = _public_load_copilot_context(db, payload.session_token)
    remaining = _copilot_invite_remaining(invite)
    return {
        "organization_name": org.company_name,
        "logo_url": _resolve_enterprise_logo_url(org),
        "client_first_name": (client.full_name or "there").split(" ")[0],
        "destination_country": client.destination_country_name,
        "visa_type": client.visa_type,
        "remaining_messages": remaining,
        "exhausted": remaining <= 0,
    }


@router.post("/public/copilot/chat")
def public_copilot_chat(payload: PublicCopilotChatRequest, request: Request, db: Session = Depends(get_db)):
    """One client copilot turn. Stateless like the staff extension surface — the
    page resends its history each turn; nothing is stored server-side.

    Order is load-bearing: liveness → provider → free guardrail → RESERVE a
    capped message atomically (so parallel requests can't all pass a stale cap
    check and burst past allowed_messages) → generate → release the reservation
    if the model fails (a failed call must not consume the client's messages)."""
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.copilot_public",
        limit=ENTERPRISE_PUBLIC_RATE_LIMIT, window_seconds=ENTERPRISE_PUBLIC_RATE_WINDOW,
    )
    invite, client, org = _public_load_copilot_context(db, payload.session_token)

    if not enterprise_copilot.is_provider_available():
        raise HTTPException(status_code=503, detail="Your copilot is temporarily unavailable. Please try again shortly.")
    if not invite.unlocked_at:
        # Session tokens are only minted after verify (which unlocks), so this is
        # a defensive guard — e.g. staff re-sent a link mid-session.
        raise HTTPException(status_code=401, detail="This copilot link is no longer active.")

    remaining = _copilot_invite_remaining(invite)
    if remaining <= 0:
        raise HTTPException(status_code=403, detail="You've used all the messages included with your copilot. Please contact your consultancy.")

    message = (payload.message or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="Please type a message.")
    history = [{"role": t.role, "content": t.content} for t in (payload.history or [])]

    # Free off-topic guardrail — refusals never consume a capped message. The
    # newest user history turn is screened TOGETHER with the message (not
    # instead of it) so a short "continue" can't smuggle an off-topic ask,
    # while an on-topic follow-up after a refused turn isn't re-refused: the
    # guardrail's on-topic whitelist wins over off-topic terms, so the new
    # message's own topicality always counts.
    latest_user_turn = ""
    for turn in reversed(history):
        if (turn.get("role") or "") != "assistant":
            latest_user_turn = str(turn.get("content") or "")
            break
    if ai_guardrails.is_off_topic(message) or (
        len(message) < 200 and latest_user_turn
        and ai_guardrails.is_off_topic(f"{message}\n{latest_user_turn}")
    ):
        ai_guardrails.record_block(source=enterprise_copilot.CLIENT_USAGE_SOURCE, detail="enterprise_client_link")
        return {"reply": ai_guardrails.OFF_TOPIC_REFUSAL, "remaining_messages": remaining}

    # Reserve the message BEFORE the model call: a conditional atomic increment
    # is the actual cap — the read-only check above is just a fast client-safe
    # error. Without this, N parallel requests all see remaining==1 and each
    # gets a full (Rilono-funded) generation.
    reserved = db.query(models.EnterpriseCopilotInvite).filter(
        models.EnterpriseCopilotInvite.id == invite.id,
        models.EnterpriseCopilotInvite.used_messages < models.EnterpriseCopilotInvite.allowed_messages,
    ).update({
        "used_messages": models.EnterpriseCopilotInvite.used_messages + 1,
        "last_message_at": datetime.utcnow(),
    }, synchronize_session=False)
    if not reserved:
        db.rollback()
        raise HTTPException(status_code=403, detail="You've used all the messages included with your copilot. Please contact your consultancy.")
    db.commit()

    try:
        turn_result = enterprise_copilot.run_enterprise_copilot_client_chat(
            db,
            organization=org,
            client=client,
            message=message,
            conversation_history=history,
        )
    except Exception:
        # Release the reservation — a failed model call must not consume one of
        # the client's capped messages (house invariant: never charge for a
        # failed AI action). Floor at 0 defensively.
        try:
            db.query(models.EnterpriseCopilotInvite).filter(
                models.EnterpriseCopilotInvite.id == invite.id,
                models.EnterpriseCopilotInvite.used_messages > 0,
            ).update({
                "used_messages": models.EnterpriseCopilotInvite.used_messages - 1,
            }, synchronize_session=False)
            db.commit()
        except Exception:
            logger.exception("Failed to release copilot message reservation (invite_id=%s)", invite.id)
        logger.exception("Client copilot chat failed (org_id=%s client_id=%s)", org.id, client.id)
        raise HTTPException(status_code=502, detail="Your copilot ran into a problem answering that. Please try again.")

    # Fresh remaining after the pre-reserved increment.
    db.expire(invite)
    return {"reply": turn_result.answer, "remaining_messages": _copilot_invite_remaining(invite)}


# ===========================================================================
# Course Finder (workspace section)
#
# Browse Rilono's shared universities/courses catalog (maintained by the
# background course-catalog agent — fees, intakes, deadlines, score cutoffs,
# each row stamped last_verified_at) and generate billed AI course shortlists,
# optionally personalized to one of the org's clients. Browsing is FREE (pure
# DB reads); only the AI shortlist debits the credits wallet. Results are
# persisted (like Deep Scans) so a paid shortlist can never be lost.
# ===========================================================================

COURSE_FINDER_ACTION_KEY = "course_finder"
# Browse page size. Deliberately independent of the agent's per-country target: the
# catalog holds ~1.5k universities per country, and one response must stay small.
COURSE_CATALOG_BROWSE_PAGE_SIZE = max(10, min(200, int(os.getenv("COURSE_CATALOG_BROWSE_PAGE_SIZE", "60") or "60")))


def _serialize_course_finder_rec(row: models.EnterpriseCourseFinderRec, *, include_items: bool = True) -> dict:
    try:
        items = json.loads(row.recommendations) if row.recommendations else []
        if not isinstance(items, list):
            items = []
    except Exception:
        items = []
    try:
        query = json.loads(row.query) if row.query else {}
        if not isinstance(query, dict):
            query = {}
    except Exception:
        query = {}
    data = {
        "id": int(row.id),
        "client_id": int(row.client_id) if row.client_id else None,
        "client_name": row.client_name,
        "country_code": row.country_code,
        "degree_level": row.degree_level,
        "discipline": row.discipline,
        "query": query,
        "summary": row.summary,
        "catalog_based": bool(row.catalog_based),
        "grounded": bool(row.grounded),
        "credits_charged": int(row.credits_charged or 0),
        "created_by_name": row.created_by_name,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "count": len(items),
    }
    if include_items:
        data["recommendations"] = items
    return data


@router.get("/course-catalog/meta")
def enterprise_course_catalog_meta(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Filter options + per-country catalog stats for the Course Finder section."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="coursefinder.view")
    from app import course_catalog

    stats = course_catalog.catalog_stats(db)
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "countries": stats["countries"],
        "disciplines": course_catalog.DISCIPLINES,
        "degree_levels": course_catalog.DEGREE_LEVELS,
        # Advanced-filter vocabularies. Served rather than duplicated in the frontend so
        # every option the UI offers is an option the browse query is known to accept.
        "intakes": course_catalog.INTAKE_BUCKETS,
        "sorts": course_catalog.CATALOG_SORTS,
        "university_types": course_catalog.UNIVERSITY_TYPES,
        "gre_filters": course_catalog.GRE_FILTERS,
        "cost_credits": credits.action_cost(COURSE_FINDER_ACTION_KEY),
        "ai_available": course_catalog.ai_available(),
    }


@router.get("/course-catalog")
def enterprise_course_catalog_browse(
    request: Request,
    country: str,
    level: Optional[str] = None,
    discipline: Optional[str] = None,
    q: Optional[str] = None,
    max_tuition: Optional[int] = None,
    # --- advanced filters (all optional; validated + clamped by course_catalog) ---
    min_tuition: Optional[int] = None,
    require_tuition: Optional[bool] = None,
    max_ielts: Optional[float] = None,
    max_toefl: Optional[int] = None,
    tests_include_unknown: Optional[bool] = None,
    gre: Optional[str] = None,
    intake: Optional[str] = None,
    max_duration: Optional[int] = None,
    no_app_fee: Optional[bool] = None,
    has_deadline: Optional[bool] = None,
    max_qs_rank: Optional[int] = None,
    uni_type: Optional[str] = None,
    city: Optional[str] = None,
    verified_only: Optional[bool] = None,
    scholarships_only: Optional[bool] = None,
    sort: Optional[str] = None,
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """FREE catalog browse — universities with their matching courses (paged)."""
    _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="coursefinder.view")
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.course_catalog_browse",
        limit=120, window_seconds=600, extra_key=str(current_user.id),
    )
    from app import course_catalog

    code = (country or "").strip().upper()
    if not course_catalog.country_name(code):
        raise HTTPException(status_code=400, detail="Unknown destination country.")
    safe_max_tuition = None
    if max_tuition is not None:
        try:
            safe_max_tuition = max(0, min(int(max_tuition), 10_000_000)) or None
        except Exception:
            safe_max_tuition = None
    # The catalog runs to ~1.5k universities per country, so the page size is fixed
    # and the tail is reached by paging — it must NOT track the agent's target, which
    # would put the entire country in one response.
    page_size = COURSE_CATALOG_BROWSE_PAGE_SIZE
    safe_offset = max(0, min(int(offset or 0), 100_000))
    # One validation gate for the deep filters: unknown enum values and out-of-range
    # numbers are dropped (broader search) instead of 400-ing a stale bookmark.
    advanced = course_catalog.normalize_catalog_filters({
        "min_tuition": min_tuition,
        "require_tuition": require_tuition,
        "max_ielts": max_ielts,
        "max_toefl": max_toefl,
        "tests_include_unknown": tests_include_unknown,
        "gre": gre,
        "intake": intake,
        "max_duration_months": max_duration,
        "no_app_fee": no_app_fee,
        "has_deadline": has_deadline,
        "max_qs_rank": max_qs_rank,
        "university_type": uni_type,
        "city": city,
        "verified_only": verified_only,
        "scholarships_only": scholarships_only,
        "sort": sort,
    })

    rows, total = course_catalog.query_catalog(
        db,
        country_code=code,
        degree_level=(level or "").strip().lower() or None,
        discipline=(discipline or "").strip() or None,
        q=(q or "").strip() or None,
        max_tuition=safe_max_tuition,
        advanced=advanced,
        limit_universities=page_size,
        offset_universities=safe_offset,
    )
    return {
        "country_code": code,
        "universities": [course_catalog.serialize_university(u, courses) for u, courses in rows],
        "total_universities": total,          # all matches, across pages
        "page_courses": sum(len(courses) for _u, courses in rows),   # this page only
        "offset": safe_offset,
        "page_size": page_size,
        "has_more": (safe_offset + len(rows)) < total,
        # Echo of the filters as the server understood them (post validation/clamping) —
        # the one place to look when a result set doesn't match what the UI is showing.
        "applied_filters": advanced,
        "sort": advanced.get("sort") or "rank",
    }


class EnterpriseCourseFinderRecommend(BaseModel):
    client_id: Optional[int] = None
    country_code: Optional[str] = Field(None, max_length=8)
    degree_level: Optional[str] = Field(None, max_length=20)
    discipline: Optional[str] = Field(None, max_length=80)
    field_of_study: Optional[str] = Field(None, max_length=120)
    budget: Optional[str] = Field(None, max_length=60)
    notes: Optional[str] = Field(None, max_length=300)
    max_results: int = Field(6, ge=1, le=8)


@router.post("/course-finder/recommend")
def enterprise_course_finder_recommend(
    payload: EnterpriseCourseFinderRecommend,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Billed AI course shortlist from the verified catalog (grounded live-search
    fallback while a destination is still seeding). Charge choreography mirrors
    Deep Scan: rate-limit → wallet pre-check → generate → fail WITHOUT charging →
    persist result + debit atomically."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability=("ai.coursefinder", "credits.spend")
    )
    client = None
    if payload.client_id:
        client = _get_org_client_or_404(db, organization.id, payload.client_id, ctx=role.ctx)

    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.course_finder",
        limit=20, window_seconds=600, extra_key=str(current_user.id),
    )

    from app import course_catalog

    if not course_catalog.ai_available():
        raise HTTPException(status_code=503, detail="AI recommendations are not configured.")

    code = (payload.country_code or "").strip().upper()
    if client is not None and not code:
        code = (client.destination_country_code or "").strip().upper()
    if not course_catalog.country_name(code):
        raise HTTPException(
            status_code=400,
            detail="Pick a destination country (or set one on the client) first.",
        )

    field_query = (payload.field_of_study or "").strip()
    if not field_query and not (payload.discipline or "").strip():
        raise HTTPException(status_code=400, detail="Tell Rilono AI the field of study to shortlist for.")

    # Hard-block a broke wallet BEFORE spending any Gemini tokens.
    credits.enforce_action_or_402(db, organization.id, COURSE_FINDER_ACTION_KEY)

    catalog_rows, _total = course_catalog.query_catalog(
        db,
        country_code=code,
        degree_level=(payload.degree_level or "").strip().lower() or None,
        discipline=(payload.discipline or "").strip() or None,
        q=field_query or None,
        limit_universities=20,
    )
    if field_query and sum(len(c) for _u, c in catalog_rows) < course_catalog.RECOMMEND_MIN_CATALOG_ROWS:
        # A literal name search can be too narrow ("ML" won't match "Machine Learning"
        # rows verbatim) — widen to the discipline/level slice and let the model pick.
        catalog_rows, _total = course_catalog.query_catalog(
            db,
            country_code=code,
            degree_level=(payload.degree_level or "").strip().lower() or None,
            discipline=(payload.discipline or "").strip() or None,
            limit_universities=20,
        )

    usage_token = ai_usage.set_usage_account(organization_id=organization.id)
    try:
        result = course_catalog.recommend_courses(
            destination_country=course_catalog.country_name(code),
            catalog_rows=catalog_rows,
            client=client,
            field_of_study=field_query or None,
            degree_level=(payload.degree_level or "").strip().lower() or None,
            discipline=(payload.discipline or "").strip() or None,
            budget=payload.budget,
            notes=payload.notes,
            max_results=payload.max_results,
            usage_source="enterprise_course_finder",
        )
    finally:
        ai_usage.reset_usage_account(usage_token)

    if not result.get("available"):
        raise HTTPException(status_code=503, detail=result.get("message") or "Recommendations are unavailable right now.")
    recommendations = result.get("recommendations") or []
    if not recommendations:
        # Nothing usable came back — never bill the consultancy for an empty result.
        raise HTTPException(
            status_code=502,
            detail="Couldn't generate a shortlist. Refine the field of study or filters and retry.",
        )

    rec_row = models.EnterpriseCourseFinderRec(
        organization_id=organization.id,
        client_id=client.id if client else None,
        client_name=client.full_name if client else None,
        country_code=code,
        degree_level=(payload.degree_level or "").strip().lower() or None,
        discipline=(payload.discipline or "").strip() or None,
        query=json.dumps({
            "field_of_study": field_query or None,
            "budget": (payload.budget or "").strip() or None,
            "notes": (payload.notes or "").strip() or None,
            "max_results": payload.max_results,
        }),
        summary=result.get("summary"),
        recommendations=json.dumps(recommendations),
        catalog_based=bool(result.get("catalog_based")),
        grounded=bool(result.get("grounded")),
        model_used=result.get("model"),
        credits_charged=0,
        created_by_user_id=current_user.id,
        created_by_name=current_user.full_name or current_user.email,
    )
    db.add(rec_row)
    # Assign rec_row.id now: a client-less debit must still point at the artifact it
    # paid for, or the ledger row is untraceable in a billing dispute.
    db.flush()

    charge_target = f"{client.full_name}" if client else course_catalog.country_name(code)
    txn = credits.charge_action(
        db, organization.id, COURSE_FINDER_ACTION_KEY,
        user=current_user,
        reference_type="client" if client else "course_finder",
        reference_id=client.id if client else rec_row.id,
        description=f"Course Finder shortlist — {charge_target}",
        commit=False,
    )
    charged = credits.action_cost(COURSE_FINDER_ACTION_KEY) if txn else 0
    rec_row.credits_charged = charged
    # The shortlist row and the wallet debit commit atomically (deep-scan pattern).
    db.commit()
    db.refresh(rec_row)

    return {
        "permissions": _enterprise_permissions_for_role(role),
        "rec": _serialize_course_finder_rec(rec_row),
        "credits_charged": charged,
        "wallet": credits.wallet_state(db, organization.id, currency=_resolve_charge_currency(None, organization)),
    }


@router.get("/course-finder/recs")
def enterprise_course_finder_recs_list(
    request: Request,
    client_id: Optional[int] = None,
    limit: int = 20,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="coursefinder.view")
    query = (
        db.query(models.EnterpriseCourseFinderRec)
        .filter(models.EnterpriseCourseFinderRec.organization_id == organization.id)
    )
    # Stored shortlists carry the client's name and who ran them, and `?client_id=` is a
    # caller-supplied probe — so scope before that filter is applied. Shortlists run without a
    # client attached are workspace-level and stay visible to everyone.
    if not role.ctx.is_org_scope:
        query = query.filter(
            or_(
                models.EnterpriseCourseFinderRec.client_id.is_(None),
                models.EnterpriseCourseFinderRec.client_id.in_(
                    scoped_client_ids_subq(db, organization.id, role.ctx)
                ),
            )
        )
    if client_id:
        query = query.filter(models.EnterpriseCourseFinderRec.client_id == int(client_id))
    rows = (
        query.order_by(models.EnterpriseCourseFinderRec.created_at.desc(), models.EnterpriseCourseFinderRec.id.desc())
        .limit(max(1, min(int(limit or 20), 50)))
        .all()
    )
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "recs": [_serialize_course_finder_rec(r, include_items=False) for r in rows],
    }


@router.get("/course-finder/recs/{rec_id}")
def enterprise_course_finder_rec_detail(
    rec_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request, require_capability="coursefinder.view")
    row = (
        db.query(models.EnterpriseCourseFinderRec)
        .filter(
            models.EnterpriseCourseFinderRec.id == rec_id,
            models.EnterpriseCourseFinderRec.organization_id == organization.id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Shortlist not found.")
    # Re-resolve the client this shortlist belongs to, so a scope-limited member cannot read a
    # colleague's shortlist (and copy it onto their own client) by walking rec ids.
    if row.client_id:
        _get_org_client_or_404(db, organization.id, row.client_id, ctx=role.ctx)
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "rec": _serialize_course_finder_rec(row),
    }


class EnterpriseCourseFinderSave(BaseModel):
    index: int = Field(..., ge=0, le=15)
    client_id: Optional[int] = None  # required when the shortlist wasn't built for a client


@router.post("/course-finder/recs/{rec_id}/save-to-client")
def enterprise_course_finder_save_to_client(
    rec_id: int,
    payload: EnterpriseCourseFinderSave,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Copy one recommendation onto a client's Universities shortlist (free)."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="universities.manage"
    )
    row = (
        db.query(models.EnterpriseCourseFinderRec)
        .filter(
            models.EnterpriseCourseFinderRec.id == rec_id,
            models.EnterpriseCourseFinderRec.organization_id == organization.id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Shortlist not found.")
    target_client_id = payload.client_id or row.client_id
    if not target_client_id:
        raise HTTPException(status_code=400, detail="Pick a client to save this recommendation to.")
    client = _get_org_client_or_404(db, organization.id, int(target_client_id), ctx=role.ctx)

    try:
        items = json.loads(row.recommendations) if row.recommendations else []
    except Exception:
        items = []
    if not isinstance(items, list) or payload.index >= len(items):
        raise HTTPException(status_code=400, detail="That recommendation no longer exists.")
    item = items[payload.index] if isinstance(items[payload.index], dict) else {}
    uni_name = str(item.get("university_name") or "").strip()[:200]
    if not uni_name:
        raise HTTPException(status_code=400, detail="That recommendation no longer exists.")

    program = str(item.get("course_name") or "").strip()[:200] or None
    # Idempotent: re-clicking "Add" must not stack duplicate shortlist rows.
    for prior in _client_universities_query(db, organization.id, client.id).all():
        if (prior.university_name or "").strip().lower() == uni_name.lower() and (
            (prior.program or "").strip().lower() == (program or "").strip().lower()
        ):
            return {
                "permissions": _enterprise_permissions_for_role(role),
                "entry": _serialize_client_university(prior),
                "client_id": client.id,
                "already_saved": True,
            }

    requirements = item.get("key_requirements")
    requirements = [str(r).strip()[:140] for r in requirements if str(r).strip()][:6] if isinstance(requirements, list) else []
    difficulty = str(item.get("fit_level") or "").strip().lower()
    entry = models.EnterpriseClientUniversity(
        organization_id=organization.id,
        client_id=client.id,
        country_code=row.country_code,
        university_name=uni_name,
        program=program,
        location=str(item.get("location") or "").strip()[:160] or None,
        status="considering",
        source="ai",
        est_tuition=str(item.get("annual_tuition") or "").strip()[:80] or None,
        rationale=str(item.get("why_recommended") or "").strip()[:600] or None,
        qs_world_rank=str(item.get("qs_world_rank") or "").strip()[:20] or None,
        admission_difficulty=difficulty if difficulty in _UNIVERSITY_DIFFICULTY else None,
        key_requirements=json.dumps(requirements) if requirements else None,
        application_fee=str(item.get("application_fee") or "").strip()[:60] or None,
        website_url=_safe_external_url(item.get("website_url")),
        admissions_url=_safe_external_url(item.get("course_url")),
        added_by_user_id=current_user.id,
        added_by_name=current_user.full_name or current_user.email,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "entry": _serialize_client_university(entry),
        "client_id": client.id,
    }


# ============================================================================
# ACCESS CONTROL: offices, roles, member access, audit log
#
# Thin endpoint wrappers only. Every rule — validation, privilege-escalation checks, scope
# clamping, the owner and last-admin guards, audit writes and notifications — lives in
# app/enterprise_team.py, so there is exactly one place to read (and one place to fix) the
# question "who is allowed to change whose access".
#
# One transaction per request: the service functions never commit, the endpoint does.
# ============================================================================


class EnterpriseBranchWriteRequest(BaseModel):
    name: Optional[str] = Field(default=None, max_length=120)
    code: Optional[str] = Field(default=None, max_length=16)
    city: Optional[str] = Field(default=None, max_length=80)
    state_region: Optional[str] = Field(default=None, max_length=80)
    country_code: Optional[str] = Field(default=None, max_length=4)
    address_line: Optional[str] = Field(default=None, max_length=240)
    phone: Optional[str] = Field(default=None, max_length=40)
    email: Optional[str] = Field(default=None, max_length=200)
    timezone: Optional[str] = Field(default=None, max_length=60)


class EnterpriseBranchArchiveRequest(BaseModel):
    reassign_clients_to_branch_id: Optional[int] = None
    reassign_members_to_branch_id: Optional[int] = None


class EnterpriseBranchReassignRequest(BaseModel):
    target_branch_id: int
    client_ids: Optional[List[int]] = None


class EnterpriseRoleWriteRequest(BaseModel):
    name: Optional[str] = Field(default=None, max_length=80)
    description: Optional[str] = Field(default=None, max_length=240)
    capabilities: Optional[List[str]] = None
    data_scope: Optional[str] = Field(default=None, max_length=12)
    based_on_role_key: Optional[str] = Field(default=None, max_length=40)


class EnterpriseRoleArchiveRequest(BaseModel):
    move_members_to: Optional[str] = Field(default=None, max_length=40)


class EnterpriseRoleDuplicateRequest(BaseModel):
    name: str = Field(..., min_length=2, max_length=80)


class EnterpriseMemberAccessRequest(BaseModel):
    role_key: Optional[str] = Field(default=None, max_length=40)
    custom_role_id: Optional[int] = None
    data_scope: Optional[str] = Field(default=None, max_length=12)
    capability_grants: Optional[List[str]] = None
    capability_denies: Optional[List[str]] = None
    branch_ids: Optional[List[int]] = None
    primary_branch_id: Optional[int] = None


class EnterpriseMemberProfileRequest(BaseModel):
    full_name: Optional[str] = Field(default=None, max_length=120)
    job_title: Optional[str] = Field(default=None, max_length=120)
    phone: Optional[str] = Field(default=None, max_length=40)


class EnterpriseMemberDeactivateRequest(BaseModel):
    reason: Optional[str] = Field(default=None, max_length=240)
    reassign_clients_to_user_id: Optional[int] = None
    reassign_events_to_user_id: Optional[int] = None


class EnterpriseMemberReactivateRequest(BaseModel):
    role_key: Optional[str] = Field(default=None, max_length=40)
    custom_role_id: Optional[int] = None
    data_scope: Optional[str] = Field(default=None, max_length=12)
    branch_ids: Optional[List[int]] = None


class EnterpriseTeamBulkRequest(BaseModel):
    user_ids: List[int] = Field(..., min_length=1, max_length=100)
    action: str = Field(..., min_length=3, max_length=24)
    role_key: Optional[str] = Field(default=None, max_length=40)
    custom_role_id: Optional[int] = None
    data_scope: Optional[str] = Field(default=None, max_length=12)
    branch_ids: Optional[List[int]] = None


class EnterpriseTransferOwnershipRequest(BaseModel):
    target_user_id: int
    confirm_email: str = Field(..., min_length=3, max_length=200)
    # The 6-digit code from /organization/transfer-ownership/send-code. Required — see the
    # endpoint docstring for why typing an email address is not, on its own, a confirmation.
    code: str = Field(..., min_length=4, max_length=10)


class EnterpriseTransferOwnershipCodeRequest(BaseModel):
    target_user_id: int
    confirm_email: str = Field(..., min_length=3, max_length=200)


def _branches_payload(db: Session, organization, ctx, *, include_archived: bool = False) -> list[dict]:
    """Offices with their member and client counts.

    The counts are withheld for offices outside a scope-limited member's own set: exact client
    volumes per office are precisely what the office partition exists to keep separate.
    """
    rows = team_svc.list_branches(db, organization.id, include_archived=include_archived)
    client_counts = team_svc.branch_client_counts(db, organization.id)
    member_counts = team_svc.branch_member_counts(db, organization.id)
    visible_counts = ctx is None or ctx.is_org_scope
    out = []
    for branch in rows:
        own = visible_counts or (ctx is not None and int(branch.id) in ctx.branch_ids)
        out.append(team_svc.serialize_branch(
            branch,
            member_count=member_counts.get(int(branch.id), 0) if own else 0,
            client_count=client_counts.get(int(branch.id), 0) if own else 0,
        ))
    return out


def _with_capability_aliases(payload: dict) -> dict:
    """Add flat capability lists alongside `capability_states`.

    `capability_states` is the authoritative shape — one state per capability key
    (inherited / granted / blocked / off) — but the UI also wants the three lists directly to
    render "from your role", "added for you" and "switched off for you" without walking the map on
    every paint. Derived here rather than duplicated in the service layer, so the two can never
    disagree.
    """
    states = payload.get("capability_states") or {}
    inherited = sorted(k for k, v in states.items() if v == "inherited")
    granted = sorted(k for k, v in states.items() if v == "granted")
    blocked = sorted(k for k, v in states.items() if v == "blocked")
    payload["role_capabilities"] = inherited
    payload["inherited"] = inherited
    payload["added"] = granted
    payload["capability_grants"] = granted
    payload["blocked"] = blocked
    payload["capability_denies"] = blocked
    role = payload.get("role") or {}
    payload["role_description"] = role.get("description")
    payload["role_label"] = role.get("label")
    return payload


@router.get("/team/meta")
def enterprise_team_meta(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Everything the access-control UI is built from: the capability registry, the role presets,
    the workspace's own roles, its offices, and what the CALLER is allowed to hand out."""
    membership, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="team.view"
    )
    return team_svc.team_meta_payload(
        db, organization=organization, membership=membership, ctx=role.ctx
    )


@router.get("/team/my-access")
def enterprise_team_my_access(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """What the signed-in member's own access actually is.

    Deliberately available to EVERY member, with no capability gate: telling someone "ask an
    admin for access" is useless if they can't see what they already have, which permissions were
    switched off for them individually, or which clients they're limited to.
    """
    membership, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request
    )
    payload = _with_capability_aliases(team_svc.my_access_payload(
        db, organization=organization, membership=membership, user=current_user, ctx=role.ctx
    ))
    payload["access"] = access.access_payload(role.ctx)
    return payload


# ---- Offices ---------------------------------------------------------------


@router.get("/branches")
def enterprise_list_branches(
    request: Request,
    include_archived: bool = False,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="branches.view"
    )
    show_archived = bool(include_archived) and role.ctx.has("branches.manage")
    return {
        "branches": _branches_payload(db, organization, role.ctx, include_archived=show_archived),
        "permissions": _enterprise_permissions_for_role(role),
        "access": access.access_payload(role.ctx),
        "limits": {"max_branches": access.MAX_BRANCHES},
    }


@router.post("/branches")
def enterprise_create_branch(
    payload: EnterpriseBranchWriteRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="branches.manage"
    )
    branch = team_svc.create_branch(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        data=payload.model_dump(exclude_unset=True),
    )
    db.commit()
    return {
        "message": f"{branch.name} added.",
        "branch": team_svc.serialize_branch(branch),
        "branches": _branches_payload(db, organization, role.ctx),
    }


@router.patch("/branches/{branch_id}")
def enterprise_update_branch(
    branch_id: int,
    payload: EnterpriseBranchWriteRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="branches.manage"
    )
    branch = team_svc.get_org_branch_or_404(db, organization.id, branch_id)
    branch = team_svc.update_branch(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        branch=branch, data=payload.model_dump(exclude_unset=True),
    )
    db.commit()
    return {
        "message": "Office updated.",
        "branch": team_svc.serialize_branch(branch),
        "branches": _branches_payload(db, organization, role.ctx),
    }


@router.post("/branches/{branch_id}/set-default")
def enterprise_set_default_branch(
    branch_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="branches.manage"
    )
    branch = team_svc.get_org_branch_or_404(db, organization.id, branch_id)
    team_svc.set_default_branch(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx, branch=branch,
    )
    db.commit()
    return {
        "message": f"{branch.name} is now the default office.",
        "branches": _branches_payload(db, organization, role.ctx),
    }


@router.post("/branches/{branch_id}/archive")
def enterprise_archive_branch(
    branch_id: int,
    payload: EnterpriseBranchArchiveRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Archive an office. Refuses with a 409 carrying the counts when it still holds clients or
    staff and no destination was chosen, so the UI can offer a picker instead of orphaning them."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="branches.manage"
    )
    branch = team_svc.get_org_branch_or_404(db, organization.id, branch_id)
    result = team_svc.archive_branch(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx, branch=branch,
        reassign_clients_to_branch_id=payload.reassign_clients_to_branch_id,
        reassign_members_to_branch_id=payload.reassign_members_to_branch_id,
    )
    db.commit()
    return {
        "message": f"{branch.name} archived.",
        "branches": _branches_payload(db, organization, role.ctx, include_archived=True),
        **result,
    }


@router.post("/branches/{branch_id}/reactivate")
def enterprise_reactivate_branch(
    branch_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="branches.manage"
    )
    branch = team_svc.get_org_branch_or_404(db, organization.id, branch_id)
    branch = team_svc.reactivate_branch(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx, branch=branch,
    )
    db.commit()
    return {
        "message": f"{branch.name} reopened.",
        "branches": _branches_payload(db, organization, role.ctx, include_archived=True),
    }


@router.post("/branches/{branch_id}/reassign-clients")
def enterprise_reassign_branch_clients(
    branch_id: int,
    payload: EnterpriseBranchReassignRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Move clients between offices. Also the merge operation: move everything, then archive."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request,
        require_capability=("branches.manage", "clients.set_branch"),
    )
    branch = team_svc.get_org_branch_or_404(db, organization.id, branch_id)
    target = team_svc.get_org_branch_or_404(db, organization.id, payload.target_branch_id)
    moved = team_svc.reassign_branch_clients(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        branch=branch, target_branch=target, client_ids=payload.client_ids,
        # A scope-limited actor may only move clients they can actually see.
        scope_filter=(lambda q: scope_client_query(q, role.ctx)),
    )
    db.commit()
    return {
        "message": (f"{moved} client{'s' if moved != 1 else ''} moved to {target.name}."
                    if moved else "No clients needed moving."),
        "moved": moved,
        "branches": _branches_payload(db, organization, role.ctx),
    }


@router.get("/branches/{branch_id}/members")
def enterprise_branch_members(
    branch_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="branches.view"
    )
    branch = team_svc.get_org_branch_or_404(db, organization.id, branch_id)
    return {
        "branch": team_svc.serialize_branch(branch),
        "members": team_svc.branch_members(db, organization.id, branch),
    }


# ---- Roles ----------------------------------------------------------------


@router.get("/roles")
def enterprise_list_roles(
    request: Request,
    include_archived: bool = False,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="team.view"
    )
    show_archived = bool(include_archived) and role.ctx.has("roles.manage")
    registry = access.capability_registry_payload()
    return {
        "presets": registry.get("role_presets", []),
        "custom_roles": team_svc.list_roles_payload(
            db, organization.id, include_archived=show_archived
        ),
        "capabilities": registry.get("capabilities", []),
        "sections": registry.get("sections", []),
        "limits": {"max_custom_roles": access.MAX_CUSTOM_ROLES},
        "actor_capabilities": sorted(
            access.CAPABILITY_KEYS if role.ctx.is_owner else role.ctx.capabilities
        ),
        "access": access.access_payload(role.ctx),
    }


@router.post("/roles")
def enterprise_create_role(
    payload: EnterpriseRoleWriteRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="roles.manage"
    )
    created = team_svc.create_role(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        data=payload.model_dump(exclude_unset=True),
    )
    db.commit()
    return {
        "message": f"{created.name} created.",
        "role": team_svc.serialize_role(created),
        "custom_roles": team_svc.list_roles_payload(db, organization.id),
    }


@router.patch("/roles/{role_id}")
def enterprise_update_role(
    role_id: int,
    payload: EnterpriseRoleWriteRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="roles.manage"
    )
    target = team_svc.get_org_role_or_404(db, organization.id, role_id)
    target = team_svc.update_role(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        role=target, data=payload.model_dump(exclude_unset=True),
    )
    db.commit()
    return {
        "message": f"{target.name} updated.",
        "role": team_svc.serialize_role(target),
        "custom_roles": team_svc.list_roles_payload(db, organization.id),
        "members": _list_organization_members(db, organization.id, ctx=role.ctx),
    }


@router.post("/roles/{role_id}/duplicate")
def enterprise_duplicate_role(
    role_id: int,
    payload: EnterpriseRoleDuplicateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="roles.manage"
    )
    source = team_svc.get_org_role_or_404(db, organization.id, role_id)
    created = team_svc.duplicate_role(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        role=source, name=payload.name,
    )
    db.commit()
    return {
        "message": f"{created.name} created.",
        "role": team_svc.serialize_role(created),
        "custom_roles": team_svc.list_roles_payload(db, organization.id),
    }


@router.post("/roles/{role_id}/archive")
def enterprise_archive_role(
    role_id: int,
    payload: EnterpriseRoleArchiveRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Archive a role. 409s with the member count while anyone is still on it, unless a
    destination role is supplied — a role is never hard-deleted, so history stays readable."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="roles.manage"
    )
    target = team_svc.get_org_role_or_404(db, organization.id, role_id)
    result = team_svc.archive_role(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        role=target, move_members_to=payload.move_members_to,
    )
    db.commit()
    return {
        "message": f"{target.name} archived.",
        "custom_roles": team_svc.list_roles_payload(db, organization.id),
        "members": _list_organization_members(db, organization.id, ctx=role.ctx),
        **result,
    }


# ---- Member access --------------------------------------------------------


@router.get("/team/users/{member_user_id}/access")
def enterprise_member_access(
    member_user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """One member's full effective access, with each permission marked inherited / added / blocked.
    This is where the heavy capability list lives, so /team can stay light."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="team.view"
    )
    membership = team_svc.get_org_membership_or_404(db, organization.id, member_user_id)
    user = db.query(models.User).filter(models.User.id == int(member_user_id)).first()
    if not user:
        raise HTTPException(status_code=404, detail="Organization member not found.")
    payload = _with_capability_aliases(team_svc.member_access_payload(
        db, organization=organization, membership=membership, user=user
    ))
    payload["actor_capabilities"] = sorted(
        access.CAPABILITY_KEYS if role.ctx.is_owner else role.ctx.capabilities
    )
    payload["can_edit"] = role.ctx.has("roles.manage")
    return payload


@router.patch("/team/users/{member_user_id}/access")
def enterprise_update_member_access(
    member_user_id: int,
    payload: EnterpriseMemberAccessRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="roles.manage"
    )
    membership = team_svc.get_org_membership_or_404(db, organization.id, member_user_id)
    result = team_svc.update_member_access(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        membership=membership, data=payload.model_dump(exclude_unset=True),
    )
    db.commit()
    changed = result.get("changed") or []
    return {
        "message": "Access updated." if changed else "Nothing to change.",
        "member": result.get("member"),
        "changed": changed,
        "members": _list_organization_members(db, organization.id, ctx=role.ctx),
    }


@router.patch("/team/users/{member_user_id}/profile")
def enterprise_update_member_profile(
    member_user_id: int,
    payload: EnterpriseMemberProfileRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Name, job title and phone. Anyone may edit their OWN entry; editing a colleague's needs
    team.manage."""
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request)
    is_self = int(member_user_id) == int(current_user.id)
    if not is_self and not role.ctx.has("team.manage"):
        raise HTTPException(status_code=403, detail=access.denied_detail("team.manage"))
    membership = team_svc.get_org_membership_or_404(db, organization.id, member_user_id)
    user = db.query(models.User).filter(models.User.id == int(member_user_id)).first()
    if not user:
        raise HTTPException(status_code=404, detail="Organization member not found.")
    result = team_svc.update_member_profile(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        membership=membership, user=user, data=payload.model_dump(exclude_unset=True),
    )
    db.commit()
    return {
        "message": "Details saved.",
        "member": result.get("member"),
        "members": _list_organization_members(db, organization.id, ctx=role.ctx),
    }


@router.post("/team/users/{member_user_id}/deactivate")
def enterprise_deactivate_member(
    member_user_id: int,
    payload: EnterpriseMemberDeactivateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Turn off someone's access and sign them out.

    409s with their open record counts when they still own clients or reminders and no successor
    was named — offboarding must not silently orphan a caseload.
    """
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="team.manage"
    )
    membership = team_svc.get_org_membership_or_404(db, organization.id, member_user_id)
    user = db.query(models.User).filter(models.User.id == int(member_user_id)).first()
    if not user:
        raise HTTPException(status_code=404, detail="Organization member not found.")
    result = team_svc.deactivate_member(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        membership=membership, user=user, reason=payload.reason,
        reassign_clients_to_user_id=payload.reassign_clients_to_user_id,
        reassign_events_to_user_id=payload.reassign_events_to_user_id,
    )
    db.commit()
    return {
        "message": f"{user.full_name or user.email} no longer has access.",
        "members": _list_organization_members(db, organization.id, ctx=role.ctx),
        **result,
    }


@router.get("/team/users/{member_user_id}/offboard-preview")
def enterprise_member_offboard_preview(
    member_user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """What this person still owns, so the deactivate dialog can ask about it up front rather
    than after a failed attempt."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="team.manage"
    )
    team_svc.get_org_membership_or_404(db, organization.id, member_user_id)
    return team_svc.member_offboard_counts(db, organization.id, member_user_id)


@router.post("/team/users/{member_user_id}/reactivate")
def enterprise_reactivate_member(
    member_user_id: int,
    payload: EnterpriseMemberReactivateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="team.manage"
    )
    membership = team_svc.get_org_membership_or_404(db, organization.id, member_user_id)
    user = db.query(models.User).filter(models.User.id == int(member_user_id)).first()
    if not user:
        raise HTTPException(status_code=404, detail="Organization member not found.")
    result = team_svc.reactivate_member(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        membership=membership, user=user, data=payload.model_dump(exclude_unset=True),
    )
    db.commit()
    return {
        "message": f"{user.full_name or user.email} can sign in again.",
        "members": _list_organization_members(db, organization.id, ctx=role.ctx),
        **result,
    }


@router.post("/team/users/{member_user_id}/resend-invite")
def enterprise_resend_member_invite(
    member_user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Mint a fresh password-setup link and email the invitation again."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="team.manage"
    )
    _enforce_rate_limit_or_429(
        request=request,
        scope="enterprise.team_invite",
        limit=ENTERPRISE_TEAM_INVITE_RATE_LIMIT,
        window_seconds=ENTERPRISE_TEAM_INVITE_RATE_WINDOW_SECONDS,
        extra_key=f"org:{organization.id}:user:{current_user.id}",
    )
    membership = team_svc.get_org_membership_or_404(db, organization.id, member_user_id)
    user = db.query(models.User).filter(models.User.id == int(member_user_id)).first()
    if not user:
        raise HTTPException(status_code=404, detail="Organization member not found.")
    if not membership.is_active:
        raise HTTPException(
            status_code=400,
            detail="Restore their access first, then resend the invitation.",
        )

    password_setup_token = generate_verification_token()
    user.password_reset_token = hash_token(password_setup_token)
    user.password_reset_token_expires = datetime.utcnow() + timedelta(
        hours=ENTERPRISE_INVITE_PASSWORD_SETUP_EXPIRES_HOURS
    )
    db.commit()

    portal_url = _build_enterprise_portal_url(organization.subdomain_slug, request)
    password_setup_url = (
        f"{ENTERPRISE_PASSWORD_SETUP_BASE_URL}/reset-password"
        f"?token={quote(password_setup_token, safe='')}"
    )
    sent = False
    try:
        sent = send_enterprise_team_invite_email(
            invitee_email=user.email,
            invitee_name=user.full_name,
            organization_name=organization.company_name,
            role=access.role_label_for(
                access.normalize_role_key(
                    getattr(membership, "role_key", None), getattr(membership, "role", None)
                )
            ),
            portal_url=portal_url,
            set_password_url=password_setup_url,
            password_setup_expires_hours=ENTERPRISE_INVITE_PASSWORD_SETUP_EXPIRES_HOURS,
            invited_by_name=current_user.full_name,
            invited_by_email=current_user.email,
        )
    except Exception:
        logger.exception(
            "Failed to resend enterprise invite (org_id=%s, user_id=%s)", organization.id, user.id
        )
    return {
        "message": (
            f"Invitation resent to {user.email}."
            if sent
            else "Couldn't send the email just now. They can request a password link from the sign-in screen."
        ),
        "invite_email_sent": sent,
    }


@router.post("/team/bulk")
def enterprise_team_bulk(
    payload: EnterpriseTeamBulkRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Apply one change to several members.

    Every per-member guard from the single-member endpoints runs again for each id, and anything
    that fails is reported in `skipped` with a reason rather than failing the whole batch or —
    worse — succeeding silently for some and not others.
    """
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="team.manage"
    )
    action = (payload.action or "").strip().lower()
    if action in {"set_role", "set_scope"} and not role.ctx.has("roles.manage"):
        raise HTTPException(status_code=403, detail=access.denied_detail("roles.manage"))
    if action in {"set_branches", "add_branch"} and not role.ctx.has("branches.manage"):
        raise HTTPException(status_code=403, detail=access.denied_detail("branches.manage"))
    if action not in {"set_role", "set_scope", "set_branches", "add_branch", "deactivate"}:
        raise HTTPException(status_code=400, detail="Unknown bulk action.")

    applied: list[int] = []
    skipped: list[dict] = []
    for user_id in list(dict.fromkeys(payload.user_ids)):
        try:
            membership = team_svc.get_org_membership_or_404(db, organization.id, user_id)
            if action == "deactivate":
                user = db.query(models.User).filter(models.User.id == int(user_id)).first()
                if not user:
                    raise HTTPException(status_code=404, detail="Member not found.")
                team_svc.deactivate_member(
                    db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
                    membership=membership, user=user,
                )
            else:
                data: dict = {}
                if action == "set_role":
                    data["role_key"] = payload.role_key
                    if payload.custom_role_id is not None:
                        data["custom_role_id"] = payload.custom_role_id
                elif action == "set_scope":
                    data["data_scope"] = payload.data_scope
                    if payload.branch_ids is not None:
                        data["branch_ids"] = payload.branch_ids
                elif action == "set_branches":
                    data["branch_ids"] = payload.branch_ids or []
                elif action == "add_branch":
                    existing = team_svc.member_branch_ids(db, organization.id).get(
                        int(membership.id), []
                    )
                    data["branch_ids"] = sorted(set(existing) | set(payload.branch_ids or []))
                team_svc.update_member_access(
                    db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
                    membership=membership, data={k: v for k, v in data.items() if v is not None},
                )
            # Flush per member: the service layer runs bulk UPDATE statements, and this session
            # has autoflush off — a later statement would otherwise not see this member's
            # pending writes and could undo them.
            db.flush()
            applied.append(int(user_id))
        except HTTPException as exc:
            detail = exc.detail
            if isinstance(detail, dict):
                detail = detail.get("message") or "Couldn't apply this change."
            skipped.append({"user_id": int(user_id), "reason": str(detail)})
        except Exception:
            logger.exception("Bulk team action failed (org=%s, user=%s)", organization.id, user_id)
            skipped.append({"user_id": int(user_id), "reason": "Something went wrong for this member."})

    if applied:
        db.commit()
    else:
        db.rollback()
    return {
        "message": (
            f"Updated {len(applied)} member{'s' if len(applied) != 1 else ''}."
            + (f" {len(skipped)} skipped." if skipped else "")
        ),
        "applied": applied,
        "skipped": skipped,
        "members": _list_organization_members(db, organization.id, ctx=role.ctx),
    }


def _person_label(user) -> str:
    return (getattr(user, "full_name", None) or getattr(user, "email", None) or "A team member").strip()


def _resolve_owner_transfer_target(db: Session, *, organization, role, target_user_id: int, confirm_email: str):
    """Resolve and fully validate a pending ownership transfer.

    Both steps of the flow run this — the same capability gate, the same typed-email check, the
    same owner-or-recovery rule — so the emailed code can only ever confirm a transfer that was
    already permitted at the moment it was requested AND at the moment it is used.
    """
    target_membership = team_svc.get_org_membership_or_404(db, organization.id, target_user_id)
    target_user = db.query(models.User).filter(models.User.id == int(target_user_id)).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="Organization member not found.")
    if not (role.ctx.has("org.transfer_ownership") or role.ctx.has("team.manage")):
        raise HTTPException(status_code=403, detail=access.denied_detail("org.transfer_ownership"))
    team_svc.assert_transfer_ownership_allowed(
        db, organization=organization, actor_ctx=role.ctx,
        target_membership=target_membership, target_user=target_user,
        confirm_email=confirm_email,
    )
    return target_membership, target_user


def _send_owner_transfer_notices(db: Session, *, organization, actor_user, new_owner_user, previous_owner_user_id) -> None:
    """Tell the incoming and outgoing owners by email, after the transfer has committed.

    The in-app bell only reaches someone who signs in; an inbox reaches them tonight. Strictly
    best-effort — the transfer is already durable, and a bounced notification must never turn a
    completed handover into a 500 the caller can't interpret.
    """
    org_name = getattr(organization, "company_name", None) or "your workspace"
    actor_label = _person_label(actor_user)
    new_owner_label = _person_label(new_owner_user)

    previous_owner = None
    prev_id = int(previous_owner_user_id or 0)
    if prev_id and prev_id != int(getattr(new_owner_user, "id", 0) or 0):
        previous_owner = db.query(models.User).filter(models.User.id == prev_id).first()

    for recipient, is_new_owner in ((new_owner_user, True), (previous_owner, False)):
        email = (getattr(recipient, "email", "") or "").strip() if recipient else ""
        if not email:
            continue
        try:
            send_enterprise_owner_transfer_notice_email(
                to_email=email,
                recipient_name=getattr(recipient, "full_name", None),
                organization_name=org_name,
                new_owner_label=new_owner_label,
                actor_label=actor_label,
                is_new_owner=is_new_owner,
            )
        except Exception:
            logger.exception("enterprise: ownership-transfer notice email failed (org=%s)", organization.id)


@router.post("/organization/transfer-ownership/send-code")
def enterprise_transfer_ownership_send_code(
    payload: EnterpriseTransferOwnershipCodeRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Step 1 of a transfer: email the ACTOR a 6-digit code confirming this exact handover.

    Typing the new owner's address proves only that the person at the keyboard can read the
    members table. Ownership carries refunds and the payout bank account and cannot be taken
    back, so the second factor is an inbox: a hijacked session has the cookie, not the mail.

    The whole transfer is validated here first — capability, typed email, target still active —
    so a caller who could never complete it can't make Rilono send a security email about it.
    """
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.owner_transfer_code", limit=6, window_seconds=3600,
        extra_key=str(getattr(current_user, "id", "") or ""),
    )
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request)
    _membership, target_user = _resolve_owner_transfer_target(
        db, organization=organization, role=role,
        target_user_id=payload.target_user_id, confirm_email=payload.confirm_email,
    )

    actor_email = (current_user.email or "").strip().lower()
    if not actor_email:
        raise HTTPException(
            status_code=400,
            detail="Your account has no email address, so we can't send a confirmation code.",
        )

    code = step_up.issue_code(
        db,
        user_id=current_user.id,
        organization_id=organization.id,
        purpose=step_up.PURPOSE_OWNER_TRANSFER,
        context_key=step_up.context_key_for_target_user(target_user.id),
    )
    access.audit(
        db,
        organization_id=organization.id,
        actor=current_user,
        action="owner_transfer_code_sent",
        summary=f"Requested a confirmation code to transfer ownership to {_person_label(target_user)}",
        target_user=target_user,
        ip_address=extract_client_ip(request),
    )
    db.commit()

    sent, _mid, err = send_enterprise_owner_transfer_code_email(
        to_email=actor_email,
        actor_name=current_user.full_name,
        organization_name=getattr(organization, "company_name", None) or "your workspace",
        new_owner_label=_person_label(target_user),
        code=code,
        expires_in_minutes=step_up.CODE_EXPIRES_MINUTES,
    )
    result = {
        "message": f"We've emailed a 6-digit code to {_mask_email(actor_email)}.",
        "masked_email": _mask_email(actor_email),
        "expires_in_minutes": step_up.CODE_EXPIRES_MINUTES,
    }
    if not sent:
        if _is_development_env():
            # Local sandbox without an email provider — surface the code so the flow is testable.
            result["dev_code"] = code
        else:
            logger.warning("enterprise: ownership-transfer code email failed (org=%s): %s", organization.id, err)
            raise HTTPException(
                status_code=502,
                detail="We couldn't send the confirmation email right now. Please try again in a moment.",
            )
    return result


@router.post("/organization/transfer-ownership")
def enterprise_transfer_ownership(
    payload: EnterpriseTransferOwnershipRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Hand the workspace to someone else. Step 2: the emailed code, then the handover.

    Gated on the owner-only capability OR — the recovery path — on a team.manage holder acting on
    an owner whose access has already been switched off. Without that second route, a workspace
    whose founder has left could never transfer ownership, accept a new DPA version or set a
    payout account again.

    On top of that gate, the caller must present the code from /transfer-ownership/send-code. It
    is bound to this target member, single-use, and spent only if the handover itself commits.
    """
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.owner_transfer", limit=12, window_seconds=3600,
        extra_key=str(getattr(current_user, "id", "") or ""),
    )
    _, organization, role = _require_enterprise_membership(db=db, user=current_user, request=request)
    target_membership, target_user = _resolve_owner_transfer_target(
        db, organization=organization, role=role,
        target_user_id=payload.target_user_id, confirm_email=payload.confirm_email,
    )
    # Second proof, before anything is written: the actor still holds the inbox they signed up
    # with. A wrong code burns an attempt (and commits that counter) rather than failing free.
    step_up.verify_code_or_400(
        db,
        user_id=current_user.id,
        organization_id=organization.id,
        purpose=step_up.PURPOSE_OWNER_TRANSFER,
        context_key=step_up.context_key_for_target_user(target_user.id),
        code=payload.code,
    )
    result = team_svc.transfer_ownership(
        db, organization=organization, actor_user=current_user, actor_ctx=role.ctx,
        target_membership=target_membership, target_user=target_user,
        confirm_email=payload.confirm_email,
    )
    db.commit()
    _send_owner_transfer_notices(
        db, organization=organization, actor_user=current_user,
        new_owner_user=target_user, previous_owner_user_id=result.get("previous_owner_user_id"),
    )
    return {
        "message": f"{target_user.full_name or target_user.email} is now the workspace owner.",
        "members": _list_organization_members(db, organization.id, ctx=role.ctx),
        **result,
    }


@router.get("/team/access-log")
def enterprise_access_log(
    request: Request,
    limit: int = 50,
    offset: int = 0,
    action: Optional[str] = None,
    target_user_id: Optional[int] = None,
    days: int = 90,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Who changed whose access, and what it was before."""
    _, organization, role = _require_enterprise_membership(
        db=db, user=current_user, request=request, require_capability="audit.view"
    )
    return team_svc.access_log_payload(
        db, organization=organization, ctx=role.ctx, limit=limit, offset=offset,
        action=action, target_user_id=target_user_id, days=days,
    )


# ============================================================================
# Lead collection forms — org-branded public forms + a per-org leads inbox.
#
# The shareable link (/f/{token}) stores its token RAW (unlike portal/pay
# hashes) on purpose: it is public-by-design — it reveals only the form
# definition + org branding and grants no data access — and the org must be
# able to re-copy the exact same link from the console at any time. Pausing
# the form or rotating the link kills the old URL.
# ============================================================================

_LEAD_FIELD_TYPES = {"text", "email", "phone", "textarea", "select", "date", "number", "checkbox", "file"}
_LEAD_STATUSES = ("new", "contacted", "converted", "closed")
_LEAD_FORM_MAX_FIELDS = 20
# Flood control, deliberately fail-SOFT: past the daily cap a form keeps accepting
# (a genuine prospect must never be turned away because a spammer got there first)
# but stops firing alert emails, so a proxy-rotating bot can't spend the shared
# transactional sender's reputation. Only the far-higher hard ceiling 429s.
_LEAD_FORM_DAILY_CAP = int(os.getenv("ENTERPRISE_LEAD_FORM_DAILY_CAP", "300"))
_LEAD_FORM_HARD_CAP = int(os.getenv("ENTERPRISE_LEAD_FORM_HARD_CAP", "3000"))
_LEAD_EMAIL_RE = re.compile(r"[^@\s]+@[^@\s]+\.[^@\s]+")

# ---- File-upload fields ----------------------------------------------------
# A lead-form upload is the only place in the product where a wholly ANONYMOUS
# visitor writes bytes into the org's bucket, so every limit here is deliberately
# tighter than the staff document uploader's:
#   * a smaller per-file ceiling (a prospect sends a passport page, not a dataset),
#   * a bounded number of file FIELDS per form, so an org can't publish a surface
#     that lets one visitor stage 20 × 5 files,
#   * a per-form 24h upload ceiling on top of the per-IP rate limit, so rotating
#     proxies can't run up an org's storage bill,
#   * a short TTL on staged (never-submitted) files, swept opportunistically.
# The extension allowlist is shared with client documents on purpose — these files
# become client documents on conversion and must pass the same bar.
_LEAD_UPLOAD_MAX_BYTES = int(os.getenv("ENTERPRISE_LEAD_UPLOAD_MAX_BYTES", str(10 * 1024 * 1024)))
_LEAD_UPLOAD_MAX_PER_FIELD = int(os.getenv("ENTERPRISE_LEAD_UPLOAD_MAX_PER_FIELD", "5"))
_LEAD_FORM_MAX_FILE_FIELDS = int(os.getenv("ENTERPRISE_LEAD_FORM_MAX_FILE_FIELDS", "5"))
_LEAD_UPLOAD_FORM_DAILY_CAP = int(os.getenv("ENTERPRISE_LEAD_UPLOAD_FORM_DAILY_CAP", "400"))
_LEAD_UPLOAD_STAGED_TTL_HOURS = int(os.getenv("ENTERPRISE_LEAD_UPLOAD_STAGED_TTL_HOURS", "6"))


def _build_lead_form_url(subdomain_slug, token: str, request: Request | None) -> str:
    subdomain = str(subdomain_slug or "").strip().lower()
    base = None
    if subdomain:
        host = f"{subdomain}.{ENTERPRISE_ROOT_DOMAIN}"
        port = _request_port_for_local_enterprise_url(request)
        if port:
            host = f"{host}:{port}"
        base = f"{ENTERPRISE_PORTAL_SCHEME}://{host}"
    if not base:
        base = ENTERPRISE_PASSWORD_SETUP_BASE_URL
    return f"{base.rstrip('/')}/f/{token}"


def _normalize_lead_form_fields(raw) -> list[dict]:
    """Validate + normalize a staff-built field list. Keys are derived server-side
    from labels (never client-supplied) so the submit endpoint can trust them."""
    if not isinstance(raw, list) or not raw:
        raise HTTPException(status_code=400, detail="Add at least one field to the form.")
    if len(raw) > _LEAD_FORM_MAX_FIELDS:
        raise HTTPException(status_code=400, detail=f"A form can have at most {_LEAD_FORM_MAX_FIELDS} fields.")
    fields: list[dict] = []
    seen: set[str] = set()
    file_fields = 0
    for item in raw:
        if not isinstance(item, dict):
            raise HTTPException(status_code=400, detail="Invalid field definition.")
        label = str(item.get("label") or "").strip()[:120]
        if not label:
            raise HTTPException(status_code=400, detail="Every field needs a label.")
        ftype = str(item.get("type") or "text").strip().lower()
        if ftype not in _LEAD_FIELD_TYPES:
            ftype = "text"
        key = re.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_")[:40] or "field"
        base_key, n = key, 2
        while key in seen:
            key = f"{base_key}_{n}"
            n += 1
        seen.add(key)
        field: dict = {"key": key, "label": label, "type": ftype, "required": bool(item.get("required"))}
        placeholder = str(item.get("placeholder") or "").strip()[:120]
        if placeholder:
            field["placeholder"] = placeholder
        if ftype == "select":
            options = [str(o or "").strip()[:80] for o in (item.get("options") or []) if str(o or "").strip()]
            options = options[:24]
            if len(options) < 2:
                raise HTTPException(status_code=400, detail=f'Dropdown "{label}" needs at least 2 options.')
            field["options"] = options
        if ftype == "file":
            file_fields += 1
            if file_fields > _LEAD_FORM_MAX_FILE_FIELDS:
                raise HTTPException(
                    status_code=400,
                    detail=f"A form can ask for at most {_LEAD_FORM_MAX_FILE_FIELDS} file uploads.",
                )
            try:
                max_files = int(item.get("max_files") or 1)
            except (TypeError, ValueError):
                max_files = 1
            field["max_files"] = max(1, min(max_files, _LEAD_UPLOAD_MAX_PER_FIELD))
        fields.append(field)
    return fields


def _drop_storage_keys(keys, *, context: str) -> None:
    """Delete blobs whose rows are already gone. Always called AFTER the commit:
    dropping bytes first would, on a failed commit, leave surviving rows pointing
    at nothing. delete_document is best-effort and never raises."""
    for key in keys or []:
        if key:
            enterprise_storage.delete_document(key)
    if keys:
        logger.info("Dropped %s lead storage object(s) for %s", len(keys), context)


def _lead_file_fields(fields: list[dict]) -> dict[str, dict]:
    """The form's file fields, keyed. Used by both the staged upload endpoint (to
    reject a key that isn't actually a file field) and the submit validator."""
    return {f.get("key"): f for f in fields if (f.get("type") or "") == "file" and f.get("key")}


def _lead_field_max_files(field: dict) -> int:
    try:
        return max(1, min(int(field.get("max_files") or 1), _LEAD_UPLOAD_MAX_PER_FIELD))
    except (TypeError, ValueError):
        return 1


def _parse_lead_form_fields(form: models.EnterpriseLeadForm) -> list[dict]:
    try:
        fields = json.loads(form.fields_json or "[]")
        return fields if isinstance(fields, list) else []
    except Exception:
        return []


def _validate_lead_answers(
    fields: list[dict],
    raw_answers,
    uploads_by_key: Optional[dict] = None,
    *,
    uploads_enabled: bool = True,
) -> tuple[list[dict], dict]:
    """Check a public submission against the form's field spec. Returns the ordered
    answer list to store and the denormalized contact info (name/email/phone).

    File answers are NOT read out of `raw_answers` — the browser only ever sends
    opaque upload tokens, and the caller resolves those to real staged rows before
    calling here. `uploads_by_key` is that resolved map, so a token the visitor
    invented can never become an answer. When storage is unconfigured the whole
    upload surface is down; a required file field then degrades to optional rather
    than making a paying org's form unsubmittable (same fail-soft rule as the
    daily flood cap)."""
    answers = raw_answers if isinstance(raw_answers, dict) else {}
    uploads_by_key = uploads_by_key or {}
    out: list[dict] = []
    contact = {"name": None, "email": None, "phone": None}
    first_text_value = None
    for f in fields:
        key = f.get("key")
        ftype = f.get("type") or "text"
        raw_v = answers.get(key)
        if ftype == "file":
            rows = uploads_by_key.get(key) or []
            if not rows:
                if f.get("required") and uploads_enabled:
                    raise HTTPException(status_code=400, detail=f'Please attach a file for "{f.get("label")}".')
                continue
            # The stored value is the human-readable filename list: it is what the
            # alert email, the inbox preview and the convert-to-client note render.
            # The files themselves are addressed through enterprise_lead_uploads.
            out.append({
                "key": key,
                "label": f.get("label"),
                "type": ftype,
                "value": ", ".join(r.original_filename for r in rows)[:500],
            })
            continue
        if ftype == "checkbox":
            value = "Yes" if raw_v in (True, "true", "on", "yes", "Yes", 1, "1") else ""
        else:
            value = str(raw_v if raw_v is not None else "").strip()
        value = value[:5000 if ftype == "textarea" else 300]
        if f.get("required") and not value:
            raise HTTPException(status_code=400, detail=f'"{f.get("label")}" is required.')
        if not value:
            continue
        if ftype == "email" and not _LEAD_EMAIL_RE.fullmatch(value):
            raise HTTPException(status_code=400, detail=f'Please enter a valid email for "{f.get("label")}".')
        # A bot posting straight at the endpoint doesn't see the <select>; hold answers
        # to the choices the org actually published, or the inbox fills with free text
        # in a column the consultancy believes is constrained.
        if ftype == "select" and value not in (f.get("options") or []):
            raise HTTPException(status_code=400, detail=f'Please choose one of the listed options for "{f.get("label")}".')
        if ftype == "number":
            try:
                float(value.replace(",", ""))
            except ValueError:
                raise HTTPException(status_code=400, detail=f'"{f.get("label")}" must be a number.')
        out.append({"key": key, "label": f.get("label"), "type": ftype, "value": value})
        if ftype == "email" and not contact["email"]:
            contact["email"] = value.lower()[:200]
        elif ftype == "phone" and not contact["phone"]:
            contact["phone"] = value[:40]
        elif ftype == "text":
            if first_text_value is None:
                first_text_value = value
            if not contact["name"] and "name" in (str(key) + str(f.get("label") or "").lower()):
                contact["name"] = value[:120]
    if not out:
        raise HTTPException(status_code=400, detail="Please fill in the form before submitting.")
    if not contact["name"] and first_text_value:
        contact["name"] = first_text_value[:120]
    return out, contact


def _require_lead_access(db, user, request, *, capability: str):
    """Membership gate for every lead surface.

    A lead is an unassigned enquiry — it has no branch and no counselor yet, so
    there is nothing to scope it by. Rather than quietly hand a branch- or
    caseload-scoped member the whole workspace's contact details (every other
    client surface narrows to their scope), the inbox is workspace-scope only.
    """
    membership, organization, role = _require_enterprise_membership(
        db=db, user=user, request=request, require_capability=capability
    )
    ctx = role.ctx
    if not (ctx.is_admin_like or ctx.is_org_scope):
        raise HTTPException(
            status_code=403,
            detail="Leads cover the whole workspace, so they're only visible to members whose access scope is the entire workspace. Ask a workspace admin.",
        )
    return membership, organization, role


def _get_org_lead_form_or_404(db: Session, organization_id: int, form_id: int) -> models.EnterpriseLeadForm:
    form = (
        db.query(models.EnterpriseLeadForm)
        .filter(
            models.EnterpriseLeadForm.id == int(form_id),
            models.EnterpriseLeadForm.organization_id == int(organization_id),
        )
        .first()
    )
    if not form:
        raise HTTPException(status_code=404, detail="Form not found.")
    return form


def _get_org_lead_or_404(db: Session, organization_id: int, lead_id: int) -> models.EnterpriseLead:
    lead = (
        db.query(models.EnterpriseLead)
        .filter(
            models.EnterpriseLead.id == int(lead_id),
            models.EnterpriseLead.organization_id == int(organization_id),
        )
        .first()
    )
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found.")
    return lead


def _serialize_lead_form(
    form: models.EnterpriseLeadForm,
    organization: models.EnterpriseOrganization,
    request: Request | None,
    *,
    total_leads: int = 0,
    new_leads: int = 0,
) -> dict:
    return {
        "id": form.id,
        "name": form.name,
        "title": form.title or form.name,
        "intro_text": form.intro_text,
        "fields": _parse_lead_form_fields(form),
        "is_active": bool(form.is_active),
        "submit_label": form.submit_label,
        "success_message": form.success_message,
        "notify_email": form.notify_email,
        "link": _build_lead_form_url(organization.subdomain_slug, form.public_token, request),
        "total_leads": int(total_leads or 0),
        "new_leads": int(new_leads or 0),
        "created_by_name": form.created_by_name,
        "created_at": _iso(form.created_at),
    }


def _serialize_lead(lead: models.EnterpriseLead, files: Optional[list] = None) -> dict:
    try:
        answers = json.loads(lead.answers_json or "[]")
        if not isinstance(answers, list):
            answers = []
    except Exception:
        answers = []
    return {
        "id": lead.id,
        "form_id": lead.form_id,
        "form_name": lead.form_name,
        "full_name": lead.full_name,
        "email": lead.email,
        "phone": lead.phone,
        "answers": answers,
        # Never the storage key — the bytes are only reachable through the
        # capability-gated download endpoint, same rule as client documents.
        "files": [
            {
                "id": u.id,
                "field_key": u.field_key,
                "field_label": u.field_label,
                "filename": u.original_filename,
                "file_size": u.file_size,
                "converted_document_id": u.converted_document_id,
            }
            for u in (files or [])
        ],
        "status": lead.status or "new",
        "converted_client_id": lead.converted_client_id,
        "source": lead.source,
        "created_at": _iso(lead.created_at),
    }


def _lead_files_map(db: Session, organization_id: int, lead_ids: list[int]) -> dict:
    """Attached files for a page of leads, in one query (the inbox lists up to 200)."""
    if not lead_ids:
        return {}
    rows = (
        db.query(models.EnterpriseLeadUpload)
        .filter(
            models.EnterpriseLeadUpload.organization_id == int(organization_id),
            models.EnterpriseLeadUpload.lead_id.in_(lead_ids),
        )
        .order_by(models.EnterpriseLeadUpload.id.asc())
        .all()
    )
    out: dict = {}
    for row in rows:
        out.setdefault(row.lead_id, []).append(row)
    return out


def _lead_files(db: Session, lead: models.EnterpriseLead) -> list:
    return _lead_files_map(db, lead.organization_id, [lead.id]).get(lead.id, [])


def _copy_lead_files_to_client(
    db: Session,
    *,
    organization: models.EnterpriseOrganization,
    lead: models.EnterpriseLead,
    client: models.EnterpriseClient,
) -> list[dict]:
    """Copy a converted lead's attachments into the client's document locker.

    The bytes are COPIED to a client-scoped key rather than re-pointed at the
    lead's: deleting the lead afterwards drops its blobs, and a client document
    must never dangle. `converted_document_id` makes this idempotent, so a lead
    re-linked to another client can't duplicate the same file.

    Returns the text-extraction jobs to start AFTER the caller commits (the worker
    opens its own session and cannot see uncommitted rows). One unreadable file is
    logged and skipped — a storage hiccup must not block the conversion itself.
    """
    if not enterprise_storage.is_configured():
        return []
    jobs: list[dict] = []
    rows = (
        db.query(models.EnterpriseLeadUpload)
        .filter(
            models.EnterpriseLeadUpload.organization_id == organization.id,
            models.EnterpriseLeadUpload.lead_id == lead.id,
            models.EnterpriseLeadUpload.converted_document_id.is_(None),
        )
        .order_by(models.EnterpriseLeadUpload.id.asc())
        .all()
    )
    for row in rows:
        try:
            data = enterprise_storage.fetch_document(row.storage_key)
        except Exception:
            logger.exception("Lead file copy: could not read upload id=%s (lead=%s)", row.id, lead.id)
            continue
        ext = os.path.splitext(row.original_filename)[1].lower()
        storage_key = f"enterprise/{organization.id}/clients/{client.id}/{uuid.uuid4().hex}{ext}"
        try:
            enterprise_storage.store_document(storage_key, data, content_type=row.mime_type)
        except Exception:
            logger.exception("Lead file copy: could not store copy of upload id=%s (lead=%s)", row.id, lead.id)
            continue
        doc = models.EnterpriseClientDocument(
            organization_id=organization.id,
            client_id=client.id,
            document_type=catalog.normalize_document_type(row.field_label),
            original_filename=row.original_filename,
            storage_key=storage_key,
            file_size=row.file_size,
            mime_type=row.mime_type,
            uploaded_by_user_id=None,
            uploaded_by_name=f"{client.full_name} (attached to their enquiry)",
        )
        db.add(doc)
        db.flush()
        row.converted_document_id = doc.id
        # Text extraction only — never the billed scan. Nobody chose to spend the
        # org's credits here, exactly as with the client secure-link upload.
        jobs.append({
            "document_id": doc.id,
            "data": data,
            "filename": row.original_filename,
            "mime_type": row.mime_type,
        })
    return jobs


class EnterpriseLeadFormBody(BaseModel):
    name: str = Field(..., min_length=2, max_length=80)
    title: Optional[str] = Field(default=None, max_length=120)
    intro_text: Optional[str] = Field(default=None, max_length=600)
    fields: list = Field(default_factory=list)
    submit_label: Optional[str] = Field(default=None, max_length=40)
    success_message: Optional[str] = Field(default=None, max_length=400)
    notify_email: Optional[str] = Field(default=None, max_length=200)
    is_active: Optional[bool] = None


class EnterpriseLeadFormPatchBody(BaseModel):
    """Partial update — every field optional, so a pause toggle can send only
    `is_active`. (Reusing the create body here made `name` required and 422'd
    every partial PATCH before the handler ever ran.)"""
    name: Optional[str] = Field(default=None, min_length=2, max_length=80)
    title: Optional[str] = Field(default=None, max_length=120)
    intro_text: Optional[str] = Field(default=None, max_length=600)
    fields: Optional[list] = None
    submit_label: Optional[str] = Field(default=None, max_length=40)
    success_message: Optional[str] = Field(default=None, max_length=400)
    notify_email: Optional[str] = Field(default=None, max_length=200)
    is_active: Optional[bool] = None


class EnterpriseLeadFormShareEmailBody(BaseModel):
    to_email: EmailStr
    note: Optional[str] = Field(default=None, max_length=400)


class EnterpriseLeadStatusBody(BaseModel):
    status: str = Field(..., max_length=20)


class EnterpriseLeadMarkConvertedBody(BaseModel):
    client_id: int


class PublicLeadSubmitBody(BaseModel):
    answers: dict = Field(default_factory=dict)
    # Honeypot — real visitors never see or fill this input. Bots that do get a
    # cheerful success response and nothing stored.
    website: Optional[str] = Field(default=None, max_length=200)
    source: Optional[str] = Field(default=None, max_length=200)
    cf_turnstile_token: Optional[str] = Field(default=None, max_length=4000)


def _clean_lead_notify_email(value) -> Optional[str]:
    email = str(value or "").strip().lower()
    if not email:
        return None
    if not _LEAD_EMAIL_RE.fullmatch(email):
        raise HTTPException(status_code=400, detail="The notification email address doesn't look valid.")
    return email[:200]


@router.get("/lead-forms")
def enterprise_list_lead_forms(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_lead_access(db, current_user, request, capability="clients.view")
    forms = (
        db.query(models.EnterpriseLeadForm)
        .filter(models.EnterpriseLeadForm.organization_id == organization.id)
        .order_by(models.EnterpriseLeadForm.created_at.desc(), models.EnterpriseLeadForm.id.desc())
        .all()
    )
    totals = dict(
        db.query(models.EnterpriseLead.form_id, func.count(models.EnterpriseLead.id))
        .filter(models.EnterpriseLead.organization_id == organization.id, models.EnterpriseLead.form_id.isnot(None))
        .group_by(models.EnterpriseLead.form_id)
        .all()
    )
    fresh = dict(
        db.query(models.EnterpriseLead.form_id, func.count(models.EnterpriseLead.id))
        .filter(
            models.EnterpriseLead.organization_id == organization.id,
            models.EnterpriseLead.form_id.isnot(None),
            models.EnterpriseLead.status == "new",
        )
        .group_by(models.EnterpriseLead.form_id)
        .all()
    )
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "forms": [
            _serialize_lead_form(
                f, organization, request,
                total_leads=totals.get(f.id, 0), new_leads=fresh.get(f.id, 0),
            )
            for f in forms
        ],
    }


@router.post("/lead-forms")
def enterprise_create_lead_form(
    payload: EnterpriseLeadFormBody,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_lead_access(db, current_user, request, capability="clients.edit")
    fields = _normalize_lead_form_fields(payload.fields)
    form = models.EnterpriseLeadForm(
        organization_id=organization.id,
        name=(payload.name or "").strip()[:80],
        title=(payload.title or "").strip()[:120] or None,
        intro_text=(payload.intro_text or "").strip()[:600] or None,
        fields_json=json.dumps(fields),
        public_token=generate_verification_token(),
        is_active=payload.is_active if payload.is_active is not None else True,
        submit_label=(payload.submit_label or "").strip()[:40] or None,
        success_message=(payload.success_message or "").strip()[:400] or None,
        notify_email=_clean_lead_notify_email(payload.notify_email),
        created_by_user_id=current_user.id,
        created_by_name=(current_user.full_name or current_user.email or "")[:120] or None,
    )
    db.add(form)
    db.commit()
    db.refresh(form)
    return {
        "message": "Form created. Share the link to start collecting leads.",
        "permissions": _enterprise_permissions_for_role(role),
        "form": _serialize_lead_form(form, organization, request),
    }


@router.patch("/lead-forms/{form_id}")
def enterprise_update_lead_form(
    form_id: int,
    payload: EnterpriseLeadFormPatchBody,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_lead_access(db, current_user, request, capability="clients.edit")
    form = _get_org_lead_form_or_404(db, organization.id, form_id)
    data = payload.model_dump(exclude_unset=True)
    if "name" in data:
        form.name = (payload.name or "").strip()[:80] or form.name
    if "title" in data:
        form.title = (payload.title or "").strip()[:120] or None
    if "intro_text" in data:
        form.intro_text = (payload.intro_text or "").strip()[:600] or None
    if "fields" in data:
        form.fields_json = json.dumps(_normalize_lead_form_fields(payload.fields))
    if "submit_label" in data:
        form.submit_label = (payload.submit_label or "").strip()[:40] or None
    if "success_message" in data:
        form.success_message = (payload.success_message or "").strip()[:400] or None
    if "notify_email" in data:
        form.notify_email = _clean_lead_notify_email(payload.notify_email)
    if "is_active" in data and payload.is_active is not None:
        form.is_active = bool(payload.is_active)
    db.commit()
    db.refresh(form)
    return {
        "message": "Form updated.",
        "permissions": _enterprise_permissions_for_role(role),
        "form": _serialize_lead_form(form, organization, request),
    }


@router.post("/lead-forms/{form_id}/rotate-link")
def enterprise_rotate_lead_form_link(
    form_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Mint a fresh link. The old URL stops resolving immediately — the recovery
    path if a link ends up somewhere the org regrets."""
    _, organization, role = _require_lead_access(db, current_user, request, capability="clients.edit")
    form = _get_org_lead_form_or_404(db, organization.id, form_id)
    form.public_token = generate_verification_token()
    db.commit()
    db.refresh(form)
    return {
        "message": "New link generated. The old link no longer works.",
        "permissions": _enterprise_permissions_for_role(role),
        "form": _serialize_lead_form(form, organization, request),
    }


@router.delete("/lead-forms/{form_id}")
def enterprise_delete_lead_form(
    form_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_lead_access(db, current_user, request, capability="clients.delete")
    form = _get_org_lead_form_or_404(db, organization.id, form_id)
    # Detach leads explicitly (not via FK ON DELETE) — the sqlite sandbox never
    # enforces foreign keys, and collected leads must survive their form.
    db.query(models.EnterpriseLead).filter(
        models.EnterpriseLead.organization_id == organization.id,
        models.EnterpriseLead.form_id == form.id,
    ).update({"form_id": None}, synchronize_session=False)
    # Files already attached to a lead are detached with it. Files still STAGED on
    # this form belong to a submission that will now never arrive, so they are the
    # one thing here that is genuinely orphaned — take them with the form.
    staged = (
        db.query(models.EnterpriseLeadUpload)
        .filter(
            models.EnterpriseLeadUpload.organization_id == organization.id,
            models.EnterpriseLeadUpload.form_id == form.id,
            models.EnterpriseLeadUpload.lead_id.is_(None),
        )
        .all()
    )
    storage_keys = [u.storage_key for u in staged]
    for u in staged:
        db.delete(u)
    db.query(models.EnterpriseLeadUpload).filter(
        models.EnterpriseLeadUpload.organization_id == organization.id,
        models.EnterpriseLeadUpload.form_id == form.id,
    ).update({"form_id": None}, synchronize_session=False)
    db.delete(form)
    db.commit()
    _drop_storage_keys(storage_keys, context=f"deleted form {form_id}")
    return {"message": "Form deleted. Collected leads were kept.", "permissions": _enterprise_permissions_for_role(role)}


@router.post("/lead-forms/{form_id}/share-email")
def enterprise_share_lead_form_email(
    form_id: int,
    payload: EnterpriseLeadFormShareEmailBody,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_lead_access(db, current_user, request, capability=("clients.edit", "emails.send"))
    form = _get_org_lead_form_or_404(db, organization.id, form_id)
    if not form.is_active:
        raise HTTPException(status_code=400, detail="This form is paused — resume it before sharing.")
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.lead_form_share", limit=30, window_seconds=3600,
        extra_key=f"org:{organization.id}",
    )
    link = _build_lead_form_url(organization.subdomain_slug, form.public_token, request)
    sent, _mid, err = send_enterprise_lead_form_link_email(
        to_email=str(payload.to_email),
        organization_name=organization.company_name,
        # Same rule as the public page: the internal name never reaches an outsider's
        # subject line.
        form_title=(form.title or "").strip() or "Get in touch",
        form_url=link,
        sender_name=current_user.full_name or current_user.email,
        note=payload.note,
        logo_url=_absolute_enterprise_logo_url(organization),
        reply_to=current_user.email,
    )
    message = (f"Form link sent to {str(payload.to_email).strip().lower()}."
               if sent else f"The email could not be sent right now. {err or ''}".strip())
    return {"message": message, "email_sent": sent, "permissions": _enterprise_permissions_for_role(role)}


@router.get("/leads")
def enterprise_list_leads(
    request: Request,
    status_filter: Optional[str] = None,
    form_id: Optional[int] = None,
    q: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_lead_access(db, current_user, request, capability="clients.view")
    base = db.query(models.EnterpriseLead).filter(models.EnterpriseLead.organization_id == organization.id)
    counts = dict(
        db.query(models.EnterpriseLead.status, func.count(models.EnterpriseLead.id))
        .filter(models.EnterpriseLead.organization_id == organization.id)
        .group_by(models.EnterpriseLead.status)
        .all()
    )
    if status_filter and status_filter in _LEAD_STATUSES:
        base = base.filter(models.EnterpriseLead.status == status_filter)
    if form_id:
        base = base.filter(models.EnterpriseLead.form_id == int(form_id))
    needle = (q or "").strip()
    if needle:
        like = f"%{needle}%"
        base = base.filter(or_(
            models.EnterpriseLead.full_name.ilike(like),
            models.EnterpriseLead.email.ilike(like),
            models.EnterpriseLead.phone.ilike(like),
        ))
    total = base.count()
    leads = (
        base.order_by(models.EnterpriseLead.created_at.desc(), models.EnterpriseLead.id.desc())
        .offset(max(0, int(offset)))
        .limit(min(max(1, int(limit)), 200))
        .all()
    )
    files_by_lead = _lead_files_map(db, organization.id, [l.id for l in leads])
    return {
        "permissions": _enterprise_permissions_for_role(role),
        "leads": [_serialize_lead(l, files_by_lead.get(l.id)) for l in leads],
        "total": int(total),
        "counts": {s: int(counts.get(s, 0)) for s in _LEAD_STATUSES},
    }


@router.patch("/leads/{lead_id}/status")
def enterprise_update_lead_status(
    lead_id: int,
    payload: EnterpriseLeadStatusBody,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_lead_access(db, current_user, request, capability="clients.edit")
    lead = _get_org_lead_or_404(db, organization.id, lead_id)
    new_status = (payload.status or "").strip().lower()
    if new_status not in _LEAD_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid lead status.")
    lead.status = new_status
    db.commit()
    return {
        "message": "Lead updated.",
        "permissions": _enterprise_permissions_for_role(role),
        "lead": _serialize_lead(lead, _lead_files(db, lead)),
    }


@router.post("/leads/{lead_id}/mark-converted")
def enterprise_mark_lead_converted(
    lead_id: int,
    payload: EnterpriseLeadMarkConvertedBody,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Link a lead to the client it became. The client itself is created through the
    standard POST /clients flow (billing gates, consent attestation and all) — this
    endpoint only records the outcome on the lead, and carries any files the
    prospect attached into the new client's document locker."""
    _, organization, role = _require_lead_access(db, current_user, request, capability="clients.create")
    lead = _get_org_lead_or_404(db, organization.id, lead_id)
    client = _get_org_client_or_404(db, organization.id, payload.client_id, ctx=role.ctx)
    lead.status = "converted"
    lead.converted_client_id = client.id
    jobs = _copy_lead_files_to_client(db, organization=organization, lead=lead, client=client)
    db.commit()
    for job in jobs:
        _start_document_processing(
            job["document_id"], job["data"], job["filename"], job["mime_type"], validate=False,
        )
    copied = len(jobs)
    message = "Lead marked as converted."
    if copied:
        message += f" {copied} attached file{'s' if copied != 1 else ''} moved into the client's documents."
    return {
        "message": message,
        "permissions": _enterprise_permissions_for_role(role),
        "lead": _serialize_lead(lead, _lead_files(db, lead)),
        "documents_copied": copied,
    }


@router.get("/leads/{lead_id}/files/{file_id}/download")
def enterprise_download_lead_file(
    lead_id: int,
    file_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Stream a file a prospect attached to their enquiry.

    Gated on documents.download, not the inbox's clients.view: these are passport
    pages and bank letters exactly like client documents, so a member who may read
    the inbox but not pull raw files still can't. Resolving the LEAD first is what
    stops sequential file ids being walked against another lead."""
    _, organization, role = _require_lead_access(db, current_user, request, capability="documents.download")
    lead = _get_org_lead_or_404(db, organization.id, lead_id)
    row = (
        db.query(models.EnterpriseLeadUpload)
        .filter(
            models.EnterpriseLeadUpload.id == int(file_id),
            models.EnterpriseLeadUpload.lead_id == lead.id,
            models.EnterpriseLeadUpload.organization_id == organization.id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="File not found.")

    try:
        data = enterprise_storage.fetch_document(row.storage_key)
    except Exception:
        logger.exception("Failed to fetch lead upload id=%s", row.id)
        raise HTTPException(status_code=502, detail="Could not retrieve the file right now.")

    # Content-Type from the VALIDATED extension, never the uploader-supplied mime —
    # and here the uploader is an anonymous stranger, so this matters more than
    # anywhere else in the product. Only known-safe visual types render inline.
    ext = os.path.splitext(row.original_filename)[1].lower()
    if ext in ENTERPRISE_DOC_INLINE_EXT:
        disposition = "inline"
        media_type = ENTERPRISE_DOC_EXT_MIME.get(ext, "application/octet-stream")
    else:
        disposition = "attachment"
        media_type = "application/octet-stream"
    filename = _safe_filename(row.original_filename)
    return Response(
        content=data,
        media_type=media_type,
        headers={
            "Content-Disposition": f'{disposition}; filename="{filename}"',
            "X-Content-Type-Options": "nosniff",
            "Cache-Control": "private, no-store",
        },
    )


@router.delete("/leads/{lead_id}")
def enterprise_delete_lead(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    _, organization, role = _require_lead_access(db, current_user, request, capability="clients.delete")
    lead = _get_org_lead_or_404(db, organization.id, lead_id)
    # Drop the attachment rows explicitly rather than relying on ON DELETE CASCADE
    # (the sqlite sandbox never enforces FKs) and collect their keys — the blobs go
    # only AFTER the commit, so a failed delete can't strand rows pointing at
    # bytes that no longer exist. Files already copied into a client's locker are
    # untouched: that copy is the client's document now.
    uploads = _lead_files(db, lead)
    storage_keys = [u.storage_key for u in uploads]
    for u in uploads:
        db.delete(u)
    db.delete(lead)
    db.commit()
    _drop_storage_keys(storage_keys, context=f"lead {lead_id}")
    return {"message": "Lead deleted.", "permissions": _enterprise_permissions_for_role(role)}


# ---- Public (anonymous, token-scoped) --------------------------------------

def _public_lead_form_or_404(db: Session, token: str) -> models.EnterpriseLeadForm:
    clean = (token or "").strip()
    if not clean or len(clean) > 80:
        raise HTTPException(status_code=404, detail="This form link is invalid.")
    form = (
        db.query(models.EnterpriseLeadForm)
        .filter(models.EnterpriseLeadForm.public_token == clean)
        .first()
    )
    if not form:
        raise HTTPException(status_code=404, detail="This form link is invalid.")
    return form


@router.get("/public/forms/{token}")
def public_lead_form_info(token: str, request: Request, db: Session = Depends(get_db)):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.lead_form_view", limit=30, window_seconds=300,
        extra_key=hash_token((token or "").strip())[:16],
    )
    form = _public_lead_form_or_404(db, token)
    org = (
        db.query(models.EnterpriseOrganization)
        .filter(models.EnterpriseOrganization.id == form.organization_id)
        .first()
    )
    if not org:
        raise HTTPException(status_code=404, detail="This form link is invalid.")
    closed = not bool(form.is_active)
    return {
        "organization_name": org.company_name,
        "logo_url": _resolve_enterprise_logo_url(org),
        "closed": closed,
        # Only the site key (public by definition) — never the secret.
        "turnstile_site_key": (os.getenv("TURNSTILE_SITE_KEY", "").strip() if _is_turnstile_required() else ""),
        # Lets the page say "uploads are unavailable" instead of silently failing on
        # every file pick when storage isn't configured. The form stays submittable.
        "uploads_enabled": enterprise_storage.is_configured(),
        "upload_max_bytes": _LEAD_UPLOAD_MAX_BYTES,
        "upload_allowed_ext": sorted(ENTERPRISE_DOC_ALLOWED_EXT),
        "form": {
            # The form's `name` is the org's internal label ("Instagram junk leads") and
            # must never surface to a prospect — an unset public title falls back to a
            # neutral heading instead.
            "title": (form.title or "").strip() or "Get in touch",
            "intro_text": form.intro_text,
            "fields": [] if closed else _parse_lead_form_fields(form),
            "submit_label": form.submit_label or "Submit",
            "success_message": form.success_message,
        },
    }


def _sweep_staged_lead_uploads(db: Session, *, form_id: int) -> list[str]:
    """Mark this form's abandoned staged uploads for deletion (the visitor picked a
    file and then closed the tab) and return their storage keys.

    Opportunistic — it runs on the upload path, which is the only path that creates
    these rows, so a form that stops receiving uploads has nothing left to sweep.
    Rows only; the caller drops the blobs after the commit."""
    cutoff = datetime.now(dt_timezone.utc) - timedelta(hours=_LEAD_UPLOAD_STAGED_TTL_HOURS)
    keys: list[str] = []
    try:
        stale = (
            db.query(models.EnterpriseLeadUpload)
            .filter(
                models.EnterpriseLeadUpload.form_id == int(form_id),
                models.EnterpriseLeadUpload.lead_id.is_(None),
                models.EnterpriseLeadUpload.created_at < cutoff,
            )
            .limit(50)
            .all()
        )
        for row in stale:
            keys.append(row.storage_key)
            db.delete(row)
    except Exception:
        logger.exception("Staged lead-upload sweep failed (form_id=%s)", form_id)
    return keys


@router.post("/public/forms/{token}/upload")
async def public_lead_form_upload(
    token: str,
    request: Request,
    field_key: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Stage one file for a lead-form file field, before the form is submitted.

    This is the only endpoint in the product where a wholly anonymous visitor
    writes bytes into an org's bucket, so it is deliberately narrow: the field_key
    must name a `file` field on THIS form, the extension and size bars are the
    tight lead-form ones, and both a per-IP rate limit and a per-form 24h ceiling
    apply. The row is created unbound (lead_id null) and is worthless until a
    submission claims it with the opaque token returned here; anything never
    claimed is swept after a few hours.

    Turnstile deliberately does NOT gate this call — its token is single-use and is
    spent at submit, which is the step that actually creates a record. The caps
    above are what hold this endpoint.
    """
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.lead_form_upload", limit=25, window_seconds=1800,
        extra_key=hash_token((token or "").strip())[:16],
    )
    form = _public_lead_form_or_404(db, token)
    if not form.is_active:
        raise HTTPException(status_code=400, detail="This form is no longer accepting responses.")

    field = _lead_file_fields(_parse_lead_form_fields(form)).get((field_key or "").strip())
    if not field:
        raise HTTPException(status_code=400, detail="This form doesn't ask for that file.")

    if not enterprise_storage.is_configured():
        raise HTTPException(status_code=503, detail="File uploads aren't available right now. Please submit without attaching.")

    # Per-form ceiling on top of the per-IP limit: rotating proxies must not be able
    # to run up a paying org's storage bill one IP at a time.
    since = datetime.utcnow() - timedelta(hours=24)
    recent = int((
        db.query(func.count(models.EnterpriseLeadUpload.id))
        .filter(
            models.EnterpriseLeadUpload.form_id == form.id,
            models.EnterpriseLeadUpload.created_at >= since,
        )
        .scalar()
    ) or 0)
    if recent >= _LEAD_UPLOAD_FORM_DAILY_CAP:
        raise HTTPException(status_code=429, detail="This form is receiving too many files right now. Please try again later.")

    original = _safe_filename(file.filename)
    ext = os.path.splitext(original)[1].lower()
    if ext not in ENTERPRISE_DOC_ALLOWED_EXT:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Allowed: PDF, images, Word/Excel, CSV, or text.",
        )

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="That file is empty.")
    if len(data) > _LEAD_UPLOAD_MAX_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"That file is too large. Maximum size is {_LEAD_UPLOAD_MAX_BYTES // (1024 * 1024)} MB.",
        )

    swept = _sweep_staged_lead_uploads(db, form_id=form.id)

    storage_key = f"enterprise/{form.organization_id}/leads/{uuid.uuid4().hex}{ext}"
    try:
        await run_in_threadpool(enterprise_storage.store_document, storage_key, data, content_type=file.content_type)
    except Exception:
        logger.exception("Failed to store lead-form upload (org_id=%s, form_id=%s)", form.organization_id, form.id)
        raise HTTPException(status_code=502, detail="Could not upload that file right now. Please try again.")

    row = models.EnterpriseLeadUpload(
        organization_id=form.organization_id,
        form_id=form.id,
        lead_id=None,
        field_key=field["key"],
        field_label=(field.get("label") or "")[:120] or None,
        upload_token=secrets.token_urlsafe(32),
        original_filename=original,
        storage_key=storage_key,
        file_size=len(data),
        mime_type=(file.content_type or None),
        ip_address=(extract_client_ip(request) if request else None),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    _drop_storage_keys(swept, context=f"staged sweep on form {form.id}")

    return {
        "upload_token": row.upload_token,
        "filename": row.original_filename,
        "file_size": row.file_size,
        "field_key": row.field_key,
    }


def _resolve_lead_upload_tokens(
    db: Session,
    *,
    form: models.EnterpriseLeadForm,
    fields: list[dict],
    raw_answers,
) -> tuple[dict, list[int]]:
    """Turn the submission's opaque upload tokens into the staged rows they name.

    Nothing here trusts the browser beyond the token itself: a row only counts if
    it belongs to THIS form, to THIS field, and has not already been claimed by an
    earlier submission. Returns the {field_key: [rows]} map for the answer
    validator and the flat id list to bind once the lead exists.
    """
    answers = raw_answers if isinstance(raw_answers, dict) else {}
    file_fields = _lead_file_fields(fields)
    by_key: dict = {}
    all_ids: list[int] = []
    for key, field in file_fields.items():
        raw_v = answers.get(key)
        if raw_v is None or raw_v == "":
            continue
        tokens = raw_v if isinstance(raw_v, list) else [raw_v]
        tokens = list(dict.fromkeys(str(t or "").strip() for t in tokens if str(t or "").strip()))
        if not tokens:
            continue
        if len(tokens) > _lead_field_max_files(field):
            raise HTTPException(
                status_code=400,
                detail=f'You can attach at most {_lead_field_max_files(field)} file(s) to "{field.get("label")}".',
            )
        rows = (
            db.query(models.EnterpriseLeadUpload)
            .filter(
                models.EnterpriseLeadUpload.upload_token.in_(tokens),
                models.EnterpriseLeadUpload.form_id == form.id,
                models.EnterpriseLeadUpload.field_key == key,
                models.EnterpriseLeadUpload.lead_id.is_(None),
            )
            .order_by(models.EnterpriseLeadUpload.id.asc())
            .all()
        )
        if len(rows) != len(tokens):
            raise HTTPException(
                status_code=400,
                detail=f'One of the files you attached to "{field.get("label")}" is no longer available. Please attach it again.',
            )
        by_key[key] = rows
        all_ids += [r.id for r in rows]
    return by_key, all_ids


@router.post("/public/forms/{token}/submit")
def public_lead_form_submit(
    token: str,
    payload: PublicLeadSubmitBody,
    request: Request,
    db: Session = Depends(get_db),
):
    _enforce_rate_limit_or_429(
        request=request, scope="enterprise.lead_form_submit", limit=10, window_seconds=3600,
        extra_key=hash_token((token or "").strip())[:16],
    )
    form = _public_lead_form_or_404(db, token)
    if not form.is_active:
        raise HTTPException(status_code=400, detail="This form is no longer accepting responses.")
    org = (
        db.query(models.EnterpriseOrganization)
        .filter(models.EnterpriseOrganization.id == form.organization_id)
        .first()
    )
    if not org:
        raise HTTPException(status_code=404, detail="This form link is invalid.")

    success_message = form.success_message or "Thanks! Your details were sent — the team will get back to you shortly."

    # Honeypot tripped: pretend success, store nothing.
    if (payload.website or "").strip():
        return {"message": success_message}

    turnstile_token = (payload.cf_turnstile_token or "").strip()
    if _is_turnstile_required() and not is_request_ip_whitelisted(request):
        if not turnstile_token or not verify_turnstile_token(turnstile_token, extract_client_ip(request) if request else None):
            raise HTTPException(status_code=400, detail="Security verification failed. Please reload the page and try again.")

    # Rolling 24h volume check. Deliberately fail-soft: a proxy-rotating bot must not
    # be able to shut a paying org's form to real prospects, so past the daily cap we
    # keep accepting and only mute the alert emails (which go out over the shared
    # transactional domain). Only the far-higher hard ceiling refuses outright.
    since = datetime.utcnow() - timedelta(hours=24)
    recent = int((
        db.query(func.count(models.EnterpriseLead.id))
        .filter(models.EnterpriseLead.form_id == form.id, models.EnterpriseLead.created_at >= since)
        .scalar()
    ) or 0)
    if recent >= _LEAD_FORM_HARD_CAP:
        raise HTTPException(status_code=429, detail="This form is receiving too many responses right now. Please try again later.")
    flooded = recent >= _LEAD_FORM_DAILY_CAP

    fields = _parse_lead_form_fields(form)
    uploads_by_key, upload_ids = _resolve_lead_upload_tokens(
        db, form=form, fields=fields, raw_answers=payload.answers
    )
    answers, contact = _validate_lead_answers(
        fields, payload.answers, uploads_by_key,
        uploads_enabled=enterprise_storage.is_configured(),
    )

    lead = models.EnterpriseLead(
        organization_id=org.id,
        form_id=form.id,
        form_name=form.name,
        full_name=contact["name"],
        email=contact["email"],
        phone=contact["phone"],
        answers_json=json.dumps(answers),
        status="new",
        ip_address=(extract_client_ip(request) if request else None),
        source=(payload.source or "").strip()[:200] or None,
    )
    db.add(lead)
    if upload_ids:
        # Claim the staged files in the SAME transaction as the lead, filtering on
        # lead_id IS NULL again so two submissions racing on the same tokens can't
        # both take them. A short count means the other one won — roll back rather
        # than store a lead whose answers name files it doesn't own.
        db.flush()
        bound = (
            db.query(models.EnterpriseLeadUpload)
            .filter(
                models.EnterpriseLeadUpload.id.in_(upload_ids),
                models.EnterpriseLeadUpload.lead_id.is_(None),
            )
            .update({"lead_id": lead.id}, synchronize_session=False)
        )
        if bound != len(upload_ids):
            db.rollback()
            raise HTTPException(
                status_code=400,
                detail="Your attached files are no longer available. Please attach them again and resubmit.",
            )
    db.commit()
    db.refresh(lead)

    # Everything below is best-effort — the lead is already safe in the inbox.
    notif.notify_org(
        db, org.id, type="lead_received",
        title=f"New lead: {lead.full_name or 'Someone'} via {form.name}",
        body=(lead.email or lead.phone or None),
        actor_user_id=None, reference_type="lead", reference_id=lead.id,
        recipient_capability="clients.view", commit=True,
    )
    if form.notify_email and not flooded:
        try:
            leads_url = (
                (_build_enterprise_portal_url(org.subdomain_slug) or f"{DEFAULT_PUBLIC_BASE_URL.rstrip('/')}/enterprise")
                + "/leads"
            )
            send_enterprise_new_lead_email(
                to_email=form.notify_email,
                organization_name=org.company_name,
                form_name=form.name,
                lead_name=lead.full_name,
                answers=[(a.get("label"), a.get("value")) for a in answers],
                leads_url=leads_url,
                logo_url=_absolute_enterprise_logo_url(org),
                lead_email=lead.email,
            )
        except Exception:
            logger.exception("New-lead alert email failed (lead=%s)", lead.id)

    return {"message": success_message}

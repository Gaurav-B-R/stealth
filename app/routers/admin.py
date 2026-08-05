import calendar
import os
import logging
import secrets
import string
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from fastapi import APIRouter, Body, Depends, HTTPException, Query, Request, Response, status
from jose import JWTError, jwt
from sqlalchemy import bindparam, case, desc, func, inspect, or_, select, text
from sqlalchemy.orm import Session, aliased

from app import models, schemas
from app import ai_usage
from app import acquisition
from app import enterprise_credits
from app import enterprise_billing
from app import enterprise_coupons
from app import enterprise_refunds
from app import email_service
from app import fx
from app import money
from app import visa_pass
from app import ai_guardrails
from app.auth import (
    ALGORITHM,
    AUTH_COOKIE_DOMAIN,
    AUTH_COOKIE_SAMESITE,
    SECRET_KEY,
    _resolve_auth_cookie_secure,
    get_password_hash,
    get_current_admin_user,
)
from app.database import get_db
from app.utils.rate_limiter import check_ip_rate_limit, extract_client_ip, is_request_ip_whitelisted
from app.utils.turnstile import is_turnstile_enabled, verify_turnstile_token

router = APIRouter(prefix="/api/admin", tags=["admin"])
logger = logging.getLogger(__name__)

VALID_USER_PLAN_FILTERS = {"all", "free", "visa_pass"}
VALID_USER_ROLE_FILTERS = {"all", "student", "staff", "admin", "developer"}
ENTERPRISE_ROLE_ADMIN = "admin"
ENTERPRISE_ROOT_DOMAIN = (
    os.getenv("ENTERPRISE_ROOT_DOMAIN", "rilono.com").strip().lower() or "rilono.com"
).lstrip(".")
ENTERPRISE_PORTAL_SCHEME = os.getenv("ENTERPRISE_PORTAL_SCHEME", "https").strip().lower() or "https"
ENTERPRISE_PORTAL_PORT = os.getenv("ENTERPRISE_PORTAL_PORT", "").strip().lstrip(":")
ADMIN_TURNSTILE_COOKIE_NAME = (
    os.getenv("ADMIN_TURNSTILE_COOKIE_NAME", "rilono_admin_turnstile").strip()
    or "rilono_admin_turnstile"
)
ADMIN_TURNSTILE_COOKIE_TTL_MINUTES = max(
    5,
    int(os.getenv("ADMIN_TURNSTILE_COOKIE_TTL_MINUTES", "45") or "45"),
)
ADMIN_TURNSTILE_VERIFY_RATE_LIMIT = max(
    5,
    int(os.getenv("ADMIN_TURNSTILE_VERIFY_RATE_LIMIT", "15") or "15"),
)
ADMIN_TURNSTILE_VERIFY_RATE_WINDOW_SECONDS = max(
    10,
    int(os.getenv("ADMIN_TURNSTILE_VERIFY_RATE_WINDOW_SECONDS", "60") or "60"),
)
ADMIN_ENDPOINT_RATE_LIMIT = max(
    20,
    int(os.getenv("ADMIN_ENDPOINT_RATE_LIMIT", "240") or "240"),
)
ADMIN_ENDPOINT_RATE_WINDOW_SECONDS = max(
    10,
    int(os.getenv("ADMIN_ENDPOINT_RATE_WINDOW_SECONDS", "60") or "60"),
)
# Last-resort INR → USD rate for the company-finance report (≈ ₹83.3/$). Used only when
# the shared FX layer has no rate at all — see _inr_to_usd_rate().
ADMIN_ANALYTICS_INR_TO_USD_FALLBACK = Decimal("0.012")


def _inr_to_usd_rate(fx_state: dict | None = None) -> Decimal:
    """INR → USD, inverted from the ONE live USD → INR rate in app/fx.py.

    This used to be a hardcoded ADMIN_ANALYTICS_INR_TO_USD=0.012 constant, which meant the
    company-finance screen converted rupees at ₹83.3/$ while the enterprise-revenue screen
    (app/enterprise_credits.usd_to_inr_paise → fx.get_usd_to_inr) converted the same rupee
    at the live rate — two admin screens quoting different numbers for the same money.
    Routing both through app/fx.py is the point: it is cached, live, and falls back to the
    static USD_TO_INR env value on its own, so this function never has to guess.

    Pass the caller's already-read `fx_state` so a single report is built at a single rate
    and reports the rate it actually used — see _payment_amount_usd.
    """
    raw = fx_state.get("rate") if isinstance(fx_state, dict) else fx.get_usd_to_inr()
    try:
        usd_to_inr = Decimal(str(raw or 0))
    except (InvalidOperation, TypeError, ValueError):
        usd_to_inr = Decimal("0")
    if usd_to_inr <= 0:
        return ADMIN_ANALYTICS_INR_TO_USD_FALLBACK
    return Decimal(1) / usd_to_inr


def _assert_manageable_target(target_user: models.User, acting_user: models.User) -> None:
    """Guardrails for privileged account management actions."""
    if target_user.id == acting_user.id:
        raise HTTPException(status_code=400, detail="You cannot perform this action on your own account.")

    if target_user.is_developer and not acting_user.is_developer:
        raise HTTPException(status_code=403, detail="Only a developer can manage developer accounts.")

    if target_user.is_admin and not acting_user.is_developer:
        raise HTTPException(status_code=403, detail="Only a developer can manage admin accounts.")


def _is_development_env() -> bool:
    return os.getenv("ENVIRONMENT", "production").strip().lower() == "development"


def _is_admin_turnstile_required(request: Request) -> bool:
    if not is_turnstile_enabled():
        return False
    if _is_development_env():
        return False
    if is_request_ip_whitelisted(request):
        return False
    return True


def _request_port_for_local_enterprise_url(request: Request | None) -> str | None:
    if ENTERPRISE_PORTAL_PORT:
        return ENTERPRISE_PORTAL_PORT
    if not request or not _is_development_env():
        return None
    port = request.url.port
    if not port or int(port) in {80, 443}:
        return None
    return str(port)


def _build_enterprise_portal_url(
    subdomain_slug: str | None,
    request: Request | None = None,
) -> str | None:
    subdomain = str(subdomain_slug or "").strip().lower()
    if not subdomain:
        return None
    host = f"{subdomain}.{ENTERPRISE_ROOT_DOMAIN}"
    port = _request_port_for_local_enterprise_url(request)
    if port:
        host = f"{host}:{port}"
    return f"{ENTERPRISE_PORTAL_SCHEME}://{host}/enterprise"


def _generate_enterprise_temp_password(length: int = 16) -> str:
    """
    Generate a strong temporary password with upper/lowercase letters, numbers, and symbols.
    """
    normalized_length = max(int(length), 12)
    rng = secrets.SystemRandom()
    required = [
        rng.choice(string.ascii_lowercase),
        rng.choice(string.ascii_uppercase),
        rng.choice(string.digits),
        rng.choice("!@#$%^&*()-_=+"),
    ]
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*()-_=+"
    remaining = [rng.choice(alphabet) for _ in range(max(normalized_length - len(required), 0))]
    password_chars = required + remaining
    rng.shuffle(password_chars)
    return "".join(password_chars)


def _enforce_rate_limit_or_429(
    *,
    request: Request,
    scope: str,
    limit: int,
    window_seconds: int,
    extra_key: str | None = None,
) -> None:
    allowed, retry_after = check_ip_rate_limit(
        request=request,
        scope=scope,
        limit=limit,
        window_seconds=window_seconds,
        extra_key=extra_key,
    )
    if not allowed:
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
            headers={"Retry-After": str(retry_after)},
        )


def _build_turnstile_proof_token(user_id: int) -> str:
    now = datetime.now(timezone.utc)
    expires_at = now + timedelta(minutes=ADMIN_TURNSTILE_COOKIE_TTL_MINUTES)
    payload = {
        "purpose": "admin_console",
        "user_id": int(user_id),
        "iat": int(now.timestamp()),
        "exp": int(expires_at.timestamp()),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def _set_admin_turnstile_cookie(*, request: Request, response: Response, user_id: int) -> None:
    response.set_cookie(
        key=ADMIN_TURNSTILE_COOKIE_NAME,
        value=_build_turnstile_proof_token(user_id),
        max_age=ADMIN_TURNSTILE_COOKIE_TTL_MINUTES * 60,
        httponly=True,
        secure=_resolve_auth_cookie_secure(request),
        samesite=AUTH_COOKIE_SAMESITE,
        domain=AUTH_COOKIE_DOMAIN,
        path="/",
    )


def _clear_admin_turnstile_cookie(response: Response) -> None:
    response.delete_cookie(
        key=ADMIN_TURNSTILE_COOKIE_NAME,
        domain=AUTH_COOKIE_DOMAIN,
        path="/",
    )


def require_admin_turnstile_proof(
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
) -> None:
    """
    Enforce a successful Cloudflare Turnstile challenge before admin API access.
    """
    if not _is_admin_turnstile_required(request):
        return

    signed_token = request.cookies.get(ADMIN_TURNSTILE_COOKIE_NAME)
    if not signed_token:
        raise HTTPException(
            status_code=403,
            detail="Cloudflare verification required. Complete the admin protection check.",
        )

    try:
        payload = jwt.decode(signed_token, SECRET_KEY, algorithms=[ALGORITHM])
        purpose = str(payload.get("purpose") or "")
        user_id = int(payload.get("user_id") or 0)
    except (JWTError, ValueError, TypeError):
        raise HTTPException(
            status_code=403,
            detail="Cloudflare verification expired. Complete the admin protection check again.",
        )

    if purpose != "admin_console" or user_id != int(current_user.id):
        raise HTTPException(
            status_code=403,
            detail="Cloudflare verification expired. Complete the admin protection check again.",
        )


@router.post("/turnstile/verify")
def verify_admin_turnstile(
    payload: schemas.AdminTurnstileVerifyRequest,
    request: Request,
    response: Response,
    current_user: models.User = Depends(get_current_admin_user),
):
    """
    Verify Cloudflare Turnstile and mint a short-lived admin protection cookie.
    """
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.turnstile.verify",
        limit=ADMIN_TURNSTILE_VERIFY_RATE_LIMIT,
        window_seconds=ADMIN_TURNSTILE_VERIFY_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )

    if _is_admin_turnstile_required(request):
        token = (payload.token or "").strip()
        if not token:
            raise HTTPException(status_code=400, detail="Turnstile token is required.")
        if not verify_turnstile_token(token, extract_client_ip(request)):
            raise HTTPException(
                status_code=400,
                detail="Cloudflare verification failed. Please try again.",
            )

    _set_admin_turnstile_cookie(request=request, response=response, user_id=current_user.id)
    return {
        "message": "Cloudflare verification completed.",
        "expires_in_seconds": ADMIN_TURNSTILE_COOKIE_TTL_MINUTES * 60,
    }


@router.post("/turnstile/clear")
def clear_admin_turnstile(
    response: Response,
    current_user: models.User = Depends(get_current_admin_user),
):
    """Clear admin protection cookie during logout."""
    del current_user
    _clear_admin_turnstile_cookie(response)
    return {"message": "Cloudflare verification cleared."}


def _build_plan_metrics_for_filtered_users(db: Session, filtered_user_ids_subquery) -> dict[str, int]:
    """
    Plan metrics for the filtered user population under the current B2C model
    (one-time Visa Success Pass replaced the old recurring Pro / Journey plans):
    - visa_pass_users: users with an active premium (Visa Success Pass) subscription
    - free_users: everyone else in the filtered set
    """
    filtered_user_ids_select = select(filtered_user_ids_subquery.c.id)

    total_users = (
        db.query(func.count()).select_from(filtered_user_ids_subquery).scalar() or 0
    )
    visa_pass_users = (
        db.query(func.count(func.distinct(models.Subscription.user_id)))
        .filter(
            models.Subscription.user_id.in_(filtered_user_ids_select),
            models.Subscription.plan == "pro",
            models.Subscription.status == "active",
        )
        .scalar()
        or 0
    )
    return {
        "visa_pass_users": int(visa_pass_users),
        "free_users": max(int(total_users) - int(visa_pass_users), 0),
    }


def _money_decimal(value) -> Decimal:
    try:
        return Decimal(str(value or "0"))
    except InvalidOperation:
        return Decimal("0")


def _money_float(value: Decimal) -> float:
    rounded = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return float(rounded)


def _coerce_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.now(timezone.utc).date()


def _month_key(value) -> str:
    normalized = _coerce_date(value)
    return f"{normalized.year:04d}-{normalized.month:02d}"


def _payment_settled_inr_paise(payment: models.SubscriptionPayment) -> tuple[Decimal, bool]:
    """(INR paise settled, is_settled) for one payment.

    `amount_paise` is an integer in the MINOR UNIT OF `payment.currency` — cents on a USD
    row — so it is not a rupee figure and must never be summed across rows. Razorpay's own
    converted INR settlement lands in `base_amount_paise`; that is the only summable
    number. Falling back to `amount_paise` is valid in exactly two cases:

      * INR payments, where charged == settled (Razorpay omits base_amount entirely), and
      * pre-migration rows, where the column is NULL and every row is genuinely INR.

    A NON-INR row with no settlement figure gets neither: it contributes 0 and comes back
    `is_settled=False`. Reading its 1299 cents as ₹12.99 would book roughly 1% of a $12.99
    sale into total_returns, ROI and break-even — a wrong number wearing a flag is still a
    wrong number — and it would put this screen at odds with the identical rule already
    enforced in visa_pass.build_revenue_analytics, enterprise_credits._kind_money and
    _sum_collected below. Understating is recoverable the moment the webhook backfills
    base_amount; silently mixing units is not. The caller keeps the row VISIBLE at zero
    with the flag attached rather than dropping it, so it reads as "not counted yet"
    instead of "no revenue".
    """
    base = payment.base_amount_paise
    if base is not None:
        return _money_decimal(base), True
    currency = str(payment.currency or "INR").strip().upper()
    if currency == "INR":
        return _money_decimal(payment.amount_paise), True
    return Decimal("0"), False


def _payment_amount_usd(
    payment: models.SubscriptionPayment, inr_to_usd: Decimal
) -> tuple[Decimal, bool]:
    """(USD value of the INR settlement, is_settled). See _payment_settled_inr_paise.

    The rate is passed IN rather than re-read per row: one report has to be built at one
    rate. app/fx.py refreshes on a TTL, so a per-row lookup can straddle a refresh and
    leave the `usd_to_inr` figure stamped on the summary — and on the CSV/XLSX export —
    something other than the rate the USD column was actually computed with, which makes
    the export impossible to reconcile against the INR books it came from.
    """
    settled_paise, is_settled = _payment_settled_inr_paise(payment)
    return (settled_paise / Decimal("100")) * inr_to_usd, is_settled


def _add_month_bucket(monthly_buckets: dict, when, *, investment: Decimal = Decimal("0"), returns: Decimal = Decimal("0")) -> None:
    key = _month_key(when)
    monthly_buckets[key]["investment_usd"] += investment
    monthly_buckets[key]["returns_usd"] += returns


def _get_table_columns(inspector, table_name: str) -> dict[str, dict]:
    try:
        return {str(col["name"]): col for col in inspector.get_columns(table_name)}
    except Exception:
        return {}


def _cleanup_legacy_marketplace_rows(db: Session, user_id: int) -> None:
    """
    Best-effort cleanup for legacy marketplace tables that may still exist in some DBs.
    These tables are not represented in current ORM models but can block user deletion
    through FK constraints (e.g. test buyer/seller accounts).
    """
    bind = db.get_bind()
    if bind is None:
        return

    inspector = inspect(bind)
    table_names = set(inspector.get_table_names())
    if not {"items", "messages", "item_images"} & table_names:
        return

    item_columns = _get_table_columns(inspector, "items") if "items" in table_names else {}
    message_columns = _get_table_columns(inspector, "messages") if "messages" in table_names else {}
    item_image_columns = _get_table_columns(inspector, "item_images") if "item_images" in table_names else {}

    # If marketplace items are owned by this user, delete dependent rows first.
    owner_columns = [
        col
        for col in ("seller_id", "user_id", "owner_id", "created_by_id", "creator_id")
        if col in item_columns
    ]
    owned_item_ids: list[int] = []
    if owner_columns:
        owner_where = " OR ".join(f'"{col}" = :uid' for col in owner_columns)
        rows = db.execute(
            text(f'SELECT "id" FROM "items" WHERE {owner_where}'),
            {"uid": int(user_id)},
        ).fetchall()
        owned_item_ids = [int(row[0]) for row in rows if row and row[0] is not None]

    if owned_item_ids:
        if item_image_columns and "item_id" in item_image_columns:
            db.execute(
                text('DELETE FROM "item_images" WHERE "item_id" IN :item_ids').bindparams(
                    bindparam("item_ids", expanding=True)
                ),
                {"item_ids": owned_item_ids},
            )
        if message_columns and "item_id" in message_columns:
            db.execute(
                text('DELETE FROM "messages" WHERE "item_id" IN :item_ids').bindparams(
                    bindparam("item_ids", expanding=True)
                ),
                {"item_ids": owned_item_ids},
            )

    # Remove messages directly tied to the user.
    for column in ("sender_id", "receiver_id", "user_id", "seller_id", "buyer_id"):
        if column in message_columns:
            db.execute(
                text(f'DELETE FROM "messages" WHERE "{column}" = :uid'),
                {"uid": int(user_id)},
            )

    # Null out buyer-like references when possible; otherwise delete matching rows.
    for column in ("buyer_id", "sold_to_user_id", "reserved_by_user_id"):
        if column not in item_columns:
            continue
        if bool(item_columns[column].get("nullable", True)):
            db.execute(
                text(f'UPDATE "items" SET "{column}" = NULL WHERE "{column}" = :uid'),
                {"uid": int(user_id)},
            )
        else:
            db.execute(
                text(f'DELETE FROM "items" WHERE "{column}" = :uid'),
                {"uid": int(user_id)},
            )

    # Finally delete items owned by the user.
    if owner_columns:
        owner_where = " OR ".join(f'"{col}" = :uid' for col in owner_columns)
        db.execute(
            text(f'DELETE FROM "items" WHERE {owner_where}'),
            {"uid": int(user_id)},
        )


@router.get("/users", response_model=schemas.AdminUserListResponse)
def list_users_admin(
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: str | None = Query(None, min_length=1, max_length=120),
    plan_filter: str = Query("all", alias="plan"),
    role_filter: str = Query("all", alias="role"),
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """
    List users for admin dashboard with search + filters + pagination.
    """
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.users.list",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )

    normalized_plan = (plan_filter or "all").strip().lower()
    if normalized_plan not in VALID_USER_PLAN_FILTERS:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid plan filter. Use one of: {', '.join(sorted(VALID_USER_PLAN_FILTERS))}",
        )

    normalized_role = (role_filter or "all").strip().lower()
    if normalized_role not in VALID_USER_ROLE_FILTERS:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid role filter. Use one of: {', '.join(sorted(VALID_USER_ROLE_FILTERS))}",
        )

    query = db.query(models.User)

    search_term = (search or "").strip()
    if search_term:
        token = f"%{search_term}%"
        query = query.filter(
            or_(
                models.User.email.ilike(token),
                models.User.username.ilike(token),
                models.User.full_name.ilike(token),
                models.User.university.ilike(token),
            )
        )

    if normalized_role == "student":
        query = query.filter(models.User.is_admin.is_(False), models.User.is_developer.is_(False))
    elif normalized_role == "staff":
        query = query.filter(or_(models.User.is_admin.is_(True), models.User.is_developer.is_(True)))
    elif normalized_role == "admin":
        query = query.filter(models.User.is_admin.is_(True))
    elif normalized_role == "developer":
        query = query.filter(models.User.is_developer.is_(True))

    active_pass_user_ids_select = select(models.Subscription.user_id).where(
        models.Subscription.plan == "pro",
        models.Subscription.status == "active",
    )

    if normalized_plan == "free":
        query = query.filter(~models.User.id.in_(active_pass_user_ids_select))
    elif normalized_plan == "visa_pass":
        query = query.filter(models.User.id.in_(active_pass_user_ids_select))

    total = query.count()
    filtered_user_ids_subquery = query.with_entities(models.User.id).subquery()
    metrics = _build_plan_metrics_for_filtered_users(db=db, filtered_user_ids_subquery=filtered_user_ids_subquery)
    users = (
        query.order_by(desc(models.User.created_at), desc(models.User.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return {
        "users": users,
        "total": total,
        "page": page,
        "page_size": page_size,
        "metrics": metrics,
    }


def _compute_company_finance_analytics(db: Session, *, ledger_limit: int | None = 250) -> dict:
    """
    Build the company finance analytics payload.

    Shared by the console endpoint (capped ledger) and the spreadsheet export
    (ledger_limit=None → every row, so the download is the complete book).
    """
    entries = (
        db.query(models.CompanyFinanceEntry)
        .order_by(desc(models.CompanyFinanceEntry.occurred_on), desc(models.CompanyFinanceEntry.id))
        .all()
    )
    verified_payments = (
        db.query(models.SubscriptionPayment)
        .filter(models.SubscriptionPayment.status == "verified")
        .order_by(
            desc(models.SubscriptionPayment.verified_at),
            desc(models.SubscriptionPayment.created_at),
            desc(models.SubscriptionPayment.id),
        )
        .all()
    )

    total_invested = Decimal("0")
    total_returns = Decimal("0")
    investment_entry_count = 0
    return_entry_count = 0
    monthly_buckets = defaultdict(lambda: {"investment_usd": Decimal("0"), "returns_usd": Decimal("0")})
    expense_breakdown: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    contributor_spend: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    ledger: list[dict] = []

    for entry in entries:
        amount = _money_decimal(entry.amount_usd)
        occurred_on = _coerce_date(entry.occurred_on)
        label = (entry.vendor or entry.category or "Uncategorized").strip() or "Uncategorized"
        paid_by = (entry.paid_by or "Unassigned").strip() or "Unassigned"
        if amount < 0:
            absolute_amount = abs(amount)
            total_invested += absolute_amount
            investment_entry_count += 1
            expense_breakdown[label] += absolute_amount
            contributor_spend[paid_by] += absolute_amount
            _add_month_bucket(monthly_buckets, occurred_on, investment=absolute_amount)
            kind = "Investment"
        else:
            total_returns += amount
            return_entry_count += 1
            _add_month_bucket(monthly_buckets, occurred_on, returns=amount)
            kind = "Return"

        ledger.append(
            {
                "id": f"finance-{entry.id}",
                "kind": kind,
                "category": entry.category or "Operations",
                "vendor": entry.vendor or "Unknown",
                "paid_by": paid_by,
                "description": entry.description,
                "amount_usd": _money_float(amount),
                "occurred_on": occurred_on.isoformat(),
                "source": entry.source or "manual",
            }
        )

    # Read ONCE, before the loop, and reuse for every row and for the rate reported on the
    # summary/export. app/fx.py refreshes on a TTL, so a per-row lookup could convert the
    # first payments at one rate and the last at another, and stamp a third on the export.
    fx_state = fx.get_state()
    inr_to_usd = _inr_to_usd_rate(fx_state)

    estimated_settlement_count = 0
    for payment in verified_payments:
        amount_usd, settlement_is_exact = _payment_amount_usd(payment, inr_to_usd)
        # A zero-value row (100%-off coupon, or a payment with no INR settlement figure
        # yet) used to be `continue`d here, which removed it from total_returns, ROI,
        # monthly_series AND the ledger — the exact opposite of the "kept visible as zero"
        # promise. Every verified payment is now emitted; a foreign row whose settlement
        # has not landed contributes 0 and carries `settlement_estimated`, so it reads as
        # "not counted yet" rather than either vanishing or booking cents as rupees.
        if not settlement_is_exact:
            estimated_settlement_count += 1

        payment_date = _coerce_date(payment.verified_at or payment.signature_verified_at or payment.created_at)
        total_returns += amount_usd
        return_entry_count += 1
        _add_month_bucket(monthly_buckets, payment_date, returns=amount_usd)
        charged_currency = str(payment.currency or "INR").strip().upper()
        charged_display = money.format_money(payment.amount_paise, charged_currency)
        ledger.append(
            {
                "id": f"payment-{payment.id}",
                "kind": "Return",
                "category": "Subscriptions",
                "vendor": (payment.provider or "Payment").title(),
                "paid_by": "Customer Revenue",
                "description": (
                    f"Verified subscription revenue · charged {charged_display}"
                    + (
                        ""
                        if settlement_is_exact
                        else " (not counted — no INR settlement figure on this row yet)"
                    )
                ),
                "amount_usd": _money_float(amount_usd),
                "occurred_on": payment_date.isoformat(),
                "source": "subscription_payments",
                "charged_currency": charged_currency,
                "charged_display": charged_display,
                "settlement_estimated": not settlement_is_exact,
            }
        )

    net_usd = total_returns - total_invested
    roi_percent = Decimal("0")
    if total_invested > 0:
        roi_percent = (net_usd / total_invested) * Decimal("100")

    monthly_series = []
    for month in sorted(monthly_buckets.keys()):
        investment_usd = monthly_buckets[month]["investment_usd"]
        returns_usd = monthly_buckets[month]["returns_usd"]
        monthly_series.append(
            {
                "month": month,
                "investment_usd": _money_float(investment_usd),
                "returns_usd": _money_float(returns_usd),
                "net_usd": _money_float(returns_usd - investment_usd),
            }
        )

    breakdown_items = []
    for label, amount in sorted(expense_breakdown.items(), key=lambda item: item[1], reverse=True):
        percentage = Decimal("0")
        if total_invested > 0:
            percentage = (amount / total_invested) * Decimal("100")
        breakdown_items.append(
            {
                "label": label,
                "amount_usd": _money_float(amount),
                "percentage": _money_float(percentage),
            }
        )

    contributor_items = []
    for label, amount in sorted(contributor_spend.items(), key=lambda item: item[1], reverse=True):
        percentage = Decimal("0")
        if total_invested > 0:
            percentage = (amount / total_invested) * Decimal("100")
        contributor_items.append(
            {
                "label": label,
                "amount_usd": _money_float(amount),
                "percentage": _money_float(percentage),
            }
        )

    ledger.sort(key=lambda item: (item["occurred_on"], item["id"]), reverse=True)

    # `fx_state` was read before the payment loop; reuse it so the rate reported here is
    # provably the rate every row above was converted at.
    usd_to_inr = _money_float(_money_decimal(fx_state.get("rate") or 0))
    notes = [
        "Investment rows are stored in company_finance_entries.",
        "Founder spend uses the paid_by column on each investment row.",
        "Returns are calculated live from verified subscription payments.",
        "Payments are booked at their INR settlement figure (base_amount_paise), never at "
        "the charged amount — a $12.99 charge stores 1299 cents, not rupees.",
        f"INR is converted to USD at the shared live rate ($1 = ₹{usd_to_inr}, "
        f"source: {fx_state.get('source') or 'fallback'}).",
    ]
    if estimated_settlement_count:
        notes.append(
            f"{estimated_settlement_count} foreign-currency payment(s) have no INR "
            "settlement figure yet and are NOT counted in returns — they are listed at "
            "zero and flagged, never booked at their charged amount. Total returns is a "
            "known understatement until Razorpay's base_amount lands on those rows."
        )

    return {
        "summary": {
            "total_invested_usd": _money_float(total_invested),
            "total_returns_usd": _money_float(total_returns),
            "net_usd": _money_float(net_usd),
            "roi_percent": _money_float(roi_percent),
            "break_even_gap_usd": _money_float(max(total_invested - total_returns, Decimal("0"))),
            "investment_entry_count": investment_entry_count,
            "return_entry_count": return_entry_count,
            "estimated_settlement_count": estimated_settlement_count,
            "usd_to_inr": usd_to_inr,
            "fx_source": fx_state.get("source") or "fallback",
        },
        "monthly_series": monthly_series,
        "expense_breakdown": breakdown_items,
        "contributor_breakdown": contributor_items,
        "ledger": ledger[:ledger_limit] if ledger_limit else ledger,
        "notes": notes,
    }


@router.get(
    "/company-finance/analytics",
    response_model=schemas.AdminCompanyFinanceAnalyticsResponse,
)
def company_finance_analytics_admin(
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """
    Company-level finance analytics for the admin console.

    Investment rows come from company_finance_entries. Returns are calculated live from
    verified subscription payments, booked at their INR settlement figure and converted to
    USD at the shared live FX rate (app/fx.py) — the same rate the enterprise revenue
    screen uses, so the two screens can't disagree.
    """
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.company_finance.analytics",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    return _compute_company_finance_analytics(db)


# --- Company finance spreadsheet export (Excel .xlsx, CSV fallback) ---

# "Charged" and "Settlement" are appended (not inserted) so the existing column positions
# — and the number-format/auto-filter indices below that hardcode them — stay put.
FINANCE_EXPORT_LEDGER_HEADERS = [
    "Date (UTC)", "Type", "Vendor", "Category", "Paid By", "Description", "Amount (USD)", "Source",
    "Charged", "Settlement",
]
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _finance_export_rows(payload: dict) -> list[list]:
    """Ledger rows in export order (dates as `date` so Excel sorts/filters them properly)."""
    rows: list[list] = []
    for item in payload.get("ledger", []):
        rows.append([
            _finance_parse_date_safe(item.get("occurred_on")),
            item.get("kind") or "",
            item.get("vendor") or "",
            item.get("category") or "",
            item.get("paid_by") or "",
            item.get("description") or "",
            float(item.get("amount_usd") or 0),
            item.get("source") or "",
            item.get("charged_display") or "",
            "not counted" if item.get("settlement_estimated") else "settled",
        ])
    return rows


def _finance_parse_date_safe(value):
    try:
        return datetime.strptime(str(value).strip()[:10], "%Y-%m-%d").date()
    except Exception:
        return value or ""


def _build_finance_csv(payload: dict) -> bytes:
    """CSV fallback (also the explicit ?format=csv output). BOM so Excel reads UTF-8."""
    import csv
    import io as _io

    buffer = _io.StringIO()
    writer = csv.writer(buffer)
    summary = payload.get("summary", {})
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    writer.writerow(["Rilono - Company Finance Export"])
    writer.writerow(["Generated (UTC)", generated])
    writer.writerow([])
    writer.writerow(["Total invested (USD)", summary.get("total_invested_usd", 0)])
    writer.writerow(["Total returns (USD)", summary.get("total_returns_usd", 0)])
    writer.writerow(["Net (USD)", summary.get("net_usd", 0)])
    writer.writerow(["ROI (%)", summary.get("roi_percent", 0)])
    writer.writerow(["Break-even gap (USD)", summary.get("break_even_gap_usd", 0)])
    # The rate the USD column was built with — without it the export can't be reconciled
    # against the INR books it was derived from.
    writer.writerow([f"USD -> INR rate ({summary.get('fx_source', 'fallback')})", summary.get("usd_to_inr", 0)])
    # Foreign payments left OUT of "Total returns" because their INR settlement figure has
    # not landed yet. Non-zero means the totals above are a known understatement, and the
    # export is unreadable as a "complete book" without saying so.
    writer.writerow(
        ["Payments not counted (no INR settlement)", summary.get("estimated_settlement_count", 0)]
    )
    writer.writerow([])
    writer.writerow(FINANCE_EXPORT_LEDGER_HEADERS)
    for row in _finance_export_rows(payload):
        writer.writerow([value.isoformat() if isinstance(value, date) else value for value in row])

    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def _build_finance_xlsx(payload: dict) -> bytes | None:
    """
    Real .xlsx workbook (Summary / Ledger / Spend by vendor / Spend by founder / Monthly).
    Returns None when openpyxl isn't installed so the caller can fall back to CSV.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:  # pragma: no cover - depends on the deployed environment
        logger.warning("openpyxl unavailable; falling back to CSV for the finance export")
        return None

    from io import BytesIO

    USD_FORMAT = '"$"#,##0.00'
    PCT_FORMAT = '0.00"%"'
    DATE_FORMAT = "yyyy-mm-dd"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4338CA")
    title_font = Font(bold=True, size=14)

    summary = payload.get("summary", {})
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    workbook = Workbook()

    def style_header(sheet, row_index: int, column_count: int) -> None:
        for column_index in range(1, column_count + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def set_widths(sheet, widths: list[int]) -> None:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    # --- Summary ---
    overview = workbook.active
    overview.title = "Summary"
    overview["A1"] = "Rilono — Company Finance"
    overview["A1"].font = title_font
    overview["A2"] = "Generated (UTC)"
    overview["B2"] = generated
    overview["A4"] = "Metric"
    overview["B4"] = "Value"
    style_header(overview, 4, 2)
    metrics = [
        ("Total invested (USD)", summary.get("total_invested_usd", 0), USD_FORMAT),
        ("Total returns (USD)", summary.get("total_returns_usd", 0), USD_FORMAT),
        ("Net (USD)", summary.get("net_usd", 0), USD_FORMAT),
        ("ROI (%)", summary.get("roi_percent", 0), PCT_FORMAT),
        ("Break-even gap (USD)", summary.get("break_even_gap_usd", 0), USD_FORMAT),
        ("Investment entries", summary.get("investment_entry_count", 0), "0"),
        ("Return entries", summary.get("return_entry_count", 0), "0"),
        # The rate the USD column was built with, so the export reconciles against the
        # INR books it was derived from.
        (f"USD → INR rate ({summary.get('fx_source', 'fallback')})", summary.get("usd_to_inr", 0), "0.00"),
        # Foreign payments excluded from returns because Razorpay's INR settlement figure
        # has not landed on the row yet. Non-zero means the total above understates.
        ("Payments not counted (no INR settlement)", summary.get("estimated_settlement_count", 0), "0"),
    ]
    for offset, (label, value, number_format) in enumerate(metrics):
        row_index = 5 + offset
        overview.cell(row=row_index, column=1, value=label)
        value_cell = overview.cell(row=row_index, column=2, value=value)
        value_cell.number_format = number_format
    notes_row = 5 + len(metrics) + 1
    overview.cell(row=notes_row, column=1, value="Notes").font = Font(bold=True)
    for offset, note in enumerate(payload.get("notes", []), start=1):
        overview.cell(row=notes_row + offset, column=1, value=note)
    set_widths(overview, [34, 26])

    # --- Ledger ---
    ledger_sheet = workbook.create_sheet("Ledger")
    ledger_sheet.append(FINANCE_EXPORT_LEDGER_HEADERS)
    style_header(ledger_sheet, 1, len(FINANCE_EXPORT_LEDGER_HEADERS))
    for row in _finance_export_rows(payload):
        ledger_sheet.append(row)
    for row_index in range(2, ledger_sheet.max_row + 1):
        ledger_sheet.cell(row=row_index, column=1).number_format = DATE_FORMAT
        ledger_sheet.cell(row=row_index, column=7).number_format = USD_FORMAT
    ledger_sheet.freeze_panes = "A2"
    if ledger_sheet.max_row > 1:
        ledger_sheet.auto_filter.ref = f"A1:J{ledger_sheet.max_row}"
    set_widths(ledger_sheet, [13, 12, 18, 22, 16, 42, 15, 22, 14, 12])

    # --- Breakdowns ---
    def write_breakdown(title: str, label_header: str, items: list[dict]) -> None:
        sheet = workbook.create_sheet(title)
        sheet.append([label_header, "Amount (USD)", "% of invested"])
        style_header(sheet, 1, 3)
        for item in items:
            sheet.append([
                item.get("label") or "",
                float(item.get("amount_usd") or 0),
                float(item.get("percentage") or 0),
            ])
        for row_index in range(2, sheet.max_row + 1):
            sheet.cell(row=row_index, column=2).number_format = USD_FORMAT
            sheet.cell(row=row_index, column=3).number_format = PCT_FORMAT
        sheet.freeze_panes = "A2"
        set_widths(sheet, [28, 16, 16])

    write_breakdown("Spend by vendor", "Vendor", payload.get("expense_breakdown", []))
    write_breakdown("Spend by founder", "Paid by", payload.get("contributor_breakdown", []))

    # --- Monthly ---
    monthly_sheet = workbook.create_sheet("Monthly")
    monthly_sheet.append(["Month", "Investment (USD)", "Returns (USD)", "Net (USD)"])
    style_header(monthly_sheet, 1, 4)
    for item in payload.get("monthly_series", []):
        monthly_sheet.append([
            item.get("month") or "",
            float(item.get("investment_usd") or 0),
            float(item.get("returns_usd") or 0),
            float(item.get("net_usd") or 0),
        ])
    for row_index in range(2, monthly_sheet.max_row + 1):
        for column_index in (2, 3, 4):
            monthly_sheet.cell(row=row_index, column=column_index).number_format = USD_FORMAT
    monthly_sheet.freeze_panes = "A2"
    set_widths(monthly_sheet, [14, 18, 18, 16])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


@router.get("/company-finance/export")
def export_company_finance_admin(
    request: Request,
    export_format: str = Query("xlsx", alias="format", pattern="^(xlsx|csv)$"),
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Download the full finance book (ledger + summary + breakdowns) as Excel or CSV."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.company_finance.export",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )

    payload = _compute_company_finance_analytics(db, ledger_limit=None)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    content: bytes | None = None
    if export_format == "xlsx":
        content = _build_finance_xlsx(payload)

    if content is None:  # explicit CSV request, or openpyxl missing on this host
        content = _build_finance_csv(payload)
        filename = f"rilono-finance-{stamp}.csv"
        media_type = "text/csv; charset=utf-8"
    else:
        filename = f"rilono-finance-{stamp}.xlsx"
        media_type = XLSX_MEDIA_TYPE

    return Response(
        content=content,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )


# --- Company finance entries: admin CRUD (the DB is the source of truth) ---
# Entries live in company_finance_entries; the code seed only bootstraps a fresh DB once
# (see schema_patch), so anything created/edited/deleted here persists across deploys.

def _finance_parse_date(value) -> date:
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip()[:10], "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(status_code=400, detail="Enter a valid date (YYYY-MM-DD).")


def _finance_signed_amount(entry_type: str | None, amount) -> tuple[str, Decimal]:
    """Return (normalized entry_type, signed amount). Expense = negative, return = positive."""
    try:
        magnitude = Decimal(str(amount)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Enter a valid amount.")
    if magnitude <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than 0.")
    if magnitude > Decimal("100000000"):
        raise HTTPException(status_code=400, detail="Amount is too large.")
    kind = (entry_type or "expense").strip().lower()
    if kind not in {"expense", "return"}:
        raise HTTPException(status_code=400, detail="Type must be 'expense' or 'return'.")
    return kind, (magnitude if kind == "return" else -magnitude)


def _serialize_finance_entry(entry: models.CompanyFinanceEntry) -> dict:
    amount = _money_decimal(entry.amount_usd)
    return {
        "id": entry.id,
        "ledger_id": f"finance-{entry.id}",
        "entry_type": entry.entry_type,
        "kind": "Return" if amount >= 0 else "Investment",
        "category": entry.category,
        "vendor": entry.vendor,
        "paid_by": entry.paid_by,
        "description": entry.description,
        "amount_usd": _money_float(amount),
        "occurred_on": _coerce_date(entry.occurred_on).isoformat(),
        "source": entry.source or "manual",
    }


@router.post("/company-finance/entries")
def create_company_finance_entry_admin(
    request: Request,
    payload: schemas.AdminCompanyFinanceEntryCreateRequest,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Add a company investment/return entry (stored in the DB with source='manual')."""
    _enforce_rate_limit_or_429(
        request=request, scope="admin.company_finance.create",
        limit=ADMIN_ENDPOINT_RATE_LIMIT, window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    category = (payload.category or "").strip()
    vendor = (payload.vendor or "").strip()
    if not vendor or not category:
        raise HTTPException(status_code=400, detail="Vendor and category are required.")
    paid_by = (payload.paid_by or "Gaurav").strip() or "Gaurav"
    occurred_on = _finance_parse_date(payload.occurred_on)
    entry_type, signed_amount = _finance_signed_amount(payload.entry_type, payload.amount_usd)

    row = models.CompanyFinanceEntry(
        seed_key=None,
        entry_type=entry_type,
        category=category,
        vendor=vendor,
        description=((payload.description or "").strip() or None),
        amount_usd=signed_amount,
        occurred_on=occurred_on,
        paid_by=paid_by,
        source="manual",
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"entry": _serialize_finance_entry(row)}


@router.patch("/company-finance/entries/{entry_id}")
def update_company_finance_entry_admin(
    entry_id: int,
    request: Request,
    payload: schemas.AdminCompanyFinanceEntryUpdateRequest,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Edit any finance entry (including a re-attribution of paid_by / amount split)."""
    _enforce_rate_limit_or_429(
        request=request, scope="admin.company_finance.update",
        limit=ADMIN_ENDPOINT_RATE_LIMIT, window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    row = db.query(models.CompanyFinanceEntry).filter(models.CompanyFinanceEntry.id == entry_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Finance entry not found.")

    if payload.category is not None:
        category = payload.category.strip()
        if not category:
            raise HTTPException(status_code=400, detail="Category cannot be empty.")
        row.category = category
    if payload.vendor is not None:
        vendor = payload.vendor.strip()
        if not vendor:
            raise HTTPException(status_code=400, detail="Vendor cannot be empty.")
        row.vendor = vendor
    if payload.paid_by is not None:
        row.paid_by = payload.paid_by.strip() or row.paid_by
    if payload.description is not None:
        row.description = payload.description.strip() or None
    if payload.occurred_on is not None:
        row.occurred_on = _finance_parse_date(payload.occurred_on)
    if payload.amount_usd is not None or payload.entry_type is not None:
        entry_type = payload.entry_type if payload.entry_type is not None else row.entry_type
        magnitude = payload.amount_usd if payload.amount_usd is not None else abs(_money_float(_money_decimal(row.amount_usd)))
        row.entry_type, row.amount_usd = _finance_signed_amount(entry_type, magnitude)

    db.commit()
    db.refresh(row)
    return {"entry": _serialize_finance_entry(row)}


@router.delete("/company-finance/entries/{entry_id}")
def delete_company_finance_entry_admin(
    entry_id: int,
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Delete a finance entry. Subscription-revenue returns aren't in this table, so they're safe."""
    _enforce_rate_limit_or_429(
        request=request, scope="admin.company_finance.delete",
        limit=ADMIN_ENDPOINT_RATE_LIMIT, window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    row = db.query(models.CompanyFinanceEntry).filter(models.CompanyFinanceEntry.id == entry_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Finance entry not found.")
    db.delete(row)
    db.commit()
    return {"deleted": True, "id": entry_id}


@router.get("/ai-usage/analytics")
def ai_usage_analytics_admin(
    request: Request,
    start: str | None = None,
    end: str | None = None,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Gemini API usage & estimated cost analytics for the admin console.

    Optional `start`/`end` (YYYY-MM-DD, inclusive, UTC) filter the range summary,
    timeline, per-day table and breakdowns. Omitted = the last 30 days.
    """
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.ai_usage.analytics",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    start_date = _finance_parse_date(start) if (start or "").strip() else None
    end_date = _finance_parse_date(end) if (end or "").strip() else None
    return ai_usage.build_ai_usage_analytics(db, start=start_date, end=end_date)


@router.get("/acquisition/analytics")
def acquisition_analytics_admin(
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Where signups come from — first-party traffic-source breakdown for the admin console."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.acquisition.analytics",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    return acquisition.build_analytics(db)


@router.post("/growth/conversion-analysis")
def growth_conversion_analysis_admin(
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
    limit: int = 30,
):
    """Internal Growth Agent — which free accounts to promote the Visa Success Pass to, and how.

    Admin-only. Runs on demand (not automatically) to bound AI cost. Deterministic usage
    scoring in Python + Gemini for the per-account conversion strategy.
    """
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.growth.conversion",
        limit=int(os.getenv("ADMIN_GROWTH_RATE_LIMIT", "12")),
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    from app import growth_agent
    try:
        return growth_agent.run_conversion_analysis(db, limit=limit)
    except Exception:
        logger.exception("Growth conversion analysis failed")
        raise HTTPException(status_code=502, detail="The growth analysis could not be completed. Please try again.")


@router.get("/growth/funnel")
def growth_funnel_admin(
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
    days: int = 0,
):
    """B2C activation funnel: signup -> activated -> ran red-flag scan -> purchased Visa Pass.
    `days` > 0 windows by signup recency; 0 = all-time. Cheap aggregate queries (no AI)."""
    _enforce_rate_limit_or_429(
        request=request, scope="admin.growth.funnel",
        limit=ADMIN_ENDPOINT_RATE_LIMIT, window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    from app import growth_funnel
    return growth_funnel.build_activation_funnel(db, days=days or None)


@router.get("/growth/outcomes")
def growth_outcomes_admin(
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Visa-decision approval rate from the outcome-capture loop, segmented by red-flag-scan usage.
    Powers the pricing-power stat ('approved at X% vs. the ~35-48% market baseline')."""
    _enforce_rate_limit_or_429(
        request=request, scope="admin.growth.outcomes",
        limit=ADMIN_ENDPOINT_RATE_LIMIT, window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    from app import growth_funnel
    return growth_funnel.build_outcome_analytics(db)


def _admin_iso(value):
    return value.isoformat() if value else None


def _serialize_coupon_code(db: Session, row: models.CouponCode) -> dict:
    """B2C coupon code + live usage stats for the admin console."""
    verified_uses = int(
        db.query(func.count(models.SubscriptionPayment.id))
        .filter(
            models.SubscriptionPayment.status == "verified",
            func.upper(models.SubscriptionPayment.coupon_code) == str(row.coupon_code or "").upper(),
        )
        .scalar() or 0
    )
    restricted_email = None
    restricted_to = getattr(row, "restricted_to_user_id", None)
    if restricted_to is not None:
        u = db.query(models.User.email).filter(models.User.id == int(restricted_to)).first()
        restricted_email = u[0] if u else None
    return {
        "code": row.coupon_code,
        "percent_off": float(row.percent_off) if row.percent_off is not None else None,
        "max_uses_per_user": row.max_uses_per_user,
        "restricted_to_user_id": restricted_to,
        "restricted_to_email": restricted_email,
        "verified_uses": verified_uses,
        "created_at": _admin_iso(getattr(row, "created_at", None)),
    }


@router.get("/users/{user_id}/detail")
def admin_user_detail(
    user_id: int,
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Full per-account view for the admin 'Manage account' screen: profile, plan + usage,
    coupons applied (and when), payments, referral, and a deterministic intent snapshot."""
    _enforce_rate_limit_or_429(
        request=request, scope="admin.user.detail",
        limit=ADMIN_ENDPOINT_RATE_LIMIT, window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    from app import subscriptions as subs
    from app import visa_pass
    from app import visa_catalog
    from app import acquisition as acq

    user = db.query(models.User).filter(models.User.id == int(user_id)).first()
    if not user:
        raise HTTPException(status_code=404, detail="Account not found.")
    sub = db.query(models.Subscription).filter(models.Subscription.user_id == user.id).first()

    dest = getattr(user, "destination_country_code", None)
    role = "developer" if user.is_developer else ("admin" if user.is_admin else "student")
    referred_by = None
    if getattr(user, "referred_by_user_id", None):
        rb = db.query(models.User).filter(models.User.id == user.referred_by_user_id).first()
        referred_by = {"id": rb.id, "email": rb.email} if rb else None
    referrals_made = db.query(models.User).filter(models.User.referred_by_user_id == user.id).count()

    is_pass = bool(sub and visa_pass.has_active_pass(sub))
    usage = []
    if sub:
        # "group" tells the admin UI which surface each quota belongs to. Every counter here
        # is consumed in the web app; the Chrome extension (Rilono Copilot) has no counters
        # of its own — Copilot chat is pass-gated and shares the "Rilono AI messages" quota.
        # Red-flag scan, voice interview and university shortlist are the
        # Visa-Success-Pass-metered features (see app/visa_pass.py FEATURES); "pass_limit"
        # is what a pass holder gets (-1 = unlimited), so the UI can render truthful limits
        # (e.g. voice interviews are capped at 3 even on the pass, not unlimited).
        usage = [
            {"key": "ai_messages", "label": "Rilono AI messages (web + Copilot chat)", "group": "Web app", "used": sub.ai_messages_used or 0, "free_limit": subs.FREE_AI_MESSAGE_LIMIT, "pass_limit": -1},
            {"key": "document_uploads", "label": "Document uploads", "group": "Web app", "used": sub.document_uploads_used or 0, "free_limit": subs.FREE_DOCUMENT_UPLOAD_LIMIT, "pass_limit": -1},
            {"key": "prep_sessions", "label": "Interview-prep sessions", "group": "Web app", "used": sub.prep_sessions_used or 0, "free_limit": subs.FREE_PREP_SESSION_LIMIT, "pass_limit": -1},
            {"key": "mock_interviews", "label": "Mock interviews", "group": "Web app", "used": sub.mock_interviews_used or 0, "free_limit": subs.FREE_MOCK_INTERVIEW_LIMIT, "pass_limit": -1},
            {"key": "red_flag_scans", "label": "Red-flag scan", "group": "Visa Success Pass features", "used": sub.red_flag_scans_used or 0, "free_limit": visa_pass.FREE_RED_FLAG_SCANS, "pass_limit": visa_pass.FEATURES["red_flag_scan"]["pass_limit"]},
            {"key": "voice_interviews", "label": "Voice mock interview", "group": "Visa Success Pass features", "used": sub.pass_voice_interviews_used or 0, "free_limit": 0, "pass_limit": visa_pass.FEATURES["voice_interview"]["pass_limit"]},
            {"key": "university_shortlists", "label": "AI university shortlist", "group": "Visa Success Pass features", "used": getattr(sub, "university_recommendations_used", 0) or 0, "free_limit": visa_pass.FREE_UNIVERSITY_RECOMMENDATIONS, "pass_limit": visa_pass.FEATURES["university_shortlist"]["pass_limit"]},
        ]

    payments = (
        db.query(models.SubscriptionPayment)
        .filter(models.SubscriptionPayment.user_id == user.id)
        .order_by(models.SubscriptionPayment.created_at.desc())
        .limit(50).all()
    )
    # `amount_paise` is minor units of the row's OWN currency, so the display string has to
    # be formatted here where the currency is in hand — the console can't assume rupees.
    payment_rows = [{
        "amount_paise": p.amount_paise, "currency": p.currency, "status": p.status,
        "amount_display": money.format_money(p.amount_paise, p.currency or "INR"),
        "settled_inr_display": (
            money.format_money(p.base_amount_paise, "INR") if p.base_amount_paise is not None else None
        ),
        "is_international": bool(p.is_international),
        "coupon_code": p.coupon_code, "coupon_percent_off": float(p.coupon_percent_off) if p.coupon_percent_off is not None else None,
        "pricing_model": p.pricing_model, "created_at": _admin_iso(p.created_at), "verified_at": _admin_iso(p.verified_at),
    } for p in payments]
    # Coupons this account has applied (from payment history), newest first.
    coupons = [{
        "code": p.coupon_code, "percent_off": float(p.coupon_percent_off) if p.coupon_percent_off is not None else None,
        "applied_at": _admin_iso(p.verified_at or p.created_at), "status": p.status,
        "amount_paise": p.amount_paise, "currency": p.currency,
        "amount_display": money.format_money(p.amount_paise, p.currency or "INR"),
        "on": "Visa Success Pass" if p.pricing_model == "visa_pass" else (p.pricing_model or "subscription"),
    } for p in payments if p.coupon_code]

    feat = growth_agent_features_safe(db, user, sub)

    # Coupon codes issued FOR this account (admin "conversion play" offers).
    coupon_offer_rows = (
        db.query(models.CouponCode)
        .filter(models.CouponCode.restricted_to_user_id == user.id)
        .all()
    )
    coupon_offers = sorted(
        (_serialize_coupon_code(db, r) for r in coupon_offer_rows),
        key=lambda c: c["created_at"] or "", reverse=True,
    )

    return {
        "account": {
            "id": user.id, "name": user.full_name or user.username or user.email, "email": user.email,
            "username": user.username, "role": role, "is_active": bool(user.is_active),
            "email_verified": bool(user.email_verified), "auth_provider": user.auth_provider or "password",
            "country_code": dest, "country_name": (visa_catalog.country_meta(dest) or {}).get("name") if dest else None,
            "visa_type_label": visa_catalog.visa_type_label(dest, getattr(user, "visa_type_key", None)) if dest else None,
            "university": user.university, "university_email": getattr(user, "university_email", None),
            "home_country": user.current_residence_country,
            "created_at": _admin_iso(user.created_at), "first_login_at": _admin_iso(getattr(user, "first_login_at", None)),
            "last_login_at": _admin_iso(getattr(user, "last_login_at", None)),
            "onboarding_completed_at": _admin_iso(getattr(user, "onboarding_completed_at", None)),
            "marketing_consent": bool(getattr(user, "marketing_emails_consent", False)),
            "heard_about_us": getattr(user, "heard_about_us", None),
            "heard_about_label": acq.heard_about_label(getattr(user, "heard_about_us", None)),
            "acquisition": {
                "channel": getattr(user, "acquisition_channel", None), "source": getattr(user, "acquisition_source", None),
                "medium": getattr(user, "acquisition_medium", None), "campaign": getattr(user, "acquisition_campaign", None),
                "landing_page": getattr(user, "acquisition_landing_page", None),
            },
            "referral": {
                "code": getattr(user, "referral_code", None), "referred_by": referred_by,
                "referrals_made": referrals_made, "reward_granted_at": _admin_iso(getattr(user, "referral_reward_granted_at", None)),
            },
            "visa_decision": {
                "decision": getattr(user, "visa_decision", None),
                "decision_at": _admin_iso(getattr(user, "visa_decision_at", None)),
                "source": getattr(user, "visa_decision_source", None),
            },
        },
        "subscription": {
            "plan": (sub.plan if sub else "free"), "status": (sub.status if sub else None),
            "is_pass_active": is_pass, "ends_at": _admin_iso(sub.ends_at) if sub else None,
            "pass_days_left": visa_pass.pass_days_left(sub) if (sub and is_pass) else None,
            "usage": usage,
        },
        "coupons": coupons,
        "coupon_offers": coupon_offers,
        "payments": payment_rows,
        "intent": {
            "score": feat["intent_score"] if feat else None,
            "hit_any_limit": feat["hit_any_limit"] if feat else None,
        },
    }


@router.get("/coupons")
def list_coupon_codes_admin(
    request: Request,
    user_id: int | None = None,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """List B2C coupon codes (all, or only those issued to one account via ?user_id=)."""
    _enforce_rate_limit_or_429(
        request=request, scope="admin.coupons.list",
        limit=ADMIN_ENDPOINT_RATE_LIMIT, window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    q = db.query(models.CouponCode)
    if user_id is not None:
        q = q.filter(models.CouponCode.restricted_to_user_id == int(user_id))
    rows = q.all()
    coupons = sorted(
        (_serialize_coupon_code(db, r) for r in rows),
        key=lambda c: c["created_at"] or "", reverse=True,
    )
    return {"coupons": coupons}


@router.post("/coupons")
def create_coupon_code_admin(
    request: Request,
    payload: schemas.AdminCouponCodeCreateRequest,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Create a B2C coupon code — optionally restricted to one student account."""
    _enforce_rate_limit_or_429(
        request=request, scope="admin.coupons.create",
        limit=ADMIN_ENDPOINT_RATE_LIMIT, window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    from app.routers.subscription import _normalize_coupon_code

    code = _normalize_coupon_code(payload.code)
    if not code or len(code) < 3 or len(code) > 32:
        raise HTTPException(status_code=400, detail="Enter a code of 3–32 letters/numbers (-, _ allowed).")

    try:
        percent = Decimal(str(payload.percent_off)).quantize(Decimal("0.01"))
    except Exception:
        raise HTTPException(status_code=400, detail="Enter a valid discount percentage.")
    if not percent.is_finite():  # NaN survives quantize and would poison comparisons
        raise HTTPException(status_code=400, detail="Enter a valid discount percentage.")
    # Discounts are capped to the 1–100% band; out-of-range values clamp.
    percent = min(max(percent, Decimal("1")), Decimal("100"))

    max_uses = payload.max_uses_per_user
    if max_uses is not None:
        max_uses = int(max_uses)
        if max_uses < 1:
            raise HTTPException(status_code=400, detail="Max uses per user must be at least 1 (or empty for unlimited).")

    restricted_to = payload.restricted_to_user_id
    if restricted_to is not None:
        target = db.query(models.User).filter(models.User.id == int(restricted_to)).first()
        if not target:
            raise HTTPException(status_code=404, detail="Account to restrict this coupon to was not found.")

    existing = db.query(models.CouponCode).filter(
        func.upper(models.CouponCode.coupon_code) == code
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"Coupon code '{code}' already exists.")

    row = models.CouponCode(
        coupon_code=code,
        percent_off=percent,
        max_uses_per_user=max_uses,
        restricted_to_user_id=int(restricted_to) if restricted_to is not None else None,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"coupon": _serialize_coupon_code(db, row)}


@router.delete("/coupons/{code}")
def delete_coupon_code_admin(
    code: str,
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Delete a B2C coupon code (past payments keep their recorded code/discount)."""
    _enforce_rate_limit_or_429(
        request=request, scope="admin.coupons.delete",
        limit=ADMIN_ENDPOINT_RATE_LIMIT, window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    from app.routers.subscription import _normalize_coupon_code

    normalized = _normalize_coupon_code(code)
    row = db.query(models.CouponCode).filter(
        func.upper(models.CouponCode.coupon_code) == normalized
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Coupon code not found.")
    db.delete(row)
    db.commit()
    return {"deleted": True, "code": normalized}


def growth_agent_features_safe(db, user, sub):
    if not sub:
        return None
    try:
        from app import growth_agent
        return growth_agent._features_for(user, sub)
    except Exception:
        return None


@router.post("/users/{user_id}/conversion-reco")
def admin_user_conversion_reco(
    user_id: int,
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Run the Gemini growth agent for a single account (the 'AI recommendation' button)."""
    _enforce_rate_limit_or_429(
        request=request, scope="admin.user.reco",
        limit=int(os.getenv("ADMIN_GROWTH_RATE_LIMIT", "12")), window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    from app import growth_agent
    try:
        result = growth_agent.analyze_account(db, user_id)
    except Exception:
        logger.exception("Single-account conversion reco failed (user_id=%s)", user_id)
        raise HTTPException(status_code=502, detail="Could not generate a recommendation. Please try again.")
    if not result.get("found"):
        raise HTTPException(status_code=404, detail="Account not found.")
    return result


@router.post("/users/{user_id}/visa-decision")
def admin_set_visa_decision(
    user_id: int,
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
    payload: dict = Body(...),
):
    """Admin override to record/clear a student's final visa decision from the Manage view.
    Feeds the same approval-rate analytics as the student's self-report."""
    _enforce_rate_limit_or_429(
        request=request, scope="admin.user.decision",
        limit=ADMIN_ENDPOINT_RATE_LIMIT, window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    user = db.query(models.User).filter(models.User.id == int(user_id)).first()
    if not user:
        raise HTTPException(status_code=404, detail="Account not found.")
    decision = str(payload.get("decision") or "").strip().lower()
    valid = ("approved", "refused", "withdrawn", "deferred")
    if decision in ("", "none", "clear"):
        user.visa_decision = None
        user.visa_decision_at = None
        user.visa_decision_source = None
    elif decision in valid:
        user.visa_decision = decision
        user.visa_decision_at = datetime.utcnow()
        user.visa_decision_source = "admin"
    else:
        raise HTTPException(status_code=422, detail="Invalid decision.")
    db.commit()
    return {
        "decision": user.visa_decision,
        "decision_at": _admin_iso(user.visa_decision_at),
        "source": user.visa_decision_source,
    }


@router.get("/enterprise/revenue")
def enterprise_revenue_analytics_admin(
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Prepaid-credit revenue model: credit/infra revenue vs real Gemini cost = margin."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.enterprise.revenue",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    return enterprise_credits.build_revenue_analytics(db)


@router.get("/b2c/revenue")
def b2c_revenue_analytics_admin(
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """B2C 'Visa Success Pass' economics: one-time pass revenue vs real Gemini cost."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.b2c.revenue",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    return visa_pass.build_revenue_analytics(db)


@router.get("/b2c/accounts")
def b2c_account_cost_profit_admin(
    request: Request,
    limit: int = Query(100, ge=1, le=500),
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Per-account B2C economics: each student's attributed Gemini cost vs pass revenue = profit."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.b2c.accounts",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    return visa_pass.build_account_cost_profit(db, limit=limit)


@router.post("/enterprise/calendar-reminders/run")
def run_enterprise_calendar_reminders_admin(
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
):
    """Manually trigger the daily enterprise calendar-reminder email job (force run)."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.enterprise.calendar_reminders",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    from app.services.enterprise_calendar_reminders import run_calendar_reminder_job
    return run_calendar_reminder_job(force=True)


@router.get("/ai-optimization")
def ai_optimization_analytics_admin(
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """GCP cost-optimization metrics: off-topic prompts blocked + context-cache savings."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.ai_optimization",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    return ai_guardrails.build_optimization_analytics(db)


@router.get("/enterprise/accounts", response_model=schemas.AdminEnterpriseAccountListResponse)
def list_enterprise_accounts_admin(
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: str | None = Query(None, min_length=1, max_length=120),
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """
    List enterprise organizations for admin dashboard with search + pagination.
    """
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.enterprise.list",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )

    creator_user = aliased(models.User)
    query = db.query(models.EnterpriseOrganization).outerjoin(
        creator_user,
        creator_user.id == models.EnterpriseOrganization.created_by_user_id,
    )

    search_term = (search or "").strip()
    if search_term:
        token = f"%{search_term}%"
        query = query.filter(
            or_(
                models.EnterpriseOrganization.company_name.ilike(token),
                models.EnterpriseOrganization.subdomain_slug.ilike(token),
                creator_user.email.ilike(token),
                creator_user.full_name.ilike(token),
            )
        )

    total = int(query.count() or 0)
    filtered_org_ids_subquery = query.with_entities(models.EnterpriseOrganization.id).subquery()
    filtered_org_ids_select = select(filtered_org_ids_subquery.c.id)

    active_members = int(
        db.query(func.count(models.EnterpriseOrganizationMember.id))
        .filter(
            models.EnterpriseOrganizationMember.organization_id.in_(filtered_org_ids_select),
            models.EnterpriseOrganizationMember.is_active.is_(True),
        )
        .scalar()
        or 0
    )
    active_admins = int(
        db.query(func.count(models.EnterpriseOrganizationMember.id))
        .filter(
            models.EnterpriseOrganizationMember.organization_id.in_(filtered_org_ids_select),
            models.EnterpriseOrganizationMember.is_active.is_(True),
            func.lower(models.EnterpriseOrganizationMember.role) == ENTERPRISE_ROLE_ADMIN,
        )
        .scalar()
        or 0
    )

    organizations = (
        query.order_by(desc(models.EnterpriseOrganization.created_at), desc(models.EnterpriseOrganization.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    if not organizations:
        return {
            "accounts": [],
            "total": total,
            "page": page,
            "page_size": page_size,
            "metrics": {
                "active_members": active_members,
                "active_admins": active_admins,
            },
        }

    organization_ids = [int(org.id) for org in organizations if org and org.id is not None]
    creator_ids = [
        int(org.created_by_user_id)
        for org in organizations
        if org and org.created_by_user_id is not None
    ]

    member_rows = (
        db.query(
            models.EnterpriseOrganizationMember.organization_id.label("organization_id"),
            func.count(models.EnterpriseOrganizationMember.id).label("total_members"),
            func.sum(
                case(
                    (models.EnterpriseOrganizationMember.is_active.is_(True), 1),
                    else_=0,
                )
            ).label("active_members"),
            func.sum(
                case(
                    (
                        (models.EnterpriseOrganizationMember.is_active.is_(True))
                        & (func.lower(models.EnterpriseOrganizationMember.role) == ENTERPRISE_ROLE_ADMIN),
                        1,
                    ),
                    else_=0,
                )
            ).label("active_admins"),
        )
        .filter(models.EnterpriseOrganizationMember.organization_id.in_(organization_ids))
        .group_by(models.EnterpriseOrganizationMember.organization_id)
        .all()
    )
    member_stats_by_org_id = {
        int(row.organization_id): {
            "total_members": int(row.total_members or 0),
            "active_members": int(row.active_members or 0),
            "active_admins": int(row.active_admins or 0),
        }
        for row in member_rows
        if row and row.organization_id is not None
    }

    creator_rows = (
        db.query(models.User.id, models.User.email, models.User.full_name)
        .filter(models.User.id.in_(creator_ids))
        .all()
        if creator_ids
        else []
    )
    creators_by_id = {
        int(row.id): {
            "email": row.email,
            "full_name": row.full_name,
        }
        for row in creator_rows
        if row and row.id is not None
    }

    accounts: list[dict] = []
    for org in organizations:
        org_id = int(org.id)
        member_stats = member_stats_by_org_id.get(org_id, {})
        creator = creators_by_id.get(int(org.created_by_user_id), {}) if org.created_by_user_id else {}
        accounts.append(
            {
                "organization_id": org_id,
                "company_name": (org.company_name or "").strip() or "Untitled Organization",
                "subdomain_slug": (org.subdomain_slug or "").strip().lower() or None,
                "portal_url": _build_enterprise_portal_url(org.subdomain_slug, request),
                "created_at": org.created_at,
                "created_by_user_id": int(org.created_by_user_id) if org.created_by_user_id is not None else None,
                "created_by_email": creator.get("email"),
                "created_by_name": creator.get("full_name"),
                "total_members": int(member_stats.get("total_members", 0)),
                "active_members": int(member_stats.get("active_members", 0)),
                "active_admins": int(member_stats.get("active_admins", 0)),
            }
        )

    return {
        "accounts": accounts,
        "total": total,
        "page": page,
        "page_size": page_size,
        "metrics": {
            "active_members": active_members,
            "active_admins": active_admins,
        },
    }


@router.post(
    "/enterprise/credentials",
    response_model=schemas.AdminEnterpriseCredentialCreateResponse,
)
def create_enterprise_credentials_admin(
    payload: schemas.AdminEnterpriseCredentialCreateRequest,
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """
    Create or update enterprise login access.
    Existing main-platform users keep their current password; new users receive a temporary password.
    """
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.enterprise.credentials.create",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )

    target_email = str(payload.email or "").strip().lower()
    full_name = str(payload.full_name or "").strip()
    if len(full_name) < 2:
        raise HTTPException(status_code=400, detail="Name must be at least 2 characters.")
    if len(full_name) > 120:
        raise HTTPException(status_code=400, detail="Name must be at most 120 characters.")

    existing_credential = (
        db.query(models.EnterpriseCredential)
        .filter(models.EnterpriseCredential.email == target_email)
        .first()
    )
    existing_user = db.query(models.User).filter(models.User.email == target_email).first()

    if existing_user and existing_user.is_developer:
        raise HTTPException(
            status_code=403,
            detail="Developer accounts cannot be converted into enterprise credentials.",
        )

    # If user already exists in main platform, preserve existing password and enable enterprise login reuse.
    if existing_user:
        credential_created = False
        if not str(existing_user.full_name or "").strip():
            existing_user.full_name = full_name
        if not existing_user.is_active:
            existing_user.is_active = True

        if existing_credential:
            existing_credential.password_hash = existing_user.hashed_password
            existing_credential.full_name = full_name
            existing_credential.is_active = True
        else:
            credential_created = True
            db.add(
                models.EnterpriseCredential(
                    email=target_email,
                    password_hash=existing_user.hashed_password,
                    full_name=full_name,
                    is_active=True,
                )
            )

        try:
            db.commit()
        except Exception:
            db.rollback()
            logger.exception("Failed to grant enterprise access for existing email=%s", target_email)
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to grant enterprise access. Please try again.",
            )

        return {
            "email": target_email,
            "full_name": full_name,
            "temporary_password": None,
            "uses_existing_main_password": True,
            "credential_created": credential_created,
            "user_created": False,
            "message": (
                "Enterprise access granted. User can login with their existing main platform password."
            ),
        }

    temp_password = _generate_enterprise_temp_password()
    hashed_password = get_password_hash(temp_password)

    credential_created = False
    if existing_credential:
        existing_credential.password_hash = hashed_password
        existing_credential.full_name = full_name
        existing_credential.is_active = True
    else:
        credential_created = True
        db.add(
            models.EnterpriseCredential(
                email=target_email,
                password_hash=hashed_password,
                full_name=full_name,
                is_active=True,
            )
        )

    user_created = True
    db.add(
        models.User(
            email=target_email,
            username=None,
            hashed_password=hashed_password,
            full_name=full_name,
            university="Enterprise Account",
            is_active=True,
            email_verified=True,
            is_admin=True,
            accepted_terms_privacy_at=datetime.utcnow(),
        )
    )

    try:
        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Failed to create enterprise credentials for email=%s", target_email)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create enterprise credentials. Please try again.",
        )

    return {
        "email": target_email,
        "full_name": full_name,
        "temporary_password": temp_password,
        "uses_existing_main_password": False,
        "credential_created": credential_created,
        "user_created": user_created,
        "message": (
            "Enterprise credentials created successfully."
            if credential_created
            else "Enterprise credentials updated and password rotated successfully."
        ),
    }


# ===========================================================================
# Per-account enterprise discount codes (admin-managed)
# ===========================================================================

def _get_enterprise_org_or_404(db: Session, organization_id: int) -> models.EnterpriseOrganization:
    org = (
        db.query(models.EnterpriseOrganization)
        .filter(models.EnterpriseOrganization.id == organization_id)
        .first()
    )
    if not org:
        raise HTTPException(status_code=404, detail="Enterprise account not found.")
    return org


def _admin_iso(value):
    if not value:
        return None
    try:
        return value.isoformat()
    except Exception:
        return None


_CREDIT_PAYMENT_KIND_LABELS = {"credits": "Credit top-up", "infra_fee": "Infrastructure fee"}

# Razorpay will not process a refund more than 6 months after the payment was captured.
# Checked here so the admin sees the deadline on the screen (and a clear message on the
# POST) instead of typing an amount, clicking a button that moves real money, and getting
# a raw gateway 502 back.
# https://razorpay.com/docs/payments/refunds/?preferred-country=IN
REFUND_WINDOW_MONTHS = 6


def _refund_deadline(captured_at: datetime | None) -> datetime | None:
    """Calendar 6-month deadline for refunding a payment captured at `captured_at`.

    Calendar months, not 180 days — the two differ by up to 4 days, which is exactly the
    window in which we'd tell an admin a refund is fine and Razorpay would reject it.
    """
    if not captured_at:
        return None
    month_index = captured_at.month - 1 + REFUND_WINDOW_MONTHS
    year = captured_at.year + month_index // 12
    month = month_index % 12 + 1
    day = min(captured_at.day, calendar.monthrange(year, month)[1])
    return captured_at.replace(year=year, month=month, day=day)


def _refund_window_state(payment: models.EnterpriseCreditPayment) -> tuple[datetime | None, bool]:
    """(deadline, is_open) for one payment. Unknown capture time is treated as open —
    Razorpay stays the authority, we just avoid promising a refund we can't deliver."""
    deadline = _refund_deadline(payment.verified_at or payment.created_at)
    if deadline is None:
        return None, True
    now = datetime.now(deadline.tzinfo) if deadline.tzinfo else datetime.utcnow()
    return deadline, now <= deadline


@router.get("/enterprise/accounts/{organization_id}/details")
def enterprise_account_details_admin(
    request: Request,
    organization_id: int,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Full account snapshot for one enterprise org — wallet balance, lifetime
    credits purchased/spent, recent Razorpay purchases (with any discount applied),
    recent credit usage, and the configured discount codes. Read-only."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.enterprise.account.details",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    org = _get_enterprise_org_or_404(db, organization_id)

    # --- organization + members ---
    members = (
        db.query(models.EnterpriseOrganizationMember)
        .filter(models.EnterpriseOrganizationMember.organization_id == org.id)
        .all()
    )
    members_active = sum(1 for m in members if m.is_active)
    admins = sum(1 for m in members if m.is_active and (m.role or "") == "admin")
    creator = (
        db.query(models.User).filter(models.User.id == org.created_by_user_id).first()
        if org.created_by_user_id else None
    )

    # --- wallet (balance, lifetime purchased/spent, infra-fee state) ---
    wallet = enterprise_credits.wallet_state(db, org.id)

    # --- credit/infra purchases (Razorpay charges) ---
    P = models.EnterpriseCreditPayment
    payment_rows = (
        db.query(P)
        .filter(P.organization_id == org.id)
        .order_by(desc(P.created_at), desc(P.id))
        .limit(25)
        .all()
    )
    purchases = []
    for p in payment_rows:
        refundable_paise = enterprise_refunds.payment_refundable_paise(p)
        # Every *_paise field on this row is minor units of THIS payment's currency, so it
        # is formatted with that currency — format_inr() here printed "₹12.99" against a
        # $12.99 charge. `refund_deadline` is Razorpay's 6-month window (see above).
        pay_currency = money.normalize_currency(p.currency, strict=False)
        refund_deadline, refund_window_open = _refund_window_state(p)
        purchases.append({
            "id": p.id,
            "kind": p.kind,
            "kind_label": _CREDIT_PAYMENT_KIND_LABELS.get(p.kind, p.kind),
            "package_key": p.package_key,
            "credits": int(p.credits or 0),
            "bonus_credits": int(p.bonus_credits or 0),
            "total_credits": int(p.credits or 0) + int(p.bonus_credits or 0),
            "amount_paise": int(p.amount_paise or 0),
            "amount_display": money.format_money(p.amount_paise, pay_currency),
            "currency": pay_currency,
            "currency_symbol": money.symbol_for(pay_currency),
            "currency_exponent": money.MINOR_UNIT_EXPONENT.get(pay_currency, 2),
            # Razorpay's INR settlement for a foreign charge — the only figure that may be
            # summed with other rows. NULL on pre-migration rows (all genuinely INR).
            "base_amount_paise": int(p.base_amount_paise) if p.base_amount_paise is not None else None,
            "settled_inr_display": (
                enterprise_credits.format_inr(p.base_amount_paise) if p.base_amount_paise is not None else None
            ),
            "is_international": bool(p.is_international),
            "status": p.status,
            "provider": p.provider,
            "razorpay_payment_id": p.razorpay_payment_id,
            "coupon_code": p.coupon_code,
            "coupon_percent_off": float(p.coupon_percent_off) if p.coupon_percent_off is not None else None,
            "original_amount_display": (
                money.format_money(p.original_amount_paise, pay_currency) if p.original_amount_paise else None
            ),
            "refunded_amount_paise": int(p.refunded_amount_paise or 0),
            "refunded_amount_display": money.format_money(p.refunded_amount_paise, pay_currency),
            "refundable_paise": refundable_paise,
            "refundable_display": money.format_money(refundable_paise, pay_currency),
            "is_refundable": bool(refundable_paise > 0 and p.razorpay_payment_id and refund_window_open),
            "refund_window_open": refund_window_open,
            "refund_deadline": _admin_iso(refund_deadline),
            "suggested_clawback_credits": enterprise_refunds.proportional_credits(p, refundable_paise),
            "created_at": _admin_iso(p.created_at),
            "verified_at": _admin_iso(p.verified_at),
        })

    # Gross collected per kind. A refund flips status to refunded/partially_refunded,
    # so include those and net out refunds separately (total refunded is computed below).
    #
    # Sums the INR SETTLEMENT figure, never the charged amount: adding a $39 top-up's
    # amount_paise (3900 cents) to a ₹2,999 row's 299900 produces a number that is neither
    # currency. The fallback to amount_paise is valid only where charged == settled, i.e.
    # INR rows (Razorpay omits base_amount on domestic payments) and pre-migration rows.
    # A foreign row whose settlement figure has not landed yet contributes 0 rather than
    # having its cents added as paise — and `unsettled_payment_count` below reports how
    # many, because a total that silently drops rows reads as "they never paid".
    _settled_inr = case(
        (P.base_amount_paise.isnot(None), P.base_amount_paise),
        (func.upper(func.coalesce(P.currency, "INR")) == "INR", P.amount_paise),
        else_=0,
    )

    def _sum_collected(kind: str) -> int:
        return int(
            db.query(func.coalesce(func.sum(_settled_inr), 0))
            .filter(
                P.organization_id == org.id,
                P.status.in_(enterprise_credits.REVENUE_PAYMENT_STATUSES),
                P.kind == kind,
            )
            .scalar() or 0
        )

    credit_paid = _sum_collected("credits")
    infra_paid = _sum_collected("infra_fee")
    # Plan subscriptions live in their OWN table, so a founder looking at a paying Growth
    # account used to see "Total paid: ₹0" — the tiers are now the largest line and were
    # invisible here. INR-only (the plan book is), and gross of GST because that is what
    # actually left the customer's account.
    SP = models.EnterpriseSubscriptionPayment
    plan_paid = int(
        db.query(func.coalesce(func.sum(SP.amount_paise - func.coalesce(SP.refunded_amount_paise, 0)), 0))
        .filter(
            SP.organization_id == org.id,
            SP.status.in_(enterprise_credits.REVENUE_PAYMENT_STATUSES),
            func.upper(func.coalesce(SP.currency, "INR")) == "INR",
        )
        .scalar() or 0
    )
    plan_payment_count = int(
        db.query(func.count(SP.id))
        .filter(
            SP.organization_id == org.id,
            SP.status.in_(enterprise_credits.REVENUE_PAYMENT_STATUSES),
            # Same INR filter as plan_paid above, or the count claims rows whose money was
            # excluded from the total and the two figures disagree.
            func.upper(func.coalesce(SP.currency, "INR")) == "INR",
        )
        .scalar() or 0
    )
    verified_count = int(
        db.query(func.count(P.id))
        .filter(
            P.organization_id == org.id,
            P.status.in_(enterprise_credits.REVENUE_PAYMENT_STATUSES),
        )
        .scalar() or 0
    )

    # --- recent credit usage (ledger) ---
    T = models.EnterpriseCreditTransaction
    txn_rows = (
        db.query(T)
        .filter(T.organization_id == org.id)
        .order_by(desc(T.created_at), desc(T.id))
        .limit(15)
        .all()
    )
    recent_activity = []
    for t in txn_rows:
        action = enterprise_credits.get_action(t.action_key) if t.action_key else None
        recent_activity.append({
            "id": t.id,
            "type": t.type,
            "action_key": t.action_key,
            "action_label": action["label"] if action else None,
            "credits": int(t.credits),
            "balance_after": int(t.balance_after),
            "description": t.description,
            "created_by_name": t.created_by_name,
            "created_at": _admin_iso(t.created_at),
        })

    # --- discount codes configured for this account ---
    coupon_rows = (
        db.query(models.EnterpriseCoupon)
        .filter(models.EnterpriseCoupon.organization_id == org.id)
        .order_by(desc(models.EnterpriseCoupon.created_at), desc(models.EnterpriseCoupon.id))
        .all()
    )
    coupons = [enterprise_coupons.serialize(db, c) for c in coupon_rows]

    # --- refunds issued to this account ---
    refunds = enterprise_refunds.list_refunds(db, org.id)
    refunded_paise = enterprise_refunds.total_refunded_paise(db, org.id)
    total_paid = credit_paid + infra_paid + plan_paid

    # `enterprise_refunds.total_refunded_paise` sums EnterpriseRefund.amount_paise, and a
    # refund is recorded in the ORIGINAL payment's currency — so the moment this org has a
    # non-INR refund that total is a mix of paise and cents. `credit_paid`/`infra_paid` are
    # settled INR, so subtracting a mixed total from them would print a confident wrong
    # number. Flag it and let the console say "mixed" instead.
    refund_currencies = {
        str(row[0] or "INR").strip().upper()
        for row in db.query(models.EnterpriseRefund.currency)
        .filter(
            models.EnterpriseRefund.organization_id == org.id,
            models.EnterpriseRefund.kind == "money",
            models.EnterpriseRefund.status != "failed",
        )
        .distinct()
        .all()
    }
    refunds_mixed_currency = bool(refund_currencies - {"INR"})
    payment_currencies = sorted({
        money.normalize_currency(row[0], strict=False)
        for row in db.query(P.currency).filter(P.organization_id == org.id).distinct().all()
    })
    # Revenue-status foreign payments that `_settled_inr` scored as 0 because Razorpay's
    # INR settlement figure has not landed on the row. They are missing from
    # credit_revenue/infra_revenue/total_paid, so the count has to travel with the totals
    # — otherwise the screen shows a confident rupee number that is quietly short.
    unsettled_payment_count = int(
        db.query(func.count(P.id))
        .filter(
            P.organization_id == org.id,
            P.status.in_(enterprise_credits.REVENUE_PAYMENT_STATUSES),
            P.base_amount_paise.is_(None),
            func.upper(func.coalesce(P.currency, "INR")) != "INR",
        )
        .scalar() or 0
    )

    return {
        "organization": {
            "id": int(org.id),
            "company_name": (org.company_name or "").strip() or "Untitled Organization",
            "subdomain_slug": org.subdomain_slug,
            "created_by_name": (creator.full_name or creator.email) if creator else None,
            "created_by_email": creator.email if creator else None,
            "created_at": _admin_iso(org.created_at),
            "members_total": len(members),
            "members_active": members_active,
            "admins": admins,
        },
        "wallet": wallet,
        # Which tier this org is on, its caps, and where it sits against them. The console
        # showed only the retired infra fee before, so a founder looking at an account had
        # no way to see what it actually pays every month.
        "subscription": enterprise_billing.build_subscription_state(db, org.id),
        # Every *_paise total here is INR SETTLEMENT (base_amount_paise), which is what
        # makes it summable at all. `currency` says so explicitly so the console never has
        # to infer it. The refund-derived figures go null when they can't be stated in one
        # currency — see refunds_mixed_currency above.
        "totals": {
            "currency": "INR",
            "credit_revenue_paise": credit_paid,
            "credit_revenue_display": enterprise_credits.format_inr(credit_paid),
            "plan_revenue_paise": plan_paid,
            "plan_revenue_display": enterprise_credits.format_inr(plan_paid),
            "plan_payment_count": plan_payment_count,
            "infra_revenue_paise": infra_paid,
            "infra_revenue_display": enterprise_credits.format_inr(infra_paid),
            "total_paid_paise": total_paid,
            "total_paid_display": enterprise_credits.format_inr(total_paid),
            "refunded_paise": None if refunds_mixed_currency else refunded_paise,
            "refunded_display": None if refunds_mixed_currency else enterprise_credits.format_inr(refunded_paise),
            "net_paid_paise": None if refunds_mixed_currency else total_paid - refunded_paise,
            "net_paid_display": (
                None if refunds_mixed_currency
                else enterprise_credits.format_inr(total_paid - refunded_paise)
            ),
            "refunds_mixed_currency": refunds_mixed_currency,
            "refund_currencies": sorted(refund_currencies),
            "payment_currencies": payment_currencies,
            # >0 means the INR totals above exclude that many foreign payments whose
            # settlement figure hasn't arrived — a known understatement, not zero revenue.
            "unsettled_payment_count": unsettled_payment_count,
            "verified_payment_count": verified_count,
        },
        "credit_value_inr": enterprise_credits.paise_to_rupees(enterprise_credits.PAISE_PER_CREDIT),
        "razorpay_enabled": enterprise_refunds.razorpay_enabled(),
        "purchases": purchases,
        "recent_activity": recent_activity,
        "coupons": coupons,
        "refunds": refunds,
    }


@router.post("/enterprise/accounts/{organization_id}/refunds")
def issue_enterprise_account_refund_admin(
    request: Request,
    organization_id: int,
    payload: schemas.AdminEnterpriseRefundRequest,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Issue a refund for one enterprise account — either a goodwill credit grant
    or a real Razorpay money refund (with optional credit claw-back). Recorded in
    the enterprise_refunds audit table."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.enterprise.refund",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    org = _get_enterprise_org_or_404(db, organization_id)
    kind = (payload.kind or "").strip().lower()
    reason = (payload.reason or "").strip() or None

    if kind == "credits":
        refund = enterprise_refunds.issue_credit_refund(
            db, org.id, int(payload.credits or 0), reason, current_user
        )
    elif kind == "money":
        if not payload.payment_id:
            raise HTTPException(status_code=400, detail="Select a payment to refund.")
        payment = (
            db.query(models.EnterpriseCreditPayment)
            .filter(
                models.EnterpriseCreditPayment.id == int(payload.payment_id),
                models.EnterpriseCreditPayment.organization_id == org.id,
            )
            .first()
        )
        if not payment:
            raise HTTPException(status_code=404, detail="Payment not found for this account.")

        # A Razorpay refund is issued in the ORIGINAL payment's currency and its `amount`
        # is minor units OF THAT CURRENCY. The client therefore sends the minor amount plus
        # the currency it believes it is refunding, and we assert the two agree BEFORE any
        # money moves — a console left open across a currency change must fail loudly, not
        # send $50.00 back against a ₹5,000 charge (which is what `amount_rupees * 100` did).
        payment_currency = money.normalize_currency(payment.currency, strict=False)
        requested_currency = str(payload.amount_currency or "").strip().upper()
        if not requested_currency:
            raise HTTPException(
                status_code=400,
                detail=f"Refund currency is required; this payment was charged in {payment_currency}.",
            )
        if requested_currency != payment_currency:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"This payment was charged in {payment_currency}, so it can only be "
                    f"refunded in {payment_currency} (you sent {requested_currency})."
                ),
            )
        amount_minor = int(payload.amount_minor or 0)
        if amount_minor <= 0:
            raise HTTPException(status_code=400, detail="Refund amount must be greater than zero.")
        # enterprise_refunds re-checks this, but its message formats the balance as INR, so
        # the over-refund error on a USD payment reads "₹12.99". Same guard, right currency.
        refundable_minor = enterprise_refunds.payment_refundable_paise(payment)
        if amount_minor > refundable_minor:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Refund amount exceeds the refundable balance "
                    f"({money.format_money(refundable_minor, payment_currency)})."
                ),
            )

        # Razorpay's own 6-month cap. Checked here so the admin gets the deadline rather
        # than a gateway rejection after clicking a button that moves real money.
        deadline, window_open = _refund_window_state(payment)
        if not window_open:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Razorpay only allows refunds within 6 months of the payment; this one "
                    f"closed on {deadline.date().isoformat()}. Issue a goodwill credit refund instead."
                ),
            )

        refund = enterprise_refunds.issue_money_refund(
            db, org.id, payment, amount_minor, int(payload.clawback_credits or 0), reason, current_user
        )
    else:
        raise HTTPException(status_code=400, detail="Unknown refund type.")

    logger.info(
        "Admin %s issued a %s refund for org %s (refund_id=%s, amount_minor=%s %s, credits_delta=%s)",
        current_user.id, kind, org.id, refund.id, refund.amount_paise, refund.currency, refund.credits_delta,
    )
    return {
        "refund": enterprise_refunds.serialize_refund(refund),
        "wallet": enterprise_credits.wallet_state(db, org.id),
    }


@router.get("/enterprise/accounts/{organization_id}/coupons")
def list_enterprise_account_coupons_admin(
    request: Request,
    organization_id: int,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """List the discount codes configured for one enterprise account."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.enterprise.coupons.list",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    org = _get_enterprise_org_or_404(db, organization_id)
    rows = (
        db.query(models.EnterpriseCoupon)
        .filter(models.EnterpriseCoupon.organization_id == org.id)
        .order_by(desc(models.EnterpriseCoupon.created_at), desc(models.EnterpriseCoupon.id))
        .all()
    )
    return {
        "organization_id": int(org.id),
        "company_name": (org.company_name or "").strip() or "Untitled Organization",
        "coupons": [enterprise_coupons.serialize(db, row) for row in rows],
    }


@router.post("/enterprise/accounts/{organization_id}/coupons")
def create_enterprise_account_coupon_admin(
    request: Request,
    organization_id: int,
    payload: schemas.AdminEnterpriseCouponCreateRequest,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Create a discount code for one enterprise account."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.enterprise.coupons.create",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    org = _get_enterprise_org_or_404(db, organization_id)

    code = enterprise_coupons.normalize_code(payload.code)
    if not code:
        raise HTTPException(status_code=400, detail="Enter a discount code (letters, numbers, - or _).")
    percent = enterprise_coupons.parse_percent_off(payload.percent_off)
    applies_to = enterprise_coupons.normalize_applies_to(payload.applies_to)
    max_redemptions = enterprise_coupons.parse_max_redemptions(payload.max_redemptions)

    existing = enterprise_coupons.find_coupon(db, org.id, code)
    if existing:
        raise HTTPException(status_code=409, detail=f"Code '{code}' already exists for this account.")

    note = (payload.note or "").strip() or None
    coupon = models.EnterpriseCoupon(
        organization_id=org.id,
        code=code,
        percent_off=percent,
        is_active=bool(payload.is_active),
        applies_to=applies_to,
        max_redemptions=max_redemptions,
        note=note,
        created_by_user_id=current_user.id,
    )
    db.add(coupon)
    db.commit()
    db.refresh(coupon)
    return {"coupon": enterprise_coupons.serialize(db, coupon)}


@router.patch("/enterprise/accounts/{organization_id}/coupons/{coupon_id}")
def update_enterprise_account_coupon_admin(
    request: Request,
    organization_id: int,
    coupon_id: int,
    payload: schemas.AdminEnterpriseCouponUpdateRequest,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Update a discount code (percent, scope, cap, note, or active toggle)."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.enterprise.coupons.update",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    org = _get_enterprise_org_or_404(db, organization_id)
    coupon = (
        db.query(models.EnterpriseCoupon)
        .filter(
            models.EnterpriseCoupon.id == coupon_id,
            models.EnterpriseCoupon.organization_id == org.id,
        )
        .first()
    )
    if not coupon:
        raise HTTPException(status_code=404, detail="Discount code not found.")

    provided = payload.model_dump(exclude_unset=True)
    if "percent_off" in provided and provided["percent_off"] is not None:
        coupon.percent_off = enterprise_coupons.parse_percent_off(provided["percent_off"])
    if "applies_to" in provided and provided["applies_to"] is not None:
        coupon.applies_to = enterprise_coupons.normalize_applies_to(provided["applies_to"])
    if "max_redemptions" in provided:
        coupon.max_redemptions = enterprise_coupons.parse_max_redemptions(provided["max_redemptions"])
    if "note" in provided:
        coupon.note = (provided["note"] or "").strip() or None
    if "is_active" in provided and provided["is_active"] is not None:
        coupon.is_active = bool(provided["is_active"])

    db.commit()
    db.refresh(coupon)
    return {"coupon": enterprise_coupons.serialize(db, coupon)}


@router.delete("/enterprise/accounts/{organization_id}/coupons/{coupon_id}")
def delete_enterprise_account_coupon_admin(
    request: Request,
    organization_id: int,
    coupon_id: int,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Delete a discount code from one enterprise account."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.enterprise.coupons.delete",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    org = _get_enterprise_org_or_404(db, organization_id)
    coupon = (
        db.query(models.EnterpriseCoupon)
        .filter(
            models.EnterpriseCoupon.id == coupon_id,
            models.EnterpriseCoupon.organization_id == org.id,
        )
        .first()
    )
    if not coupon:
        raise HTTPException(status_code=404, detail="Discount code not found.")
    db.delete(coupon)
    db.commit()
    return {"deleted": True, "id": coupon_id}


_COUPON_APPLIES_EMAIL_LABELS = {
    "all": "credit top-ups & plan billing",
    "credits": "credit top-ups",
    "billing": "plan billing",
}


@router.post("/enterprise/accounts/{organization_id}/coupons/{coupon_id}/send-email")
def send_enterprise_account_coupon_email_admin(
    request: Request,
    organization_id: int,
    coupon_id: int,
    payload: schemas.AdminEnterpriseCouponSendEmailRequest,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """Email a discount-code promotion to the account's team (or one address)."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.enterprise.coupons.send_email",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    org = _get_enterprise_org_or_404(db, organization_id)
    coupon = (
        db.query(models.EnterpriseCoupon)
        .filter(
            models.EnterpriseCoupon.id == coupon_id,
            models.EnterpriseCoupon.organization_id == org.id,
        )
        .first()
    )
    if not coupon:
        raise HTTPException(status_code=404, detail="Discount code not found.")

    # Resolve recipients: an explicit override, else every active member, else
    # the account creator as a last resort.
    recipients: list[tuple[str, str | None]] = []
    override = (payload.email or "").strip().lower() if payload and payload.email else ""
    if override:
        existing = db.query(models.User).filter(func.lower(models.User.email) == override).first()
        recipients.append((override, existing.full_name if existing else None))
    else:
        seen: set[str] = set()
        member_rows = (
            db.query(models.User.email, models.User.full_name)
            .join(
                models.EnterpriseOrganizationMember,
                models.EnterpriseOrganizationMember.user_id == models.User.id,
            )
            .filter(
                models.EnterpriseOrganizationMember.organization_id == org.id,
                models.EnterpriseOrganizationMember.is_active.is_(True),
            )
            .all()
        )
        for email, name in member_rows:
            normalized = (email or "").strip().lower()
            if normalized and normalized not in seen:
                seen.add(normalized)
                recipients.append((normalized, name))
        if not recipients and org.created_by_user_id:
            creator = db.query(models.User).filter(models.User.id == org.created_by_user_id).first()
            if creator and creator.email:
                recipients.append((creator.email.strip().lower(), creator.full_name))

    if not recipients:
        raise HTTPException(status_code=400, detail="No recipient email found for this account.")

    percent_display = enterprise_coupons.format_percent_off(coupon.percent_off) + "%"
    applies_label = _COUPON_APPLIES_EMAIL_LABELS.get(str(coupon.applies_to or "all"), "your purchases")
    portal_url = _build_enterprise_portal_url(org.subdomain_slug, request)
    code = enterprise_coupons.normalize_code(coupon.code)

    sent = 0
    failed: list[str] = []
    for email, name in recipients:
        ok = email_service.send_enterprise_discount_promo_email(
            recipient_email=email,
            recipient_name=name,
            organization_name=org.company_name or "your organization",
            code=code,
            percent_display=percent_display,
            applies_to_label=applies_label,
            note=coupon.note,
            portal_url=portal_url,
        )
        if ok:
            sent += 1
        else:
            failed.append(email)

    if sent == 0:
        raise HTTPException(
            status_code=502,
            detail="Could not send the promotion email. Please check the email service configuration.",
        )

    return {
        "sent": sent,
        "failed": failed,
        "recipients": [email for email, _ in recipients],
        "message": f"Promotion email sent to {sent} recipient{'s' if sent != 1 else ''}.",
    }


@router.patch("/users/{user_id}/status", response_model=schemas.AdminUserSummary)
def update_user_status_admin(
    request: Request,
    user_id: int,
    payload: schemas.AdminUserStatusUpdateRequest,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """
    Activate or deactivate a user account (admin/developer only).
    """
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.users.update_status",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )

    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if not payload.is_active:
        _assert_manageable_target(user, current_user)

    user.is_active = bool(payload.is_active)
    db.commit()
    db.refresh(user)
    return user


@router.delete("/users/{user_id}")
def delete_user_admin(
    request: Request,
    user_id: int,
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
    db: Session = Depends(get_db),
):
    """
    Permanently delete a user account (admin/developer only).
    """
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.users.delete",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )

    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    _assert_manageable_target(user, current_user)

    try:
        # Clean up referral links from other users to avoid FK constraint failures.
        db.query(models.User).filter(models.User.referred_by_user_id == user.id).update(
            {models.User.referred_by_user_id: None},
            synchronize_session=False,
        )

        # This table currently has no ORM relationship on User, so clear explicitly.
        db.query(models.RilonoAiChatUploadEvent).filter(
            models.RilonoAiChatUploadEvent.user_id == user.id
        ).delete(synchronize_session=False)

        # Some upgraded DBs still include legacy marketplace tables (items/messages/item_images)
        # where user-linked rows must be removed first to satisfy FK constraints.
        _cleanup_legacy_marketplace_rows(db=db, user_id=user.id)

        db.delete(user)
        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Failed to delete user_id=%s from admin console", user_id)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete user account. Please try again.",
        )

    return {"message": "User account deleted successfully."}


# ===========================================================================
# Course Finder catalog agent — admin controls
# ===========================================================================

@router.post("/course-catalog/refresh-now")
def run_course_catalog_refresh_admin(
    request: Request,
    limit: int | None = Query(None, ge=1, le=100),
    countries: str | None = Query(None, max_length=40),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user),
    _: None = Depends(require_admin_turnstile_proof),
):
    """Force a catalog agent run NOW, in the background (a batch of grounded Gemini
    calls takes minutes — too long for a request). `limit` bursts past the daily
    batch (e.g. limit=50 to seed a country fast); `countries` = CSV like "US,UK".
    Progress/outcome lands in course_catalog_refresh_runs (see /course-catalog/status)."""
    import threading as _threading

    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.course_catalog.refresh",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    from datetime import datetime as _dt, timezone as _tz

    from app.services.course_catalog_refresh import run_course_catalog_refresh_job, run_in_progress

    # Immediate feedback beats a silent background skip: the job itself also refuses
    # to stack onto a live run, but a 409 here stops the double-click → double-spend
    # pattern at the source.
    existing_run = (
        db.query(models.CourseCatalogRefreshRun)
        .filter(models.CourseCatalogRefreshRun.run_date == _dt.now(_tz.utc).date())
        .first()
    )
    if run_in_progress(existing_run):
        raise HTTPException(
            status_code=409,
            detail="A catalog refresh is already running — check /api/admin/course-catalog/status.",
        )

    country_list = [c.strip().upper() for c in countries.split(",") if c.strip()] if countries else None

    worker = _threading.Thread(
        target=run_course_catalog_refresh_job,
        kwargs={"force": True, "limit": limit, "countries": country_list},
        name="course-catalog-refresh-manual",
        daemon=True,
    )
    worker.start()
    return {
        "status": "started",
        "limit": limit,
        "countries": country_list,
        "note": "Running in the background — check /api/admin/course-catalog/status for the run row.",
    }


@router.get("/course-catalog/status")
def course_catalog_status_admin(
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    """Catalog size/freshness per country + recent agent runs."""
    _enforce_rate_limit_or_429(
        request=request,
        scope="admin.course_catalog.status",
        limit=ADMIN_ENDPOINT_RATE_LIMIT,
        window_seconds=ADMIN_ENDPOINT_RATE_WINDOW_SECONDS,
        extra_key=f"user:{current_user.id}",
    )
    from app import course_catalog

    runs = (
        db.query(models.CourseCatalogRefreshRun)
        .order_by(models.CourseCatalogRefreshRun.run_date.desc())
        .limit(14)
        .all()
    )
    return {
        **course_catalog.catalog_stats(db),
        "runs": [
            {
                "run_date": r.run_date.isoformat() if r.run_date else None,
                "status": r.status,
                "started_at": r.started_at.isoformat() if r.started_at else None,
                "completed_at": r.completed_at.isoformat() if r.completed_at else None,
                "universities_discovered": r.universities_discovered,
                "universities_refreshed": r.universities_refreshed,
                "courses_upserted": r.courses_upserted,
                "error_message": r.error_message,
            }
            for r in runs
        ],
    }
